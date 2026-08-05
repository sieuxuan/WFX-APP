from datetime import date

import pytest
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

from wfx_panel import prefs
from wfx_panel.automation import sale_asn_create
from wfx_panel.automation.sale_asn_create import (
    _auto_add_po,
    _buyer_options,
    _refresh_existing_new_form,
    _set_style_goods_description_cell,
    _set_style_hts_cell,
)
from wfx_panel.panel_api import PanelAPI
from wfx_panel.sale_asn_workbook import (
    SALE_ASN_COLUMNS,
    SaleASNWorkbookError,
    read_sale_asn_workbook,
    write_sale_asn_price_check_workbook,
    write_sale_asn_template,
)


def _input_workbook(path, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(SALE_ASN_COLUMNS)
    for row in rows:
        sheet.append([row.get(column) for column in SALE_ASN_COLUMNS])
    workbook.save(path)
    workbook.close()


def _valid_rows():
    return [
        {
            "Invoice No": "INV-001",
            "Invoice Date": date(2026, 8, 1),
            "Shipping Bill No": "",
            "Shipping Bill Date": date(2026, 8, 2),
            "Style No": "STYLE A MEN",
            "PO No": "PO-001",
            "HS CODE": 62014010,
            "Goods Description": "Men's jacket",
            "Qty": 10,
            "Price": 95,
            "Carton": 2,
            "NW": 9.5,
            "GW": 10.5,
            "CBM": 1.25,
            "Destination": "Germany",
            "FTY": "PRO SPORTS GIAO THUY JSC",
            "Cargo Ready Date": date(2026, 8, 3),
            "Consignee Address": "PUMA EUROPE GMBH",
            "Ship To": "PUMA CENTRAL WAREHOUSE",
            "Shipping Mode": " air ",
        },
        {
            "Style No": "STYLE B WOMEN",
            "PO No": "PO-002",
            "Destination": "",
            "FTY": "",
        },
    ]


def test_template_keeps_reference_schema_and_readable_format(tmp_path):
    target = write_sale_asn_template(tmp_path / "sale-asn.xlsx")

    workbook = load_workbook(target)
    sheet = workbook["SALE ASN"]
    assert tuple(cell.value for cell in sheet[1]) == SALE_ASN_COLUMNS
    assert sheet.freeze_panes == "A2"
    assert sheet.sheet_view.showGridLines is False
    assert sheet.tables["SaleASNInput"].ref == "A1:V21"
    assert sheet.auto_filter.ref is None
    assert sheet["A1"].fill.fgColor.rgb == "00FDE68A"
    assert sheet["G1"].fill.fgColor.rgb == "00DBEAFE"
    assert sheet["V1"].fill.fgColor.rgb == "00FDE68A"
    assert SALE_ASN_COLUMNS == (
        "Style No",
        "PO No",
        "Qty",
        "Price",
        "Carton",
        "NW",
        "GW",
        "CBM",
        "FOB Price",
        "Service Price",
        "Cargo Ready Date",
        "HS CODE",
        "Goods Description",
        "Invoice No",
        "Invoice Date",
        "Shipping Bill No",
        "Shipping Bill Date",
        "Destination",
        "FTY",
        "Consignee Address",
        "Ship To",
        "Shipping Mode",
    )
    assert sheet["K2"].number_format == "dd/mm/yyyy"
    assert sheet["O2"].number_format == "dd/mm/yyyy"
    assert sheet["Q2"].number_format == "dd/mm/yyyy"
    assert "SEASON" not in SALE_ASN_COLUMNS
    assert "DESCRIPTION" not in SALE_ASN_COLUMNS
    assert SALE_ASN_COLUMNS[-3:] == (
        "Consignee Address",
        "Ship To",
        "Shipping Mode",
    )
    validations = list(sheet.data_validations.dataValidation)
    assert len(validations) == 4
    shipping_validation = next(item for item in validations if item.type == "list")
    assert shipping_validation.formula1 == '"AIR,SEA,COURIER"'
    assert str(shipping_validation.sqref) == "V2"
    assert shipping_validation.allow_blank is False
    date_validations = [item for item in validations if item.type == "date"]
    assert {str(item.sqref) for item in date_validations} == {
        "K2:K2001",
        "O2:O2001",
        "Q2:Q2001",
    }
    assert all(item.allow_blank is True for item in date_validations)
    workbook.close()


def test_reader_preserves_row_order_and_applies_business_fallbacks(tmp_path):
    source = tmp_path / "input.xlsx"
    _input_workbook(source, _valid_rows())

    document = read_sale_asn_workbook(source)

    assert document["invoice_no"] == "INV-001"
    assert document["po_count"] == 2
    assert document["style_count"] == 2
    assert [row["po_no"] for row in document["rows"]] == ["PO-001", "PO-002"]
    assert document["rows"][0]["shipping_bill_no"] == "INV-001"
    assert document["rows"][1]["invoice_date"] == "2026-08-01"
    assert document["rows"][1]["shipping_bill_date"] == "2026-08-02"
    assert document["rows"][1]["cargo_ready_date"] == "2026-08-03"
    assert document["rows"][1]["destination"] == "Germany"
    assert document["rows"][0]["hs_code"] == "62014010"
    assert document["rows"][0]["goods_description"] == "Men's jacket"
    assert document["rows"][0]["price"] == "95"
    assert document["rows"][0]["shipping_mode"] == "AIR"
    assert document["rows"][1]["shipping_mode"] == "AIR"
    assert document["rows"][1]["consignee_address"] == "PUMA EUROPE GMBH"
    assert document["rows"][1]["ship_to"] == "PUMA CENTRAL WAREHOUSE"


def test_reader_allows_blank_destination_and_preserves_it_for_wfx_defaults(tmp_path):
    rows = _valid_rows()
    for row in rows:
        row["Destination"] = ""
    source = tmp_path / "blank-destination.xlsx"
    _input_workbook(source, rows)

    document = read_sale_asn_workbook(source)

    assert [row["destination"] for row in document["rows"]] == ["", ""]


def test_reader_reports_missing_headers_by_sheet(tmp_path):
    source = tmp_path / "missing-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nhập liệu"
    sheet.append(["PO No", "Style No"])
    sheet.append(["PO-1", "STYLE-1"])
    workbook.save(source)
    workbook.close()

    with pytest.raises(SaleASNWorkbookError) as raised:
        read_sale_asn_workbook(source)

    assert raised.value.code == "SALE_ASN_FILE_HEADERS_INVALID"
    assert any(
        'Sheet "Nhập liệu": thiếu cột' in error and "Invoice No" in error
        for error in raised.value.errors
    )


def test_reader_uses_shipping_mode_only_from_first_data_row(tmp_path):
    rows = _valid_rows()
    rows[1]["Shipping Mode"] = "TRUCK"
    source = tmp_path / "first-row-shipping-mode.xlsx"
    _input_workbook(source, rows)

    document = read_sale_asn_workbook(source)

    assert [row["shipping_mode"] for row in document["rows"]] == ["AIR", "AIR"]


@pytest.mark.parametrize("shipping_mode", ["", "TRUCK"])
def test_reader_requires_supported_shipping_mode_for_shipping_info(
    tmp_path,
    shipping_mode,
):
    rows = _valid_rows()
    rows[0]["Shipping Mode"] = shipping_mode
    source = tmp_path / "invalid-shipping-mode.xlsx"
    _input_workbook(source, rows)

    with pytest.raises(SaleASNWorkbookError) as raised:
        read_sale_asn_workbook(source)

    assert raised.value.code == "SALE_ASN_FILE_VALIDATION_FAILED"
    assert any("Shipping Mode" in error for error in raised.value.errors)


def test_reader_ignores_every_row_without_po_number(tmp_path):
    rows = _valid_rows()
    rows.append(
        {
            "Style No": "TOTAL / GHI CHÚ",
            "PO No": "",
            "Qty": "không phải số",
            "Destination": "",
            "FTY": "",
        }
    )
    source = tmp_path / "ignore-no-po.xlsx"
    _input_workbook(source, rows)

    document = read_sale_asn_workbook(source)

    assert document["po_count"] == 2
    assert [row["source_row"] for row in document["rows"]] == [2, 3]


def test_reader_allows_one_po_to_have_multiple_styles(tmp_path):
    rows = _valid_rows()
    rows[1]["PO No"] = "PO-001"
    source = tmp_path / "same-po-multiple-styles.xlsx"
    _input_workbook(source, rows)

    document = read_sale_asn_workbook(source)

    assert [row["po_no"] for row in document["rows"]] == ["PO-001", "PO-001"]
    assert [row["style_no"] for row in document["rows"]] == [
        rows[0]["Style No"],
        rows[1]["Style No"],
    ]


def test_reader_rejects_only_duplicate_po_and_style_pair(tmp_path):
    rows = _valid_rows()
    rows[1]["PO No"] = rows[0]["PO No"]
    rows[1]["Style No"] = rows[0]["Style No"]
    source = tmp_path / "invalid.xlsx"
    _input_workbook(source, rows)

    with pytest.raises(SaleASNWorkbookError) as raised:
        read_sale_asn_workbook(source)

    assert raised.value.code == "SALE_ASN_FILE_VALIDATION_FAILED"
    assert any("PO No + Style No bị trùng" in error for error in raised.value.errors)


def test_reader_relaxes_fields_when_continuing_only_order_details(tmp_path):
    source = tmp_path / "continue-order.xlsx"
    _input_workbook(
        source,
        [{"PO No": "PO-001", "Carton": 2, "NW": 9.5}],
    )

    document = read_sale_asn_workbook(
        source,
        required_stages=["order_details"],
    )

    assert document["po_count"] == 1
    assert document["style_count"] == 0
    assert document["rows"][0]["po_no"] == "PO-001"
    assert document["rows"][0]["carton"] == "2"
    assert document["rows"][0]["cargo_ready_date"] == ""

    with pytest.raises(SaleASNWorkbookError) as raised:
        read_sale_asn_workbook(source, required_stages=["style_details"])
    assert any("Style No bắt buộc" in error for error in raised.value.errors)


def test_full_template_can_prefill_current_order_details(tmp_path):
    target = write_sale_asn_template(
        tmp_path / "continue.xlsx",
        [
            {
                "po_no": "PO-001",
                "carton": "2",
                "nw": "9.5",
                "gw": "10.5",
                "cbm": "1.25",
                "fob_price": "12.75",
                "service_price": "0.5",
                "cargo_ready_date": "03 Aug 2026",
            }
        ],
    )

    workbook = load_workbook(target)
    sheet = workbook["SALE ASN"]
    assert sheet["B2"].value == "PO-001"
    assert sheet["E2"].value == 2
    assert sheet["F2"].value == 9.5
    assert sheet["I2"].value == 12.75
    assert sheet["K2"].value.date() == date(2026, 8, 3)
    assert sheet.tables["SaleASNInput"].ref == "A1:V21"
    workbook.close()


def test_buyer_scan_waits_for_lazy_bound_select_options(monkeypatch):
    class FakeCell:
        def __init__(self):
            self.calls = 0

        def evaluate(self, _script):
            self.calls += 1
            if self.calls == 1:
                return []
            return [{"label": "  PUMA  ", "value": "77540"}]

    cell = FakeCell()
    frame = object()
    monkeypatch.setattr(sale_asn_create, "_buyer_cell", lambda _frame: cell)
    monkeypatch.setattr(sale_asn_create, "_wait", lambda _frame, _milliseconds: None)

    assert _buyer_options(frame, timeout_s=1) == [
        {"label": "PUMA", "value": "77540"}
    ]
    assert cell.calls == 2


def test_start_refreshes_existing_new_form_before_selecting_buyer(monkeypatch):
    calls = []

    class FakeFrame:
        url = "https://prosports.worldfashionexchange.com/WFXBase4.0/WFXSalesASN.aspx"

        def goto(self, url, **kwargs):
            calls.append(("goto", url, kwargs))

    frame = FakeFrame()
    page = type("FakePage", (), {"context": object()})()
    logs = []
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        lambda _context, _selector, timeout_s: (page, frame),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_ensure_select_value",
        lambda _page, selector, value, label, _log: calls.append(
            ("select", selector, value, label)
        ),
    )

    assert _refresh_existing_new_form(page, logs.append) is frame
    assert calls == [
        (
            "goto",
            frame.url,
            {"wait_until": "domcontentloaded", "timeout": 15_000},
        ),
        ("select", "#ddlASNType", "1", "ASN Type"),
        (
            "select",
            "#ddlASNAgainst",
            "BuyerOrderDispatch",
            "ASN Against",
        ),
    ]
    assert logs == [
        "[SALE ASN] Đã refresh form New đang mở trước khi chọn Buyer."
    ]


@pytest.mark.parametrize(
    ("final", "expected_action_selector"),
    (
        (False, sale_asn_create.PO_CONTINUE_SELECTOR),
        (True, sale_asn_create.PO_OK_SELECTOR),
    ),
)
def test_auto_add_po_selects_checkbox_by_exact_identity(
    monkeypatch,
    final,
    expected_action_selector,
):
    captured = {}

    class FakeLocator:
        @property
        def first(self):
            return self

        def evaluate(self, _script, spec=None):
            if spec is None:
                captured["continued_dom"] = True
                captured["action_script"] = _script
                return {"ok": True, "tag": "A", "id": ""}
            captured.update(spec)
            return {"ok": True, "value": spec["selection_value"]}

        def wait_for(self, **_kwargs):
            return None

        def click(self, **_kwargs):
            captured["continued"] = True

    class FakeFrame:
        def locator(self, selector):
            captured.setdefault("selectors", []).append(selector)
            return FakeLocator()

    candidate = {
        "row_index": 3,
        "selection_name": "optShipmentId",
        "selection_value": "7740025278_32043",
        "selection_order_id": "220106328",
        "po_no": "PO-1",
    }
    monkeypatch.setattr(
        sale_asn_create,
        "_search_po",
        lambda *_args, **_kwargs: [candidate],
    )
    waits = []
    monkeypatch.setattr(
        sale_asn_create,
        "_wait",
        lambda *_args: waits.append(_args),
    )

    added, _candidates, _reason = _auto_add_po(
        FakeFrame(),
        {"source_row": 2, "po_no": "PO-1"},
        lambda _message: None,
        final=final,
    )

    assert added is True
    assert captured["row_index"] == 3
    assert captured["selection_name"] == "optShipmentId"
    assert captured["selection_value"] == "7740025278_32043"
    assert captured["selection_order_id"] == "220106328"
    assert captured["continued_dom"] is True
    assert "querySelector('a, button, input, [onclick]')" in captured["action_script"]
    assert captured["selectors"] == [
        sale_asn_create.PO_RESULTS_TABLE_SELECTOR,
        expected_action_selector,
    ]
    assert len(waits) == (0 if final else 1)


def test_auto_add_po_narrows_by_enabled_fields_in_fixed_order(monkeypatch):
    searches = []
    selection_specs = []
    candidates = [
        {"row_index": 1, "selection_value": "A", "po_no": "PO-1"},
        {"row_index": 2, "selection_value": "B", "po_no": "PO-1"},
    ]

    def fake_search(_frame, _row, *, fields):
        searches.append(tuple(fields))
        return candidates[:1] if fields == ("po", "style", "destination") else candidates

    class FakeLocator:
        first = property(lambda self: self)

        def evaluate(self, _script, spec=None):
            if spec is None:
                return {"ok": True, "tag": "A"}
            selection_specs.append(spec)
            return {"ok": True, "value": spec["selection_value"]}

        def wait_for(self, **_kwargs):
            return None

    class FakeFrame:
        def locator(self, _selector):
            return FakeLocator()

    monkeypatch.setattr(sale_asn_create, "_search_po", fake_search)
    monkeypatch.setattr(sale_asn_create, "_wait", lambda *_args: None)

    added, _candidates, _reason = _auto_add_po(
        FakeFrame(),
        {
            "source_row": 2,
            "po_no": "PO-1",
            "style_no": "S1",
            "destination": "DE",
        },
        lambda _message: None,
    )

    assert added is True
    assert searches == [
        ("po",),
        ("po", "style"),
        ("po", "style", "destination"),
    ]
    assert [spec["selection_value"] for spec in selection_specs] == ["A"]


def test_auto_add_po_respects_disabled_fields_and_selects_all_at_end(monkeypatch):
    searches = []
    selected = []
    actions = []
    candidates = [
        {"row_index": 1, "selection_value": "A", "po_no": "PO-1", "dispatched_qty": "4"},
        {"row_index": 2, "selection_value": "B", "po_no": "PO-1", "dispatched_qty": "3"},
        {"row_index": 3, "selection_value": "C", "po_no": "PO-1", "dispatched_qty": "3"},
    ]

    def fake_search(_frame, _row, *, fields):
        searches.append(tuple(fields))
        return candidates

    class FakeLocator:
        first = property(lambda self: self)

        def __init__(self, selector):
            self.selector = selector

        def evaluate(self, _script, spec=None):
            if spec is None:
                actions.append(self.selector)
                return {"ok": True, "tag": "A"}
            selected.append(spec["selection_value"])
            return {"ok": True, "value": spec["selection_value"]}

        def wait_for(self, **_kwargs):
            return None

    class FakeFrame:
        def locator(self, selector):
            return FakeLocator(selector)

    monkeypatch.setattr(sale_asn_create, "_search_po", fake_search)
    monkeypatch.setattr(sale_asn_create, "_wait", lambda *_args: None)

    added, returned, reason = _auto_add_po(
        FakeFrame(),
        {"source_row": 2, "style_no": "S1", "destination": "DE", "qty": "10"},
        lambda _message: None,
        search_fields=("style", "destination"),
    )

    assert added is True
    assert returned == candidates
    assert reason == "Style + Destination"
    assert searches == [("style",), ("style", "destination")]
    assert selected == ["A", "B", "C"]
    assert actions == [sale_asn_create.PO_CONTINUE_SELECTOR]


def test_auto_add_po_asks_user_when_ambiguous_rows_have_wrong_total_qty(monkeypatch):
    candidates = [
        {"row_index": 1, "selection_value": "A", "dispatched_qty": "4"},
        {"row_index": 2, "selection_value": "B", "dispatched_qty": "5"},
    ]
    logs = []

    monkeypatch.setattr(
        sale_asn_create,
        "_search_po",
        lambda _frame, _row, *, fields: candidates,
    )

    added, returned, reason = _auto_add_po(
        object(),
        {"source_row": 2, "po_no": "PO-1", "style_no": "S1", "qty": "10"},
        logs.append,
    )

    assert added is False
    assert returned == candidates
    assert reason.startswith("qty_mismatch:")
    assert any("chờ user xác nhận" in message for message in logs)


def test_auto_add_po_recovers_once_when_popup_document_changes(monkeypatch):
    context = object()
    stale_frame = object()
    fresh_frame = object()
    calls = []
    logs = []

    def fake_auto(frame, _row, _log, *, final, search_fields):
        calls.append((frame, final, tuple(search_fields)))
        if frame is stale_frame:
            raise sale_asn_create._POFrameChanged
        return True, [{"selection_value": "A"}], "PO"

    def fake_frame(selected_context, selector, timeout_s):
        assert selected_context is context
        assert selector == f"{sale_asn_create.PO_POPUP_SELECTOR} #txtOCNo"
        assert 0 < timeout_s <= sale_asn_create.PO_POPUP_RECOVERY_TIMEOUT_SECONDS
        return object(), fresh_frame

    monkeypatch.setattr(sale_asn_create, "_auto_add_po", fake_auto)
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        fake_frame,
    )

    added, candidates, reason, returned_frame = (
        sale_asn_create._auto_add_po_with_frame_retry(
            context,
            stale_frame,
            {"source_row": 2, "po_no": "PO-1"},
            logs.append,
            final=False,
            search_fields=("po", "style"),
        )
    )

    assert added is True
    assert candidates == [{"selection_value": "A"}]
    assert reason == "PO"
    assert returned_frame is fresh_frame
    assert [call[0] for call in calls] == [stale_frame, fresh_frame]
    assert any("nhận kết quả từ frame mới" in message for message in logs)


def test_auto_add_po_uses_results_already_rendered_after_document_change(monkeypatch):
    context = object()
    stale_frame = object()
    result_frame = object()
    candidates = [{"selection_value": "A", "po_no": "PO-1"}]
    calls = []

    def fake_auto(
        frame,
        _row,
        _log,
        *,
        final,
        search_fields,
        recovered_search=None,
    ):
        calls.append((frame, recovered_search))
        if frame is stale_frame:
            raise sale_asn_create._POFrameChanged(
                fields=("po",),
                search_submitted=True,
            )
        assert recovered_search == (("po",), candidates)
        return True, candidates, "PO"

    monkeypatch.setattr(sale_asn_create, "_auto_add_po", fake_auto)
    monkeypatch.setattr(
        sale_asn_create,
        "_recover_submitted_po_results",
        lambda selected_context, *, timeout_s: (
            result_frame,
            candidates,
        ),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        lambda *_args, **_kwargs: pytest.fail("không được Search lại"),
    )

    added, returned, reason, returned_frame = (
        sale_asn_create._auto_add_po_with_frame_retry(
            context,
            stale_frame,
            {"source_row": 2, "po_no": "PO-1"},
            lambda _message: None,
            final=False,
            search_fields=("po", "style"),
        )
    )

    assert added is True
    assert returned == candidates
    assert reason == "PO"
    assert returned_frame is result_frame
    assert calls == [(stale_frame, None), (result_frame, (("po",), candidates))]


def test_auto_add_po_skips_blank_optional_destination(monkeypatch):
    searches = []
    logs = []

    def fake_search(_frame, _row, *, fields):
        searches.append(tuple(fields))
        return [{"row_index": 1, "selection_value": "A", "po_no": "PO-1"}]

    class FakeLocator:
        first = property(lambda self: self)

        def evaluate(self, _script, spec=None):
            if spec is None:
                return {"ok": True, "tag": "A"}
            return {"ok": True, "value": spec["selection_value"]}

        def wait_for(self, **_kwargs):
            return None

    class FakeFrame:
        def locator(self, _selector):
            return FakeLocator()

    monkeypatch.setattr(sale_asn_create, "_search_po", fake_search)
    monkeypatch.setattr(sale_asn_create, "_wait", lambda *_args: None)

    added, _returned, reason = _auto_add_po(
        FakeFrame(),
        {"source_row": 2, "po_no": "PO-1", "style_no": "S1", "destination": ""},
        logs.append,
    )

    assert added is True
    assert reason == "PO"
    assert searches == [("po",)]
    assert any("Destination trống" in message for message in logs)


def test_sale_asn_po_results_use_exact_wfx_table():
    assert sale_asn_create.PO_RESULTS_TABLE_XPATH == (
        '//*[@id="wfx_GMPOAsnSearch"]/div[3]/table'
    )
    assert sale_asn_create.PO_RESULTS_TABLE_SELECTOR == (
        'xpath=//*[@id="wfx_GMPOAsnSearch"]/div[3]/table'
    )
    assert sale_asn_create.PO_OK_SELECTOR == (
        'xpath=//*[@id="wfx_GMPOAsnSearch"]'
        "/table[1]/tbody/tr/td[3]/table/tbody/tr/td[3]"
    )


def test_sale_asn_order_grid_uses_exact_wfx_columns():
    assert sale_asn_create.ORDER_GRID_SELECTOR == "#gridOrderDetails_tblGridContent"
    assert sale_asn_create.ORDER_FIELD_COLUMNS == {
        "carton": "colTotalNoOfCartons",
        "gw": "colTotalGrossWeight",
        "nw": "colTotalNetWeight",
        "cbm": "colTotalVolume",
        "fob_price": "colFFTextField1",
        "service_price": "colFFTextField2",
        "cargo_ready_date": "colFFDate1",
    }


def test_sale_asn_order_grid_disambiguates_same_po_by_style():
    markup = """
      <table id="gridOrderDetails_tblGridContent"><tbody>
        <tr class="trContent" id="row-a">
          <td id="colOrderRefNum">PO-001</td>
          <td id="colStyle">JLD-SMOW17905-M ACEL JACKET-MEN</td>
          <td id="colTotalNoOfCartons">1</td>
        </tr>
        <tr class="trContent" id="row-b">
          <td id="colOrderRefNum">PO-001</td>
          <td id="colStyle">JLD-STYLE-B-M CLAES PANT (SWV002/STYLE-B)</td>
          <td id="colTotalNoOfCartons">2</td>
        </tr>
      </tbody></table>
    """
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(channel="chrome", headless=True)
        try:
            page = browser.new_page()
            page.set_content(markup)
            result = page.evaluate(
                sale_asn_create._MARK_ORDER_GRID_CELL_JS,
                {
                    "table": sale_asn_create.ORDER_GRID_SELECTOR,
                    "po_no": "PO-001",
                    "style_no": "M Acel Jacket",
                    "column_id": "colTotalNoOfCartons",
                },
            )
            assert result["ok"] is True
            assert result["row_id"] == "row-a"
            assert page.locator(
                "#row-a #colTotalNoOfCartons"
            ).get_attribute("data-wfx-sale-asn-target") == "1"
        finally:
            browser.close()


def test_missing_order_rows_uses_po_and_style_identity():
    rows = [
        {"po_no": "PO-001", "style_no": "M Acel Jacket"},
        {"po_no": "PO-001", "style_no": "STYLE-B"},
    ]
    present = {
        (
            sale_asn_create._fold("PO-001"),
            sale_asn_create._fold("JLD-SMOW17905-M ACEL JACKET-MEN"),
        )
    }

    assert sale_asn_create._missing_order_rows(rows, present) == [rows[1]]


def test_sale_asn_style_details_targets_exact_hts_cell(monkeypatch):
    captured = {}

    class FakeFrame:
        def evaluate(self, script, spec):
            captured["script"] = script
            captured["spec"] = spec
            return {"ok": True, "style": "RVR-STYLE A", "column_id": "colHTSCode"}

    monkeypatch.setattr(
        sale_asn_create,
        "_edit_marked_table_cell",
        lambda frame, value: captured.update(frame=frame, value=value),
    )
    frame = FakeFrame()

    _set_style_hts_cell(frame, "STYLE A", "62014010")

    assert captured["spec"] == {"style": "STYLE A"}
    assert "#gridStyleDetails_tblGridContent" in captured["script"]
    assert "td#colStyle" in captured["script"]
    assert "td#colHTSCode" in captured["script"]
    assert captured["frame"] is frame
    assert captured["value"] == "62014010"


def test_sale_asn_style_details_targets_exact_goods_description_cell(monkeypatch):
    captured = {}

    class FakeFrame:
        def evaluate(self, script, spec):
            captured["script"] = script
            captured["spec"] = spec
            return {
                "ok": True,
                "style": "RVR-STYLE A",
                "column_id": "colGoodsDescription",
            }

    monkeypatch.setattr(
        sale_asn_create,
        "_edit_marked_table_cell",
        lambda frame, value: captured.update(frame=frame, value=value),
    )
    frame = FakeFrame()

    _set_style_goods_description_cell(frame, "STYLE A", "Men's jacket")

    assert captured["spec"] == {"style": "STYLE A"}
    assert "td#colGoodsDescription" in captured["script"]
    assert captured["frame"] is frame
    assert captured["value"] == "Men's jacket"


def test_sale_asn_order_grid_retries_only_rows_missing_after_final_ok(monkeypatch):
    rows = [
        {"source_row": 2, "po_no": "PO005500-DE-1"},
        {"source_row": 3, "po_no": "PO005501-DE-1"},
    ]
    main_frame = object()
    popup_frame = object()
    refreshed_frame = object()
    context = object()
    calls = {"wait": [], "retried": []}

    class FakeAdd:
        @property
        def first(self):
            return self

        def wait_for(self, **_kwargs):
            return None

        def evaluate(self, _script):
            calls["add_clicked"] = True
            return {"ok": True, "tag": "DIV", "id": ""}

    class FakeMainFrame:
        def locator(self, selector):
            assert selector == f"xpath={sale_asn_create.ADD_ORDER_XPATH}"
            return FakeAdd()

    main_frame = FakeMainFrame()

    def fake_wait(frame, waited_rows, timeout_s, *, allow_incomplete=False):
        calls["wait"].append((frame, timeout_s, allow_incomplete))
        assert list(waited_rows) == rows
        if allow_incomplete:
            return {sale_asn_create._fold("PO005500-DE-1")}
        return {sale_asn_create._fold(row["po_no"]) for row in rows}

    def fake_frame_with_selector(_context, selector, timeout_s):
        assert _context is context
        if selector == sale_asn_create.PO_POPUP_SELECTOR:
            if timeout_s < 1:
                raise sale_asn_create.PlaywrightTimeoutError("popup closed")
            return object(), popup_frame
        assert selector == "#sectionOrderDetails"
        return object(), refreshed_frame

    def fake_auto_add(frame, row, _log, *, final, search_fields):
        calls["retried"].append(
            (frame, row["po_no"], final, tuple(search_fields))
        )
        return True, [], "PO No"

    monkeypatch.setattr(sale_asn_create, "_wait_order_grid", fake_wait)
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        fake_frame_with_selector,
    )
    monkeypatch.setattr(sale_asn_create, "_auto_add_po", fake_auto_add)
    logs = []

    result = sale_asn_create._ensure_order_grid_rows(
        context,
        main_frame,
        rows,
        logs.append,
    )

    assert result is refreshed_frame
    assert calls["add_clicked"] is True
    assert calls["retried"] == [
        (
            popup_frame,
            "PO005501-DE-1",
            True,
            ("po", "style", "destination"),
        )
    ]
    assert calls["wait"] == [
        (
            main_frame,
            sale_asn_create.ORDER_GRID_SYNC_TIMEOUT_SECONDS,
            True,
        ),
        (
            refreshed_frame,
            sale_asn_create.ORDER_GRID_SYNC_TIMEOUT_SECONDS,
            False,
        ),
    ]
    assert any("còn thiếu 1 PO" in message for message in logs)
    assert logs[-1] == "[SALE ASN] Đã xác nhận đủ PO trong Order Details."


def test_reopens_add_order_popup_when_wfx_closes_it_between_rows(monkeypatch):
    rows_added = [
        {"source_row": 2, "po_no": "PO-1", "style_no": "STYLE A"},
        {"source_row": 3, "po_no": "PO-2", "style_no": "STYLE B"},
    ]
    context = object()
    page = object()
    popup_frame = object()
    calls = []

    class FakeAdd:
        @property
        def first(self):
            return self

        def wait_for(self, **kwargs):
            calls.append(("add-wait", kwargs))

        def evaluate(self, _script):
            calls.append(("add-dom-click",))
            return {"ok": True, "tag": "DIV", "id": ""}

    class FakeMainFrame:
        def locator(self, selector):
            assert selector == f"xpath={sale_asn_create.ADD_ORDER_XPATH}"
            return FakeAdd()

    main_frame = FakeMainFrame()
    popup_checks = 0

    def fake_frame_with_selector(selected_context, selector, timeout_s):
        nonlocal popup_checks
        assert selected_context is context
        calls.append(("frame", selector, timeout_s))
        if selector == sale_asn_create.PO_POPUP_SELECTOR:
            popup_checks += 1
            if popup_checks == 1:
                raise sale_asn_create.PlaywrightTimeoutError("popup closed")
            return page, popup_frame
        assert selector == "#sectionOrderDetails"
        return page, main_frame

    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        fake_frame_with_selector,
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_wait_order_grid",
        lambda frame, rows, timeout_s: calls.append(
            ("grid", frame, list(rows), timeout_s)
        ),
    )
    logs = []

    result = sale_asn_create._ensure_po_popup_for_next_row(
        context,
        rows_added,
        logs.append,
    )

    assert result is popup_frame
    assert (
        "grid",
        main_frame,
        rows_added,
        sale_asn_create.ORDER_GRID_SYNC_TIMEOUT_SECONDS,
    ) in calls
    assert ("add-dom-click",) in calls
    assert any("đã mở lại Add Order Details" in message for message in logs)


def test_sale_asn_resume_after_manual_final_ok_skips_closed_popup(monkeypatch):
    rows = [
        {"source_row": 2, "po_no": "PO-1", "invoice_no": "INV-1"},
    ]
    context = object()
    page = type("FakePage", (), {"context": context})()
    main_frame = object()
    selectors = []

    class FakePlaywright:
        def stop(self):
            return None

    class FakePlaywrightStarter:
        def start(self):
            return FakePlaywright()

    def fake_frame_with_selector(_context, selector, timeout_s):
        assert _context is context
        selectors.append((selector, timeout_s))
        assert selector == "#sectionOrderDetails"
        return page, main_frame

    monkeypatch.setattr(
        sale_asn_create,
        "sync_playwright",
        lambda: FakePlaywrightStarter(),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_active_wfx_page",
        lambda _playwright, _log: (object(), page),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        fake_frame_with_selector,
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_ensure_order_grid_rows",
        lambda _context, frame, _rows, _log, _search_fields: frame,
    )
    monkeypatch.setattr(sale_asn_create, "_fill_order_details", lambda *_args: None)
    monkeypatch.setattr(sale_asn_create, "_fill_style_details", lambda *_args: None)
    monkeypatch.setattr(sale_asn_create, "_fill_shipping", lambda *_args: None)
    monkeypatch.setattr(
        sale_asn_create,
        "_check_sale_asn_price_on_page",
        lambda *_args: {"ok": True, "code": "SALE_ASN_PRICE_CHECKED", "message": "checked"},
    )

    result = sale_asn_create.run_sale_asn_create(
        "menu-xpath",
        "BUYER A",
        rows,
        start_index=len(rows),
        log=lambda _message: None,
    )

    assert result["code"] == "SALE_ASN_FORM_COMPLETED"
    assert selectors == [("#sectionOrderDetails", 15)]


def test_sale_asn_resume_shipping_uses_visible_shipping_tab_only(monkeypatch):
    rows = [{"source_row": 2, "po_no": "PO-1", "invoice_no": "INV-1"}]
    context = object()
    page = type("FakePage", (), {"context": context})()
    shipping_frame = object()
    selectors = []
    calls = []

    class FakePlaywright:
        def stop(self):
            return None

    class FakePlaywrightStarter:
        def start(self):
            return FakePlaywright()

    def fake_frame_with_selector(_context, selector, timeout_s):
        assert _context is context
        selectors.append((selector, timeout_s))
        return page, shipping_frame

    monkeypatch.setattr(
        sale_asn_create,
        "sync_playwright",
        lambda: FakePlaywrightStarter(),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_active_wfx_page",
        lambda _playwright, _log: (object(), page),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        fake_frame_with_selector,
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_fill_order_details",
        lambda *_args: calls.append("order"),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_fill_style_details",
        lambda *_args: calls.append("style"),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_fill_shipping",
        lambda *_args: ["Factory: option-not-found"],
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_check_sale_asn_price_on_page",
        lambda *_args: {"ok": True, "code": "SALE_ASN_PRICE_CHECKED", "message": "checked"},
    )

    result = sale_asn_create.run_sale_asn_create(
        "menu-xpath",
        "BUYER A",
        rows,
        start_index=len(rows),
        stage="shipping_info",
        log=lambda _message: None,
    )

    assert result["code"] == "SALE_ASN_FORM_COMPLETED"
    assert result["warning_count"] == 1
    assert result["warnings"] == ["Factory: option-not-found"]
    assert "Shipping Info đã bỏ qua 1 trường" in result["message"]
    assert selectors == [("#tabShippingInfo", 15)]
    assert calls == []


def test_sale_asn_can_skip_add_po_and_start_from_existing_order_grid(monkeypatch):
    rows = [{"source_row": 2, "po_no": "PO-1", "carton": "2"}]
    context = object()
    page = type("FakePage", (), {"context": context})()
    order_frame = object()
    calls = []

    class FakePlaywright:
        def stop(self):
            return None

    class FakePlaywrightStarter:
        def start(self):
            return FakePlaywright()

    monkeypatch.setattr(
        sale_asn_create,
        "sync_playwright",
        lambda: FakePlaywrightStarter(),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_active_wfx_page",
        lambda _playwright, _log: (object(), page),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        lambda selected_context, selector, timeout_s: (
            calls.append(("frame", selected_context, selector, timeout_s))
            or (page, order_frame)
        ),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_wait_order_grid",
        lambda frame, selected_rows, **kwargs: (
            calls.append(("wait", frame, list(selected_rows), kwargs))
            or {sale_asn_create._fold("PO-1")}
        ),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_ensure_order_grid_rows",
        lambda *_args: pytest.fail("Không được mở hoặc thêm lại PO"),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_fill_order_details",
        lambda frame, selected_rows, _log, _progress=None: calls.append(
            ("fill-order", frame, list(selected_rows))
        ),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_check_sale_asn_price_on_page",
        lambda *_args: {"ok": True, "code": "SALE_ASN_PRICE_CHECKED", "message": "checked"},
    )

    result = sale_asn_create.run_sale_asn_create(
        "menu-xpath",
        "",
        rows,
        stage="order_details",
        skip_stages=("style_details", "shipping_info"),
        log=lambda _message: None,
    )

    assert result["code"] == "SALE_ASN_FORM_COMPLETED"
    assert result["add_po_selected"] is False
    assert ("fill-order", order_frame, rows) in calls
    assert calls[0][2] == "#sectionOrderDetails"


def test_shipping_info_skips_failed_field_and_continues(monkeypatch):
    calls = []
    logs = []

    class FakeTab:
        @property
        def first(self):
            return self

        def wait_for(self, **_kwargs):
            return None

        def click(self, **_kwargs):
            return None

    class FakeFrame:
        def locator(self, selector):
            assert selector == "#tabShippingInfo"
            return FakeTab()

    def fake_set_control(_frame, selector, value, mode, timeout_s):
        calls.append((selector, value, mode, timeout_s))
        if selector == "#ddlFactory":
            return {"ok": False, "reason": "option-not-found"}
        return {"ok": True}

    monkeypatch.setattr(sale_asn_create, "_set_control", fake_set_control)
    monkeypatch.setattr(sale_asn_create, "_wait", lambda *_args: None)

    warnings = sale_asn_create._fill_shipping(
        FakeFrame(),
        {
            "invoice_no": "INV-1",
            "invoice_date": "2026-08-01",
            "shipping_bill_no": "SB-1",
            "shipping_bill_date": "2026-08-02",
            "destination": "Germany",
            "factory": "PRO SPORTS GIAO THUY JSC",
            "consignee_address": "PUMA EUROPE",
            "ship_to": "PUMA CENTRAL",
            "shipping_mode": "AIR",
        },
        logs.append,
    )

    assert warnings == [
        'Factory: WFX không có lựa chọn "PRO SPORTS GIAO THUY JSC"'
    ]
    assert any(
        selector == sale_asn_create.CONSIGNEE_ADDRESS_SELECTOR
        and value == "PUMA EUROPE"
        and mode == "closest"
        for selector, value, mode, _timeout in calls
    )
    assert any(
        selector == sale_asn_create.SHIP_TO_SELECTOR
        and value == "PUMA CENTRAL"
        and mode == "closest"
        for selector, value, mode, _timeout in calls
    )
    for port_selector in sale_asn_create.PORT_OF_LOADING_SELECTORS:
        assert (port_selector, "HAN - Hanoi", "exact", 6) in calls
    assert ("#ddlShipmentMode", "AIR", "exact", 6) in calls
    shipment_index = next(
        index for index, call in enumerate(calls) if call[0] == "#ddlShipmentMode"
    )
    port_index = next(
        index
        for index, call in enumerate(calls)
        if call[0] in sale_asn_create.PORT_OF_LOADING_SELECTORS
    )
    assert shipment_index < port_index
    assert (
        "#ddlDeliveryTerms",
        "FCA HANOI, VIET NAM",
        "exact",
        6,
    ) in calls
    assert (
        "#ddlFactory",
        "PRO SPORTS GIAO THUY JSC",
        "factory_first",
        6,
    ) in calls
    assert all(selector != "#ddlNotify1" for selector, *_rest in calls)
    assert any("Shipping Info bỏ qua Factory: WFX không có lựa chọn" in item for item in logs)
    assert logs[-1] == (
        "[SALE ASN] Đã điền Shipping Info; bỏ qua 1 trường và chưa bấm Save."
    )


def test_shipping_modes_map_to_port_and_delivery_terms():
    assert sale_asn_create.SHIPPING_MODE_VALUES == {
        "AIR": {
            "port_of_loading": "HAN - Hanoi",
            "delivery_terms": "FCA HANOI, VIET NAM",
        },
        "SEA": {
            "port_of_loading": "HPH - Haiphong",
            "delivery_terms": "FOB HAIPHONG, VIETNAM",
        },
        "COURIER": {
            "port_of_loading": "HAN - Hanoi",
            "delivery_terms": "EXW",
        },
    }


def test_shipping_keeps_default_final_destination_when_country_is_not_found(
    monkeypatch,
):
    calls = []
    logs = []

    class FakeTab:
        @property
        def first(self):
            return self

        def wait_for(self, **_kwargs):
            return None

        def click(self, **_kwargs):
            return None

    class FakeFrame:
        def locator(self, selector):
            assert selector == "#tabShippingInfo"
            return FakeTab()

    def fake_set_control(_frame, selector, value, mode, timeout_s):
        calls.append((selector, value, mode, timeout_s))
        if selector == "#Cell_DestinationCountry":
            return {"ok": False, "reason": "option-not-found"}
        return {"ok": True}

    monkeypatch.setattr(sale_asn_create, "_set_control", fake_set_control)
    monkeypatch.setattr(sale_asn_create, "_wait", lambda *_args: None)

    sale_asn_create._fill_shipping(
        FakeFrame(),
        {
            "invoice_no": "INV-1",
            "invoice_date": "2026-08-01",
            "shipping_bill_no": "SB-1",
            "shipping_bill_date": "2026-08-02",
            "destination": "DE",
            "factory": "FACTORY",
            "consignee_address": "CONSIGNEE",
            "ship_to": "SHIP TO",
            "shipping_mode": "AIR",
        },
        logs.append,
    )

    assert any(call[0] == "#Cell_DestinationCountry" for call in calls)
    assert all(call[0] != "#Cell_FinalDestination" for call in calls)
    assert any("Giữ nguyên Final Destination" in line for line in logs)


def test_shipping_keeps_destination_defaults_when_file_destination_is_blank(monkeypatch):
    calls = []

    class FakeTab:
        @property
        def first(self):
            return self

        def wait_for(self, **_kwargs):
            return None

        def click(self, **_kwargs):
            return None

    class FakeFrame:
        def locator(self, _selector):
            return FakeTab()

    monkeypatch.setattr(
        sale_asn_create,
        "_set_control",
        lambda _frame, selector, value, mode, timeout_s: calls.append(selector) or {"ok": True},
    )
    monkeypatch.setattr(sale_asn_create, "_wait", lambda *_args: None)

    sale_asn_create._fill_shipping(
        FakeFrame(),
        {
            "invoice_no": "INV-1", "invoice_date": "2026-08-01",
            "shipping_bill_no": "SB-1", "shipping_bill_date": "2026-08-02",
            "destination": "", "factory": "FACTORY", "shipping_mode": "AIR",
        },
        lambda _message: None,
    )

    assert "#Cell_DestinationCountry" not in calls
    assert "#Cell_FinalDestination" not in calls


def test_fuzzy_dropdown_requires_one_unique_best_match():
    assert sale_asn_create._best_dropdown_label(
        ["PUMA EUROPE GMBH", "OTHER COMPANY"],
        "puma europe",
    ) == "PUMA EUROPE GMBH"
    assert sale_asn_create._best_dropdown_label(
        ["ALPHA ONE", "ALPHA TWO"],
        "alpha",
    ) is None
    assert sale_asn_create._best_dropdown_label(["OTHER"], "missing") is None


def test_factory_dropdown_is_fuzzy_case_insensitive_and_skips_dotted_rows():
    assert sale_asn_create._best_factory_label(
        [
            "PRO SPORTS GIAO THUY JSC.",
            "PRO SPORTS GIAO THUY JSC",
            "GIAO THUY GARMENT FACTORY",
        ],
        "GIAO THUY",
    ) == "PRO SPORTS GIAO THUY JSC"
    assert sale_asn_create._best_factory_label(
        ["PROSPORTS GIAO THUY JOINT STOCK COMPANY", "OTHER FACTORY"],
        "pro sports giao thuy",
    ) == "PROSPORTS GIAO THUY JOINT STOCK COMPANY"
    assert sale_asn_create._best_factory_label(
        ["PRO SPORTS NAM DINH", "OTHER FACTORY"],
        "GIAO THUY",
    ) is None


def test_shipping_control_chooses_visible_host_for_alternative_port_cells():
    source = sale_asn_create._SET_CONTROL_JS
    assert "document.querySelectorAll(spec.selector)" in source
    assert "hosts.find(shown)" in source


def test_sale_asn_table_value_confirmation_handles_wfx_formats():
    assert sale_asn_create._number_for_wfx("110", integer=True) == "110"
    assert sale_asn_create._number_for_wfx("498.99999999999994") == "499"
    assert sale_asn_create._table_value_matches("1085", "1,085.0000")
    assert sale_asn_create._table_value_matches("03/08/2026", "03 Aug 2026")
    assert sale_asn_create._table_value_matches(
        "Men's pants (1 layer)", "Mens pants (1 layer)"
    )
    assert not sale_asn_create._table_value_matches("110", "0")


def test_price_check_compares_po_style_and_all_summary_totals():
    comparisons, totals = sale_asn_create._price_check_rows(
        [
            {
                "source_row": 2,
                "po_no": "NT 12-2026",
                "style_no": "Striker XT Gen.3 Pants Multicam",
                "qty": "70",
                "price": "95",
            }
        ],
        [
            {
                "order_no": "PSW-UFPRO-26-4957.12/NT 12-2026",
                "article": "UFPRO-Striker XT Gen.3 Pants Multicam-TENDER-MEN PANTS",
                "qty": "70.0000",
                "price": "95.0000",
            }
        ],
    )
    summary = sale_asn_create._summary_price_check(
        totals,
        {
            "total_quantity": "70.00",
            "value_in_doc_currency": "6650.00",
            "net_value_in_doc_currency": "6650.00",
        },
    )

    assert comparisons[0]["status"] == "ok"
    assert comparisons[0]["system_qty"] == "70"
    assert comparisons[0]["system_prices"] == ["95"]
    assert comparisons[0]["system_order_nos"] == ["PSW-UFPRO-26-4957.12/NT 12-2026"]
    assert summary["ok"] is True
    assert all(item["ok"] for item in summary["checks"].values())


def test_price_check_flags_qty_price_and_summary_mismatches():
    comparisons, totals = sale_asn_create._price_check_rows(
        [
            {
                "source_row": 2,
                "po_no": "PO-1",
                "style_no": "STYLE-A",
                "qty": "10",
                "price": "20",
            }
        ],
        [{"po_no": "PO-1", "article": "STYLE-A", "qty": "9", "price": "21"}],
    )
    summary = sale_asn_create._summary_price_check(
        totals,
        {
            "total_quantity": "9",
            "value_in_doc_currency": "189",
            "net_value_in_doc_currency": "189",
        },
    )

    assert comparisons[0]["status"] == "mismatch"
    assert summary["ok"] is False
    assert all(not item["ok"] for item in summary["checks"].values())


def test_price_check_export_workbook_has_detail_and_summary_sheets(tmp_path):
    target = write_sale_asn_price_check_workbook(
        tmp_path / "price-check",
        {
            "message": "Đã check 1 PO + Style: 1 khớp.",
            "comparisons": [
                {
                    "source_rows": [2],
                    "po_no": "NT 12-2026",
                    "style_no": "Striker XT",
                    "system_order_nos": ["PSW-26/NT 12-2026"],
                    "file_qty": "70",
                    "file_price": "95",
                    "system_qty": "70",
                    "system_prices": ["95"],
                    "status": "ok",
                    "message": "Khớp Qty và Price.",
                }
            ],
            "summary": {
                "checks": {
                    key: {"expected": value, "actual": value, "ok": True}
                    for key, value in {
                        "total_quantity": "70",
                        "value_in_doc_currency": "6650",
                        "net_value_in_doc_currency": "6650",
                    }.items()
                }
            },
        },
    )

    workbook = load_workbook(target)
    assert target.suffix == ".xlsx"
    assert workbook.sheetnames == ["PO + Style", "Summary Total"]
    assert workbook["PO + Style"]["D5"].value == "PSW-26/NT 12-2026"
    assert workbook["PO + Style"]["I5"].value == "KHỚP"
    assert workbook["Summary Total"]["D4"].value == "KHỚP"
    workbook.close()


def test_sale_asn_dates_use_english_wfx_format_independent_of_system_locale(monkeypatch):
    real_datetime = sale_asn_create.datetime

    class LocalizedDate:
        day = 3
        month = 8
        year = 2026

        def strftime(self, _pattern):
            return "03 thg 8 2026"

    class LocalizedDatetime:
        @classmethod
        def strptime(cls, value, pattern):
            assert (value, pattern) == ("2026-08-03", "%Y-%m-%d")
            return LocalizedDate()

    monkeypatch.setattr(sale_asn_create, "datetime", LocalizedDatetime)

    assert sale_asn_create._date_for_wfx("2026-08-03") == "03 Aug 2026"

    monkeypatch.setattr(sale_asn_create, "datetime", real_datetime)


def test_sale_asn_dates_support_every_english_wfx_month():
    expected = (
        "02 Jan 2026",
        "02 Feb 2026",
        "02 Mar 2026",
        "02 Apr 2026",
        "02 May 2026",
        "02 Jun 2026",
        "02 Jul 2026",
        "02 Aug 2026",
        "02 Sep 2026",
        "02 Oct 2026",
        "02 Nov 2026",
        "02 Dec 2026",
    )

    assert tuple(
        sale_asn_create._date_for_wfx(f"2026-{month:02d}-02")
        for month in range(1, 13)
    ) == expected


class _FakeSaleASNLogin:
    COMPANY_ID = "psh"
    CATALOG_XPATH = "catalog"

    def __init__(self):
        self.calls = []

    def scan_sale_asn_buyers(self, xpath, log=print):
        self.calls.append(("scan", xpath))
        return {
            "ok": True,
            "code": "SALE_ASN_BUYERS_SCANNED",
            "message": "buyers",
            "buyers": [{"label": "BUYER A", "value": "A"}],
        }

    def run_sale_asn_create(
        self,
        xpath,
        buyer,
        rows,
        start_index,
        log=print,
        *,
        stage="po",
        skip_stages=(),
        search_fields=("po", "style", "destination"),
        progress=None,
    ):
        self.calls.append(
            (
                "create",
                xpath,
                buyer,
                len(rows),
                start_index,
                stage,
                tuple(skip_stages),
                tuple(search_fields),
            )
        )
        if start_index == 0:
            return {
                "ok": True,
                "code": "SALE_ASN_PO_SELECTION_REQUIRED",
                "message": "manual",
                "next_index": 1,
            }
        return {
            "ok": True,
            "code": "SALE_ASN_FORM_COMPLETED",
            "message": "done",
        }


def test_panel_api_keeps_review_across_manual_po_checkpoint(tmp_path):
    source = tmp_path / "input.xlsx"
    _input_workbook(source, _valid_rows())
    login = _FakeSaleASNLogin()
    api = PanelAPI(login_module=login, prefs_module=prefs, base_dir=tmp_path / "data")

    scanned = api.scan_sale_asn_buyers()
    reviewed = api.prepare_sale_asn_create(str(source), "BUYER A")
    pending = api.start_sale_asn_create(reviewed["review_token"])
    completed = api.continue_sale_asn_create(reviewed["review_token"])
    expired = api.continue_sale_asn_create(reviewed["review_token"])

    assert scanned["buyers"] == [{"label": "BUYER A", "value": "A"}]
    assert reviewed["code"] == "SALE_ASN_CREATE_REVIEW_READY"
    assert pending["code"] == "SALE_ASN_PO_SELECTION_REQUIRED"
    assert completed["code"] == "SALE_ASN_FORM_COMPLETED"
    assert expired["code"] == "SALE_ASN_CREATE_REVIEW_EXPIRED"
    assert [call[4] for call in login.calls if call[0] == "create"] == [0, 1]


def test_panel_api_returns_automatic_price_check_with_completed_sale_asn(tmp_path):
    source = tmp_path / "input.xlsx"
    _input_workbook(source, _valid_rows())
    login = _FakeSaleASNLogin()
    api = PanelAPI(login_module=login, prefs_module=prefs, base_dir=tmp_path / "data")

    reviewed = api.prepare_sale_asn_create(str(source), "BUYER A")
    api.start_sale_asn_create(reviewed["review_token"])
    completed = api.continue_sale_asn_create(reviewed["review_token"])

    assert completed["code"] == "SALE_ASN_FORM_COMPLETED"
    assert "price_check_token" not in completed


def test_panel_api_retries_or_skips_failed_sale_asn_stage(tmp_path):
    source = tmp_path / "input.xlsx"
    _input_workbook(source, _valid_rows())

    class FailingStageLogin(_FakeSaleASNLogin):
        def run_sale_asn_create(
            self,
            xpath,
            buyer,
            rows,
            start_index,
            log=print,
            *,
            stage="po",
            skip_stages=(),
            search_fields=("po", "style", "destination"),
            progress=None,
        ):
            self.calls.append(
                (
                    "create",
                    xpath,
                    buyer,
                    len(rows),
                    start_index,
                    stage,
                    tuple(skip_stages),
                    tuple(search_fields),
                )
            )
            if "style_details" not in skip_stages:
                return {
                    "ok": False,
                    "code": "SALE_ASN_FIELD_NOT_EDITABLE",
                    "message": "style failed",
                    "resumable": True,
                    "resume_stage": "style_details",
                    "stage_label": "Style Details",
                    "can_skip": True,
                }
            return {
                "ok": True,
                "code": "SALE_ASN_FORM_COMPLETED",
                "message": "done",
            }

    login = FailingStageLogin()
    api = PanelAPI(login_module=login, prefs_module=prefs, base_dir=tmp_path / "data")
    reviewed = api.prepare_sale_asn_create(str(source), "BUYER A")

    failed = api.start_sale_asn_create(reviewed["review_token"])
    completed = api.skip_sale_asn_create_step(reviewed["review_token"])

    assert failed["resume_stage"] == "style_details"
    assert failed["review_token"] == reviewed["review_token"]
    assert completed["code"] == "SALE_ASN_FORM_COMPLETED"
    assert login.calls[-1][5:7] == ("style_details", ("style_details",))


def test_panel_api_starts_at_first_selected_stage_without_buyer(tmp_path):
    source = tmp_path / "continue.xlsx"
    _input_workbook(source, [{"PO No": "PO-001", "Carton": 2}])

    class ContinueLogin(_FakeSaleASNLogin):
        def run_sale_asn_create(
            self,
            xpath,
            buyer,
            rows,
            start_index,
            log=print,
            *,
            stage="po",
            skip_stages=(),
            search_fields=("po", "style", "destination"),
            progress=None,
        ):
            self.calls.append(
                (
                    "continue-selected",
                    xpath,
                    buyer,
                    list(rows),
                    start_index,
                    stage,
                    tuple(skip_stages),
                    tuple(search_fields),
                )
            )
            return {
                "ok": True,
                "code": "SALE_ASN_FORM_COMPLETED",
                "message": "done",
            }

    login = ContinueLogin()
    api = PanelAPI(login_module=login, prefs_module=prefs, base_dir=tmp_path / "data")

    reviewed = api.prepare_sale_asn_create(
        str(source),
        "",
        ["order_details"],
    )
    completed = api.start_sale_asn_create(reviewed["review_token"])

    assert reviewed["selected_stages"] == ["order_details"]
    assert completed["code"] == "SALE_ASN_FORM_COMPLETED"
    call = login.calls[0]
    assert call[2] == ""
    assert call[5] == "order_details"
    assert call[6] == ("po", "style_details", "shipping_info")


def test_stage_progress_numbers_follow_the_five_fixed_stages():
    seen = []

    def sink(stage, message, step, total, *, state="active"):
        seen.append((stage, message, step, total, state))

    for stage in sale_asn_create.SALE_ASN_STAGE_ORDER:
        sale_asn_create._emit_stage_progress(sink, stage, f"Đang điền {stage}")
    sale_asn_create._emit_stage_progress(sink, "style_details", "bỏ", state="skipped")

    # Bộ đếm luôn tính trên năm bước cố định, không theo số bước user chọn.
    assert [(item[0], item[2], item[3]) for item in seen] == [
        ("po", 1, 5),
        ("order_details", 2, 5),
        ("style_details", 3, 5),
        ("shipping_info", 4, 5),
        ("price_check", 5, 5),
        ("style_details", 3, 5),
    ]
    assert seen[-1][4] == "skipped"


def test_stage_progress_is_optional_and_never_breaks_the_flow():
    def broken(*_args, **_kwargs):
        raise ValueError("callback hỏng")

    # Không có callback, và callback lỗi, đều không được ném ra ngoài.
    sale_asn_create._emit_stage_progress(None, "po", "x")
    sale_asn_create._emit_stage_progress(broken, "po", "x")


def test_panel_api_streams_sale_asn_progress_with_its_own_method(tmp_path):
    source = tmp_path / "input.xlsx"
    _input_workbook(source, _valid_rows())
    payloads = []

    class ProgressLogin(_FakeSaleASNLogin):
        def run_sale_asn_create(
            self,
            xpath,
            buyer,
            rows,
            start_index,
            log=print,
            *,
            stage="po",
            skip_stages=(),
            search_fields=("po", "style", "destination"),
            progress=None,
        ):
            progress("order_details", "Đang điền Order Details", 2, 4)
            return {
                "ok": True,
                "code": "SALE_ASN_FORM_COMPLETED",
                "message": "done",
            }

    api = PanelAPI(
        login_module=ProgressLogin(),
        prefs_module=prefs,
        base_dir=tmp_path / "data",
    )
    api.set_progress_sink(payloads.append)
    reviewed = api.prepare_sale_asn_create(str(source), "BUYER A")
    api.start_sale_asn_create(reviewed["review_token"])

    # Payload phải mang method của chính flow Sale ASN để UI không đè thẻ GDN.
    assert [item["method"] for item in payloads] == ["start_sale_asn_create"]
    assert payloads[0]["stage"] == "order_details"
    assert payloads[0]["step"] == 2
    assert payloads[0]["total"] == 4
    assert payloads[0]["state"] == "active"


def test_order_and_style_stages_report_row_counters(monkeypatch):
    """Hai bước điền chạy vòng lặp từng dòng nên phải có bộ đếm như bước PO."""
    seen = []

    def sink(stage, message, step, total, *, state="active"):
        seen.append((stage, message))

    rows = [
        {"po_no": "PO-1", "style_no": "S1", "hs_code": "6109", "carton": "2"},
        {"po_no": "PO-2", "style_no": "S2", "hs_code": "6110", "carton": "3"},
        {"po_no": "PO-3", "style_no": "S1", "hs_code": "6109", "carton": "4"},
    ]

    class FakeLocator:
        first = property(lambda self: self)

        def wait_for(self, **_kwargs):
            return None

        def click(self, **_kwargs):
            return None

    class FakeFrame:
        def locator(self, _selector):
            return FakeLocator()

    monkeypatch.setattr(
        sale_asn_create, "_set_order_grid_cell", lambda *a, **k: {"ok": True}
    )
    monkeypatch.setattr(
        sale_asn_create, "_set_style_hts_cell", lambda *a, **k: {"ok": True}
    )
    monkeypatch.setattr(sale_asn_create, "_wait", lambda *a, **k: None)

    frame = FakeFrame()
    sale_asn_create._fill_order_details(frame, rows, lambda _m: None, sink)
    sale_asn_create._fill_style_details(frame, rows, lambda _m: None, sink)

    order = [message for stage, message in seen if stage == "order_details"]
    style = [message for stage, message in seen if stage == "style_details"]
    assert order == ["Order Details 1/3", "Order Details 2/3", "Order Details 3/3"]
    # Style gom theo Style No. nên chỉ có 2 lượt cho 3 dòng.
    assert style == ["Style Details 1/2", "Style Details 2/2"]


def test_panel_api_remembers_selected_sale_asn_stages(tmp_path):
    base = tmp_path / "data"
    base.mkdir(parents=True, exist_ok=True)
    api = PanelAPI(
        login_module=_FakeSaleASNLogin(),
        prefs_module=prefs,
        base_dir=base,
    )

    assert api.get_initial_state()["sale_asn_stages"] == [
        "po",
        "order_details",
        "style_details",
        "shipping_info",
    ]

    saved = api.set_sale_asn_stages(["order_details", "shipping_info"])
    assert saved["sale_asn_stages"] == ["order_details", "shipping_info"]
    assert api.get_initial_state()["sale_asn_stages"] == [
        "order_details",
        "shipping_info",
    ]

    # Bỏ hết bước sẽ khiến user mở app ra mà không chạy được gì → quay về đủ bốn.
    assert api.set_sale_asn_stages([])["sale_asn_stages"] == [
        "po",
        "order_details",
        "style_details",
        "shipping_info",
    ]


def test_panel_api_snapshots_sale_asn_po_search_fields_for_each_review(tmp_path):
    source = tmp_path / "input.xlsx"
    _input_workbook(source, _valid_rows())
    login = _FakeSaleASNLogin()
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    api = PanelAPI(
        login_module=login,
        prefs_module=prefs,
        base_dir=data_dir,
    )

    assert api.get_initial_state()["sale_asn_po_search_fields"] == [
        "po",
        "style",
        "destination",
    ]
    saved = api.set_sale_asn_po_search_fields(["po", "destination"])
    assert saved["sale_asn_po_search_fields"] == ["po", "destination"]
    reviewed = api.prepare_sale_asn_create(str(source), "BUYER A")

    api.set_sale_asn_po_search_fields(["style"])
    api.start_sale_asn_create(reviewed["review_token"])

    assert login.calls[-1][7] == ("po", "destination")


def test_number_out_of_decimal_range_is_a_file_error_not_an_automation_crash(
    tmp_path,
):
    """Ô số quá lớn phải chặn ở bước kiểm tra file.

    Nếu để lọt, ``_number_for_wfx`` gọi ``quantize()`` vượt precision của
    Decimal context và ném ``InvalidOperation`` lúc đang điền lên WFX — lỗi
    nhập liệu của user biến thành lỗi automation có gửi telemetry.
    """
    rows = _valid_rows()
    rows[0]["Carton"] = "1E+50"
    source = tmp_path / "input.xlsx"
    _input_workbook(source, rows)

    with pytest.raises(SaleASNWorkbookError) as error:
        read_sale_asn_workbook(source)

    assert error.value.code == "SALE_ASN_FILE_VALIDATION_FAILED"
    assert any("quá lớn" in item for item in error.value.errors)


def test_number_for_wfx_never_raises_on_out_of_range_input():
    assert sale_asn_create._number_for_wfx("1E+50", integer=True) == "1E+50"
    assert sale_asn_create._number_for_wfx("2", integer=True) == "2"
    assert sale_asn_create._number_for_wfx("9.50") == "9.5"


def test_skipped_stage_never_waits_for_its_own_tab(monkeypatch):
    """Bỏ qua một bước không được bắt user chờ đúng tab đang hỏng.

    Người dùng bấm Bỏ qua chính vì tab đó không dùng được; resolve frame trước
    vòng lặp làm nút Bỏ qua vô dụng.
    """
    rows = [{"source_row": 2, "po_no": "PO-1", "invoice_no": "INV-1"}]
    context = object()
    page = type("FakePage", (), {"context": context})()

    class FakePlaywright:
        def stop(self):
            return None

    monkeypatch.setattr(
        sale_asn_create,
        "sync_playwright",
        lambda: type("Starter", (), {"start": lambda _self: FakePlaywright()})(),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_active_wfx_page",
        lambda _playwright, _log: (object(), page),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        lambda *_args, **_kwargs: pytest.fail(
            "Không được resolve frame của bước đã bỏ qua"
        ),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_check_sale_asn_price_on_page",
        lambda *_args: {"ok": True, "code": "SALE_ASN_PRICE_CHECKED", "message": "checked"},
    )

    result = sale_asn_create.run_sale_asn_create(
        "menu-xpath",
        "",
        rows,
        start_index=len(rows),
        stage="shipping_info",
        skip_stages=("shipping_info",),
        log=lambda _message: None,
    )

    assert result["code"] == "SALE_ASN_FORM_COMPLETED"
    assert result["warnings"] == []


def test_recovering_missing_po_asks_the_user_instead_of_failing(monkeypatch):
    """Popup đang mở chờ user chọn dòng thì không được trả lỗi kỹ thuật."""
    rows = [
        {"source_row": 2, "po_no": "PO-1", "style_no": "STYLE A"},
        {"source_row": 3, "po_no": "PO-2", "style_no": "STYLE B"},
    ]
    context = object()
    page = type("FakePage", (), {"context": context})()
    candidates = [
        {"po_no": "PO-2", "style_no": "STYLE B1", "dispatched_qty": "10"},
        {"po_no": "PO-2", "style_no": "STYLE B2", "dispatched_qty": "20"},
    ]

    class FakePlaywright:
        def stop(self):
            return None

    monkeypatch.setattr(
        sale_asn_create,
        "sync_playwright",
        lambda: type("Starter", (), {"start": lambda _self: FakePlaywright()})(),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_active_wfx_page",
        lambda _playwright, _log: (object(), page),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_frame_with_selector",
        lambda _context, _selector, timeout_s=15: (page, object()),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_wait_order_grid",
        lambda _frame, _rows, timeout_s=10, **_kwargs: {
            sale_asn_create._fold("PO-1")
        },
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_auto_add_po",
        lambda _frame, _row, _log, *, final, search_fields: (
            False,
            candidates,
            "ambiguous",
        ),
    )
    monkeypatch.setattr(
        sale_asn_create,
        "_fill_order_details",
        lambda *_args: pytest.fail("Chưa đủ PO thì không được điền"),
    )

    result = sale_asn_create.run_sale_asn_create(
        "menu-xpath",
        "BUYER A",
        rows,
        start_index=len(rows),
        stage="po",
        log=lambda _message: None,
    )

    assert result["code"] == "SALE_ASN_PO_SELECTION_REQUIRED"
    assert result["po_no"] == "PO-2"
    assert result["candidates"] == candidates
    # Lượt Tiếp tục phải bỏ qua vòng thêm PO và để _ensure_order_grid_rows dò
    # lại, nếu không các PO đã vào grid sẽ bị thêm trùng.
    assert result["next_index"] == len(rows)


def test_auto_add_po_stops_narrowing_once_wfx_returns_nothing(monkeypatch):
    """Ba lượt thử là thu hẹp dần nên 0 kết quả ở lượt đầu là 0 ở mọi lượt."""
    searches = []

    def fake_search(_frame, _row, *, fields):
        searches.append(tuple(fields))
        return []

    monkeypatch.setattr(sale_asn_create, "_search_po", fake_search)

    added, candidates, reason = _auto_add_po(
        object(),
        {"source_row": 2, "po_no": "PO-1"},
        lambda _message: None,
    )

    assert added is False
    assert candidates == []
    assert reason == "not_found"
    assert searches == [("po",)]


def test_po_search_clears_disabled_fields_before_each_row(monkeypatch):
    filled = []
    destinations = []

    def fake_fill(_frame, selector, value):
        filled.append((selector, value))
        return selector in {"#txtOCNo", sale_asn_create.STYLE_INPUT_SELECTORS[0]}

    class FakeTable:
        first = property(lambda self: self)

        def wait_for(self, **_kwargs):
            return None

        def evaluate(self, _script):
            return []

    class FakeFrame:
        def locator(self, selector):
            assert selector == sale_asn_create.PO_RESULTS_TABLE_SELECTOR
            return FakeTable()

    monkeypatch.setattr(sale_asn_create, "_fill_popup_input", fake_fill)
    monkeypatch.setattr(
        sale_asn_create,
        "_select_popup_destination",
        lambda _frame, value: destinations.append(value) or True,
    )
    monkeypatch.setattr(sale_asn_create, "_click_search", lambda _frame: None)

    sale_asn_create._search_po(
        FakeFrame(),
        {"po_no": "PO-1", "style_no": "STYLE-1", "destination": "DE"},
        fields=("po",),
    )

    assert filled[:2] == [
        ("#txtOCNo", "PO-1"),
        (sale_asn_create.STYLE_INPUT_SELECTORS[0], ""),
    ]
    assert destinations == [""]


