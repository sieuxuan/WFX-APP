from openpyxl import Workbook, load_workbook

from wfx_panel.automation.dispatch import (
    choose_latest_pending_row,
    reload_dispatch_workbook,
)


def test_choose_latest_pending_dispatch_by_processed_on():
    rows = [
        {
            "row_id": "12",
            "package_name": "DecisionOne_BuyerOrderDispatch",
            "processed_on": "7/30/2026 4:55:05 PM",
            "transaction_detail": "InProgress",
        },
        {
            "row_id": "10",
            "package_name": "DecisionOne_BuyerOrderDispatch",
            "processed_on": "7/30/2026 4:03:20 PM",
            "transaction_detail": "Pending",
        },
        {
            "row_id": "11",
            "package_name": "DecisionOne_BuyerOrderDispatch",
            "processed_on": "7/30/2026 4:44:20 PM",
            "transaction_detail": "Pending",
        },
        {
            "row_id": "99",
            "package_name": "OtherPackage",
            "processed_on": "8/1/2026 8:00:00 AM",
            "transaction_detail": "Pending",
        },
    ]

    selected = choose_latest_pending_row(rows)

    assert selected is not None
    assert selected["row_id"] == "11"


def test_choose_latest_pending_ignores_rows_seen_before_process():
    rows = [
        {
            "row_id": "20",
            "package_name": "DecisionOne_BuyerOrderDispatch",
            "processed_on": "8/2/2026 1:00:00 PM",
            "transaction_detail": "Pending",
        },
        {
            "row_id": "21",
            "package_name": "DecisionOne_BuyerOrderDispatch",
            "processed_on": "8/2/2026 1:01:00 PM",
            "transaction_detail": "Pending",
        },
    ]

    selected = choose_latest_pending_row(rows, excluded_ids={"20"})

    assert selected is not None
    assert selected["row_id"] == "21"


def test_reload_dispatch_workbook_resaves_valid_xlsx(tmp_path):
    source = tmp_path / "report.xlsx"
    target = tmp_path / "reload.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["OCNo", "DispatchQty"])
    sheet.append(["PSW-001", 123])
    workbook.save(source)
    workbook.close()

    reload_dispatch_workbook(source, target)

    reloaded = load_workbook(target, data_only=False)
    try:
        assert reloaded.sheetnames == ["Summary"]
        assert reloaded["Summary"]["A2"].value == "PSW-001"
        assert reloaded["Summary"]["B2"].value == 123
        assert reloaded.calculation.fullCalcOnLoad is True
    finally:
        reloaded.close()
