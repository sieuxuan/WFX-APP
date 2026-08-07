from openpyxl import load_workbook

from wfx_panel.style_workbook import (
    GUIDE_SHEET,
    LIST_SHEET,
    STYLE_COLUMNS,
    STYLE_SHEET,
    StyleWorkbookError,
    read_style_workbook,
    write_style_template,
)


def test_style_template_has_two_user_facing_sheets_and_hidden_lists(tmp_path):
    target = write_style_template(tmp_path / "styles.xlsx")

    workbook = load_workbook(target)
    assert workbook.sheetnames == [GUIDE_SHEET, STYLE_SHEET, LIST_SHEET]
    assert workbook[LIST_SHEET].sheet_state == "veryHidden"
    sheet = workbook[STYLE_SHEET]
    assert tuple(cell.value for cell in sheet[1]) == STYLE_COLUMNS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is None
    validations = list(sheet.data_validations.dataValidation)
    assert any(item.formula1 == '"New,Copy"' for item in validations)
    assert any(item.formula1 == "=StyleMaterialType" for item in validations)
    workbook.close()


def test_style_template_has_server_dropdowns_and_dependent_subcategory(tmp_path):
    target = write_style_template(
        tmp_path / "styles.xlsx",
        options={
            "fields": {
                "style_copy": ["Style Apparel A", "Style Apparel B"],
                "material_type": ["KNIT", "WOVEN"],
                "buyer": ["Buyer A"],
                "division": ["Division A"],
                "product_group": ["Top", "Bottom"],
                "color_card": ["Color A"],
                "size_range": ["Size A"],
                "season": ["FW27"],
            },
            "subcategories_by_product_group": {
                "Top": ["Jacket", "Shirt"],
                "Bottom": ["Pants"],
            },
        },
    )
    workbook = load_workbook(target)
    validations = list(workbook[STYLE_SHEET].data_validations.dataValidation)
    formulas = {item.formula1 for item in validations}
    assert "=StyleCopy" in formulas
    assert "=StyleBuyer" in formulas
    assert "=StyleProductGroup" in formulas
    assert any("VLOOKUP($F2" in formula for formula in formulas)
    workbook.close()


def test_style_workbook_reads_new_and_copy_rows(tmp_path):
    target = write_style_template(tmp_path / "styles.xlsx")
    workbook = load_workbook(target)
    sheet = workbook[STYLE_SHEET]
    sheet.append(
        [
            "New", "", "KNIT", "Buyer A", "Knit", "Top", "Polo",
            "Color A", "Size A", "SS27", "BUY-1", "INT-1",
        ]
    )
    sheet.append(
        [
            "Copy", "SWN001", "", "", "", "", "", "", "", "",
            "BUY-2", "INT-2",
        ]
    )
    workbook.save(target)
    workbook.close()

    rows = read_style_workbook(target)

    assert [row.type for row in rows] == ["New", "Copy"]
    assert rows[0].material_type == "KNIT"
    assert rows[1].style_copy == "SWN001"
    assert rows[1].buyer_style_ref == "BUY-2"


def test_new_style_requires_all_business_fields(tmp_path):
    target = write_style_template(tmp_path / "styles.xlsx")
    workbook = load_workbook(target)
    sheet = workbook[STYLE_SHEET]
    sheet.append(["New", "", "KNIT"])
    workbook.save(target)
    workbook.close()

    try:
        read_style_workbook(target)
    except StyleWorkbookError as error:
        assert error.code == "STYLE_FILE_VALIDATION_FAILED"
        assert "Buyer" in error.errors[0]
    else:
        raise AssertionError("Workbook thiếu trường phải bị từ chối")


def test_copy_style_requires_source_reference(tmp_path):
    target = write_style_template(tmp_path / "styles.xlsx")
    workbook = load_workbook(target)
    workbook[STYLE_SHEET].append(["Copy"])
    workbook.save(target)
    workbook.close()

    try:
        read_style_workbook(target)
    except StyleWorkbookError as error:
        assert error.code == "STYLE_FILE_VALIDATION_FAILED"
        assert "Style copy" in error.errors[0]
    else:
        raise AssertionError("Copy thiếu Style nguồn phải bị từ chối")
