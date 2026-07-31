import json
import threading
import time
from pathlib import Path

from openpyxl import load_workbook

from wfx_panel import (
    article_library,
    constants,
    costing_workbook,
    oc_workbook,
    prefs,
    telemetry,
)
from wfx_panel.atomic_io import write_json_atomic
from wfx_panel.automation import runtime as automation_runtime
from wfx_panel.panel_api import PanelAPI


class FakeLogin:
    COMPANY_ID = "psh"
    CATALOG_XPATH = '//*[@id="0003_6200"]/a'

    def __init__(self):
        self.calls = []

    def run(self, user_id, password, company_id="psh", log=print):
        self.calls.append(("run", user_id, password, company_id))
        log("[SESSION] fake login")
        return {
            "ok": True,
            "code": "LOGGED_IN",
            "message": "ok",
            "current_division": "woven",
            "division_label": "WOVEN",
            "division_name": "PRO SPORTS - WOVEN HANOI",
        }

    def check_session(self, log=print):
        self.calls.append(("check_session",))
        return {"ok": False, "code": "NOT_LOGGED_IN", "message": "no"}

    def start_chrome(self, log=print):
        self.calls.append(("start_chrome",))
        return {
            "ok": True,
            "code": "CHROME_OPENED",
            "message": "opened",
            "chrome_alive": True,
        }

    def open_module(self, module_name, xpath, log=print):
        self.calls.append(("open_module", module_name, xpath))
        return {"ok": True, "code": "MODULE_OPENED", "message": module_name}

    def open_sale_asn_new(self, xpath, log=print):
        self.calls.append(("open_sale_asn_new", xpath))
        return {"ok": True, "code": "SALE_ASN_NEW_READY", "message": "ready"}

    def search_oc_list(self, xpath, filter_kind, query, log=print):
        self.calls.append(("search_oc", xpath, filter_kind, query))
        return {"ok": True, "code": "MODULE_SEARCH_APPLIED", "message": "found"}

    def open_oc_revision_report(self, log=print):
        self.calls.append(("open_oc_revision_report",))
        return {
            "ok": True,
            "code": "OC_REVISION_REPORT_READY",
            "message": "ready",
        }

    def upload_oc_edi(self, path, buyer, mode, log=print):
        upload_path = Path(path)
        self.calls.append(
            ("upload_oc_edi", upload_path.is_file(), buyer, mode)
        )
        return {
            "ok": True,
            "code": "OC_TRANSACTION_CREATED",
            "message": "created",
        }

    def search_sample_list(self, xpath, filter_kind, query, log=print):
        self.calls.append(("search_sample", xpath, filter_kind, query))
        return {"ok": True, "code": "MODULE_SEARCH_APPLIED", "message": "found"}

    def open_sample_new(self, xpath, log=print):
        self.calls.append(("open_sample_new", xpath))
        return {"ok": True, "code": "SAMPLE_NEW_READY", "message": "ready"}

    def search_sale_asn_list(self, xpath, filter_kind, query, log=print):
        self.calls.append(("search_sale_asn", xpath, filter_kind, query))
        return {"ok": True, "code": "MODULE_SEARCH_APPLIED", "message": "found"}

    def search_rmpo_list(self, xpath, supplier, order_no, log=print):
        self.calls.append(("search_rmpo", xpath, supplier, order_no))
        return {"ok": True, "code": "MODULE_SEARCH_APPLIED", "message": "found"}

    def search_indent_list(
        self, xpath, module_name, supplier, article, indent_no, style, log=print
    ):
        self.calls.append((
            "search_indent",
            xpath,
            module_name,
            supplier,
            article,
            indent_no,
            style,
        ))
        return {"ok": True, "code": "MODULE_SEARCH_APPLIED", "message": "found"}

    def open_module_new(self, module_id, log=print):
        self.calls.append(("open_module_new", module_id))
        return {"ok": True, "code": "MODULE_NEW_READY", "message": "ready"}

    def toggle_company_foc(self, xpath, log=print):
        self.calls.append(("toggle_company_foc", xpath))
        return {
            "ok": True,
            "code": "COMPANY_FOC_CHANGED",
            "message": "Đổi FOC thành công. Trạng thái hiện tại: FOC cho ASN.",
            "foc_enabled": True,
            "foc_mode": "FOC cho ASN",
            "saved": True,
        }

    def open_supplier_category(self, xpath, category_name, category_value, log=print):
        self.calls.append((
            "open_supplier_category", xpath, category_name, category_value
        ))
        return {"ok": True, "code": "SUPPLIER_CATEGORY_READY", "message": "ready"}

    def find_supplier_across_categories(self, xpath, categories, query, log=print):
        self.calls.append(("find_supplier", xpath, dict(categories), query))
        return {"ok": True, "code": "SUPPLIER_FOUND", "message": "found"}

    def find_supplier_in_category(
        self, xpath, category_name, category_value, query, log=print
    ):
        self.calls.append((
            "find_supplier_in_category",
            xpath,
            category_name,
            category_value,
            query,
        ))
        return {"ok": True, "code": "SUPPLIER_FOUND", "message": "found"}

    def find_and_open_buyer(self, xpath, query, log=print):
        self.calls.append(("find_buyer", xpath, query))
        return {"ok": True, "code": "BUYER_EDIT_OPENED", "message": "opened"}

    def switch_division(self, division_key, log=print):
        self.calls.append(("switch_division", division_key))
        return {
            "ok": True,
            "code": "DIVISION_CHANGED",
            "message": "changed",
            "current_division": division_key,
            "division_label": division_key.upper(),
            "division_name": division_key,
        }

    def set_catalog_category(self, category_name, category_value, log=print):
        self.calls.append(("set_catalog_category", category_name, category_value))
        return {"ok": True, "code": "CATEGORY_SELECTED", "message": category_name}

    def prepare_catalog_master(self, category_name, category_value, log=print):
        self.calls.append(("prepare_catalog_master", category_name, category_value))
        return {
            "ok": True,
            "code": "CATEGORY_SELECTED",
            "message": category_name,
            "category": category_name,
            "value": category_value,
        }

    def scan_catalog_folders(self, category_name, category_value, log=print):
        self.calls.append(("scan_catalog_folders", category_name, category_value))
        return {
            "ok": True,
            "code": "CATALOG_FOLDERS_SCANNED",
            "message": "scanned",
            "category": category_name,
            "value": category_value,
            "folders": [
                {
                    "node_id": "101",
                    "node_code": "22_1",
                    "name": "DEV",
                    "path": ["KNIT", "DEV"],
                    "path_label": "KNIT / DEV",
                    "kind": "group",
                    "depth": 2,
                }
            ],
        }

    def open_catalog_folder(
        self, category_name, category_value, node_id, log=print
    ):
        self.calls.append(
            ("open_catalog_folder", category_name, category_value, node_id)
        )
        name = "Master" if not node_id else "KNIT / DEV"
        return {
            "ok": True,
            "code": "CATALOG_FOLDER_OPENED",
            "message": name,
            "category": category_name,
            "value": category_value,
            "folder": {
                "node_id": node_id,
                "path_label": name,
            },
        }

    def quick_find_catalog(self, category_name, category_value, filter_kind, query,
                           user_id, password, company_id="psh", log=print, destination=None):
        self.calls.append(("quick_find_catalog", category_name, category_value,
                           filter_kind, query, user_id, password, destination))
        return {
            "ok": True,
            "code": "RESULT_OPENED",
            "message": query,
            "codes": [query],
            "article_code": query,
            "category": category_name,
            "filter_kind": filter_kind,
            "query": query,
        }

    def find_in_open_catalog(self, category_name, filter_kind, query, log=print):
        self.calls.append(
            ("find_in_open_catalog", category_name, filter_kind, query)
        )
        return {
            "ok": True,
            "code": "RESULT_OPENED",
            "message": query,
            "codes": [query],
            "article_code": query,
            "category": category_name,
            "filter_kind": filter_kind,
            "query": query,
        }

    def open_catalog_destination(self, article_code, destination, log=print):
        self.calls.append(
            ("open_catalog_destination", article_code, destination)
        )
        return {
            "ok": True,
            "code": "CATALOG_DESTINATION_OPENED",
            "message": destination,
            "article_code": article_code,
            "destination": destination,
        }

    def scan_catalog_files(self, article_code, log=print):
        self.calls.append(("scan_catalog_files", article_code))
        return {
            "ok": True,
            "code": "CATALOG_FILES_SCANNED",
            "message": "1 file",
            "article_code": article_code,
            "file_count": 1,
            "sections": [
                {
                    "index": 5,
                    "name": "Techpack",
                    "available": True,
                    "file_count": 1,
                }
            ],
            "files": [
                {
                    "section": "Techpack",
                    "section_index": 5,
                    "file_name": "jacket.pdf",
                    "comments": "Final",
                    "uploaded_on": "17 Jun 2024",
                    "uploaded_by": "HR",
                    "download_url": (
                        "https://prosports.worldfashionexchange.com/"
                        "Company/77400/Documents/jacket.pdf"
                    ),
                }
            ],
        }

    def download_catalog_file(self, file_info, log=print):
        self.calls.append(
            (
                "download_catalog_file",
                file_info["file_name"],
                file_info["download_url"],
            )
        )
        return {
            "ok": True,
            "code": "CATALOG_FILE_DOWNLOADED",
            "message": "downloaded",
            "file_name": file_info["file_name"],
            "download_path": f"C:/Downloads/{file_info['file_name']}",
        }

    def scan_open_costing(
        self,
        article_code,
        style_status=None,
        require_open=True,
        scan_details=False,
        scan_article_options=False,
        log=print,
    ):
        del style_status, scan_details, log
        self.calls.append(("scan_open_costing", article_code))
        result = {
            "ok": True,
            "code": "COSTING_SCANNED",
            "message": "scanned",
            "costing": {
                "format_version": costing_workbook.FORMAT_VERSION,
                "style_code": article_code,
                "title": "Current",
                "cost_sheet_status": "Open",
                "cost_sheet_type": "Internal Cost Sheets",
                "order_execution_type": "Trading",
                "season": "SS27",
                "template": "FOB",
                "signature": "fake-live",
                "sections": [
                    {
                        "section_key": "fabric",
                        "name": "Fabric",
                        "row_order": 1,
                    }
                ],
                "items": [
                    {
                        "section_key": "fabric",
                        "section_name": "Fabric",
                        "item_key": "fabric-1",
                        "row_order": 1,
                        "action": "UPSERT",
                        "item_type": "article",
                        "article_code": "FAB-001",
                        "article_name": "Jersey",
                    }
                ],
                "fields": [
                    {
                        "scope": "item",
                        "section_key": "fabric",
                        "item_key": "fabric-1",
                        "field_key": "colConsQty",
                        "label": "Cons. Qty.",
                        "value": 1.25,
                        "data_type": "number",
                        "editable": True,
                        "required": False,
                        "options": [],
                        "row_order": 1,
                    }
                ],
            },
        }
        if scan_article_options:
            result["costing"]["sections"][0].update(
                {
                    "article_code_options": ["FAB-001", "FAB-002"],
                    "article_name_options": ["Jersey", "Rib"],
                }
            )
        return result

    def scan_active_open_costing(
        self,
        require_open=True,
        scan_details=False,
        scan_article_options=False,
        log=print,
    ):
        del scan_details, log
        self.calls.append(("scan_active_open_costing", require_open))
        return self.scan_open_costing(
            "ACTIVE0001",
            require_open=require_open,
            scan_article_options=scan_article_options,
        )

    def inspect_active_costing(self, log=print):
        del log
        self.calls.append(("inspect_active_costing",))
        return {
            "ok": True,
            "code": "COSTING_CONTEXT_INSPECTED",
            "message": "inspected",
            "article_code": "ACTIVE0001",
            "costing_status": "Open",
            "style_status": {
                "code": "ACTIVE0001",
                "season": "",
                "internal_costsheet_status": "Open",
            },
        }

    def apply_costing_plan(
        self,
        article_code,
        plan,
        source_document=None,
        article_resolutions=None,
        active_tab_only=False,
        log=print,
    ):
        if active_tab_only:
            self.calls.append(("apply_active_tab_only", article_code))
        self.calls.append(
            (
                "apply_costing_plan",
                article_code,
                len(plan.get("fields_to_set") or ()),
                bool(source_document),
                dict(article_resolutions or {}),
            )
        )
        return {
            "ok": True,
            "code": "COSTING_APPLIED",
            "message": "saved",
            "article_code": article_code,
            "applied_count": len(plan.get("fields_to_set") or ()),
            "verified": True,
        }


def make_api(tmp_path):
    fake = FakeLogin()
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)
    return api, fake


def test_stop_cancels_running_flow_at_safe_checkpoint(tmp_path):
    entered = threading.Event()

    class SlowLogin(FakeLogin):
        def open_module(self, module_name, xpath, log=print):
            entered.set()
            while True:
                automation_runtime.checkpoint()
                time.sleep(0.01)

    api = PanelAPI(
        login_module=SlowLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )
    results = []
    caller = threading.Thread(
        target=lambda: results.append(api.open_module("0004_0050_0020"))
    )
    caller.start()
    assert entered.wait(timeout=1)

    requested = api.cancel_current_action()
    caller.join(timeout=2)
    api.shutdown()

    assert requested["code"] == "CANCEL_REQUESTED"
    assert results[0]["code"] == "ACTION_CANCELLED"
    assert "checkpoint an toàn" in results[0]["message"]


def test_find_code_uses_the_prepared_catalog_grid(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.prepare_catalog("Apparel")
    result = api.find_code("Apparel", "ABC123")
    assert result["code"] == "RESULT_OPENED"
    assert (
        "find_in_open_catalog",
        "Apparel",
        "code",
        "ABC123",
    ) in fake.calls
    assert not any(call[0] == "quick_find_catalog" for call in fake.calls)


def test_find_buyer_reference_uses_buyer_kind(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.prepare_catalog("Apparel")
    api.find_buyer_reference("Apparel", "PO-9")
    assert (
        "find_in_open_catalog",
        "Apparel",
        "buyer_reference",
        "PO-9",
    ) in fake.calls


def test_non_apparel_catalog_search_uses_article_name(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.prepare_catalog("Trims")

    result = api.catalog_action("Trims", "article_name", "Zipper", None)
    invalid = api.catalog_action("Trims", "buyer_reference", "BUY-1", None)

    assert result["code"] == "RESULT_OPENED"
    assert (
        "find_in_open_catalog",
        "Trims",
        "article_name",
        "Zipper",
    ) in fake.calls
    assert invalid["code"] == "INVALID_FILTER"


def test_catalog_destination_reuses_current_search_without_reopening_catalog(
    tmp_path,
):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)

    api.prepare_catalog("Apparel")
    fake.calls.clear()
    found = api.find_code("Apparel", "ABC123")
    opened = api.open_catalog_destination("costsheet", found["article_code"])

    assert opened["code"] == "CATALOG_DESTINATION_OPENED"
    assert fake.calls == [
        (
            "find_in_open_catalog",
            "Apparel",
            "code",
            "ABC123",
        ),
        ("open_catalog_destination", "ABC123", "costsheet"),
    ]


def test_catalog_destination_requires_a_current_search_result(tmp_path):
    api, fake = make_api(tmp_path)

    result = api.open_catalog_destination("bom", "ABC123")

    assert result["code"] == "CATALOG_RESULT_REQUIRED"
    assert fake.calls == []


def test_catalog_destination_rejects_a_stale_style(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.prepare_catalog("Apparel")
    api.find_code("Apparel", "ABC123")

    result = api.open_catalog_destination("bom", "OTHER")

    assert result["code"] == "CATALOG_RESULT_CHANGED"
    assert not any(call[0] == "open_catalog_destination" for call in fake.calls)


def test_catalog_search_auto_prepares_master_without_list_step(tmp_path):
    api, fake = make_api(tmp_path)

    result = api.find_code("Apparel", "ABC123")

    assert result["code"] == "RESULT_OPENED"
    assert fake.calls == [
        ("prepare_catalog_master", "Apparel", "01"),
        ("find_in_open_catalog", "Apparel", "code", "ABC123"),
    ]


def test_non_catalog_flow_invalidates_prepared_master_context(tmp_path):
    api, fake = make_api(tmp_path)
    api.prepare_catalog("Apparel")
    fake.calls.clear()

    opened = api.open_module("0004_0050_0020")
    assert opened["code"] == "MODULE_OPENED"
    fake.calls.clear()

    result = api.find_code("Apparel", "ABC123")

    assert result["code"] == "RESULT_OPENED"
    assert fake.calls == [
        ("prepare_catalog_master", "Apparel", "01"),
        ("find_in_open_catalog", "Apparel", "code", "ABC123"),
    ]


def test_catalog_direct_costing_is_one_user_action(tmp_path):
    api, fake = make_api(tmp_path)

    result = api.catalog_action("Apparel", "code", "ABC123", "costsheet")

    assert result["code"] == "CATALOG_DESTINATION_OPENED"
    assert fake.calls == [
        ("prepare_catalog_master", "Apparel", "01"),
        ("find_in_open_catalog", "Apparel", "code", "ABC123"),
        ("open_catalog_destination", "ABC123", "costsheet"),
    ]


def test_catalog_direct_costing_prefers_atomic_search_and_popup_open(tmp_path):
    api, fake = make_api(tmp_path)

    def combined(category, kind, query, destination, log=print):
        del log
        fake.calls.append(
            (
                "find_and_open_catalog_destination",
                category,
                kind,
                query,
                destination,
            )
        )
        return {
            "ok": True,
            "code": "CATALOG_DESTINATION_OPENED",
            "message": "opened atomically",
            "article_code": query,
            "style_status": {
                "code": query,
                "season": "SS27",
                "internal_costsheet_status": "Open",
            },
            "category": category,
            "filter_kind": kind,
            "query": query,
            "destination": destination,
        }

    fake.find_and_open_catalog_destination = combined

    result = api.catalog_action(
        "Apparel",
        "code",
        "ABC123",
        "costsheet",
    )

    assert result["code"] == "CATALOG_DESTINATION_OPENED"
    assert fake.calls == [
        ("prepare_catalog_master", "Apparel", "01"),
        (
            "find_and_open_catalog_destination",
            "Apparel",
            "code",
            "ABC123",
            "costsheet",
        ),
    ]


def test_catalog_direct_destination_reuses_matching_popup(tmp_path):
    api, fake = make_api(tmp_path)
    api.catalog_action("Apparel", "code", "ABC123", None)
    fake.calls.clear()

    result = api.catalog_action("Apparel", "code", "ABC123", "bom")

    assert result["code"] == "CATALOG_DESTINATION_OPENED"
    assert fake.calls == [
        ("open_catalog_destination", "ABC123", "bom"),
    ]


def test_catalog_costing_export_scans_open_style_and_writes_xlsx(tmp_path):
    api, fake = make_api(tmp_path)
    target = tmp_path / "SWN0000001-Costing.xlsx"

    result = api.export_catalog_costing(
        "Apparel",
        "code",
        "SWN0000001",
        str(target),
    )

    assert result["code"] == "COSTING_EXPORTED"
    assert result["file_name"] == target.name
    assert target.is_file()
    loaded = costing_workbook.read_costing_xlsx(target)
    assert loaded["style_code"] == "SWN0000001"
    assert ("scan_open_costing", "SWN0000001") in fake.calls


def test_catalog_costing_export_without_query_scans_only_active_tab(tmp_path):
    api, fake = make_api(tmp_path)
    target = tmp_path / "Current-Style-Costing.xlsx"

    result = api.export_catalog_costing(
        "Apparel",
        "code",
        "",
        str(target),
    )

    assert result["code"] == "COSTING_EXPORTED"
    assert result["article_code"] == "ACTIVE0001"
    assert target.is_file()
    assert ("scan_active_open_costing", False) in fake.calls


def test_catalog_costing_article_dropdown_scan_is_opt_in_and_cached(tmp_path):
    api, _fake = make_api(tmp_path)
    api._account = lambda: {"user_id": "alice"}
    scanned_target = tmp_path / "scanned-dropdowns.xlsx"
    cached_target = tmp_path / "cached-dropdowns.xlsx"

    scanned = api.export_catalog_costing(
        "Apparel",
        "code",
        "",
        str(scanned_target),
        True,
    )
    cached = api.export_catalog_costing(
        "Apparel",
        "code",
        "",
        str(cached_target),
        False,
    )

    assert scanned["article_option_source"] == "scan"
    assert scanned["article_option_count"] == 2
    assert cached["article_option_source"] == "cache"
    assert cached["article_option_count"] == 2
    assert (tmp_path / "costing-article-options.json").is_file()


def test_server_article_library_drives_dropdowns_and_code_suggestions(tmp_path):
    write_json_atomic(
        tmp_path / "article-library.json",
        {
            "schema_version": article_library.SCHEMA_VERSION,
            "remote_version": "20260731",
            "generated_at": "2026-07-31T08:00:00Z",
            "synced_at": 100.0,
            "sha256": "a" * 64,
            "sections": [
                {
                    "section_key": "*",
                    "section_name": "All Categories",
                    "options": [
                        {
                            "article_code": "F0001",
                            "article_name": "Cotton Jersey",
                            "buyer_reference": "",
                            "article_category": "Textiles/Fabric",
                        },
                        {
                            "article_code": "T0002",
                            "article_name": "Metal Zipper",
                            "buyer_reference": "",
                            "article_category": "Trims",
                        },
                        {
                            "article_code": "SWN0001",
                            "article_name": "Cotton Dress",
                            "buyer_reference": "BUY-COTTON",
                            "article_category": "Apparel",
                        },
                    ],
                }
            ],
        },
    )
    api, _fake = make_api(tmp_path)
    target = tmp_path / "server-dropdowns.xlsx"

    suggestions = api.suggest_articles(
        "Apparel",
        "buyer_reference",
        "buy-c",
    )
    exported = api.export_catalog_costing(
        "Apparel",
        "code",
        "",
        str(target),
    )

    assert suggestions["suggestions"] == [
        {
            "article_code": "SWN0001",
            "article_name": "Cotton Dress",
            "buyer_reference": "BUY-COTTON",
            "article_category": "Apparel",
            "value": "BUY-COTTON",
        }
    ]
    assert exported["article_option_source"] == "server"
    assert exported["article_option_count"] == 1
    assert api.get_initial_state()["article_library"]["available"] is True


def test_catalog_costing_inspection_reads_active_tab_before_file_dialog(tmp_path):
    api, fake = make_api(tmp_path)

    result = api.inspect_active_catalog_costing("Apparel")

    assert result["code"] == "COSTING_CONTEXT_INSPECTED"
    assert result["article_code"] == "ACTIVE0001"
    assert result["style_status"]["internal_costsheet_status"] == "Open"
    assert ("inspect_active_costing",) in fake.calls


def test_catalog_costing_file_validation_does_not_scan_wfx(tmp_path):
    api, fake = make_api(tmp_path)
    target = tmp_path / "valid.xlsx"
    costing_workbook.write_costing_xlsx(
        fake.scan_open_costing("SWN0000001")["costing"],
        target,
    )
    fake.calls.clear()

    result = api.validate_catalog_costing_file(str(target))

    assert result["code"] == "COSTING_FILE_VALID"
    assert result["style_code"] == "SWN0000001"
    assert result["validation_errors"] == []
    assert fake.calls == []
    assert not any(
        call[0] in {
            "prepare_catalog_master",
            "find_in_open_catalog",
            "open_catalog_destination",
        }
        for call in fake.calls
    )


def test_catalog_costing_file_actions_reuse_open_costing_without_reopening(
    tmp_path,
):
    api, fake = make_api(tmp_path)
    opened = api.catalog_action(
        "Apparel",
        "code",
        "SWN0000001",
        "costsheet",
    )
    assert opened["code"] == "CATALOG_DESTINATION_OPENED"
    fake.calls.clear()

    target = tmp_path / "SWN0000001-Costing.xlsx"
    exported = api.export_catalog_costing(
        "Apparel",
        "code",
        "SWN0000001",
        str(target),
    )
    prepared = api.prepare_catalog_costing_import(
        "Apparel",
        "code",
        "SWN0000001",
        str(target),
    )
    applied = api.apply_catalog_costing(prepared["plan_token"], {})

    assert exported["code"] == "COSTING_EXPORTED"
    assert prepared["code"] == "COSTING_DRY_RUN_READY"
    assert applied["code"] == "COSTING_APPLIED"
    assert not any(
        call[0] == "open_catalog_destination" for call in fake.calls
    )


def test_catalog_costing_import_returns_dry_run_token_without_writing(tmp_path):
    api, fake = make_api(tmp_path)
    source = tmp_path / "SWN0000001-edit.xlsx"
    scanned = fake.scan_open_costing("SWN0000001")["costing"]
    scanned["fields"][0]["value"] = "Edited"
    scanned["title"] = "Edited"
    costing_workbook.write_costing_xlsx(scanned, source)
    fake.calls.clear()

    result = api.prepare_catalog_costing_import(
        "Apparel",
        "code",
        "SWN0000001",
        str(source),
    )

    assert result["code"] == "COSTING_DRY_RUN_READY"
    assert len(result["plan_token"]) == 32
    assert result["counts"]["fields_to_set"] == 1
    assert ("scan_open_costing", "SWN0000001") in fake.calls
    assert not any(call[0].startswith("apply") for call in fake.calls)

    cleared = api.clear_catalog_costing_plan(result["plan_token"])
    assert cleared["code"] == "COSTING_PLAN_CLEARED"


def test_catalog_costing_import_without_query_uses_active_tab_through_apply(
    tmp_path,
):
    api, fake = make_api(tmp_path)
    source = tmp_path / "active-edit.xlsx"
    document = fake.scan_open_costing("ACTIVE0001")["costing"]
    costing_workbook.write_costing_xlsx(document, source)
    fake.calls.clear()

    prepared = api.prepare_catalog_costing_import(
        "Apparel",
        "code",
        "",
        str(source),
    )
    applied = api.apply_catalog_costing(prepared["plan_token"], {})

    assert prepared["code"] == "COSTING_DRY_RUN_READY"
    assert prepared["article_code"] == "ACTIVE0001"
    assert applied["code"] == "COSTING_APPLIED"
    assert ("scan_active_open_costing", True) in fake.calls
    assert ("apply_active_tab_only", "ACTIVE0001") in fake.calls
    assert not any(
        call[0] in {
            "prepare_catalog_master",
            "find_in_open_catalog",
            "open_catalog_destination",
        }
        for call in fake.calls
    )


def test_catalog_costing_apply_uses_cached_server_side_plan_once(tmp_path):
    api, fake = make_api(tmp_path)
    source = tmp_path / "SWN0000001-edit.xlsx"
    scanned = fake.scan_open_costing("SWN0000001")["costing"]
    scanned["fields"][0]["value"] = "Edited"
    scanned["title"] = "Edited"
    costing_workbook.write_costing_xlsx(scanned, source)
    fake.calls.clear()
    prepared = api.prepare_catalog_costing_import(
        "Apparel",
        "code",
        "SWN0000001",
        str(source),
    )

    applied = api.apply_catalog_costing(prepared["plan_token"], {})
    repeated = api.apply_catalog_costing(prepared["plan_token"], {})

    assert applied["code"] == "COSTING_APPLIED"
    assert repeated["code"] == "COSTING_PLAN_EXPIRED"
    assert ("apply_costing_plan", "SWN0000001", 1, True, {}) in fake.calls


def test_catalog_costing_import_blocks_file_for_other_style(tmp_path):
    api, fake = make_api(tmp_path)
    source = tmp_path / "other.xlsx"
    scanned = fake.scan_open_costing("OTHER")["costing"]
    costing_workbook.write_costing_xlsx(scanned, source)
    fake.calls.clear()

    result = api.prepare_catalog_costing_import(
        "Apparel",
        "code",
        "SWN0000001",
        str(source),
    )

    assert result["code"] == "COSTING_STYLE_MISMATCH"
    assert fake.calls == []


def test_catalog_folder_scan_and_default_are_scoped_to_user(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api, fake = make_api(tmp_path)

    scanned = api.scan_catalog_folders("Apparel")
    saved = api.set_catalog_default_folder("Apparel", "101")

    assert scanned["code"] == "CATALOG_FOLDERS_SCANNED"
    assert saved["default_folder"]["path_label"] == "KNIT / DEV"
    stored = prefs.load_prefs(base_dir=tmp_path)["catalog_default_folder"]
    assert stored["node_id"] == "101"
    assert stored["user_id"] == "alice"
    assert (
        "scan_catalog_folders",
        "Apparel",
        "01",
    ) in fake.calls

    prefs.save_account("bob", "pw", base_dir=tmp_path)
    other_api, _ = make_api(tmp_path)
    assert other_api.get_initial_state()["catalog_default_folder"] is None


def test_catalog_folder_tree_is_reused_across_app_restarts(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    first_api, first_fake = make_api(tmp_path)
    first = first_api.scan_catalog_folders("Apparel")
    assert first["code"] == "CATALOG_FOLDERS_SCANNED"
    assert any(call[0] == "scan_catalog_folders" for call in first_fake.calls)

    restarted_api, restarted_fake = make_api(tmp_path)
    cached = restarted_api.scan_catalog_folders("Apparel")

    assert cached["code"] == "CATALOG_FOLDERS_CACHED"
    assert cached["folders"] == first["folders"]
    assert not any(
        call[0] == "scan_catalog_folders"
        for call in restarted_fake.calls
    )


def test_catalog_folder_refresh_forces_a_new_scan(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.scan_catalog_folders("Apparel")
    fake.calls.clear()

    refreshed = api.scan_catalog_folders("Apparel", True)

    assert refreshed["code"] == "CATALOG_FOLDERS_SCANNED"
    assert (
        "scan_catalog_folders",
        "Apparel",
        "01",
    ) in fake.calls


def test_browse_catalog_opens_saved_folder_without_master_search(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.scan_catalog_folders("Apparel")
    api.set_catalog_default_folder("Apparel", "101")
    fake.calls.clear()

    result = api.browse_catalog("Apparel")

    assert result["code"] == "CATALOG_FOLDER_OPENED"
    assert fake.calls == [
        ("open_catalog_folder", "Apparel", "01", "101"),
    ]
    assert not any(
        call[0] in {"prepare_catalog_master", "find_in_open_catalog"}
        for call in fake.calls
    )


def test_master_is_always_available_as_catalog_default(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api, fake = make_api(tmp_path)

    saved = api.set_catalog_default_folder("Apparel", "")
    result = api.browse_catalog("Apparel")

    assert saved["default_folder"]["path_label"] == "Master"
    assert ("open_catalog_folder", "Apparel", "01", "") in fake.calls
    assert result["code"] == "CATALOG_FOLDER_OPENED"


def test_stale_catalog_default_falls_back_to_master(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.scan_catalog_folders("Apparel")
    api.set_catalog_default_folder("Apparel", "101")
    calls = []

    def opener(category_name, category_value, node_id, log=print):
        calls.append(node_id)
        if node_id:
            return {
                "ok": False,
                "code": "CATALOG_FOLDER_STALE",
                "message": "stale",
            }
        return {
            "ok": True,
            "code": "CATALOG_FOLDER_OPENED",
            "message": "Master",
        }

    fake.open_catalog_folder = opener
    result = api.browse_catalog("Apparel")

    assert calls == ["101", ""]
    assert result["code"] == "CATALOG_FOLDER_FALLBACK"
    stored = prefs.load_prefs(base_dir=tmp_path)["catalog_default_folder"]
    assert stored["node_id"] == ""
    assert stored["path_label"] == "Master"


def test_division_change_keeps_scanned_catalog_folder_cache(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api, _ = make_api(tmp_path)
    api.scan_catalog_folders("Apparel")

    api.switch_division("knit")
    saved = api.set_catalog_default_folder("Apparel", "101")

    assert saved["code"] == "CATALOG_DEFAULT_FOLDER_SAVED"
    assert saved["default_folder"]["node_id"] == "101"


def test_account_change_clears_scanned_catalog_folder_cache(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api, _ = make_api(tmp_path)
    api.scan_catalog_folders("Apparel")

    api.save_account("bob", "pw2")
    result = api.set_catalog_default_folder("Apparel", "101")

    assert result["code"] == "CATALOG_FOLDER_NOT_SCANNED"


def test_catalog_default_location_is_apparel_only_but_other_category_opens(
    tmp_path,
):
    api, fake = make_api(tmp_path)

    scanned = api.scan_catalog_folders("Trims")
    saved = api.set_catalog_default_folder("Trims", "")
    browsed = api.browse_catalog("Trims")

    for result in (scanned, saved):
        assert result["code"] == "CATALOG_DEFAULT_APPAREL_ONLY"
        assert "Apparel" in result["message"]
    assert browsed["code"] == "CATALOG_FOLDER_OPENED"
    assert ("open_catalog_folder", "Trims", "05", "") in fake.calls
    assert not any(call[0] == "scan_catalog_folders" for call in fake.calls)


def test_catalog_file_action_scans_four_sections_and_hides_download_url(
    tmp_path,
):
    api, fake = make_api(tmp_path)

    result = api.catalog_action("Apparel", "code", "ABC123", "files")

    assert result["code"] == "CATALOG_FILES_SCANNED"
    assert result["article_code"] == "ABC123"
    assert result["files"][0]["file_name"] == "jacket.pdf"
    assert result["files"][0]["section"] == "Techpack"
    assert result["files"][0]["uploaded_on"] == "17 Jun 2024"
    assert result["files"][0]["uploaded_by"] == "HR"
    assert result["files"][0]["comments"] == "Final"
    assert result["files"][0]["file_id"]
    assert "download_url" not in result["files"][0]
    assert fake.calls == [
        ("prepare_catalog_master", "Apparel", "01"),
        ("find_in_open_catalog", "Apparel", "code", "ABC123"),
        ("scan_catalog_files", "ABC123"),
    ]


def test_catalog_file_download_resolves_backend_token(tmp_path):
    api, fake = make_api(tmp_path)
    scanned = api.catalog_action("Apparel", "code", "ABC123", "files")
    fake.calls.clear()

    result = api.download_catalog_file(scanned["files"][0]["file_id"])

    assert result["code"] == "CATALOG_FILE_DOWNLOADED"
    assert result["file_name"] == "jacket.pdf"
    assert fake.calls == [
        (
            "download_catalog_file",
            "jacket.pdf",
            (
                "https://prosports.worldfashionexchange.com/"
                "Company/77400/Documents/jacket.pdf"
            ),
        )
    ]


def test_catalog_file_download_rejects_unknown_token(tmp_path):
    api, fake = make_api(tmp_path)

    result = api.download_catalog_file("expired")

    assert result["code"] == "CATALOG_FILE_EXPIRED"
    assert fake.calls == []


def test_open_module_builds_xpath(tmp_path):
    api, fake = make_api(tmp_path)
    api.open_module("0004_0050_0020")
    assert ("open_module", "OC List", '//*[@id="0004_0050_0020"]/a') in fake.calls


def test_oc_revision_report_delegates_to_automation(tmp_path):
    api, fake = make_api(tmp_path)

    result = api.open_oc_revision_report()

    assert result["code"] == "OC_REVISION_REPORT_READY"
    assert ("open_oc_revision_report",) in fake.calls


def test_upload_oc_requires_review_then_confirm_before_calling_edi(tmp_path):
    source = oc_workbook.write_oc_input_template(tmp_path / "new-oc.xlsx")
    workbook = load_workbook(source)
    workbook["OC INPUT"].append(
        [
            "J.LINDEBERG",
            "SS26",
            "Confirmed",
            "USD",
            "888 COMPANY LTD",
            "PO-1",
            "SWV0004581",
            "GMPA17697",
            "PO-1",
            "PO-1",
            "08-10-2025",
            "31-12-2025",
            "05-12-2025",
            "TT After Shipment 60 Days",
            "Sweden",
            "O127",
            "Forget-Me-Not",
            "M",
            23.65,
            9,
            "1",
            "FOB",
            0,
            None,
        ]
    )
    workbook.save(source)
    workbook.close()
    api, fake = make_api(tmp_path)

    review = api.review_oc_upload("new", str(source))

    assert review["code"] == "OC_UPLOAD_REVIEW_READY"
    assert review["buyer"] == "J.LINDEBERG"
    assert review["season"] == "SS26"
    assert review["po_count"] == 1
    assert review["style_count"] == 1
    assert review["total_units"] == 9
    assert review["row_count"] == 1
    assert not any(call[0] == "upload_oc_edi" for call in fake.calls)

    result = api.confirm_oc_upload(review["review_token"])

    assert result["code"] == "OC_TRANSACTION_CREATED"
    assert ("upload_oc_edi", True, "J.LINDEBERG", "new") in fake.calls
    assert list((tmp_path / "oc-upload-cache").iterdir()) == []

    repeated = api.confirm_oc_upload(review["review_token"])
    assert repeated["code"] == "OC_UPLOAD_REVIEW_EXPIRED"
    assert sum(call[0] == "upload_oc_edi" for call in fake.calls) == 1


def test_cancel_oc_review_removes_temp_file_without_calling_edi(tmp_path):
    source = oc_workbook.write_oc_input_template(tmp_path / "cancel-oc.xlsx")
    workbook = load_workbook(source)
    workbook["OC INPUT"].append(
        [
            "J.LINDEBERG", "SS26", "Confirmed", "USD", "888 COMPANY LTD",
            "PO-1", "ARTICLE-1", "STYLE-1", "PO-1", "PO-1", "08-10-2025",
            "31-12-2025", "05-12-2025", "TT After Shipment 60 Days", "Sweden", "BLACK",
            "Black", "M", 10, 5, "1", "FOB", 0, None,
        ]
    )
    workbook.save(source)
    workbook.close()
    api, fake = make_api(tmp_path)

    review = api.review_oc_upload("new", str(source))
    result = api.cancel_oc_upload_review(review["review_token"])

    assert result["code"] == "OC_UPLOAD_REVIEW_CANCELLED"
    assert list((tmp_path / "oc-upload-cache").iterdir()) == []
    assert not any(call[0] == "upload_oc_edi" for call in fake.calls)


def test_reselecting_same_oc_path_after_cancel_reads_modified_file(tmp_path):
    source = oc_workbook.write_oc_input_template(tmp_path / "same-name.xlsx")
    workbook = load_workbook(source)
    sheet = workbook["OC INPUT"]
    sheet.append(
        [
            "J.LINDEBERG", "SS26", "Confirmed", "USD", "888 COMPANY LTD",
            "PO-1", "ARTICLE-1", "STYLE-1", "PO-1", "PO-1", "08-10-2025",
            "31-12-2025", "05-12-2025", "TT After Shipment 60 Days", "Sweden", "BLACK",
            "Black", "M", 10, 5, "1", "FOB", 0, None,
        ]
    )
    workbook.save(source)
    workbook.close()
    api, fake = make_api(tmp_path)

    first = api.review_oc_upload("new", str(source))
    assert first["total_units"] == 5
    api.cancel_oc_upload_review(first["review_token"])

    workbook = load_workbook(source)
    sheet = workbook["OC INPUT"]
    units_column = oc_workbook.INPUT_HEADERS.index("Units") + 1
    data_row = next(
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "J.LINDEBERG"
    )
    sheet.cell(data_row, units_column).value = 17
    workbook.save(source)
    workbook.close()

    second = api.review_oc_upload("new", str(source))

    assert second["total_units"] == 17
    assert second["source_sha256"] != first["source_sha256"]
    assert not any(call[0] == "upload_oc_edi" for call in fake.calls)


def test_parallel_automation_is_rejected_instead_of_queuing(tmp_path):
    api, fake = make_api(tmp_path)
    started = threading.Event()
    release = threading.Event()
    completed = []

    def slow_open(module_name, xpath, log=print):
        started.set()
        release.wait(timeout=2)
        return {
            "ok": True,
            "code": "MODULE_OPENED",
            "message": module_name,
        }

    fake.open_module = slow_open
    worker = threading.Thread(
        target=lambda: completed.append(api.open_module("0004_0050_0020")),
    )
    worker.start()
    assert started.wait(timeout=1)

    overlapping = api.open_module("0004_0056_4070")
    release.set()
    worker.join(timeout=2)

    assert overlapping["code"] == "ACTION_IN_PROGRESS"
    assert completed[0]["code"] == "MODULE_OPENED"


def test_rejected_prepare_catalog_keeps_active_catalog_context(tmp_path):
    # Bug hồi quy: prepare_catalog/browse_catalog/scan_catalog_folders từng reset
    # prepared_category TRƯỚC khi giành run lock. Một lần bấm bị từ chối
    # ACTION_IN_PROGRESS vẫn xoá mất Catalog mà workflow đang chạy đã chuẩn bị,
    # khiến find_code sau đó báo CATALOG_PREPARE_REQUIRED oan.
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    api.prepare_catalog("Apparel")
    assert api._catalog.prepared_category == "Apparel"

    # Giữ lock từ MỘT THREAD KHÁC, đúng như một workflow khác đang chạy.
    # Không tự acquire trên thread test: _run_lock là RLock để composite Costing
    # giữ khoá xuyên nhiều _run, nên tự giữ rồi tự gọi chỉ là tái nhập hợp lệ
    # chứ không giả lập được tranh chấp.
    holding = threading.Event()
    release = threading.Event()

    def hold_lock():
        api._run_lock.acquire()
        holding.set()
        release.wait(timeout=5)
        api._run_lock.release()

    holder = threading.Thread(target=hold_lock)
    holder.start()
    try:
        assert holding.wait(timeout=2)
        rejected = api.prepare_catalog("Apparel")
        rejected_browse = api.browse_catalog("Apparel")
        rejected_scan = api.scan_catalog_folders("Apparel", True)
    finally:
        release.set()
        holder.join(timeout=5)

    assert rejected["code"] == "ACTION_IN_PROGRESS"
    assert rejected_browse["code"] == "ACTION_IN_PROGRESS"
    assert rejected_scan["code"] == "ACTION_IN_PROGRESS"
    # Context của workflow đang chạy phải còn nguyên vẹn.
    assert api._catalog.prepared_category == "Apparel"


def test_sale_asn_new_uses_new_menu_xpath(tmp_path):
    api, fake = make_api(tmp_path)
    result = api.open_sale_asn_new()
    assert result["code"] == "SALE_ASN_NEW_READY"
    assert (
        "open_sale_asn_new",
        '//*[@id="0004_0070_4340"]/a',
    ) in fake.calls


def test_open_chrome_uses_login_module(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    result = api.open_chrome()
    assert result["code"] == "LOGGED_IN"
    assert result["chrome_alive"] is True
    assert ("start_chrome",) in fake.calls
    assert ("run", "u", "p", "psh") in fake.calls
    assert "đăng nhập WFX" in result["message"]


def test_prepare_catalog_opens_then_selects(tmp_path):
    api, fake = make_api(tmp_path)
    api.prepare_catalog("Apparel")
    assert fake.calls == [("prepare_catalog_master", "Apparel", "01")]


def test_log_sink_receives_lines(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, fake = make_api(tmp_path)
    lines = []
    api.set_log_sink(lines.append)
    api.login()
    assert any("fake login" in line for line in lines)
    assert ("run", "u", "p", "psh") in fake.calls


def test_get_initial_state(tmp_path):
    prefs.save_account("bob", "pw", base_dir=tmp_path)
    prefs.save_prefs(base_dir=tmp_path, theme="dark")
    api, _ = make_api(tmp_path)
    state = api.get_initial_state()
    assert state["user_id"] == "bob"
    assert state["theme"] == "dark"
    assert state["hotkey_label"] == "Ctrl + Shift + X"
    assert state["has_credentials"] is True
    assert [item["key"] for item in state["divisions"]] == [
        "woven",
        "knit",
        "pssg",
    ]


def test_get_initial_state_reads_preferences_once(tmp_path, monkeypatch):
    real_load_prefs = prefs.load_prefs
    calls = []

    def counted_load_prefs(base_dir=None):
        calls.append(base_dir)
        return real_load_prefs(base_dir=base_dir)

    monkeypatch.setattr(prefs, "load_prefs", counted_load_prefs)
    api, _ = make_api(tmp_path)

    api.get_initial_state()

    assert calls == [tmp_path]


def test_initial_state_requires_credentials_when_account_is_empty(tmp_path):
    api, _ = make_api(tmp_path)
    assert api.get_initial_state()["has_credentials"] is False


def test_switch_division_delegates_and_tracks_highlight_state(tmp_path):
    api, fake = make_api(tmp_path)
    api._session_active = True
    result = api.switch_division("knit")
    assert result["code"] == "DIVISION_CHANGED"
    assert result["current_division"] == "knit"
    assert ("switch_division", "knit") in fake.calls
    assert api.get_status()["current_division"] == "knit"


def test_save_account_persists(tmp_path):
    api, _ = make_api(tmp_path)
    api.save_account("carol", "s3cret")
    assert prefs.load_account(base_dir=tmp_path)["user_id"] == "carol"


def test_save_account_blank_password_keeps_existing(tmp_path):
    # Password input trên UI không bao giờ được điền lại khi mở Settings, nên
    # nếu người dùng chỉ đổi User ID rồi lưu, password gửi lên sẽ luôn rỗng.
    # Không được ghi đè mật khẩu đã lưu bằng chuỗi rỗng.
    api, _ = make_api(tmp_path)
    api.save_account("dave", "s3cret")
    result = api.save_account("dave2", "")
    assert result["ok"] is True
    assert result["code"] == "ACCOUNT_SAVED"
    loaded = prefs.load_account(base_dir=tmp_path)
    assert loaded == {"user_id": "dave2", "password": "s3cret"}


def test_save_account_blank_password_and_no_stored_password_fails(tmp_path):
    api, _ = make_api(tmp_path)
    result = api.save_account("erin", "   ")
    assert result["ok"] is False
    assert result["code"] == "PASSWORD_REQUIRED"
    assert "mật khẩu" in result["message"].lower()
    # Không được ghi bất cứ thứ gì xuống .env khi từ chối lưu.
    assert prefs.load_account(base_dir=tmp_path) == {"user_id": "", "password": ""}


def test_save_account_requires_user_id(tmp_path):
    api, _ = make_api(tmp_path)
    result = api.save_account("  ", "secret")
    assert result["ok"] is False
    assert result["code"] == "USER_ID_REQUIRED"
    assert prefs.load_account(base_dir=tmp_path) == {"user_id": "", "password": ""}


def test_result_sink_receives_method_result_and_elapsed(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, _ = make_api(tmp_path)
    seen = []
    api.set_result_sink(
        lambda method, result, elapsed: seen.append((method, result, elapsed))
    )
    api.prepare_catalog("Apparel")
    seen.clear()
    api.find_code("Apparel", "ABC123")
    assert len(seen) == 1
    method, result, elapsed = seen[0]
    assert method == "find_code"
    assert result["code"] == "RESULT_OPENED"
    assert isinstance(elapsed, float) and elapsed >= 0


def test_result_sink_receives_early_validation_result(tmp_path):
    api, _ = make_api(tmp_path)
    seen = []
    api.set_result_sink(
        lambda method, result, elapsed: seen.append((method, result, elapsed))
    )
    result = api.open_module("not-a-module")
    assert result["code"] == "MODULE_UNKNOWN"
    assert len(seen) == 1
    assert seen[0][0] == "open_module"
    assert seen[0][1]["code"] == "MODULE_UNKNOWN"


def test_session_state_tracks_result_codes(tmp_path):
    prefs.save_account("u", "p", base_dir=tmp_path)
    api, _ = make_api(tmp_path)
    assert api.get_status()["session_active"] is None

    api.login()
    status_after_login = api.get_status()
    assert status_after_login["session_active"] is True
    assert status_after_login["last_login_at"] is not None

    api.check_session()
    assert api.get_status()["session_active"] is False


def test_user_action_auto_relogs_once_and_retries(tmp_path):
    fake = FakeLogin()
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)
    attempts = []

    def action():
        attempts.append(True)
        if len(attempts) == 1:
            return {
                "ok": False,
                "code": "NOT_LOGGED_IN",
                "message": "expired",
            }
        return {"ok": True, "code": "MODULE_OPENED", "message": "ready"}

    result = api._run("open_module", action)

    assert result["code"] == "MODULE_OPENED"
    assert len(attempts) == 2
    assert ("run", "alice", "pw", "psh") in fake.calls
    assert api.get_status()["session_active"] is True


def test_background_session_maintenance_relogs_expired_session(tmp_path):
    fake = FakeLogin()
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)
    api._session_active = True

    assert api.should_maintain_session() is True
    result = api.maintain_session()

    assert result["code"] == "SESSION_RESTORED"
    assert fake.calls[:2] == [
        ("check_session",),
        ("run", "alice", "pw", "psh"),
    ]
    assert api.get_status()["session_active"] is True


def test_background_session_maintenance_stays_off_before_first_login(tmp_path):
    prefs.save_account("alice", "pw", base_dir=tmp_path)
    api = PanelAPI(
        login_module=FakeLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )

    assert api.should_maintain_session() is False


def test_unknown_result_code_leaves_session_state_untouched(tmp_path):
    api, _ = make_api(tmp_path)
    api._session_active = True
    api._observe("whatever", {"ok": False, "code": "SOMETHING_NEW"}, 0.1)
    assert api.get_status()["session_active"] is True


def test_set_hotkey_rejects_unsafe_combo(tmp_path):
    api, _ = make_api(tmp_path)
    result = api.set_hotkey("ctrl+backspace")
    assert result["ok"] is False
    assert result["code"] == "HOTKEY_INVALID"
    assert prefs.load_prefs(base_dir=tmp_path)["hotkey"] == "ctrl+shift+x"


def test_set_hotkey_applies_and_persists(tmp_path):
    api, _ = make_api(tmp_path)
    applied = []
    api.set_hotkey_applier(lambda spec: applied.append(spec) or None)
    result = api.set_hotkey("Alt+Shift+K")
    assert result["ok"] is True
    assert result["hotkey"] == "alt+shift+k"
    assert applied == ["alt+shift+k"]
    assert prefs.load_prefs(base_dir=tmp_path)["hotkey"] == "alt+shift+k"


def test_set_hotkey_accepts_browser_event_payload(tmp_path):
    api, _ = make_api(tmp_path)
    result = api.set_hotkey(
        {
            "ctrl": True,
            "alt": True,
            "shift": False,
            "meta": False,
            "key": "J",
            "code": "KeyJ",
        }
    )
    assert result["ok"] is True
    assert result["hotkey"] == "ctrl+alt+j"


def test_set_hotkey_rolls_back_when_registration_fails(tmp_path):
    api, _ = make_api(tmp_path)
    attempts = []

    def applier(spec):
        attempts.append(spec)
        return (
            "Phím đang bị ứng dụng khác chiếm."
            if spec == "alt+shift+k"
            else None
        )

    api.set_hotkey_applier(applier)
    result = api.set_hotkey("alt+shift+k")
    assert result["ok"] is False
    assert result["code"] == "HOTKEY_REGISTER_FAILED"
    assert attempts == ["alt+shift+k", "ctrl+shift+x"]
    assert prefs.load_prefs(base_dir=tmp_path)["hotkey"] == "ctrl+shift+x"


def test_toggle_prefs_persist(tmp_path):
    api, _ = make_api(tmp_path)
    assert api.set_start_hidden(True)["start_hidden"] is True
    assert api.set_toast_enabled(False)["toast_enabled"] is False
    loaded = prefs.load_prefs(base_dir=tmp_path)
    assert loaded["start_hidden"] is True
    assert loaded["toast_enabled"] is False


def test_return_to_list_and_module_favorites_persist(tmp_path):
    api, _ = make_api(tmp_path)
    assert api.get_initial_state()["return_to_list_after_action"] is False
    result = api.set_return_to_list_after_action(True)
    assert result["return_to_list_after_action"] is True

    pinned = api.set_module_favorite("0003_6200", True)
    assert pinned["favorite_module_ids"] == ["0003_6200"]
    assert api.get_initial_state()["favorite_module_ids"] == ["0003_6200"]
    unpinned = api.set_module_favorite("0003_6200", False)
    assert unpinned["favorite_module_ids"] == []


def test_focus_chrome_on_module_defaults_on_and_persists(tmp_path):
    api, _ = make_api(tmp_path)
    assert api.get_initial_state()["focus_chrome_on_module"] is True
    result = api.set_focus_chrome_on_module(False)
    assert result["focus_chrome_on_module"] is False
    assert prefs.load_prefs(base_dir=tmp_path)["focus_chrome_on_module"] is False


def test_costing_export_open_options_are_exposed_and_persisted(tmp_path):
    api, _ = make_api(tmp_path)
    initial = api.get_initial_state()
    assert initial["open_costing_file_after_export"] is True
    assert initial["open_costing_folder_after_export"] is False

    saved = api.set_costing_export_open_options(False, True)

    assert saved["open_costing_file_after_export"] is False
    assert saved["open_costing_folder_after_export"] is True
    loaded = prefs.load_prefs(base_dir=tmp_path)
    assert loaded["open_costing_file_after_export"] is False
    assert loaded["open_costing_folder_after_export"] is True


def test_initial_state_exposes_new_fields(tmp_path):
    api, _ = make_api(tmp_path)
    state = api.get_initial_state()
    for field in (
        "hotkey",
        "hotkey_label",
        "autostart",
        "start_hidden",
        "toast_enabled",
        "return_to_list_after_action",
        "favorite_module_ids",
        "focus_chrome_on_module",
        "open_costing_file_after_export",
        "open_costing_folder_after_export",
        "chrome_alive",
        "session_active",
        "last_login_at",
        "catalog_default_folder",
    ):
        assert field in state, field


def test_release_update_methods_delegate_and_schedule(tmp_path, monkeypatch):
    api, _ = make_api(tmp_path)
    state = {
        "ok": True,
        "code": "UPDATE_AVAILABLE",
        "message": "Có bản mới.",
        "can_update": True,
        "version": "1.1.0",
        "package_url": "https://github.com/example/update.zip",
        "checksum_url": "https://github.com/example/update.zip.sha256",
    }
    monkeypatch.setattr(
        "wfx_panel.panel_api.updater.check_for_updates",
        lambda **_kwargs: dict(state),
    )
    applied = []
    api.set_update_applier(lambda payload: applied.append(payload) or None)
    assert api.check_for_updates()["can_update"] is True
    result = api.install_update()
    assert result["code"] == "UPDATE_SCHEDULED"
    assert applied and applied[0]["version"] == "1.1.0"
    assert "tự mở lại" in result["message"]


def test_panel_update_channel_is_always_stable(tmp_path, monkeypatch):
    api, _ = make_api(tmp_path)
    calls = []
    monkeypatch.setattr(
        "wfx_panel.panel_api.updater.check_for_updates",
        lambda **kwargs: calls.append(kwargs) or {
            "ok": True,
            "code": "UP_TO_DATE",
            "can_update": False,
        },
    )
    api.set_update_channel("current")
    assert prefs.load_prefs(base_dir=tmp_path)["update_channel"] == "stable"
    api.check_for_updates()
    api.install_update()
    assert calls == [{"channel": "stable"}, {"channel": "stable"}]


def test_automation_job_has_run_id_and_can_retry(tmp_path):
    api, fake = make_api(tmp_path)
    result = api.open_module("0004_0050_0020")
    assert result["run_id"]
    history = api.get_job_history()["jobs"]
    assert history[0]["run_id"] == result["run_id"]
    retried = api.retry_job(result["run_id"])
    assert retried["ok"] is True
    assert retried["run_id"] != result["run_id"]
    assert [call[0] for call in fake.calls].count("open_module") == 2


def test_job_history_never_stores_credentials(tmp_path):
    prefs.save_account("private-user", "private-password", base_dir=tmp_path)
    api, _ = make_api(tmp_path)
    api.login()
    serialized = (tmp_path / "jobs.json").read_text(encoding="utf-8")
    assert "private-user" not in serialized
    assert "private-password" not in serialized


def test_initial_state_contains_module_classes_and_jobs(tmp_path):
    api, _ = make_api(tmp_path)
    state = api.get_initial_state()
    catalog = state["module_groups"][0]["modules"][0]
    assert catalog["kind"] == "catalog"
    assert catalog["description"]
    modules = {
        module["id"]: module
        for group in state["module_groups"]
        for module in group["modules"]
    }
    assert modules["0005_0050_0020"]["kind"] == "rmpo"
    assert modules["0005_0080_0020"]["kind"] == "indent"
    assert modules["user_indent_list"]["kind"] == "indent"
    assert modules["0063_0030_0020"]["kind"] == "list_new"
    assert modules["0065_0880_0010_0020"]["kind"] == "list_new"
    assert modules["0065_0880_0030_0020"]["kind"] == "list_new"
    assert state["jobs"] == []


def test_window_preferences_apply_and_persist(tmp_path):
    api, _ = make_api(tmp_path)
    on_top = []
    api.set_window_pref_appliers(on_top.append)
    top_result = api.set_always_on_top(False)
    assert top_result["always_on_top"] is False
    assert on_top == [False]
    loaded = prefs.load_prefs(base_dir=tmp_path)
    assert loaded["always_on_top"] is False
    assert "stick_to_browser" not in loaded
    assert not hasattr(api, "set_stick_to_browser")


def test_failed_automation_records_local_screenshot(tmp_path):
    class FailingLogin(FakeLogin):
        def open_module(self, module_name, xpath, log=print):
            return {
                "ok": False,
                "code": "MODULE_FAILED",
                "message": "fake failure",
            }

        def capture_failure_screenshot(self, path, log=print):
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"png")
            return True

    api = PanelAPI(
        login_module=FailingLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )
    result = api.open_module("0004_0050_0020")
    assert result["ok"] is False
    job = api.get_job_history()["jobs"][0]
    assert job["run_id"] == result["run_id"]
    assert job["has_screenshot"] is True
    assert "password" not in str(job).lower()


def test_expected_no_results_does_not_capture_diagnostic_screenshot(tmp_path):
    class NoResultsLogin(FakeLogin):
        def find_in_open_catalog(
            self,
            category_name,
            filter_kind,
            query,
            log=print,
        ):
            del category_name, filter_kind, query, log
            return {
                "ok": False,
                "code": "NO_RESULTS",
                "message": "not found",
            }

        def capture_failure_screenshot(self, path, log=print):
            raise AssertionError("NO_RESULTS không phải lỗi kỹ thuật")

    api = PanelAPI(
        login_module=NoResultsLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )

    result = api.find_code("Apparel", "MISSING")

    assert result["code"] == "NO_RESULTS"
    assert api.get_job_history()["jobs"][0]["has_screenshot"] is False


def test_edi_validation_failure_captures_failed_record_screenshot(tmp_path):
    class EDIValidationLogin(FakeLogin):
        def capture_failure_screenshot(self, path, log=print):
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"failed-record-popup")
            return True

    api = PanelAPI(
        login_module=EDIValidationLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )

    result = api._run(
        "confirm_oc_upload",
        lambda: {
            "ok": False,
            "code": "OC_EDI_VALIDATION_FAILED",
            "message": "Mapping Resolved: InProgress",
        },
    )

    assert result["code"] == "OC_EDI_VALIDATION_FAILED"
    assert api.get_job_history()["jobs"][0]["has_screenshot"] is True


def test_failed_automation_webhook_contains_human_readable_context(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr("wfx_panel.telemetry.DEFAULT_WEBHOOK_URL", "")

    class FailingLogin(FakeLogin):
        def open_module(self, module_name, xpath, log=print):
            return {
                "ok": False,
                "code": "MODULE_FAILED",
                "message": "PlaywrightError: frame WFX đã đổi.",
                "module": module_name,
            }

    api = PanelAPI(
        login_module=FailingLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )
    result = api.open_module("0063_0030_0020")
    assert result["ok"] is False
    payload = json.loads(
        (tmp_path / "telemetry-outbox.json").read_text(encoding="utf-8")
    )[0]
    assert payload["method_label"] == "Mở module"
    assert payload["error_title"] == "Không thể thao tác module QA List"
    assert "frame WFX đã đổi" in payload["message"]
    assert payload["suggestion"]


def test_blank_search_webhook_uses_safe_request_context(tmp_path, monkeypatch):
    monkeypatch.setattr("wfx_panel.telemetry.DEFAULT_WEBHOOK_URL", "")

    class BlankSearchFailure(FakeLogin):
        def search_oc_list(self, xpath, filter_kind, query, log=print):
            return {
                "ok": False,
                "code": "MODULE_SEARCH_NOT_READY",
                "message": "",
            }

    api = PanelAPI(
        login_module=BlankSearchFailure(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )
    result = api.search_oc("oc_no", "private-oc-query")
    assert result["ok"] is False

    payload = json.loads(
        (tmp_path / "telemetry-outbox.json").read_text(encoding="utf-8")
    )[0]
    serialized = json.dumps(payload, ensure_ascii=False)
    assert payload["module"] == "OC List"
    assert payload["filter_kind"] == "OC No."
    assert "Không tìm thấy ô OC No. trong OC List" in payload["error_detail"]
    assert "private-oc-query" not in serialized
    assert "Automation không trả về mô tả chi tiết" not in serialized


def test_automatic_report_thread_captures_disabled_endpoint(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr("wfx_panel.telemetry.DEFAULT_WEBHOOK_URL", "")
    scheduled = []

    class DeferredThread:
        def __init__(self, target, args, daemon):
            scheduled.append((target, args, daemon))

        def start(self):
            return None

    monkeypatch.setattr(
        "wfx_panel.panel_api.threading.Thread",
        DeferredThread,
    )

    class FailingLogin(FakeLogin):
        def open_module(self, module_name, xpath, log=print):
            return {
                "ok": False,
                "code": "MODULE_FAILED",
                "message": "fake failure",
                "module": module_name,
            }

    api = PanelAPI(
        login_module=FailingLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )
    api.open_module("0004_0050_0020")

    assert len(scheduled) == 1
    target, args, daemon = scheduled[0]
    assert target is telemetry.flush
    assert args == (tmp_path, "")
    assert daemon is True


def test_failed_division_switch_records_local_screenshot(tmp_path):
    class FailingDivisionLogin(FakeLogin):
        def switch_division(self, division_key, log=print):
            return {
                "ok": False,
                "code": "DIVISION_CHANGE_NOT_CONFIRMED",
                "message": f"fake failure: {division_key}",
            }

        def capture_failure_screenshot(self, path, log=print):
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"png")
            return True

    api = PanelAPI(
        login_module=FailingDivisionLogin(),
        prefs_module=prefs,
        base_dir=tmp_path,
    )

    result = api.switch_division("knit")

    assert result["ok"] is False
    assert api.get_job_history()["jobs"][0]["has_screenshot"] is True


class AuthorizedAdminLogin(FakeLogin):
    def check_module_access(self, specs, log=print):
        self.calls.append(("check_module_access", tuple(s["id"] for s in specs)))
        return {
            "ok": True,
            "code": "MODULE_ACCESS_CHECKED",
            "accessible_module_ids": [
                "0090_0250", "0004_0010_1720", "0005_0010_1290"
            ],
        }


class CompanyAuthorizedLogin(FakeLogin):
    def check_module_access(self, specs, log=print):
        return {
            "ok": True,
            "code": "MODULE_ACCESS_CHECKED",
            "accessible_module_ids": ["0090_0007"],
        }


def test_company_foc_toggle_is_permission_gated_and_delegated(tmp_path):
    denied_api, denied_fake = make_api(tmp_path)
    denied = denied_api.toggle_company_foc()
    assert denied["code"] == "ADMIN_ACCESS_DENIED"
    assert not any(
        call[0] == "toggle_company_foc" for call in denied_fake.calls
    )

    fake = CompanyAuthorizedLogin()
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)
    changed = api.toggle_company_foc()
    assert changed["code"] == "COMPANY_FOC_CHANGED"
    assert changed["foc_mode"] == "FOC cho ASN"
    assert changed["saved"] is True
    assert (
        "toggle_company_foc",
        '//*[@id="0090_0007"]/a',
    ) in fake.calls


def test_oc_sample_and_sale_asn_workflows_delegate(tmp_path):
    api, fake = make_api(tmp_path)
    assert api.search_oc("oc_no", " OC-12 ")["ok"] is True
    assert api.search_sample("created_by", " Alice ")["ok"] is True
    assert api.open_sample_new()["code"] == "SAMPLE_NEW_READY"
    assert api.search_sale_asn("invoice_no", " INV-9 ")["ok"] is True
    assert (
        "search_oc",
        '//*[@id="0004_0050_0020"]/a',
        "oc_no",
        "OC-12",
    ) in fake.calls
    assert (
        "search_sample",
        '//*[@id="0004_0056_4070"]/a',
        "created_by",
        "Alice",
    ) in fake.calls
    assert ("open_sample_new", constants.SAMPLE_NEW_XPATH) in fake.calls
    assert (
        "search_sale_asn",
        '//*[@id="0004_0070_0020"]/a',
        "invoice_no",
        "INV-9",
    ) in fake.calls


def test_rmpo_indent_and_list_new_workflows_delegate(tmp_path):
    api, fake = make_api(tmp_path)

    assert api.search_rmpo(" Acme ", " RM-42 ")["ok"] is True
    assert api.search_indent(
        "0005_0080_0020",
        " Acme ",
        " ART-1 ",
        " IN-9 ",
        " ST-2 ",
    )["ok"] is True
    assert api.search_indent(
        "user_indent_list",
        "",
        " ART-2 ",
        "",
        "",
    )["ok"] is True
    assert api.open_module_new("0063_0030_0020")["code"] == "MODULE_NEW_READY"
    assert api.open_module_new("0065_0880_0010_0020")["ok"] is True
    assert api.open_module_new("0065_0880_0030_0020")["ok"] is True

    assert (
        "search_rmpo",
        constants.MODULE_BY_ID["0005_0050_0020"]["xpath"],
        "Acme",
        "RM-42",
    ) in fake.calls
    assert (
        "search_indent",
        constants.MODULE_BY_ID["0005_0080_0020"]["xpath"],
        "Indent List",
        "Acme",
        "ART-1",
        "IN-9",
        "ST-2",
    ) in fake.calls
    assert (
        "search_indent",
        constants.MODULE_BY_ID["user_indent_list"]["xpath"],
        "User Indent",
        "",
        "ART-2",
        "",
        "",
    ) in fake.calls
    for module_id in (
        "0063_0030_0020",
        "0065_0880_0010_0020",
        "0065_0880_0030_0020",
    ):
        assert ("open_module_new", module_id) in fake.calls


def test_supplier_category_and_find_delegate_with_all_categories(tmp_path):
    fake = AuthorizedAdminLogin()
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)

    opened = api.open_supplier_category("Textiles/Fabric")
    found = api.find_supplier("  Acme  ")

    assert opened["code"] == "SUPPLIER_CATEGORY_READY"
    assert found["code"] == "SUPPLIER_FOUND"
    assert (
        "open_supplier_category",
        '//*[@id="0005_0010_1290"]/a',
        "Textiles/Fabric",
        "03",
    ) in fake.calls
    call = next(item for item in fake.calls if item[0] == "find_supplier")
    assert call[1] == '//*[@id="0005_0010_1290"]/a'
    assert call[2] == {
        "Apparel": "01",
        "Fixed Asset": "04",
        "Miscellaneous": "12",
        "Services": "06",
        "Textiles/Fabric": "03",
        "Trims": "05",
    }
    assert call[3] == "Acme"


def test_supplier_find_uses_only_selected_category(tmp_path):
    fake = AuthorizedAdminLogin()
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)

    result = api.find_supplier_in_category("Trims", "  Acme  ")

    assert result["code"] == "SUPPLIER_FOUND"
    assert (
        "find_supplier_in_category",
        '//*[@id="0005_0010_1290"]/a',
        "Trims",
        "05",
        "Acme",
    ) in fake.calls


def test_find_buyer_uses_current_buyer_list_and_trimmed_query(tmp_path):
    fake = AuthorizedAdminLogin()
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)
    result = api.find_buyer("  BirdDogs ")
    assert result["code"] == "BUYER_EDIT_OPENED"
    assert (
        "find_buyer",
        '//*[@id="0004_0010_1720"]/a',
        "BirdDogs",
    ) in fake.calls


def test_supplier_and_buyer_workflows_preserve_admin_access_gate(tmp_path):
    api, fake = make_api(tmp_path)
    assert api.open_supplier_category("Apparel")["code"] == "ADMIN_ACCESS_DENIED"
    assert api.find_supplier("Acme")["code"] == "ADMIN_ACCESS_DENIED"
    assert api.find_supplier_in_category("Apparel", "Acme")["code"] == "ADMIN_ACCESS_DENIED"
    assert api.find_buyer("BirdDogs")["code"] == "ADMIN_ACCESS_DENIED"
    assert not any(call[0] in {
        "open_supplier_category", "find_supplier", "find_supplier_in_category",
        "find_buyer"
    } for call in fake.calls)


def test_admin_mode_requires_real_wfx_access(tmp_path):
    api, fake = make_api(tmp_path)
    denied = api.set_admin_mode(True)
    assert denied["code"] == "ADMIN_ACCESS_DENIED"
    assert denied["admin_access"] is False
    assert prefs.load_prefs(base_dir=tmp_path)["admin_mode"] is False
    opened = api.open_module("0090_0250")
    assert opened["code"] == "ADMIN_ACCESS_DENIED"
    assert not any(
        call[:2] == ("open_module", "System Coding")
        for call in fake.calls
    )


def test_admin_login_exposes_only_authorized_modules(tmp_path):
    fake = AuthorizedAdminLogin()
    api = PanelAPI(login_module=fake, prefs_module=prefs, base_dir=tmp_path)
    prefs.save_account("admin", "pw", base_dir=tmp_path)
    logged_in = api.login()
    assert logged_in["admin_access"] is True
    assert logged_in["admin_mode"] is False
    assert logged_in["admin_module_ids"] == [
        "0004_0010_1720", "0005_0010_1290", "0090_0250"
    ]

    enabled = api.set_admin_mode(True)
    assert enabled["admin_mode"] is True
    opened = api.open_module("0090_0250")
    assert opened["ok"] is True
    assert ("open_module", "System Coding", '//*[@id="0090_0250"]/a') in fake.calls


def test_feedback_queues_without_exposing_webhook(tmp_path, monkeypatch):
    monkeypatch.delenv("WFX_ERROR_WEBHOOK_URL", raising=False)
    monkeypatch.setattr("wfx_panel.telemetry.DEFAULT_WEBHOOK_URL", "")
    prefs.save_account("report-user", "private-password", base_dir=tmp_path)
    api, _ = make_api(tmp_path)
    api._current_division = "woven"
    api._division_label = "WOVEN"
    api._division_name = "PRO SPORTS - WOVEN HANOI"
    result = api.submit_feedback("bug", "Nút Catalog không phản hồi", True)
    assert result["ok"] is True
    assert result["code"] == "FEEDBACK_QUEUED"
    assert result["reporting_configured"] is False
    assert "webhook" not in result
    outbox = (tmp_path / "telemetry-outbox.json").read_text(encoding="utf-8")
    assert "Nút Catalog không phản hồi" in outbox
    assert "report-user" in outbox
    assert '"company_id": "psh"' in outbox
    assert "PRO SPORTS - WOVEN HANOI" in outbox
    assert "private-password" not in outbox


def test_costing_composite_cannot_be_interrupted_between_steps(tmp_path):
    """export/import/apply Costing gồm nhiều _run liên tiếp: mở đúng Costing rồi
    mới scan/ghi. Nếu run lock được nhả giữa các bước, một flow khác có thể đổi
    module/Division và bước sau thao tác nhầm màn hình — plan token không biết."""
    api, fake = make_api(tmp_path)
    rejected: list[dict] = []
    inside_composite = threading.Event()
    may_finish = threading.Event()

    def steps() -> dict:
        # Bước 1 của composite đã chạy xong và đã nhả _run bên trong.
        api.prepare_catalog("Apparel")
        inside_composite.set()
        may_finish.wait(timeout=5)
        # Bước 2 vẫn phải chạy được trên cùng thread (RLock tái nhập).
        return api.prepare_catalog("Apparel")

    worker = threading.Thread(
        target=lambda: rejected.append(api.run_composite(steps))
    )
    worker.start()
    assert inside_composite.wait(timeout=2)

    # Đúng khe hở giữa hai bước: flow khác PHẢI bị từ chối.
    intruder = api.open_module("0004_0050_0020")
    assert api.is_action_running() is True

    may_finish.set()
    worker.join(timeout=5)

    assert intruder["code"] == "ACTION_IN_PROGRESS"
    assert rejected[0]["code"] == "CATEGORY_SELECTED"
    assert api.is_action_running() is False


def test_run_composite_rejects_when_another_thread_holds_the_lock(tmp_path):
    api, _fake = make_api(tmp_path)
    holding = threading.Event()
    release = threading.Event()

    def hold_lock():
        api._run_lock.acquire()
        holding.set()
        release.wait(timeout=5)
        api._run_lock.release()

    holder = threading.Thread(target=hold_lock)
    holder.start()
    try:
        assert holding.wait(timeout=2)
        outcome = api.run_composite(lambda: {"ok": True, "code": "NEVER_RAN"})
    finally:
        release.set()
        holder.join(timeout=5)

    assert outcome["code"] == "ACTION_IN_PROGRESS"
