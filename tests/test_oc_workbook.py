from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from wfx_panel.oc_workbook import (
    EDI_HEADERS,
    FORM_HEADERS,
    INPUT_HEADERS,
    PAYMENT_TERM_OPTIONS,
    OCWorkbookError,
    prepare_oc_workbook,
    write_oc_input_template,
)


def _input_row(**overrides):
    values = {
        "Buyer": "J.LINDEBERG",
        "Season": "SS26",
        "Order Type": "Confirmed",
        "Currency": "USD",
        "Factory": "888 COMPANY LTD",
        "Ship Under PO Ref": "PO-SS26-01",
        "Article Code": "SWV0004581",
        "Buyer Style Ref": "GMPA17697",
        "Buyer PO Num": "PO-SS26-01",
        "Summary Buyer Order Ref": "PO-SS26-01",
        "Buyer Order Date": "08-10-2025",
        "Buyer Delivery Date": "31-12-2025",
        "Raw Material ETA Date": "05-12-2025",
        "Payment Terms": "TT After Shipment 60 Days",
        "Country of Final Destination": "Sweden",
        "Color Code": "O127",
        "Color Name": "Forget-Me-Not",
        "Size Code": "M",
        "Selling Price": 23.65,
        "Units": 9,
        "Internal Lot No.": "1",
        "PO Type (Zone)": "FOB",
        "Extra Production %": 0,
        "Buyer Lot No.": "RAB",
    }
    values.update(overrides)
    return [values[header] for header in INPUT_HEADERS]


def _filled_input_file(tmp_path: Path, *rows: list[object]) -> Path:
    path = write_oc_input_template(tmp_path / "input.xlsx")
    workbook = load_workbook(path)
    sheet = workbook["OC INPUT"]
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def test_generated_template_has_one_visible_header_and_hidden_references(tmp_path):
    path = write_oc_input_template(tmp_path / "WFX-Smart-Upload-OC.xlsx")

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["OC INPUT", "REFERENCES"]
    assert workbook["REFERENCES"].sheet_state == "veryHidden"
    sheet = workbook["OC INPUT"]
    assert [cell.value for cell in sheet[1]] == list(INPUT_HEADERS)
    assert sheet.freeze_panes == "A2"
    assert len(sheet.data_validations.dataValidation) == 7
    validations = {
        next(iter(validation.sqref.ranges)).min_col: validation
        for validation in sheet.data_validations.dataValidation
    }
    assert validations[1].errorStyle == "warning"
    assert validations[5].errorStyle == "warning"
    assert validations[15].errorStyle is None
    assert validations[22].allow_blank is True
    assert validations[14].allow_blank is False
    assert sheet["A1"].comment.text.startswith("Chọn Buyer")
    assert sheet["V1"].fill.fgColor.rgb == "00F4B183"
    assert sheet["W1"].fill.fgColor.rgb == "00F4B183"
    assert sheet["X1"].fill.fgColor.rgb == "00F4B183"
    references = workbook["REFERENCES"]
    assert [references.cell(row, 3).value for row in range(2, 5)] == [
        "Confirmed",
        "Forecast",
        "SMS",
    ]
    assert [
        references.cell(row, 7).value
        for row in range(2, len(PAYMENT_TERM_OPTIONS) + 2)
    ] == list(PAYMENT_TERM_OPTIONS)
    workbook.close()


def test_new_input_is_mapped_to_value_only_edi_and_total_qty(tmp_path):
    source = _filled_input_file(
        tmp_path,
        _input_row(Units=9, **{"Size Code": "M"}),
        _input_row(Units=6, **{"Size Code": "L"}),
    )
    output = tmp_path / "edi.xlsx"

    prepared = prepare_oc_workbook(source, "new", output)

    assert prepared.buyer == "J.LINDEBERG"
    assert prepared.row_count == 2
    assert prepared.seasons == ("SS26",)
    assert prepared.po_count == 1
    assert prepared.style_count == 1
    assert prepared.total_units == 15
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Sheet1"]
    sheet = workbook["Sheet1"]
    assert [cell.value for cell in sheet[1]] == list(EDI_HEADERS)
    indexes = {header: index + 1 for index, header in enumerate(EDI_HEADERS)}
    assert sheet.cell(2, indexes["Color"]).value == "O127^Forget-Me-Not"
    assert sheet.cell(2, indexes["Market"]).value == "Europe"
    assert sheet.cell(2, indexes["Total Qty"]).value == 15
    assert sheet.cell(3, indexes["Total Qty"]).value == 15
    assert sheet.cell(2, indexes["Buyer Order Date"]).value.date() == date(2025, 10, 8)
    assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)
    workbook.close()


def test_new_input_skips_zero_units_and_applies_blank_defaults(tmp_path):
    source = _filled_input_file(
        tmp_path,
        _input_row(
            Units=0,
            **{
                "Buyer PO Num": "SKIP-PO",
                "Summary Buyer Order Ref": "SKIP-PO",
            },
        ),
        _input_row(
            **{
                "Order Type": "SMS",
                "PO Type (Zone)": None,
                "Extra Production %": None,
                "Payment Terms": "15% Deposit After Contract - 85% TT Before Shipment",
            }
        ),
    )

    prepared = prepare_oc_workbook(source, "new", tmp_path / "edi.xlsx")

    assert prepared.row_count == 1
    assert prepared.total_units == 9
    assert prepared.warnings == ("App đã bỏ qua 1 dòng có Units = 0.",)
    workbook = load_workbook(prepared.upload_path, data_only=True)
    sheet = workbook["Sheet1"]
    indexes = {header: index + 1 for index, header in enumerate(EDI_HEADERS)}
    assert sheet.cell(2, indexes["Order Type"]).value == "SMS"
    assert sheet.cell(2, indexes["Zone"]).value == "FOB"
    assert sheet.cell(2, indexes["Extra Production %"]).value == 0
    assert sheet.cell(2, indexes["Payment Terms"]).value == PAYMENT_TERM_OPTIONS[0]
    workbook.close()


def test_new_input_rejects_invalid_delivery_date_sequence(tmp_path):
    source = _filled_input_file(
        tmp_path,
        _input_row(
            **{
                "Buyer Order Date": "08-10-2025",
                "Raw Material ETA Date": "07-10-2025",
                "Buyer Delivery Date": "06-10-2025",
            }
        ),
    )

    with pytest.raises(OCWorkbookError) as caught:
        prepare_oc_workbook(source, "new", tmp_path / "edi.xlsx")

    assert caught.value.code == "OC_FILE_VALIDATION_FAILED"
    assert any(
        "Buyer Order Date < Raw Material ETA < Buyer Delivery Date" in error
        for error in caught.value.errors
    )


def test_new_input_rejects_when_every_row_has_zero_units(tmp_path):
    source = _filled_input_file(tmp_path, _input_row(Units=0))

    with pytest.raises(OCWorkbookError) as caught:
        prepare_oc_workbook(source, "new", tmp_path / "edi.xlsx")

    assert caught.value.code == "OC_FILE_EMPTY"
    assert "Units = 0" in caught.value.message


def test_new_input_rejects_more_than_one_buyer(tmp_path):
    source = _filled_input_file(
        tmp_path,
        _input_row(),
        _input_row(
            Buyer="PUMA",
            **{"Buyer PO Num": "PO-2", "Summary Buyer Order Ref": "PO-2"},
        ),
    )

    with pytest.raises(OCWorkbookError) as caught:
        prepare_oc_workbook(source, "new", tmp_path / "edi.xlsx")

    assert caught.value.code == "OC_FILE_VALIDATION_FAILED"
    assert any("nhiều Buyer" in error for error in caught.value.errors)


def test_new_input_allows_new_wfx_buyer_and_factory_with_warning(tmp_path):
    source = _filled_input_file(
        tmp_path,
        _input_row(Buyer="NEW LIVE BUYER", Factory="NEW LIVE FACTORY"),
    )

    prepared = prepare_oc_workbook(source, "new", tmp_path / "edi.xlsx")

    assert prepared.buyer == "NEW LIVE BUYER"
    assert prepared.warnings == (
        "Buyer 'NEW LIVE BUYER' chưa có trong danh sách gợi ý; app sẽ yêu cầu "
        "khớp chính xác trên WFX.",
        "Factory 'NEW LIVE FACTORY' chưa có trong danh sách gợi ý; WFX sẽ kiểm "
        "tra khi Process Package.",
    )


def test_new_input_rejects_formula_error_and_unknown_country(tmp_path):
    source = _filled_input_file(
        tmp_path,
        _input_row(
            **{
                "Color Name": "#VALUE!",
                "Country of Final Destination": "Atlantis",
            }
        ),
    )

    with pytest.raises(OCWorkbookError) as caught:
        prepare_oc_workbook(source, "new", tmp_path / "edi.xlsx")

    assert caught.value.code == "OC_FILE_VALIDATION_FAILED"
    assert any("#VALUE!" in error for error in caught.value.errors)


def test_new_input_rejects_live_formula_in_editable_sheet(tmp_path):
    source = _filled_input_file(tmp_path, _input_row())
    workbook = load_workbook(source)
    workbook["OC INPUT"]["T2"] = "=4+5"
    workbook.save(source)
    workbook.close()

    with pytest.raises(OCWorkbookError) as caught:
        prepare_oc_workbook(source, "new", tmp_path / "edi.xlsx")

    assert caught.value.code == "OC_FILE_FORMULA_ERROR"
    assert caught.value.errors == ("Ô có công thức: T2.",)


def _revise_file(tmp_path: Path) -> Path:
    path = tmp_path / "revise.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(EDI_HEADERS))
    base = dict.fromkeys(EDI_HEADERS)
    base.update(
        {
            "Factory": "888 COMPANY LTD",
            "Ship Under PO Ref": "PO-1",
            "Article": "SWV0004581",
            "Buyer": "J.LINDEBERG",
            "Currency": "USD",
            "Season": "SS26",
            "Country of Origin": "Vietnam",
            "Payment Terms": "TT After Shipment 60 Days",
            "Buyer PO Num": "PO-1",
            "Summary Buyer Order Ref": "PO-1",
            "Market Buyer Order Ref": "PO-1",
            "Destination Buyer Order Ref": "PO-1",
            "Delivery Buyer Order Ref": "PO-1",
            "Buyer Order Date": date(2025, 10, 8),
            "Order Type": "Confirmed",
            "Mode of Shipment": "AIR/SEA",
            "Buyer Delivery Date": date(2025, 12, 31),
            "OC Delivery Date": date(2025, 12, 31),
            "Raw Matetrial ETA": date(2025, 12, 5),
            "Country of Final Destination": "Sweden",
            "Final Destination": "Sweden",
            "Market": "Europe",
            "Buyer Style Ref.": "GMPA17697",
            "Color": "O127^Forget-Me-Not",
            "Size": "M",
            "Total Qty": 999,
            "Price": 23.65,
            "Units": 9,
            "Zone": "FOB",
            "Internal Lot No.": "1",
            "DeliveryOCID": "OC-DELIVERY-1",
            "Fulfillment Type": "Back Order",
            "Extra Production %": 0,
        }
    )
    sheet.append([base[header] for header in EDI_HEADERS])
    second = dict(base)
    second.update({"Size": "L", "Units": 6})
    sheet.append([second[header] for header in EDI_HEADERS])
    workbook.save(path)
    return path


def _legacy_new_file(tmp_path: Path) -> Path:
    path = tmp_path / "legacy-new.xlsx"
    workbook = Workbook()
    form = workbook.active
    form.title = "FORM"
    form["B1"] = "J.LINDEBERG"
    form["B2"] = "SS26"
    form["B3"] = "SMS"
    form["B4"] = "USD"
    form.append(list(FORM_HEADERS))
    base = {
        "Factory": "888 COMPANY LTD",
        "Ship Under PO Ref": "PO-SS26-01",
        "Article Code": "SWV0004581",
        "Buyer Style Ref": "GMPA17697",
        "Buyer PO Num": "PO-SS26-01",
        "Summary Buyer Order Ref": "PO-SS26-01",
        "Buyer Order Date": "08-10-2025",
        "Order/Buyer Delivery Date": "31-12-2025",
        "Raw Matetrial ETA Date": "05-12-2025",
        "Payment Terms": "TT After Shipment 60 Days",
        "Country of Final Destination": "Sweden",
        "Color code": "O127",
        "Color name": "Forget-Me-Not",
        "Size code": "M",
        "Selling Price": 23.65,
        "Units": 0,
        "Internal Lot No.": "1",
        "PO Type": None,
        "Extra Production": None,
        "Buyer Lot No.": None,
    }
    form.append([base[header] for header in FORM_HEADERS])
    second = dict(base)
    second.update({"Size code": "L", "Units": 6})
    form.append([second[header] for header in FORM_HEADERS])

    references = workbook.create_sheet("THONG TIN")
    references.append(["Factory", "Buyer", None, None, "Country", "Market"])
    references.append(
        ["888 COMPANY LTD", "J.LINDEBERG", None, None, "Sweden", "Europe"]
    )
    workbook.save(path)
    workbook.close()
    return path


def test_revise_recalculates_total_qty_and_keeps_delivery_id(tmp_path):
    source = _revise_file(tmp_path)
    output = tmp_path / "revise-edi.xlsx"

    prepared = prepare_oc_workbook(source, "revise", output)

    assert prepared.row_count == 2
    assert prepared.seasons == ("SS26",)
    assert prepared.po_count == 1
    assert prepared.style_count == 1
    assert prepared.total_units == 15
    assert prepared.warnings == (
        "App đã tính lại Total Qty cho 2 dòng từ cột Units.",
    )
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["Sheet1"]
    indexes = {header: index + 1 for index, header in enumerate(EDI_HEADERS)}
    assert sheet.cell(2, indexes["Total Qty"]).value == 15
    assert sheet.cell(3, indexes["Total Qty"]).value == 15
    assert sheet.cell(2, indexes["DeliveryOCID"]).value == "OC-DELIVERY-1"
    workbook.close()


def test_legacy_new_form_uses_the_same_zero_and_blank_defaults(tmp_path):
    source = _legacy_new_file(tmp_path)

    prepared = prepare_oc_workbook(source, "new", tmp_path / "legacy-edi.xlsx")

    assert prepared.row_count == 1
    assert prepared.total_units == 6
    assert prepared.warnings == ("App đã bỏ qua 1 dòng có Units = 0.",)
    workbook = load_workbook(prepared.upload_path, data_only=True)
    sheet = workbook["Sheet1"]
    indexes = {header: index + 1 for index, header in enumerate(EDI_HEADERS)}
    assert sheet.cell(2, indexes["Order Type"]).value == "SMS"
    assert sheet.cell(2, indexes["Zone"]).value == "FOB"
    assert sheet.cell(2, indexes["Extra Production %"]).value == 0
    workbook.close()


def test_revise_skips_zero_units_and_applies_blank_defaults(tmp_path):
    source = _revise_file(tmp_path)
    workbook = load_workbook(source)
    sheet = workbook["Sheet1"]
    indexes = {header: index + 1 for index, header in enumerate(EDI_HEADERS)}
    sheet.cell(2, indexes["Units"], 0)
    sheet.cell(3, indexes["Zone"], None)
    sheet.cell(3, indexes["Extra Production %"], None)
    workbook.save(source)
    workbook.close()

    prepared = prepare_oc_workbook(source, "revise", tmp_path / "revise-edi.xlsx")

    assert prepared.row_count == 1
    assert prepared.total_units == 6
    assert "App đã bỏ qua 1 dòng có Units = 0." in prepared.warnings
    workbook = load_workbook(prepared.upload_path, data_only=True)
    sheet = workbook["Sheet1"]
    assert sheet.cell(2, indexes["Zone"]).value == "FOB"
    assert sheet.cell(2, indexes["Extra Production %"]).value == 0
    workbook.close()


def test_revise_requires_delivery_date_to_equal_oc_delivery_date(tmp_path):
    source = _revise_file(tmp_path)
    workbook = load_workbook(source)
    sheet = workbook["Sheet1"]
    indexes = {header: index + 1 for index, header in enumerate(EDI_HEADERS)}
    sheet.cell(2, indexes["OC Delivery Date"], date(2026, 1, 1))
    workbook.save(source)
    workbook.close()

    with pytest.raises(OCWorkbookError) as caught:
        prepare_oc_workbook(source, "revise", tmp_path / "revise-edi.xlsx")

    assert caught.value.code == "OC_FILE_VALIDATION_FAILED"
    assert any(
        "Buyer Delivery Date phải bằng OC Delivery Date" in error
        for error in caught.value.errors
    )
