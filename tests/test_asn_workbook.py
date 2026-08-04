from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from wfx_panel.asn_workbook import merge_sale_asn_reports


def _report(path, title, value):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet["A1"] = value
    sheet["A1"].font = Font(bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["B2"] = "=1+1"
    sheet.merge_cells("A3:C3")
    sheet["A3"] = "Merged heading"
    frame = Side(style="thin", color="000000")
    sheet["A3"].border = Border(left=frame, top=frame, bottom=frame)
    sheet["C3"].border = Border(right=frame, top=frame, bottom=frame)
    sheet.column_dimensions["A"].width = 24
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.sheet_view.zoomScale = 85
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_area = "A1:C3"
    sheet.oddHeader.center.text = f"{value} header"
    workbook.save(path)
    workbook.close()


def _reports(path, titles, prefix):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, title in enumerate(titles, start=1):
        sheet = workbook.create_sheet(title)
        sheet["A1"] = f"{prefix} {index}"
    workbook.save(path)
    workbook.close()


def _wrapped_report(path, title):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.column_dimensions["A"].width = 8
    sheet["A2"] = "Goods description with enough words to require several lines"
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[2].height = 15
    workbook.save(path)
    workbook.close()


def _jl_packing_report(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "JL Packing"
    sheet.append(
        [
            "JL PO#", "Style No", "Color Code", "Net Wt", "Gross Wt",
            "No of Carton", "CBM",
        ]
    )
    sheet.append(["PO129761", "SWOW18336", "9999 BLACK", 581.5, 683.5, 68, 6.97])
    sheet.append(["PO129761", "SWOW18336", "Q142 GRAPE", 581.5, 683.5, 68, 6.97])
    sheet.append(["PO129761", "SWPA18338", "9999 BLACK", 358.7, 408.2, 33, 3.38])
    sheet.append(["PO129761", "SWPA18338", "Q142 GRAPE", 358.7, 408.2, 33, 3.38])
    workbook.save(path)
    workbook.close()


def _truewerk_packing_report(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TRUEWERK Packing"
    for reference, value in {
        "C1": "⑫ Description of Goods",
        "H1": "⑬ Qty/Unit",
        "J1": "⑭ Net-Weight",
        "L1": "⑭ Gross-Weight",
        "N1": "⑮Qty Cartons",
        "O1": "CBM",
        "C2": "Style",
        "F2": "PO No.",
        "C3": "MEN PANTS",
    }.items():
        sheet[reference] = value
    for reference in (
        "C1:G1", "H1:I1", "J1:K1", "L1:M1", "C2:E2", "F2:G2", "C3:G3"
    ):
        sheet.merge_cells(reference)

    def add_line(row, style, po, quantity, net_weight, gross_weight, cartons, cbm):
        for column, value in {
            "C": style,
            "F": po,
            "H": quantity,
            "J": net_weight,
            "L": gross_weight,
            "N": cartons,
            "O": cbm,
        }.items():
            sheet[f"{column}{row}"] = value
        for reference in (f"C{row}:E{row}", f"F{row}:G{row}", f"H{row}:I{row}", f"J{row}:K{row}", f"L{row}:M{row}"):
            sheet.merge_cells(reference)

    add_line(4, "S22-0001", "P00003739", 2700, 1221.6, 1333.2, 93, 7.79)
    add_line(5, "S22-0001", "P00003739 ADD", 90, 0, 0, 0, 0)
    add_line(6, "S22-0001", "P00003740", 2850, 1284.4, 1402, 98, 8.21)
    add_line(7, "S22-0001", "P00003740 - ADD", 90, 3, 0, 0, 0)
    add_line(8, "S22-0001", "P00009999 ADD", 90, 0, 0, 0, 0)
    add_line(9, "S22-0001", "P00009999", 2850, 500, 550, 40, 4.2)
    workbook.save(path)
    workbook.close()


def test_merge_sale_asn_reports_renames_only_colliding_sheet_titles(tmp_path):
    packing = tmp_path / "packing.xlsx"
    buyer = tmp_path / "buyer.xlsx"
    target = tmp_path / "INV-001.xlsx"
    _report(packing, "Report", "Packing data")
    _report(buyer, "Report", "Buyer data")

    output = merge_sale_asn_reports(
        packing,
        buyer,
        target,
        invoice_no="INV-001",
    )

    assert output == target.resolve()
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["INVOICE INV-001", "PKL INV-001"]
    assert workbook["PKL INV-001"]["A1"].value == "Packing data"
    assert workbook["INVOICE INV-001"]["A1"].value == "Buyer data"
    for sheet in workbook.worksheets:
        assert sheet["A1"].font.bold is True
        assert sheet["A1"].fill.fgColor.rgb == "001F4E78"
        assert sheet["B2"].value == "=1+1"
        assert "A3:C3" in {str(item) for item in sheet.merged_cells.ranges}
        assert sheet["A3"].border.left.style == "thin"
        assert sheet["C3"].border.right.style == "thin"
        assert sheet.column_dimensions["A"].width == 24
        assert sheet.row_dimensions[1].height == 30
        assert sheet.freeze_panes == "A2"
        assert sheet.sheet_view.zoomScale == 85
        assert sheet.sheet_view.showGridLines is False
        assert sheet.page_setup.orientation == "landscape"
        assert sheet.page_setup.fitToWidth == 1
        assert sheet.page_setup.paperSize == 9
        assert sheet.page_setup.fitToHeight == 0
        assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
        assert str(sheet.print_area).endswith("!$A$1:$C$3")
    assert workbook["PKL INV-001"].oddHeader.center.text == "Packing data header"
    assert workbook["INVOICE INV-001"].oddHeader.center.text == "Buyer data header"
    workbook.close()


def test_merge_sale_asn_reports_preserves_distinct_original_sheet_titles(tmp_path):
    packing = tmp_path / "packing-original.xlsx"
    buyer = tmp_path / "buyer-original.xlsx"
    target = tmp_path / "INV-002.xlsx"
    _report(packing, "WFX Packing Export", "Packing data")
    _report(buyer, "Commercial Invoice", "Buyer data")

    merge_sale_asn_reports(
        packing,
        buyer,
        target,
        invoice_no="INV-002",
    )

    workbook = load_workbook(target, data_only=False)
    assert workbook.sheetnames == ["Commercial Invoice", "WFX Packing Export"]
    assert workbook["WFX Packing Export"]["A1"].value == "Packing data"
    assert workbook["Commercial Invoice"]["A1"].value == "Buyer data"
    workbook.close()


def test_merge_sale_asn_reports_interleaves_multiple_sheets_invoice_first(tmp_path):
    packing = tmp_path / "packing-multi.xlsx"
    buyer = tmp_path / "buyer-multi.xlsx"
    target = tmp_path / "INV-003.xlsx"
    _reports(packing, ["Report 1", "Report 2", "Report 3"], "Packing")
    _reports(buyer, ["Report 1", "Report 2", "Report 3"], "Invoice")

    merge_sale_asn_reports(packing, buyer, target, invoice_no="INV-003")

    workbook = load_workbook(target, data_only=False)
    assert workbook.sheetnames == [
        "INVOICE INV-003",
        "PKL INV-003",
        "INVOICE INV-003 (2)",
        "PKL INV-003 (2)",
        "INVOICE INV-003 (3)",
        "PKL INV-003 (3)",
    ]
    assert [sheet["A1"].value for sheet in workbook.worksheets] == [
        "Invoice 1",
        "Packing 1",
        "Invoice 2",
        "Packing 2",
        "Invoice 3",
        "Packing 3",
    ]
    workbook.close()


def test_merge_sale_asn_reports_keeps_interleaving_when_counts_differ(tmp_path):
    packing = tmp_path / "packing-extra.xlsx"
    buyer = tmp_path / "buyer-short.xlsx"
    target = tmp_path / "INV-004.xlsx"
    _reports(packing, ["Packing 1", "Packing 2", "Packing 3"], "Packing")
    _reports(buyer, ["Invoice 1", "Invoice 2"], "Invoice")

    merge_sale_asn_reports(packing, buyer, target, invoice_no="INV-004")

    workbook = load_workbook(target, data_only=False)
    assert workbook.sheetnames == [
        "Invoice 1",
        "Packing 1",
        "Invoice 2",
        "Packing 2",
        "Packing 3",
    ]
    assert [sheet["A1"].value for sheet in workbook.worksheets] == [
        "Invoice 1",
        "Packing 1",
        "Invoice 2",
        "Packing 2",
        "Packing 3",
    ]
    workbook.close()


def test_merge_sale_asn_reports_expands_wrapped_rows_that_would_be_cut_off(tmp_path):
    packing = tmp_path / "packing-wrapped.xlsx"
    buyer = tmp_path / "buyer-wrapped.xlsx"
    target = tmp_path / "INV-005.xlsx"
    _wrapped_report(packing, "Packing")
    _wrapped_report(buyer, "Invoice")

    merge_sale_asn_reports(packing, buyer, target, invoice_no="INV-005")

    workbook = load_workbook(target, data_only=False)
    assert workbook["Invoice"].row_dimensions[2].height > 15
    assert workbook["Packing"].row_dimensions[2].height > 15
    workbook.close()


def test_merge_sale_asn_reports_keeps_portrait_reports_portrait_on_a4(tmp_path):
    packing = tmp_path / "packing-portrait.xlsx"
    buyer = tmp_path / "buyer-portrait.xlsx"
    target = tmp_path / "INV-006.xlsx"
    _report(packing, "Packing", "Packing data")
    _report(buyer, "Invoice", "Invoice data")
    for source in (packing, buyer):
        workbook = load_workbook(source)
        workbook.active.page_setup.orientation = "portrait"
        workbook.save(source)
        workbook.close()

    merge_sale_asn_reports(packing, buyer, target, invoice_no="INV-006")

    workbook = load_workbook(target, data_only=False)
    for sheet in workbook.worksheets:
        assert sheet.page_setup.paperSize == 9
        assert sheet.page_setup.orientation == "portrait"
        assert sheet.page_setup.fitToWidth == 1
    workbook.close()


def test_merge_sale_asn_reports_merges_jl_packing_measurements_by_po_and_style(tmp_path):
    packing = tmp_path / "jl-packing.xlsx"
    buyer = tmp_path / "buyer.xlsx"
    target = tmp_path / "1105.JLSKI.26.xlsx"
    _jl_packing_report(packing)
    _report(buyer, "Buyer Invoice", "Invoice data")

    merge_sale_asn_reports(packing, buyer, target, invoice_no="1105.JLSKI.26")

    workbook = load_workbook(target, data_only=False)
    sheet = workbook["JL Packing"]
    merged = {str(item) for item in sheet.merged_cells.ranges}
    assert {"D2:D3", "E2:E3", "F2:F3", "G2:G3"} <= merged
    assert {"D4:D5", "E4:E5", "F4:F5", "G4:G5"} <= merged
    assert sheet["D3"].value is None
    assert sheet["E5"].value is None
    assert {str(item) for item in workbook["Buyer Invoice"].merged_cells.ranges} == {
        "A3:C3"
    }
    workbook.close()


def test_merge_sale_asn_reports_merges_truewerk_measurements_from_po_to_add_row(tmp_path):
    packing = tmp_path / "truewerk-packing.xlsx"
    buyer = tmp_path / "buyer.xlsx"
    target = tmp_path / "TRUEWERK-001.xlsx"
    _truewerk_packing_report(packing)
    _report(buyer, "Buyer Invoice", "Invoice data")

    merge_sale_asn_reports(packing, buyer, target, invoice_no="TRUEWERK-001")

    workbook = load_workbook(target, data_only=False)
    sheet = workbook["TRUEWERK Packing"]
    merged = {str(item) for item in sheet.merged_cells.ranges}
    assert {"J4:K5", "L4:M5", "N4:N5", "O4:O5"} <= merged
    assert "J6:K7" not in merged
    assert "J8:K9" not in merged
    assert sheet["J5"].value is None
    assert sheet["L5"].value is None
    assert sheet["N5"].value is None
    assert sheet["O5"].value is None
    assert sheet["J7"].value == 3
    assert sheet["J9"].value == 500
    workbook.close()
