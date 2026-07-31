from pathlib import Path

import pytest
from openpyxl import load_workbook

from wfx_panel import costing_workbook


def _field(
    field_key,
    label,
    value,
    *,
    editable=True,
    item_key="fabric-1",
    section_key="fabric",
    data_type="text",
    row_order=1,
):
    return {
        "scope": "item",
        "section_key": section_key,
        "item_key": item_key,
        "field_key": field_key,
        "label": label,
        "value": value,
        "data_type": data_type,
        "editable": editable,
        "required": False,
        "options": [],
        "row_order": row_order,
    }


def sample_document():
    return {
        "format_version": costing_workbook.FORMAT_VERSION,
        "style_code": "SWN0000001",
        "title": "FOB Main",
        "cost_sheet_status": "Open",
        "cost_sheet_type": "Internal Cost Sheets",
        "order_execution_type": "Trading",
        "season": "SS27",
        "template": "FOB",
        "signature": "live-signature",
        "sections": [
            {"section_key": "fabric", "name": "Fabric", "row_order": 1},
        ],
        "items": [
            {
                "section_key": "fabric",
                "section_name": "Fabric",
                "item_key": "fabric-1",
                "row_order": 1,
                "action": "UPSERT",
                "item_type": "article",
                "article_code": "FAB-001",
                "article_name": "Cotton Jersey",
            }
        ],
        "fields": [
            _field(
                "colConsQty",
                "Cons. Qty.",
                1.25,
                data_type="number",
            ),
            _field(
                "colWastagePer",
                "Waste %",
                3,
                data_type="number",
                row_order=2,
            ),
            _field(
                "colSupplierCompanyName",
                "Supplier",
                "Supplier A",
                row_order=3,
            ),
        ],
    }


def find_field(document, field_key, item_key="fabric-1"):
    return next(
        field
        for field in document["fields"]
        if field["scope"] == "item"
        and field["field_key"] == field_key
        and field["item_key"] == item_key
    )


def _header_map(sheet):
    return {
        str(cell.value): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


def test_xlsx_round_trip_uses_two_sheet_standard_form(tmp_path):
    target = tmp_path / "SWN0000001-costing.xlsx"

    costing_workbook.write_costing_file(sample_document(), target)
    loaded = costing_workbook.read_costing_file(target)
    workbook = load_workbook(target)

    assert workbook.sheetnames == [
        costing_workbook.GUIDE_SHEET,
        costing_workbook.FORM_SHEET,
    ]
    assert loaded["style_code"] == "SWN0000001"
    assert loaded["title"] == ""
    assert [section["name"] for section in loaded["sections"]] == [
        name for _token, name in costing_workbook.STANDARD_SECTIONS
    ]
    assert loaded["sections"][0]["section_key"] == "fabric"
    assert loaded["items"][0]["article_code"] == "FAB-001"
    assert find_field(loaded, "colConsQty")["value"] == 1.25

    form = workbook[costing_workbook.FORM_SHEET]
    assert [
        cell.value
        for cell in form[1][: len(costing_workbook.FORM_COLUMNS)]
    ] == costing_workbook.FORM_COLUMNS
    assert all(
        cell.value is None
        for cell in form[1][len(costing_workbook.FORM_COLUMNS) :]
    )
    assert "Add To Cost" not in costing_workbook.FORM_COLUMNS
    assert "Cons. UOM" not in costing_workbook.FORM_COLUMNS
    header = _header_map(form)
    for name in costing_workbook.FORM_TECH_COLUMNS:
        column_letter = form.cell(1, header[name]).column_letter
        assert form.column_dimensions[column_letter].hidden is True


def test_xlsx_round_trip_keeps_duplicate_article_rows_separate(tmp_path):
    document = sample_document()
    second_key = "fabric-2"
    document["items"].append(
        {
            **document["items"][0],
            "item_key": second_key,
            "row_order": 2,
        }
    )
    document["fields"].append(
        _field(
            "colConsQty",
            "Cons. Qty.",
            2.5,
            item_key=second_key,
            data_type="number",
            row_order=2,
        )
    )
    target = tmp_path / "duplicate-article-rows.xlsx"

    costing_workbook.write_costing_file(document, target)
    loaded = costing_workbook.read_costing_file(target)

    matching_items = [
        item
        for item in loaded["items"]
        if item["article_code"] == "FAB-001"
    ]
    assert [item["item_key"] for item in matching_items] == [
        "fabric-1",
        second_key,
    ]
    assert find_field(loaded, "colConsQty", "fabric-1")["value"] == 1.25
    assert find_field(loaded, "colConsQty", second_key)["value"] == 2.5


def test_import_ignores_legacy_exported_identity_only_phantom_row(tmp_path):
    document = sample_document()
    target = tmp_path / "legacy-phantom-row.xlsx"
    costing_workbook.write_costing_file(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    phantom_row = 3
    form.cell(phantom_row, header["Section"], "Fabric")
    form.cell(phantom_row, header["Action"], "UPSERT")
    form.cell(phantom_row, header["Article Code"], "FAB-001")
    form.cell(phantom_row, header["Article Name"], "Cotton Jersey")
    form.cell(phantom_row, header["__Section Key"], "fabric")
    form.cell(phantom_row, header["__Item Key"], "FAB-001::row::3")
    form.cell(phantom_row, header["__Row Order"], 99)
    form.cell(phantom_row, header["__Item Type"], "article")
    workbook.save(target)

    loaded = costing_workbook.read_costing_file(target)

    assert [item["item_key"] for item in loaded["items"]] == ["fabric-1"]


def test_xlsx_highlights_editable_cells_and_adds_one_template_per_section(
    tmp_path,
):
    document = sample_document()
    document["sections"].append(
        {
            "section_key": "sewing-trims",
            "name": "SEWING TRIMS",
            "row_order": 2,
        }
    )
    target = tmp_path / "friendly-form.xlsx"

    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    visible_last = header["Purchase Officer"]
    yellow = costing_workbook._INPUT_FILL.fgColor.rgb
    template_yellow = costing_workbook._TEMPLATE_INPUT_FILL.fgColor.rgb
    read_only_red = costing_workbook._READ_ONLY_FILL.fgColor.rgb
    read_only_columns = {
        header["Cons. Qty. Incl. Waste"],
        header["Value in (USD)"],
    }

    template_rows = [
        row
        for row in range(2, form.max_row + 1)
        if not form.cell(row, header["Article Code"]).value
        and not form.cell(row, header["Article Name"]).value
    ]

    assert len(template_rows) == (
        6 * costing_workbook.DEFAULT_NEW_ITEM_ROWS_PER_SECTION + 1 + 1 + 2
    )
    assert all(
        form.cell(2, column).fill.fgColor.rgb == yellow
        for column in range(2, visible_last + 1)
        if column not in read_only_columns
    )
    assert all(
        form.cell(2, column).fill.fgColor.rgb == read_only_red
        for column in read_only_columns
    )
    assert all(
        form.cell(row, column).fill.fgColor.rgb == template_yellow
        for row in template_rows
        for column in range(2, visible_last + 1)
        if column not in read_only_columns
    )
    assert all(
        form.cell(row, column).fill.fgColor.rgb == read_only_red
        for row in template_rows
        for column in read_only_columns
    )
    assert all(
        form.cell(row, header["Article Code"]).comment is not None
        for row in template_rows
    )
    assert {
        form.cell(row, header["Section"]).value for row in template_rows
    } == {
        name for _token, name in costing_workbook.STANDARD_SECTIONS
    }

    guide = workbook[costing_workbook.GUIDE_SHEET]
    guide_values = {
        guide.cell(row, 1).value: guide.cell(row, 2).value
        for row in range(1, guide.max_row + 1)
    }
    assert guide_values["Các nhóm trong file"] == ", ".join(
        name for _token, name in costing_workbook.STANDARD_SECTIONS
    )
    assert "Chỉ Cost Sheet đang Open" in guide_values["Phạm vi"]


def test_xlsx_form_edits_values_and_creates_new_item(tmp_path):
    target = tmp_path / "edit.xlsx"
    costing_workbook.write_costing_xlsx(sample_document(), target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)

    form.cell(2, header["Cons. Qty."], 2.75)
    template_row = next(
        row
        for row in range(2, form.max_row + 1)
        if not form.cell(row, header["Article Code"]).value
    )
    form.cell(template_row, header["Action"], "UPSERT")
    form.cell(template_row, header["Article Code"], "FAB-002")
    form.cell(template_row, header["Article Name"], "Rib")
    form.cell(template_row, header["Cons. Qty."], 0.75)
    form.cell(template_row, header["Supplier"], "Supplier B")
    workbook.save(target)

    loaded = costing_workbook.read_costing_xlsx(target)

    assert find_field(loaded, "colConsQty")["value"] == 2.75
    new_item = next(
        item for item in loaded["items"] if item["article_code"] == "FAB-002"
    )
    assert new_item["item_key"].startswith("new:fabric:")
    assert find_field(
        loaded,
        "colConsQty",
        new_item["item_key"],
    )["value"] == 0.75
    assert find_field(
        loaded,
        "colSupplierCompanyName",
        new_item["item_key"],
    )["value"] == "Supplier B"


def test_special_cost_sections_use_fixed_rows_dropdowns_and_safe_defaults(
    tmp_path,
):
    document = sample_document()
    document["sections"].extend(
        [
            {
                "section_key": "cm",
                "name": "CM Costs",
                "row_order": 7,
                "article_options": ["FACTORY A", "FACTORY B"],
            },
            {
                "section_key": "production",
                "name": "Production Costs",
                "row_order": 8,
                "article_options": ["CM (PRODUCTIONPROCESS100001)"],
            },
            {
                "section_key": "indirect",
                "name": "Indirect Costs",
                "row_order": 9,
                "article_options": ["Air charge", "Commission Fee"],
            },
        ]
    )
    target = tmp_path / "special-costs.xlsx"
    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    rows_by_section = {
        section: [
            row
            for row in range(2, form.max_row + 1)
            if form.cell(row, header["Section"]).value == section
        ]
        for section in ("CM Costs", "Production Costs", "Indirect Costs")
    }

    assert {name: len(rows) for name, rows in rows_by_section.items()} == {
        "CM Costs": 1,
        "Production Costs": 1,
        "Indirect Costs": 2,
    }
    cm_row = rows_by_section["CM Costs"][0]
    production_row = rows_by_section["Production Costs"][0]
    indirect_row = rows_by_section["Indirect Costs"][0]
    assert form.cell(cm_row, header["Curr."]).value == "USD"
    assert form.cell(production_row, header["Minutes"]).value == 1
    assert form.cell(indirect_row, header["Curr."]).value == "USD"
    assert all(
        form.cell(row, header["Action"]).value is None
        for rows in rows_by_section.values()
        for row in rows
    )
    article_validations = [
        validation
        for validation in form.data_validations.dataValidation
        if any(
            cell_range.min_col == header["Article Name"]
            for cell_range in validation.ranges.ranges
        )
    ]
    assert len(article_validations) == 3
    hidden_values = {
        str(form.cell(row, column).value)
        for column in range(len(costing_workbook.FORM_COLUMNS) + 1, form.max_column + 1)
        for row in range(2, form.max_row + 1)
        if form.cell(row, column).value
    }
    assert {
        "FACTORY A",
        "CM (PRODUCTIONPROCESS100001)",
        "Air charge",
    } <= hidden_values

    form.cell(cm_row, header["Article Name"], "FACTORY A")
    form.cell(cm_row, header["Value"], 10)
    form.cell(production_row, header["Article Name"], "CM (PRODUCTIONPROCESS100001)")
    form.cell(production_row, header["Value"], 100)
    form.cell(production_row, header["Rate"], 2)
    form.cell(indirect_row, header["Article Name"], "Air charge")
    form.cell(indirect_row, header["Value"], 5)
    workbook.save(target)

    loaded = costing_workbook.read_costing_xlsx(target)
    special_items = [
        item for item in loaded["items"] if item["item_type"] == "cost_line"
    ]
    assert [item["article_name"] for item in special_items] == [
        "FACTORY A",
        "CM (PRODUCTIONPROCESS100001)",
        "Air charge",
    ]
    production_item = special_items[1]
    production_fields = {
        field["field_key"]: field["value"]
        for field in loaded["fields"]
        if field["item_key"] == production_item["item_key"]
    }
    assert production_fields == {
        "Minutes": 1,
        "ProductionHeaderMinutes": 1,
        "colRate1": 2,
        "ProductionValue": 100,
    }
    assert sum(
        item["section_name"] == "Indirect Costs" for item in special_items
    ) == 1


def test_form_adds_live_color_size_and_purchase_officer_dropdowns(tmp_path):
    document = sample_document()
    document["fields"].extend(
        [
            _field(
                "colMaterialColorList",
                "Material Color",
                "JL NAVY(6855)",
                row_order=4,
            ),
            _field(
                "colMaterialSizeList",
                "Material Size",
                "M",
                row_order=5,
            ),
            _field(
                "colPurchaseOfficer",
                "Purchase Officer",
                "HanhNgoThi",
                row_order=6,
            ),
        ]
    )
    target = tmp_path / "dropdowns.xlsx"

    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)

    validations = list(form.data_validations.dataValidation)
    formulas_by_column = {}
    for validation in validations:
        for cell_range in validation.ranges.ranges:
            formulas_by_column[cell_range.min_col] = validation.formula1
    assert header["Color Dep."] in formulas_by_column
    assert header["Size Dep."] in formulas_by_column
    assert header["Purchase Officer"] in formulas_by_column
    hidden_values = {
        str(form.cell(row, column).value)
        for column in range(len(costing_workbook.FORM_COLUMNS) + 1, form.max_column + 1)
        for row in range(2, form.max_row + 1)
        if form.cell(row, column).value
    }
    assert {"JL NAVY(6855)", "M", "HanhNgoThi"} <= hidden_values


def test_form_adds_cached_article_code_and_name_dropdowns_by_section(tmp_path):
    document = sample_document()
    document["sections"][0]["article_code_options"] = ["FAB-001", "FAB-002"]
    document["sections"][0]["article_name_options"] = [
        "Cotton Jersey",
        "Rib Jersey",
    ]
    target = tmp_path / "article-dropdowns.xlsx"

    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    validated_columns = {
        cell_range.min_col
        for validation in form.data_validations.dataValidation
        for cell_range in validation.ranges.ranges
    }
    hidden_values = {
        str(form.cell(row, column).value)
        for column in range(
            len(costing_workbook.FORM_COLUMNS) + 1,
            form.max_column + 1,
        )
        for row in range(2, form.max_row + 1)
        if form.cell(row, column).value
    }

    assert header["Article Code"] in validated_columns
    assert header["Article Name"] in validated_columns
    assert {"FAB-001", "FAB-002", "Cotton Jersey", "Rib Jersey"} <= hidden_values
    article_name = form.cell(2, header["Article Name"]).value
    assert isinstance(article_name, str)
    assert article_name.startswith("=IFERROR(INDEX(")
    workbook.save(target)
    loaded = costing_workbook.read_costing_xlsx(target)
    assert loaded["items"][0]["article_name"] == "Cotton Jersey"


def test_article_name_formula_follows_article_code_selected_in_excel(tmp_path):
    document = sample_document()
    document["items"][0]["article_code"] = "F0001"
    document["items"][0]["article_name"] = "Cotton Jersey"
    document["sections"][0]["article_lookup_options"] = [
        {"article_code": "F0001", "article_name": "Cotton Jersey"},
        {"article_code": "F0002", "article_name": "Rib Jersey"},
    ]
    target = tmp_path / "article-auto-name.xlsx"
    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    form.cell(2, header["Article Code"], "F0002")
    workbook.save(target)

    loaded = costing_workbook.read_costing_xlsx(target)

    assert loaded["items"][0]["article_code"] == "F0002"
    assert loaded["items"][0]["article_name"] == "Rib Jersey"


def test_excel_rewritten_article_name_formula_is_resolved_safely(tmp_path):
    document = sample_document()
    document["items"][0]["article_code"] = "F0001"
    document["items"][0]["article_name"] = "Cotton Jersey"
    document["sections"][0]["article_lookup_options"] = [
        {"article_code": "F0001", "article_name": "Cotton Jersey"},
        {"article_code": "F0002", "article_name": "Rib Jersey"},
    ]
    target = tmp_path / "article-formula-rewritten.xlsx"
    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    form.cell(2, header["Article Code"], "F0002")
    form.cell(2, header["Article Name"], "=@IFERROR(INDEX(A1:A2,1),\"\")")
    workbook.save(target)

    loaded = costing_workbook.read_costing_xlsx(target)

    assert loaded["items"][0]["article_code"] == "F0002"
    assert loaded["items"][0]["article_name"] == "Rib Jersey"


def test_article_code_follows_unique_article_name_selected_in_excel(tmp_path):
    document = sample_document()
    document["items"][0]["article_code"] = "F0001"
    document["items"][0]["article_name"] = "Cotton Jersey"
    document["sections"][0]["article_lookup_options"] = [
        {"article_code": "F0001", "article_name": "Cotton Jersey"},
        {"article_code": "F0002", "article_name": "Rib Jersey"},
    ]
    target = tmp_path / "article-auto-code.xlsx"
    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    form.cell(2, header["Article Name"], "Rib Jersey")
    workbook.save(target)

    loaded = costing_workbook.read_costing_xlsx(target)

    assert loaded["items"][0]["article_code"] == "F0002"
    assert loaded["items"][0]["article_name"] == "Rib Jersey"


def test_duplicate_article_name_requires_article_code_in_excel(tmp_path):
    document = sample_document()
    document["sections"][0]["article_lookup_options"] = [
        {"article_code": "F0001", "article_name": "Jersey"},
        {"article_code": "F0002", "article_name": "Jersey"},
    ]
    target = tmp_path / "article-name-ambiguous.xlsx"
    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    template_row = next(
        row
        for row in range(2, form.max_row + 1)
        if not form.cell(row, header["Article Code"]).value
    )
    form.cell(template_row, header["Article Name"], "Jersey")
    workbook.save(target)

    with pytest.raises(costing_workbook.CostingWorkbookError) as raised:
        costing_workbook.read_costing_xlsx(target)

    assert raised.value.code == "COSTING_VALIDATION_FAILED"
    assert "hãy chọn Article Code" in raised.value.details[0]


def test_formula_columns_are_exported_red_and_never_imported(tmp_path):
    document = sample_document()
    document["style_name"] = "KFSWPKN-S200 LN"
    document["fields"].extend(
        [
            {
                **_field(
                    "colConsPlusWastageQty",
                    "Cons. Qty. Incl. Waste",
                    1.02,
                    row_order=8,
                ),
                "editable": False,
            },
            {
                **_field(
                    "colValueInCSCurr",
                    "Value in (USD)",
                    2.04,
                    row_order=9,
                ),
                "editable": False,
            },
        ]
    )
    target = tmp_path / "formula-columns.xlsx"

    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    assert form.cell(2, header["Cons. Qty. Incl. Waste"]).value.startswith(
        "=IF("
    )
    assert form.cell(2, header["Value in (USD)"]).value.startswith("=IF(")
    assert (
        form.cell(2, header["Value in (USD)"]).fill.fgColor.rgb
        == costing_workbook._READ_ONLY_FILL.fgColor.rgb
    )
    guide = workbook[costing_workbook.GUIDE_SHEET]
    guide_values = {
        guide.cell(row, 1).value: guide.cell(row, 2).value
        for row in range(1, guide.max_row + 1)
    }
    assert guide_values["Style Name"] == "KFSWPKN-S200 LN"

    form.cell(2, header["Cons. Qty. Incl. Waste"], 999)
    form.cell(2, header["Value in (USD)"], 999)
    workbook.save(target)
    loaded = costing_workbook.read_costing_xlsx(target)
    imported_keys = {field["field_key"] for field in loaded["fields"]}
    assert "colConsPlusWastageQty" not in imported_keys
    assert "colValueInCSCurr" not in imported_keys


def test_two_adjacent_new_rows_may_use_the_same_article_for_splitter(tmp_path):
    target = tmp_path / "split.xlsx"
    costing_workbook.write_costing_xlsx(sample_document(), target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    template_rows = [
        row
        for row in range(2, form.max_row + 1)
        if form.cell(row, header["Section"]).value == "FABRIC- SHELL"
        and not form.cell(row, header["Article Code"]).value
    ][:2]
    for row in template_rows:
        form.cell(row, header["Article Code"], "FAB-SPLIT")
        form.cell(row, header["Cons. Qty."], row / 10)
    workbook.save(target)

    loaded = costing_workbook.read_costing_xlsx(target)
    matching = [
        item
        for item in loaded["items"]
        if item["article_code"] == "FAB-SPLIT"
    ]

    assert len(matching) == 2
    assert matching[0]["item_key"] != matching[1]["item_key"]


def test_csv_is_not_supported(tmp_path):
    target = tmp_path / "costing.csv"
    target.write_text("not supported", encoding="utf-8")

    with pytest.raises(costing_workbook.CostingWorkbookError) as captured:
        costing_workbook.read_costing_file(target)

    assert captured.value.code == "COSTING_FILE_TYPE_UNSUPPORTED"


def test_import_rejects_formula_in_editable_data(tmp_path):
    target = tmp_path / "formula.xlsx"
    costing_workbook.write_costing_xlsx(sample_document(), target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    form.cell(2, header["Cons. Qty."], '=HYPERLINK("x","x")')
    workbook.save(target)

    with pytest.raises(costing_workbook.CostingWorkbookError) as captured:
        costing_workbook.read_costing_xlsx(target)

    assert captured.value.code == "COSTING_FORMULA_NOT_ALLOWED"


def test_import_validation_reports_exact_row_and_column(tmp_path):
    target = tmp_path / "invalid-action.xlsx"
    costing_workbook.write_costing_xlsx(sample_document(), target)
    workbook = load_workbook(target)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)
    form.cell(2, header["Action"], "WRONG")
    workbook.save(target)

    with pytest.raises(costing_workbook.CostingWorkbookError) as captured:
        costing_workbook.read_costing_xlsx(target)

    assert captured.value.code == "COSTING_VALIDATION_FAILED"
    assert any(
        "Costing!B2" in detail
        for detail in captured.value.details
    )


def test_export_escapes_formula_like_text(tmp_path):
    document = sample_document()
    find_field(document, "colSupplierCompanyName")["value"] = "+literal"
    target = tmp_path / "safe.xlsx"

    costing_workbook.write_costing_xlsx(document, target)
    workbook = load_workbook(target, data_only=False)
    form = workbook[costing_workbook.FORM_SHEET]
    header = _header_map(form)

    assert form.cell(2, header["Supplier"]).value == "'+literal"
    loaded = costing_workbook.read_costing_xlsx(target)
    assert find_field(loaded, "colSupplierCompanyName")["value"] == "+literal"


def test_wrong_format_version_and_duplicate_keys_are_rejected():
    wrong = sample_document()
    wrong["format_version"] = "99"
    with pytest.raises(costing_workbook.CostingWorkbookError) as captured:
        costing_workbook.normalize_document(wrong)
    assert captured.value.code == "COSTING_FORMAT_UNSUPPORTED"

    duplicate = sample_document()
    duplicate["items"].append(dict(duplicate["items"][0]))
    with pytest.raises(costing_workbook.CostingWorkbookError) as captured:
        costing_workbook.normalize_document(duplicate)
    assert captured.value.code == "COSTING_VALIDATION_FAILED"


def test_only_xlsx_is_accepted(tmp_path):
    path = Path(tmp_path / "costing.json")
    path.write_text("{}", encoding="utf-8")

    with pytest.raises(costing_workbook.CostingWorkbookError) as captured:
        costing_workbook.read_costing_file(path)

    assert captured.value.code == "COSTING_FILE_TYPE_UNSUPPORTED"


def test_workbook_keeps_special_sections_and_filters_unsupported_fields(
    tmp_path,
):
    document = sample_document()
    document["sections"].extend(
        [
            {"section_key": "cm", "name": "CM Costs", "row_order": 2},
            {
                "section_key": "production",
                "name": "Production Costs",
                "row_order": 3,
            },
            {
                "section_key": "indirect",
                "name": "Indirect Costs",
                "row_order": 4,
            },
        ]
    )
    document["items"].append(
        {
            "section_key": "cm",
            "section_name": "CM Costs",
            "item_key": "cm-1",
            "row_order": 1,
            "action": "UPSERT",
            "item_type": "article",
            "article_code": "CM-001",
            "article_name": "Excluded",
        }
    )
    document["fields"].extend(
        [
            _field(
                "colDeliveryTerms",
                "Delivery Terms",
                "FOB",
                row_order=4,
            ),
            _field(
                "colRate1",
                "Rate",
                99,
                editable=False,
                data_type="number",
                row_order=5,
            ),
            _field(
                "colConsQty",
                "Cons. Qty.",
                2,
                item_key="cm-1",
                section_key="cm",
                data_type="number",
            ),
        ]
    )
    target = tmp_path / "filtered.xlsx"

    costing_workbook.write_costing_xlsx(document, target)
    loaded = costing_workbook.read_costing_xlsx(target)
    workbook = load_workbook(target, data_only=False)

    assert workbook.sheetnames == ["Hướng dẫn", "Costing"]
    assert [section["name"] for section in loaded["sections"]] == [
        name for _token, name in costing_workbook.STANDARD_SECTIONS
    ]
    assert [item["item_key"] for item in loaded["items"]] == ["fabric-1"]
    assert all(field["editable"] for field in loaded["fields"])
    assert {
        field["field_key"] for field in loaded["fields"]
    }.isdisjoint({"colDeliveryTerms"})
    visible_text = {
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "Cost Sheet" not in workbook.sheetnames
    assert {"CM Costs", "Production Costs", "Indirect Costs"} <= visible_text
    assert not any("Delivery Terms" in value for value in visible_text)


def test_empty_file_path_is_reported_as_required():
    with pytest.raises(costing_workbook.CostingWorkbookError) as captured:
        costing_workbook.read_costing_file("")

    assert captured.value.code == "COSTING_FILE_REQUIRED"
