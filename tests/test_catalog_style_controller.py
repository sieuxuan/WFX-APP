from openpyxl import load_workbook

from wfx_panel.catalog_controller import CatalogController
from wfx_panel.style_workbook import STYLE_SHEET, write_style_template


class _FakeLogin:
    def __init__(self):
        self.calls = []

    def prepare_catalog_style_row(
        self,
        category_value,
        group_id,
        row,
        copy_choice,
        log,
    ):
        self.calls.append((category_value, group_id, row, copy_choice))
        return {
            "ok": True,
            "code": "STYLE_FORM_READY",
            "message": "Dừng trước Save.",
            "requires_manual_save": True,
        }


class _FakePanel:
    def __init__(self):
        self._login = _FakeLogin()
        self._log = lambda _message: None

    def _run(self, _method, action, _request=None):
        return action()


def _new_style_file(path):
    write_style_template(path)
    workbook = load_workbook(path)
    sheet = workbook[STYLE_SHEET]
    values = [
        "New", "", "WOVEN", "Buyer A", "Woven", "Top", "Jacket",
        "Color A", "Size A", "FW27", "BUY-1", "INT-1",
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(2, column, value)
    workbook.save(path)
    workbook.close()


def test_style_controller_freezes_review_to_selected_group_and_row(tmp_path):
    source = tmp_path / "styles.xlsx"
    _new_style_file(source)
    panel = _FakePanel()
    controller = CatalogController(panel)
    controller.folder_cache["Apparel"] = [
        {
            "node_id": "774001163",
            "kind": "group",
            "name": "Jackets",
            "path_label": "Master / Jackets",
        }
    ]

    review = controller.review_style_import(str(source), "774001163")
    prepared = controller.prepare_style_row(
        review["review_token"],
        review["rows"][0]["source_row"],
    )

    assert review["code"] == "STYLE_IMPORT_REVIEW_READY"
    assert review["requires_manual_save"] is True
    assert prepared["code"] == "STYLE_FORM_READY"
    category, group_id, row, copy_choice = panel._login.calls[0]
    assert category == "01"
    assert group_id == "774001163"
    assert row["internal_style_ref"] == "INT-1"
    assert copy_choice is None


def test_style_controller_rejects_folder_that_is_not_group(tmp_path):
    source = tmp_path / "styles.xlsx"
    _new_style_file(source)
    controller = CatalogController(_FakePanel())
    controller.folder_cache["Apparel"] = [
        {
            "node_id": "42",
            "kind": "folder",
            "name": "Folder",
            "path_label": "Master / Folder",
        }
    ]

    result = controller.review_style_import(str(source), "42")

    assert result["code"] == "STYLE_GROUP_REQUIRED"
