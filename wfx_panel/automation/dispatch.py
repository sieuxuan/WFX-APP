"""Tạo (GDN) Dispatch từ Invoice GRN qua report và EDI Production Order."""

from __future__ import annotations

import re
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from wfx_panel.automation._common import (
    Any,
    Callable,
    Frame,
    Page,
    Playwright,
    PlaywrightError,
    PlaywrightTimeoutError,
    _click,
    _first_line,
    _result,
    _wait,
    _write_log,
    sync_playwright,
    time,
)
from wfx_panel.automation.modules import _active_wfx_page
from wfx_panel.automation.oc import (
    _attached_in_frames,
    _select_exact_option,
    _toolbar_link,
)
from wfx_panel.automation.runtime import (
    cancellation_deferred,
    checkpoint,
    claim_download,
)

REPORT_URL = (
    "https://prosports.worldfashionexchange.com/WFXBase4.0/"
    "WFXBICustomReportView.aspx?"
    "BICustomReportID=966ae3a5-1edb-4f60-8290-2e3e630aa41f&"
    "Path=/WFXPSHLIVE/Production%20Reports/BuyerDispatchOrder_Invoice&"
    "ReportParams="
)
REPORT_DOC_NO_SELECTOR = "#rptCustomReportViewer_ctl04_ctl03_txtValue"
REPORT_VIEW_SELECTOR = "#rptCustomReportViewer_ctl04_ctl00"
REPORT_EXPORT_LINK_SELECTOR = (
    "#rptCustomReportViewer_ctl05_ctl04_ctl00_ButtonLink"
)
REPORT_EXPORT_IMAGE_SELECTOR = (
    "#rptCustomReportViewer_ctl05_ctl04_ctl00_ButtonImg"
)
REPORT_EXPORT_MENU_SELECTOR = "#rptCustomReportViewer_ctl05_ctl04_ctl00_Menu"
REPORT_EXCEL_SELECTOR = (
    f"{REPORT_EXPORT_MENU_SELECTOR} "
    'a[title="Excel"][onclick*="EXCELOPENXML"]'
)

EDI_MENU_XPATH = '//*[@id="0040_0020_0100"]/a'
PACKAGE_TYPE_SELECTOR = "#ddlPackageType"
PACKAGE_TYPE_VALUE = "Import"
PACKAGE_LABEL = "DecisionOne_BuyerOrderDispatch"
PACKAGE_VALUE = "2"
EDI_GRID_SELECTOR = "#gridEDIProductionOrder_tblGridContent"
EDI_UPLOAD_SELECTOR = "#sectionObjectAttachment input[type='file']"
EDI_CREATE_SELECTOR = (
    'table.clsSectionTitleBar[id="sectionEDIProductionOrder"] '
    "a.ToolLink"
)

REPORT_TIMEOUT_SECONDS = 100
PACKAGE_TIMEOUT_SECONDS = 100
TRANSACTION_TIMEOUT_SECONDS = 150
GDN_PROGRESS_TOTAL = 6


def _emit_progress(
    progress: Callable[..., None] | None,
    stage: str,
    message: str,
    step: int,
    *,
    state: str = "active",
) -> None:
    if progress is None:
        return
    try:
        progress(stage, message, step, GDN_PROGRESS_TOTAL, state=state)
    except Exception:
        # Tiến độ là UX phụ trợ; lỗi WebView không được làm hỏng transaction.
        pass


_EDI_ROWS_JS = r"""() => {
  const norm = value => String(value || '').replace(/\s+/g, ' ').trim();
  const value = (row, selector) => {
    const element = row.querySelector(selector);
    return norm(element?.getAttribute('title') || element?.value ||
      element?.innerText || element?.textContent);
  };
  const table = document.querySelector('#gridEDIProductionOrder_tblGridContent');
  if (!table) return [];
  return [...table.querySelectorAll(':scope > tbody > tr')].map(row => ({
    row_id: row.getAttribute('rowid') || row.id || '',
    package_name: value(row, '#lblPackageName'),
    transaction: value(row, '#lblTransaction'),
    file_name: value(row, '#lblFileName'),
    processed_on: value(row, '#lblProcessedON'),
    status: value(row, '#lblStatus'),
    transaction_detail: value(row, '#lnkTransactionCreatedInWFX'),
    error: value(row, '#lblErrorMsg')
  })).filter(row => row.row_id);
}"""


class DispatchFlowError(RuntimeError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        errors: list[str] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.errors = list(errors or ())


def _normalise_status(value: object) -> str:
    return re.sub(r"[^a-z]", "", str(value or "").casefold())


def _status_failed(*values: object) -> bool:
    text = " ".join(_normalise_status(value) for value in values)
    return any(
        marker in text
        for marker in (
            "fail",
            "error",
            "false",
            "invalid",
            "reject",
            "cancel",
            "notprocessed",
        )
    )


def _status_complete(*values: object) -> bool:
    text = " ".join(_normalise_status(value) for value in values)
    return any(
        marker in text
        for marker in ("success", "complete", "created", "processed")
    ) and not _status_failed(*values)


def _processed_sort_key(row: dict[str, str]) -> tuple[float, int, str]:
    raw = str(row.get("processed_on") or "").strip()
    timestamp = 0.0
    for pattern in (
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            timestamp = datetime.strptime(raw, pattern).timestamp()
            break
        except ValueError:
            continue
    row_id = str(row.get("row_id") or "")
    numeric_id = int(row_id) if row_id.isdigit() else 0
    return timestamp, numeric_id, row_id


def choose_latest_pending_row(
    rows: list[dict[str, str]],
    *,
    excluded_ids: set[str] | None = None,
) -> dict[str, str] | None:
    """Chọn đúng package Dispatch Pending mới nhất theo Processed ON."""
    excluded = excluded_ids or set()
    candidates = [
        row
        for row in rows
        if str(row.get("row_id") or "") not in excluded
        and str(row.get("package_name") or "").casefold()
        == PACKAGE_LABEL.casefold()
        and _normalise_status(row.get("transaction_detail")) == "pending"
    ]
    return max(candidates, key=_processed_sort_key) if candidates else None


def reload_dispatch_workbook(source: Path, target: Path) -> None:
    """Mở và ghi lại report thành XLSX sạch để WFX import ổn định."""
    workbook = None
    try:
        workbook = load_workbook(source, data_only=False, keep_links=False)
        if not workbook.sheetnames:
            raise ValueError("Workbook không có sheet.")
        if not any(
            sheet.max_row > 0 and sheet.max_column > 0
            for sheet in workbook.worksheets
        ):
            raise ValueError("Workbook không có dữ liệu.")
        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True
            calculation.calcMode = "auto"
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target)
    except Exception as error:
        raise DispatchFlowError(
            "GDN_WORKBOOK_RELOAD_FAILED",
            "Không thể reload file report thành XLSX để import.",
            errors=[f"{type(error).__name__}: {_first_line(error)}"],
        ) from error
    finally:
        if workbook is not None:
            workbook.close()
    if not target.is_file() or target.stat().st_size <= 0:
        raise DispatchFlowError(
            "GDN_WORKBOOK_RELOAD_FAILED",
            "File XLSX sau khi reload bị rỗng.",
        )


def _wait_report_ready(report_page: Page) -> None:
    deadline = time.monotonic() + REPORT_TIMEOUT_SECONDS
    last_report_text = ""
    while time.monotonic() < deadline:
        checkpoint()
        try:
            image = report_page.locator(REPORT_EXPORT_IMAGE_SELECTOR)
            if image.count():
                source = str(image.first.get_attribute("src") or "")
                async_wait = report_page.locator(
                    "#rptCustomReportViewer_AsyncWait"
                )
                loading = async_wait.count() and async_wait.first.is_visible()
                if source and "disabled" not in source.casefold() and not loading:
                    return
            report = report_page.locator("#rptCustomReportViewer_ctl09")
            if report.count():
                last_report_text = " ".join(
                    (report.first.inner_text(timeout=1_000) or "").split()
                )
                if re.search(r"no\s+(data|rows)|không\s+có\s+dữ\s+liệu", last_report_text, re.I):
                    raise DispatchFlowError(
                        "GDN_REPORT_EMPTY",
                        "Report không có dữ liệu cho Invoice GRN đã nhập.",
                    )
        except DispatchFlowError:
            raise
        except PlaywrightError:
            pass
        _wait(report_page, 150)
    suffix = f" Chi tiết: {last_report_text[:300]}" if last_report_text else ""
    raise DispatchFlowError(
        "GDN_REPORT_NOT_READY",
        f"Report Buyer Dispatch chưa load xong.{suffix}",
    )


def _download_report(
    report_page: Page,
    target: Path,
    log: Callable[[str], None],
) -> None:
    downloads: list[Any] = []

    def receive(download: Any) -> None:
        downloads.append(download)

    report_page.on("download", receive)
    try:
        export = report_page.locator(REPORT_EXPORT_LINK_SELECTOR)
        if not export.count() or not export.first.is_visible():
            raise DispatchFlowError(
                "GDN_REPORT_NOT_READY",
                "Nút Export của report chưa sẵn sàng.",
            )
        _click(export.first)
        menu = report_page.locator(REPORT_EXPORT_MENU_SELECTOR)
        menu.wait_for(state="visible", timeout=5_000)
        excel = report_page.locator(REPORT_EXCEL_SELECTOR)
        if excel.count() != 1 or not excel.first.is_visible():
            raise DispatchFlowError(
                "GDN_REPORT_NOT_READY",
                "Lựa chọn Excel trong menu Export chưa sẵn sàng.",
            )
        _write_log(log, "[GDN] Đang export report sang Excel...")
        _click(excel.first)
        deadline = time.monotonic() + 90
        while time.monotonic() < deadline and not downloads:
            _wait(report_page, 100)
        if not downloads:
            raise DispatchFlowError(
                "GDN_REPORT_DOWNLOAD_FAILED",
                "WFX không bắt đầu tải file Excel của report.",
            )
        with cancellation_deferred():
            target.parent.mkdir(parents=True, exist_ok=True)
            claim_download(downloads[0])
            downloads[0].save_as(target)
        if not target.is_file() or target.stat().st_size <= 0:
            raise DispatchFlowError(
                "GDN_REPORT_DOWNLOAD_FAILED",
                "File Excel tải từ report bị rỗng.",
            )
    finally:
        try:
            report_page.remove_listener("download", receive)
        except Exception:
            pass


def _prepare_dispatch_workbook(
    context: Any,
    invoice: str,
    temporary: Path,
    log: Callable[[str], None],
    progress: Callable[..., None] | None = None,
) -> Path:
    report_page: Page | None = None
    try:
        report_page = context.new_page()
        report_page.goto(
            REPORT_URL,
            wait_until="domcontentloaded",
            timeout=30_000,
        )
        doc_no = report_page.locator(REPORT_DOC_NO_SELECTOR)
        doc_no.wait_for(state="visible", timeout=20_000)
        doc_no.fill(invoice)
        if doc_no.input_value() != invoice:
            raise DispatchFlowError(
                "GDN_REPORT_NOT_READY",
                "WFX chưa xác nhận Doc No. trên report.",
            )
        _write_log(log, "[GDN] Đã điền Invoice GRN vào Doc No.")
        _click(report_page.locator(REPORT_VIEW_SELECTOR))
        _write_log(log, "[GDN] Đang chờ report Buyer Dispatch load...")
        _emit_progress(
            progress,
            "report",
            "Đang tải báo cáo Buyer Dispatch…",
            1,
        )
        _wait_report_ready(report_page)
        _emit_progress(
            progress,
            "download",
            "Báo cáo đã sẵn sàng · đang tải Excel…",
            2,
        )
        raw_report = temporary / "BuyerDispatchOrder_Invoice.download.xlsx"
        _download_report(report_page, raw_report, log)
        _emit_progress(
            progress,
            "workbook",
            "Đang chuẩn hóa workbook XLSX…",
            3,
        )
        upload_path = temporary / "BuyerDispatchOrder_Invoice.reload.xlsx"
        reload_dispatch_workbook(raw_report, upload_path)
        _write_log(log, "[GDN] Đã reload và save report thành XLSX.")
        return upload_path
    except DispatchFlowError:
        raise
    except PlaywrightTimeoutError as error:
        raise DispatchFlowError(
            "GDN_REPORT_NOT_READY",
            "Report Buyer Dispatch chưa sẵn sàng.",
            errors=[_first_line(error)],
        ) from error
    except PlaywrightError as error:
        raise DispatchFlowError(
            "GDN_REPORT_DOWNLOAD_FAILED",
            "Không thể tải report Buyer Dispatch từ WFX.",
            errors=[_first_line(error)],
        ) from error
    finally:
        if report_page is not None:
            try:
                report_page.close()
            except PlaywrightError:
                pass


def _edi_frame(page: Page, timeout_s: float = 30) -> Frame:
    deadline = time.monotonic() + timeout_s
    while time.monotonic() < deadline:
        checkpoint()
        for frame in page.frames:
            try:
                if (
                    frame.locator(PACKAGE_TYPE_SELECTOR).count()
                    and frame.locator("#ddlPackage").count()
                ):
                    return frame
            except PlaywrightError:
                continue
        _wait(page, 150)
    raise DispatchFlowError(
        "GDN_EDI_NOT_READY",
        "Màn EDI Production Order chưa sẵn sàng.",
    )


def _open_edi(page: Page, log: Callable[[str], None]) -> Frame:
    try:
        _edi_frame(page, timeout_s=1)
    except DispatchFlowError:
        menu = page.locator(f"xpath={EDI_MENU_XPATH}")
        menu.wait_for(state="attached", timeout=10_000)
        _write_log(log, "[GDN] Mở EDI Production Order.")
        _click(menu)
        _edi_frame(page)
    _select_exact_option(
        page,
        PACKAGE_TYPE_SELECTOR,
        PACKAGE_TYPE_VALUE,
        PACKAGE_TYPE_VALUE,
        "PackageType",
    )
    _select_exact_option(
        page,
        "#ddlPackage",
        PACKAGE_VALUE,
        PACKAGE_LABEL,
        "Package",
    )
    _write_log(log, f"[GDN] Đã chọn Package: {PACKAGE_LABEL}.")
    return _edi_frame(page, timeout_s=5)


def _edi_rows(frame: Frame) -> list[dict[str, str]]:
    try:
        value = frame.evaluate(_EDI_ROWS_JS)
    except PlaywrightError:
        return []
    if not isinstance(value, list):
        return []
    return [
        {str(key): str(item_value or "") for key, item_value in row.items()}
        for row in value
        if isinstance(row, dict)
    ]


def _visible_message(frame: Frame) -> str:
    try:
        messages = frame.locator(
            "#sectionObjectAttachment #lblSuccessMsg, "
            "#lblSuccessMsg, [role='alert'], "
            ".clsErrorMessage, .clsSuccessMessage"
        )
        for index in range(messages.count()):
            candidate = messages.nth(index)
            if not candidate.is_visible():
                continue
            text = " ".join((candidate.text_content() or "").split())
            if text:
                return text[:1000]
    except PlaywrightError:
        pass
    return ""


def _open_import_popup(page: Page) -> tuple[Frame, Any]:
    try:
        return _attached_in_frames(page, EDI_UPLOAD_SELECTOR, timeout_s=1)
    except PlaywrightTimeoutError:
        _frame, import_link = _toolbar_link(page, "Import", timeout_s=20)
        _click(import_link)
        return _attached_in_frames(page, EDI_UPLOAD_SELECTOR, timeout_s=20)


def _process_package(
    page: Page,
    frame: Frame,
    upload_path: Path,
    known_ids: set[str],
    log: Callable[[str], None],
) -> dict[str, str]:
    _upload_frame, file_input = _open_import_popup(page)
    file_input.set_input_files(str(upload_path))
    _write_log(log, "[GDN] Đã gắn file XLSX vào Import Excel.")
    _process_frame, process_link = _toolbar_link(
        page,
        "Process Package",
        timeout_s=15,
    )
    dialog_messages: list[str] = []

    def accept_dialog(dialog: Any) -> None:
        dialog_messages.append(" ".join(str(dialog.message or "").split()))
        dialog.accept()

    page.on("dialog", accept_dialog)
    try:
        with cancellation_deferred():
            _click(process_link)
            _write_log(log, "[GDN] Đã gửi Process Package; đang chờ dòng mới.")
            deadline = time.monotonic() + PACKAGE_TIMEOUT_SECONDS
            last_new_rows: list[dict[str, str]] = []
            while time.monotonic() < deadline:
                checkpoint()
                failed_dialog = next(
                    (
                        message
                        for message in dialog_messages
                        if _status_failed(message)
                    ),
                    "",
                )
                if failed_dialog:
                    raise DispatchFlowError(
                        "GDN_PACKAGE_PROCESS_FAILED",
                        "WFX từ chối Process Package.",
                        errors=[failed_dialog],
                    )
                message = _visible_message(frame)
                if message and _status_failed(message):
                    raise DispatchFlowError(
                        "GDN_PACKAGE_PROCESS_FAILED",
                        "WFX báo lỗi khi Process Package.",
                        errors=[message],
                    )
                rows = _edi_rows(frame)
                last_new_rows = [
                    row
                    for row in rows
                    if row.get("row_id") not in known_ids
                    and row.get("package_name", "").casefold()
                    == PACKAGE_LABEL.casefold()
                ]
                for row in last_new_rows:
                    if _status_failed(
                        row.get("status"),
                        row.get("transaction_detail"),
                        row.get("error"),
                    ):
                        raise DispatchFlowError(
                            "GDN_PACKAGE_PROCESS_FAILED",
                            "Package mới bị WFX báo lỗi.",
                            errors=[row.get("error") or row.get("status") or "Failed"],
                        )
                selected = choose_latest_pending_row(
                    rows,
                    excluded_ids=known_ids,
                )
                if selected is not None:
                    _write_log(
                        log,
                        "[GDN] Package mới đã Pending; chọn theo Processed ON.",
                    )
                    return selected
                _wait(page, 250)
        detail = last_new_rows[0] if last_new_rows else {}
        raise DispatchFlowError(
            "GDN_PENDING_NOT_FOUND",
            "Không tìm thấy Transaction Detail Pending của package mới.",
            errors=[
                str(detail.get("error") or detail.get("transaction_detail") or "")
            ],
        )
    finally:
        try:
            page.remove_listener("dialog", accept_dialog)
        except Exception:
            pass


def _select_transaction(frame: Frame, row: dict[str, str]) -> None:
    row_id = str(row.get("row_id") or "")
    if not row_id:
        raise DispatchFlowError(
            "GDN_PENDING_NOT_FOUND",
            "Dòng package Pending không có mã nhận diện.",
        )
    target = frame.locator(
        f'{EDI_GRID_SELECTOR} tr[rowid="{row_id}"] input[name="rdSelector"]'
    )
    if target.count() != 1 or not target.first.is_visible():
        raise DispatchFlowError(
            "GDN_PENDING_NOT_FOUND",
            "Dòng package Pending đã thay đổi trước khi chọn.",
        )
    if not target.first.is_checked():
        target.first.check(timeout=5_000)
    if not target.first.is_checked():
        raise DispatchFlowError(
            "GDN_PENDING_NOT_FOUND",
            "WFX chưa xác nhận chọn package Pending.",
        )


def _create_transaction_link(frame: Frame) -> Any:
    links = frame.locator(EDI_CREATE_SELECTOR)
    matches: list[Any] = []
    for index in range(links.count()):
        link = links.nth(index)
        try:
            if (
                link.is_visible()
                and link.is_enabled()
                and " ".join((link.inner_text() or "").casefold().split())
                == "create transaction"
            ):
                matches.append(link)
        except PlaywrightError:
            continue
    if len(matches) != 1:
        raise DispatchFlowError(
            "GDN_EDI_NOT_READY",
            "Nút Create Transaction chưa sẵn sàng.",
        )
    return matches[0]


def _wait_transaction_result(
    page: Page,
    frame: Frame,
    row_id: str,
    dialog_messages: list[str],
) -> tuple[bool, list[str]]:
    deadline = time.monotonic() + TRANSACTION_TIMEOUT_SECONDS
    seen_row = True
    missing_polls = 0
    last_detail = "Pending"
    while time.monotonic() < deadline:
        checkpoint()
        failed_dialogs = [
            message for message in dialog_messages if _status_failed(message)
        ]
        if failed_dialogs:
            return False, failed_dialogs
        completed_dialogs = [
            message for message in dialog_messages if _status_complete(message)
        ]
        if completed_dialogs:
            return True, completed_dialogs
        message = _visible_message(frame)
        if message and _status_failed(message):
            return False, [message]
        if (
            message
            and "upload" not in message.casefold()
            and _status_complete(message)
        ):
            return True, [message]

        rows = _edi_rows(frame)
        current = next(
            (row for row in rows if row.get("row_id") == row_id),
            None,
        )
        if current is None:
            try:
                grid_ready = frame.locator(EDI_GRID_SELECTOR).count() > 0
            except PlaywrightError:
                grid_ready = False
            missing_polls = missing_polls + 1 if grid_ready else 0
            # WFX refreshes the grid while Create Transaction is running.
            # Require several stable reads before treating a removed row as done.
            if seen_row and missing_polls >= 3:
                return True, ["Package đã rời danh sách chờ xử lý."]
        else:
            seen_row = True
            missing_polls = 0
            last_detail = str(current.get("transaction_detail") or "")
            if _status_failed(
                current.get("status"),
                last_detail,
                current.get("error"),
            ):
                return False, [
                    str(
                        current.get("error")
                        or last_detail
                        or current.get("status")
                        or "Failed"
                    )
                ]
            if _status_complete(last_detail):
                return True, [last_detail]
        _wait(page, 300)
    return False, [
        f"WFX chưa xác nhận hoàn tất; Transaction Detail hiện là {last_detail or '—'}."
    ]


def run_gdn_dispatch(
    invoice: str,
    log: Callable[[str], None] = print,
    progress: Callable[..., None] | None = None,
) -> dict[str, Any]:
    """Chạy trọn flow report -> XLSX -> EDI -> Create Transaction."""
    invoice = " ".join(str(invoice or "").split())
    if not invoice:
        return _result(
            False,
            "GDN_INVOICE_REQUIRED",
            "Hãy nhập Invoice GRN trước khi Submit.",
        )
    if len(invoice) > 100 or any(ord(character) < 32 for character in invoice):
        return _result(
            False,
            "GDN_INVOICE_INVALID",
            "Invoice GRN không hợp lệ.",
        )

    playwright: Playwright | None = None
    transaction_submitted = False
    active_stage = "report"
    active_step = 1

    def stage(
        name: str,
        message: str,
        step: int,
        _total: int | None = None,
        *,
        state: str = "active",
    ) -> None:
        nonlocal active_stage, active_step
        active_stage, active_step = name, step
        _emit_progress(progress, name, message, step, state=state)

    def failure_context(code: str) -> dict[str, Any]:
        inspect_edi = transaction_submitted or active_step >= 5 or code in {
            "GDN_PACKAGE_PROCESS_FAILED",
            "GDN_PENDING_NOT_FOUND",
            "GDN_TRANSACTION_FAILED",
            "GDN_TRANSACTION_UNCONFIRMED",
        }
        return {
            "failed_stage": active_stage,
            "failed_step": active_step,
            "checkpoint": "inspect_edi" if inspect_edi else "restart_safe",
            "safe_to_retry": not inspect_edi,
        }

    try:
        stage("report", "Đang mở báo cáo Buyer Dispatch…", 1)
        playwright = sync_playwright().start()
        browser, page = _active_wfx_page(playwright, log)
        context = browser.contexts[0]
        with tempfile.TemporaryDirectory(prefix="wfx-gdn-dispatch-") as temporary:
            upload_path = _prepare_dispatch_workbook(
                context,
                invoice,
                Path(temporary),
                log,
                stage,
            )
            stage("edi", "Đang mở EDI Production Order…", 4)
            frame = _open_edi(page, log)
            known_ids = {
                str(row.get("row_id") or "") for row in _edi_rows(frame)
            }
            stage("package", "Đang upload và Process Package…", 5)
            pending = _process_package(
                page,
                frame,
                upload_path,
                known_ids,
                log,
            )
            stage("transaction", "Đang tạo transaction và chờ WFX xác nhận…", 6)
            _select_transaction(frame, pending)
            create_link = _create_transaction_link(frame)
            dialog_messages: list[str] = []

            def accept_dialog(dialog: Any) -> None:
                dialog_messages.append(
                    " ".join(str(dialog.message or "").split())
                )
                dialog.accept()

            page.on("dialog", accept_dialog)
            try:
                transaction_submitted = True
                with cancellation_deferred():
                    _click(create_link)
                    _write_log(
                        log,
                        "[GDN] Đã gửi Create Transaction; đang chờ WFX hoàn tất.",
                    )
                    confirmed, confirmations = _wait_transaction_result(
                        page,
                        frame,
                        str(pending.get("row_id") or ""),
                        dialog_messages,
                    )
            finally:
                try:
                    page.remove_listener("dialog", accept_dialog)
                except Exception:
                    pass
            if not confirmed:
                failed = any(_status_failed(value) for value in confirmations)
                code = (
                    "GDN_TRANSACTION_FAILED"
                    if failed
                    else "GDN_TRANSACTION_UNCONFIRMED"
                )
                stage(
                    "transaction",
                    (
                        "WFX báo lỗi khi tạo transaction."
                        if failed
                        else "Transaction đã gửi nhưng chưa được WFX xác nhận."
                    ),
                    6,
                    state="failed" if failed else "pending",
                )
                return _result(
                    False,
                    code,
                    (
                        "WFX báo lỗi khi tạo GDN Dispatch."
                        if failed
                        else "Đã gửi Create Transaction nhưng WFX chưa xác nhận hoàn tất. "
                        "Không tự chạy lại để tránh tạo trùng."
                    ),
                    transaction_submitted=True,
                    errors=confirmations,
                    **failure_context(code),
                )
            stage(
                "transaction",
                "WFX đã xác nhận (GDN) Dispatch hoàn tất.",
                6,
                state="completed",
            )
            return _result(
                True,
                "GDN_DISPATCH_COMPLETED",
                "(GDN) Dispatch đã được WFX xử lý thành công.",
                transaction_submitted=True,
                confirmations=confirmations,
                failed_stage="",
                checkpoint="completed",
                safe_to_retry=False,
            )
    except RuntimeError as error:
        code = str(error)
        if code in {"CHROME_CLOSED", "NOT_LOGGED_IN"}:
            result = _result(
                False,
                code,
                (
                    "Trình duyệt làm việc chưa được mở."
                    if code == "CHROME_CLOSED"
                    else "Phiên WFX chưa đăng nhập hoặc đã hết hạn."
                ),
                transaction_submitted=transaction_submitted,
            )
            stage(
                active_stage,
                str(result.get("message") or "Không thể tiếp tục GDN."),
                active_step,
                state="failed",
            )
            return {**result, **failure_context(code)}
        if isinstance(error, DispatchFlowError):
            stage(
                active_stage,
                error.message,
                active_step,
                state=(
                    "pending"
                    if error.code == "GDN_PENDING_NOT_FOUND"
                    else "failed"
                ),
            )
            return _result(
                False,
                error.code,
                error.message,
                errors=error.errors,
                transaction_submitted=transaction_submitted,
                **failure_context(error.code),
            )
        raise
    except PlaywrightTimeoutError as error:
        code = (
            "GDN_TRANSACTION_UNCONFIRMED"
            if transaction_submitted
            else "GDN_EDI_NOT_READY"
        )
        stage(
            active_stage,
            "WFX chưa phản hồi trong thời gian chờ.",
            active_step,
            state="pending" if transaction_submitted else "failed",
        )
        return _result(
            False,
            code,
            (
                "Đã gửi Create Transaction nhưng mất xác nhận từ WFX. "
                "Không tự chạy lại để tránh tạo trùng."
                if transaction_submitted
                else "EDI Production Order chưa sẵn sàng."
            ),
            errors=[_first_line(error)],
            transaction_submitted=transaction_submitted,
            **failure_context(code),
        )
    except Exception as error:
        code = (
            "GDN_TRANSACTION_UNCONFIRMED"
            if transaction_submitted
            else "GDN_DISPATCH_FAILED"
        )
        stage(
            active_stage,
            "GDN dừng do lỗi chưa xác định.",
            active_step,
            state="pending" if transaction_submitted else "failed",
        )
        return _result(
            False,
            code,
            (
                "Đã gửi Create Transaction nhưng không đọc được kết quả WFX. "
                "Không tự chạy lại để tránh tạo trùng."
                if transaction_submitted
                else "Không hoàn tất được (GDN) Dispatch."
            ),
            errors=[f"{type(error).__name__}: {_first_line(error)}"],
            transaction_submitted=transaction_submitted,
            **failure_context(code),
        )
    finally:
        if playwright is not None:
            playwright.stop()


def open_gdn_status(log: Callable[[str], None] = print) -> dict[str, Any]:
    """Mở đúng EDI package GDN để user kiểm tra mà không ghi transaction."""
    playwright: Playwright | None = None
    try:
        playwright = sync_playwright().start()
        _browser, page = _active_wfx_page(playwright, log)
        frame = _open_edi(page, log)
        rows = [
            row
            for row in _edi_rows(frame)
            if row.get("package_name", "").casefold() == PACKAGE_LABEL.casefold()
        ]
        latest = max(rows, key=_processed_sort_key) if rows else None
        detail = str((latest or {}).get("transaction_detail") or "").strip()
        return _result(
            True,
            "GDN_STATUS_READY",
            (
                f"Đã mở EDI Production Order. Package GDN mới nhất: {detail}."
                if detail
                else "Đã mở EDI Production Order để kiểm tra package GDN."
            ),
            package_count=len(rows),
            latest_status=detail,
        )
    except RuntimeError as error:
        code = str(error)
        return _result(
            False,
            code,
            (
                "Trình duyệt làm việc chưa được mở."
                if code == "CHROME_CLOSED"
                else "Phiên WFX chưa đăng nhập hoặc đã hết hạn."
            ),
        )
    except DispatchFlowError as error:
        return _result(False, error.code, error.message, errors=error.errors)
    except Exception as error:
        return _result(
            False,
            "GDN_EDI_NOT_READY",
            "Không mở được EDI Production Order để kiểm tra GDN.",
            errors=[f"{type(error).__name__}: {_first_line(error)}"],
        )
    finally:
        if playwright is not None:
            playwright.stop()
