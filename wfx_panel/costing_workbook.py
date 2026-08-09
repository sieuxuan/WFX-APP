"""Định dạng XLSX Costing hai sheet, độc lập với DOM của WFX.

Người dùng chỉ làm việc trong một form Costing có cột chuẩn. Metadata kỹ thuật
nằm ở các cột ẩn cùng sheet; workbook không chứa selector thực thi.
"""

from __future__ import annotations

import json
import re
import zipfile
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from dataclasses import field as dataclass_field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FORMAT_VERSION = "2.1"
SUPPORTED_EXTENSIONS = {".xlsx"}
MAX_FILE_BYTES = 12 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_XLSX_MEMBERS = 5_000
MAX_SECTIONS = 250
MAX_ITEMS = 20_000
MAX_FIELDS = 120_000
MAX_CELL_CHARS = 20_000
CLEAR_MARKER = "__CLEAR__"
DEFAULT_NEW_ITEM_ROWS_PER_SECTION = 3

GUIDE_SHEET = "Hướng dẫn"
FORM_SHEET = "Costing"

FIELD_SCOPES = {"cost_sheet", "section", "item"}
ITEM_ACTIONS = {"UPSERT", "DELETE"}
ITEM_TYPES = {"article", "cost_line"}
FORM_BASE_COLUMNS = [
    "Section",
    "Action",
    "Article Code",
    "Article Name",
]
FORM_TECH_COLUMNS = [
    "__Section Key",
    "__Item Key",
    "__Row Order",
    "__Item Type",
]
STANDARD_ITEM_FIELDS = (
    {
        "field_key": "colMaterialSizeList",
        "label": "Material Size",
        "data_type": "text",
    },
    {
        "field_key": "colMaterialColorList",
        "label": "Material Color",
        "data_type": "text",
    },
    {
        "field_key": "colColorDependency",
        "label": "Color Dep.",
        "data_type": "text",
    },
    {
        "field_key": "colColorDependencyMapping",
        "label": "Color Mapping",
        "data_type": "text",
    },
    {
        "field_key": "colSizeDependency",
        "label": "Size Dep.",
        "data_type": "text",
    },
    {
        "field_key": "colSizeDependencyMapping",
        "label": "Size Mapping",
        "data_type": "text",
    },
    {
        "field_key": "colShrinkagePerRemarks",
        "label": "Shrinkage %(LxW)",
        "data_type": "text",
    },
    {
        "field_key": "colConsQty",
        "label": "Cons. Qty.",
        "data_type": "number",
    },
    {
        "field_key": "colWastagePer",
        "label": "Waste %",
        "data_type": "number",
    },
    {
        "field_key": "Minutes",
        "label": "Minutes",
        "data_type": "number",
    },
    {
        "field_key": "colConsPlusWastageQty",
        "label": "Cons. Qty. Incl. Waste",
        "data_type": "number",
        "read_only": True,
    },
    {
        "field_key": "colSupplierCompanyName",
        "label": "Supplier",
        "data_type": "text",
    },
    {
        "field_key": "colCurrencyCode",
        "label": "Curr.",
        "data_type": "text",
    },
    {
        "field_key": "colRate1",
        "label": "Rate",
        "data_type": "number",
    },
    {
        "field_key": "colValue",
        "label": "Value",
        "data_type": "number",
    },
    {
        "field_key": "colValueInCSCurr",
        "label": "Value in (USD)",
        "data_type": "number",
        "read_only": True,
    },
    {
        "field_key": "colRemarks",
        "label": "Remarks",
        "data_type": "text",
    },
    {
        "field_key": "colPlacement",
        "label": "Placement",
        "data_type": "text",
    },
    {
        "field_key": "colPurchaseOfficer",
        "label": "Purchase Officer",
        "data_type": "text",
    },
)
FORM_COLUMNS = [
    *FORM_BASE_COLUMNS,
    *[str(field["label"]) for field in STANDARD_ITEM_FIELDS],
    *FORM_TECH_COLUMNS,
]
OPTIONAL_FORM_COLUMNS = {
    "Color Mapping",
    "Size Mapping",
    "Cons. Qty. Incl. Waste",
    "Value in (USD)",
}

_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_SUBHEADER_FILL = PatternFill("solid", fgColor="DFF4F2")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_TEMPLATE_INPUT_FILL = PatternFill("solid", fgColor="FFE699")
_META_FILL = PatternFill("solid", fgColor="E8EEF5")
_READ_ONLY_FILL = PatternFill("solid", fgColor="F4CCCC")
_READ_ONLY_HEADER_FILL = PatternFill("solid", fgColor="B91C1C")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(color="0F5660", bold=True, size=16)
_THIN_GRAY = Side(style="thin", color="D4DEE5")
_SECTION_SIDE = Side(style="medium", color="0F766E")
_STRUCTURE_BORDER = Border(bottom=_THIN_GRAY)
_SECTION_BORDER = Border(top=_SECTION_SIDE, bottom=_THIN_GRAY)
_DANGEROUS_EXCEL_PREFIXES = ("=", "+", "-", "@")
STANDARD_SECTIONS = (
    ("fabricshell", "FABRIC- SHELL"),
    ("fabriclining", "FABRIC - LINING"),
    ("fabricinterlining", "FABRIC - INTERLINING"),
    ("fabricpadding", "FABRIC - PADDING"),
    ("sewingtrims", "SEWING TRIMS"),
    ("packingtrims", "PACKING TRIMS"),
    ("cmcosts", "CM Costs"),
    ("productioncosts", "Production Costs"),
    ("indirectcosts", "Indirect Costs"),
)
ARTICLE_SECTION_TOKENS = frozenset(token for token, _name in STANDARD_SECTIONS[:6])
SPECIAL_COST_SECTION_ROWS = {
    "cmcosts": 1,
    "productioncosts": 1,
    "indirectcosts": 2,
}
_EXCLUDED_FIELD_TOKENS = {
    "deliveryterms",
    "processrequired",
    "bomsno",
    "bomdtsno",
    "rolllotavg",
    "destinationcountry",
    "desspecific",
    "destinationspecific",
    "materialcostincludedin",
}
_EXCLUDED_FIELD_KEYS = {
    "colbom",
    "colsno",
    "colrolllot",
    "colavg",
    "coldes",
    "colspecific",
    "colmaterialcost",
    "colincludedin",
}
_EXCLUDED_FIELD_LABELS = {
    "deliveryterms",
    "processrequired",
    "bom",
    "sno",
    "rolllot",
    "avg",
    "destinationcountry",
    "des",
    "specific",
    "materialcost",
    "includedin",
}


class CostingWorkbookError(ValueError):
    """Lỗi file có mã ổn định để UI/telemetry phân loại."""

    def __init__(self, code: str, message: str, *, details: Sequence[str] = ()):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = [str(item) for item in details]

    def as_result(self) -> dict[str, Any]:
        return {
            "ok": False,
            "code": self.code,
            "message": self.message,
            "validation_errors": list(self.details),
        }


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _text(value).casefold() in {"1", "true", "yes", "y", "x", "có"}


def _order(value: Any, fallback: int = 0) -> int:
    try:
        return int(float(_text(value)))
    except (TypeError, ValueError):
        return fallback


def _options(value: Any) -> list[str]:
    if isinstance(value, (list, tuple)):
        return [_text(item) for item in value if _text(item)]
    raw = _text(value)
    if not raw:
        return []
    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [_text(item) for item in parsed if _text(item)]
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    return [_text(item) for item in raw.split("|") if _text(item)]


def _excel_safe(value: Any) -> Any:
    """Giữ identifier là text và vô hiệu hóa formula injection khi export."""
    if not isinstance(value, str):
        return value
    if value.startswith(_DANGEROUS_EXCEL_PREFIXES):
        return "'" + value
    return value


def _excel_unescape(value: Any) -> Any:
    if (
        isinstance(value, str)
        and len(value) >= 2
        and value[0] == "'"
        and value[1] in _DANGEROUS_EXCEL_PREFIXES
    ):
        return value[1:]
    return value


def _reject_formula(value: Any, location: str) -> None:
    if isinstance(value, str) and value.startswith("="):
        raise CostingWorkbookError(
            "COSTING_FORMULA_NOT_ALLOWED",
            "File Costing không được chứa công thức trong vùng dữ liệu.",
            details=[location],
        )


def _clean_cell(value: Any, location: str) -> Any:
    _reject_formula(value, location)
    value = _excel_unescape(value)
    if isinstance(value, str) and len(value) > MAX_CELL_CHARS:
        raise CostingWorkbookError(
            "COSTING_VALIDATION_FAILED",
            "Một ô trong file Costing dài hơn giới hạn cho phép.",
            details=[location],
        )
    return value


def _normalized_field(raw: Mapping[str, Any], index: int) -> dict[str, Any]:
    scope = _text(raw.get("scope")).casefold()
    return {
        "scope": scope,
        "section_key": _text(raw.get("section_key")),
        "item_key": _text(raw.get("item_key")),
        "field_key": _text(raw.get("field_key")),
        "label": _text(raw.get("label") or raw.get("field_label")),
        "value": raw.get("value", ""),
        "data_type": _text(raw.get("data_type") or "text").casefold(),
        "editable": _bool(raw.get("editable")),
        "required": _bool(raw.get("required")),
        "options": _options(raw.get("options")),
        "row_order": _order(raw.get("row_order"), index),
    }


def _article_lookup_options(value: object) -> list[dict[str, str]]:
    if not isinstance(value, (list, tuple)):
        return []
    options = []
    seen = set()
    for raw in value:
        if not isinstance(raw, Mapping):
            continue
        code = _text(raw.get("article_code"))
        name = _text(raw.get("article_name"))
        if not code or not name:
            continue
        identity = code.casefold()
        if identity in seen:
            continue
        seen.add(identity)
        options.append({"article_code": code, "article_name": name})
    return options


def _normalized_section(raw: Mapping[str, Any], index: int) -> dict[str, Any]:
    return {
        "section_key": _text(raw.get("section_key")),
        "name": _text(raw.get("name") or raw.get("section_name")),
        "row_order": _order(raw.get("row_order"), index),
        "article_options": _options(raw.get("article_options")),
        "article_code_options": _options(raw.get("article_code_options")),
        "article_name_options": _options(raw.get("article_name_options")),
        "article_lookup_options": _article_lookup_options(
            raw.get("article_lookup_options")
        ),
    }


def _normalized_item(raw: Mapping[str, Any], index: int) -> dict[str, Any]:
    action = _text(raw.get("action") or "UPSERT").upper()
    item_type = _text(raw.get("item_type") or "article").casefold()
    return {
        "section_key": _text(raw.get("section_key")),
        "section_name": _text(raw.get("section_name")),
        "item_key": _text(raw.get("item_key")),
        "row_order": _order(raw.get("row_order"), index),
        "action": action,
        "item_type": item_type,
        "article_code": _text(raw.get("article_code")),
        "article_name": _text(raw.get("article_name")),
    }


def normalize_document(document: Mapping[str, Any]) -> dict[str, Any]:
    """Chuẩn hóa và validate document trước khi ghi file hoặc lập dry-run."""
    fields = [
        _normalized_field(field, index)
        for index, field in enumerate(document.get("fields") or ())
        if isinstance(field, Mapping)
    ]
    sections = [
        _normalized_section(section, index)
        for index, section in enumerate(document.get("sections") or ())
        if isinstance(section, Mapping)
    ]
    items = [
        _normalized_item(item, index)
        for index, item in enumerate(document.get("items") or ())
        if isinstance(item, Mapping)
    ]
    normalized = {
        "format_version": _text(
            document.get("format_version") or FORMAT_VERSION
        ),
        "style_code": _text(document.get("style_code")),
        "style_name": _text(document.get("style_name")),
        "title": _text(document.get("title")),
        "cost_sheet_status": _text(document.get("cost_sheet_status")),
        "cost_sheet_type": _text(
            document.get("cost_sheet_type") or "Internal Cost Sheets"
        ),
        "order_execution_type": _text(
            document.get("order_execution_type") or "Trading"
        ),
        "season": _text(document.get("season")),
        "template": _text(document.get("template") or "FOB"),
        "signature": _text(document.get("signature")),
        "fields": fields,
        "sections": sections,
        "items": items,
    }
    _validate_document(normalized)
    return normalized


def _semantic_token(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _text(value).casefold())


def _standard_section(
    section: Mapping[str, Any],
) -> tuple[int, str] | None:
    semantics = {
        _semantic_token(section.get("section_key")),
        _semantic_token(section.get("name")),
    }
    for index, (token, name) in enumerate(STANDARD_SECTIONS):
        if any(token in semantic for semantic in semantics):
            return index, name
    # Tương thích export cũ/test từng gọi section đầu tiên là "Fabric".
    if "fabric" in semantics:
        return 0, STANDARD_SECTIONS[0][1]
    return None


def _standard_section_token(section: Mapping[str, Any]) -> str:
    standard = _standard_section(section)
    return "" if standard is None else STANDARD_SECTIONS[standard[0]][0]


def _section_item_type(section: Mapping[str, Any]) -> str:
    return (
        "article"
        if _standard_section_token(section) in ARTICLE_SECTION_TOKENS
        else "cost_line"
    )


def _template_row_count(section: Mapping[str, Any]) -> int:
    return SPECIAL_COST_SECTION_ROWS.get(
        _standard_section_token(section),
        DEFAULT_NEW_ITEM_ROWS_PER_SECTION,
    )


def _production_summary_item(item: Mapping[str, Any]) -> bool:
    identity = _semantic_token(
        item.get("article_name") or item.get("item_key")
    )
    return identity in {"productioncosts", "otherprocessescost"}


def _form_field_key(
    definition: Mapping[str, Any],
    section: Mapping[str, Any],
) -> str:
    if (
        definition.get("label") == "Value"
        and _standard_section_token(section) == "productioncosts"
    ):
        return "ProductionValue"
    return str(definition["field_key"])


def _field_is_excluded(field: Mapping[str, Any]) -> bool:
    field_key = _semantic_token(
        re.sub(r"__\d+$", "", _text(field.get("field_key")))
    )
    label = _semantic_token(field.get("label"))
    return (
        any(token in field_key for token in _EXCLUDED_FIELD_TOKENS)
        or field_key in _EXCLUDED_FIELD_KEYS
        or label in _EXCLUDED_FIELD_LABELS
        or label in _EXCLUDED_FIELD_TOKENS
    )


def workbook_document(document: Mapping[str, Any]) -> dict[str, Any]:
    """Giữ đúng các dòng và ô người dùng được phép sửa trong file Costing."""
    normalized = normalize_document(document)
    selected_sections: dict[int, dict[str, Any]] = {}
    for section in normalized["sections"]:
        standard = _standard_section(section)
        if standard is None:
            continue
        index, name = standard
        selected_sections.setdefault(
            index,
            {
                **section,
                "name": name,
                "row_order": index + 1,
            },
        )
    for index, (token, name) in enumerate(STANDARD_SECTIONS):
        selected_sections.setdefault(
            index,
            {
                "section_key": f"standard:{token}",
                "name": name,
                "row_order": index + 1,
            },
        )
    sections = [
        selected_sections[index]
        for index in sorted(selected_sections)
    ]
    allowed_section_keys = {
        section["section_key"].casefold() for section in sections
    }
    section_by_key = {
        section["section_key"].casefold(): section for section in sections
    }
    items = [
        item
        for item in normalized["items"]
        if item["section_key"].casefold() in allowed_section_keys
        and item["item_type"]
        == _section_item_type(section_by_key[item["section_key"].casefold()])
        and not (
            _standard_section_token(
                section_by_key[item["section_key"].casefold()]
            )
            == "productioncosts"
            and _production_summary_item(item)
        )
    ]
    allowed_item_keys = {
        (item["section_key"].casefold(), item["item_key"].casefold())
        for item in items
    }
    fields = []
    standard_keys = {
        str(field["field_key"]).casefold()
        for field in STANDARD_ITEM_FIELDS
    } | {"productionvalue", "productionheaderminutes"}
    read_only_keys = {
        str(field["field_key"]).casefold()
        for field in STANDARD_ITEM_FIELDS
        if field.get("read_only")
    }
    for field in normalized["fields"]:
        base_field_key = re.sub(
            r"__\d+$",
            "",
            field["field_key"],
        ).casefold()
        if (
            field["scope"] != "item"
            or (
                not field["editable"]
                and base_field_key not in read_only_keys
            )
            or _field_is_excluded(field)
            or base_field_key not in standard_keys
        ):
            continue
        section_key = field["section_key"].casefold()
        if (
            field["scope"] in {"section", "item"}
            and allowed_section_keys
            and section_key not in allowed_section_keys
        ):
            continue
        if field["scope"] == "item" and (
            section_key,
            field["item_key"].casefold(),
        ) not in allowed_item_keys:
            continue
        fields.append(field)
    return normalize_document(
        {
            **normalized,
            "sections": sections,
            "items": items,
            "fields": fields,
        }
    )


def _document_size_errors(
    sections: Sequence[Mapping[str, Any]],
    items: Sequence[Mapping[str, Any]],
    fields: Sequence[Mapping[str, Any]],
) -> list[str]:
    errors: list[str] = []
    if len(sections) > MAX_SECTIONS:
        errors.append(f"Quá {MAX_SECTIONS} section.")
    if len(items) > MAX_ITEMS:
        errors.append(f"Quá {MAX_ITEMS} Article.")
    if len(fields) > MAX_FIELDS:
        errors.append(f"Quá {MAX_FIELDS} field.")
    return errors


def _section_validation_errors(
    sections: Sequence[Mapping[str, Any]],
) -> list[str]:
    errors: list[str] = []
    section_keys: set[str] = set()
    for section in sections:
        key = _text(section.get("section_key"))
        if not key:
            errors.append("Section thiếu Section Key.")
        elif key in section_keys:
            errors.append(f"Section Key trùng: {key}.")
        section_keys.add(key)
    return errors


def _item_validation_errors(
    items: Sequence[Mapping[str, Any]],
) -> list[str]:
    errors: list[str] = []
    item_keys: set[tuple[str, str]] = set()
    for item in items:
        section_key = _text(item.get("section_key"))
        item_key = _text(item.get("item_key"))
        action = _text(item.get("action") or "UPSERT").upper()
        if not section_key:
            errors.append("Article thiếu Section Key.")
        if not item_key:
            errors.append(
                f"Article {_text(item.get('article_code')) or '(trống)'} "
                "thiếu Item Key."
            )
        composite = (section_key.casefold(), item_key.casefold())
        if all(composite) and composite in item_keys:
            errors.append(f"Item Key trùng trong section: {item_key}.")
        item_keys.add(composite)
        if action not in ITEM_ACTIONS:
            errors.append(f"Action không hợp lệ: {action}.")
        item_type = _text(item.get("item_type") or "article").casefold()
        if item_type not in ITEM_TYPES:
            errors.append(f"Item Type không hợp lệ: {item_type}.")
        if item_type == "article" and not (
            _text(item.get("article_code"))
            or _text(item.get("article_name"))
        ):
            errors.append(f"Article {item_key or '(trống)'} thiếu Code/Name.")
        if item_type == "cost_line" and not _text(item.get("article_name")):
            errors.append(f"Dòng chi phí {item_key or '(trống)'} thiếu tên.")
    return errors


def _field_validation_errors(
    fields: Sequence[Mapping[str, Any]],
) -> list[str]:
    errors: list[str] = []
    field_keys: set[tuple[str, str, str, str]] = set()
    for field in fields:
        scope = _text(field.get("scope")).casefold()
        field_key = _text(field.get("field_key"))
        if scope not in FIELD_SCOPES:
            errors.append(f"Field scope không hợp lệ: {scope or '(trống)'}.")
        if not field_key:
            errors.append("Field thiếu Field Key.")
        composite = (
            scope,
            _text(field.get("section_key")).casefold(),
            _text(field.get("item_key")).casefold(),
            field_key.casefold(),
        )
        if field_key and composite in field_keys:
            errors.append(
                "Field Key trùng trong cùng scope: "
                + " / ".join(part or "-" for part in composite)
            )
        field_keys.add(composite)
        if scope in {"section", "item"} and not _text(
            field.get("section_key")
        ):
            errors.append(f"Field {field_key or '(trống)'} thiếu Section Key.")
        if scope == "item" and not _text(field.get("item_key")):
            errors.append(f"Item field {field_key or '(trống)'} thiếu Item Key.")
        _reject_formula(field.get("value"), f"Field {field_key or '(trống)'}")
    return errors


def _validate_document(document: Mapping[str, Any]) -> None:
    if document.get("format_version") != FORMAT_VERSION:
        raise CostingWorkbookError(
            "COSTING_FORMAT_UNSUPPORTED",
            "Phiên bản file Costing không được hỗ trợ.",
            details=[
                f"Nhận {document.get('format_version') or 'trống'}; "
                f"cần {FORMAT_VERSION}."
            ],
        )
    sections = list(document.get("sections") or ())
    items = list(document.get("items") or ())
    fields = list(document.get("fields") or ())
    errors = [] if _text(document.get("style_code")) else ["Thiếu Style Code."]
    errors.extend(_document_size_errors(sections, items, fields))
    errors.extend(_section_validation_errors(sections))
    errors.extend(_item_validation_errors(items))
    errors.extend(_field_validation_errors(fields))

    if errors:
        raise CostingWorkbookError(
            "COSTING_VALIDATION_FAILED",
            "File Costing có dữ liệu chưa hợp lệ.",
            details=errors[:100],
        )


def _preflight_path(path: str | Path, *, must_exist: bool) -> Path:
    raw_path = str(path or "").strip()
    if not raw_path:
        raise CostingWorkbookError(
            "COSTING_FILE_REQUIRED",
            "Chưa chọn file Costing.",
        )
    target = Path(raw_path).expanduser()
    suffix = target.suffix.casefold()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise CostingWorkbookError(
            "COSTING_FILE_TYPE_UNSUPPORTED",
            "Costing chỉ hỗ trợ file .xlsx.",
        )
    if must_exist:
        if not target.is_file():
            raise CostingWorkbookError(
                "COSTING_FILE_REQUIRED",
                "File Costing không còn tồn tại.",
            )
        if target.stat().st_size > MAX_FILE_BYTES:
            raise CostingWorkbookError(
                "COSTING_FILE_TOO_LARGE",
                "File Costing lớn hơn giới hạn 12 MB.",
            )
        if suffix == ".xlsx":
            _preflight_xlsx_archive(target)
    return target


def _preflight_xlsx_archive(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise CostingWorkbookError(
                    "COSTING_FILE_TOO_LARGE",
                    "Workbook có quá nhiều thành phần.",
                )
            total = sum(max(0, member.file_size) for member in members)
            if total > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise CostingWorkbookError(
                    "COSTING_FILE_TOO_LARGE",
                    "Workbook giải nén lớn hơn giới hạn an toàn.",
                )
    except CostingWorkbookError:
        raise
    except (OSError, zipfile.BadZipFile) as error:
        raise CostingWorkbookError(
            "COSTING_VALIDATION_FAILED",
            "File XLSX bị hỏng hoặc không đúng định dạng.",
        ) from error


def _set_header(ws: Any, row: int, columns: Sequence[str]) -> None:
    for column, value in enumerate(columns, 1):
        cell = ws.cell(row=row, column=column, value=value)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.auto_filter.ref = (
        f"A{row}:{get_column_letter(len(columns))}{max(row, ws.max_row)}"
    )
    ws.freeze_panes = f"A{row + 1}"


def _finish_sheet(
    ws: Any,
    widths: Mapping[int, float],
    *,
    editable_columns: Iterable[int] = (),
) -> None:
    editable = set(editable_columns)
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _STRUCTURE_BORDER
            if cell.column in editable:
                cell.fill = _INPUT_FILL
    ws.sheet_view.showGridLines = False


def _write_guide(workbook: Workbook, document: Mapping[str, Any]) -> None:
    ws = workbook.create_sheet(GUIDE_SHEET)
    ws.merge_cells("A1:F1")
    ws["A1"] = "WFX Smart · Catalog Costing"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    rows = [
        ("Style Code", document["style_code"]),
        ("Style Name", document.get("style_name") or "—"),
        ("Trạng thái khi tải", document.get("cost_sheet_status") or "—"),
        ("Phiên bản file", f"v{FORMAT_VERSION}"),
        (
            "Cách dùng",
            "Mở sheet Costing và sửa các ô màu vàng. Dòng không dùng thì để trống.",
        ),
        (
            "Phạm vi",
            "Có thể tải file ở mọi trạng thái. Chỉ Cost Sheet đang Open mới "
            "được cập nhật lại lên WFX.",
        ),
        (
            "Các nhóm trong file",
            ", ".join(
                str(section.get("name") or section.get("section_key") or "")
                for section in document.get("sections") or ()
                if str(
                    section.get("name") or section.get("section_key") or ""
                ).strip()
            )
            or "—",
        ),
        (
            "Thêm nhiều dòng",
            "Điền lần lượt các dòng vàng trong đúng nhóm. Nếu cần thêm dòng, "
            "sao chép dòng vàng cuối và chèn trước nhóm tiếp theo.",
        ),
        (
            "Tách dòng màu/size",
            "Hai dòng liền nhau cùng Article Code sẽ được tách thành hai dòng "
            "riêng trên WFX.",
        ),
        (
            "Phối màu/size",
            "Mỗi dòng phối ghi Màu/Size vật tư => Màu/Size của style. Nhiều lựa "
            "chọn được ngăn bằng dấu |.",
        ),
        (
            "Cột công thức",
            "Cons. Qty. Incl. Waste = Cons. Qty. × (1 + Waste %/100); "
            "Value in (USD) = Rate × Cons. Qty. Incl. Waste. Hai cột đỏ chỉ đọc.",
        ),
        (
            "Purchase Officer",
            "Chọn từ danh sách nếu ô này trên WFX đang trống. App sẽ báo trước "
            "khi còn thiếu.",
        ),
        (
            "Cột Action",
            "Để trống = thêm mới hoặc cập nhật. Chọn DELETE chỉ khi muốn xóa dòng.",
        ),
        (
            "Xóa giá trị",
            f"Ô trống sẽ giữ nguyên dữ liệu cũ. Muốn xóa, ghi đúng {CLEAR_MARKER}.",
        ),
        (
            "An toàn",
            "App luôn cho xem trước thay đổi và chỉ lưu sau khi điền xong.",
        ),
        (
            "CM / Production / Indirect",
            "Chọn tên trong danh sách. CM và Indirect dùng USD. Production tự "
            "đặt Minutes = 1. Dòng không chọn tên sẽ không được thêm.",
        ),
    ]
    for row, (label, value) in enumerate(rows, 3):
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=2, value=_excel_safe(value))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )
        ws.row_dimensions[row].height = 42 if row >= 6 else 24
        if label == "Cách dùng":
            ws.cell(row=row, column=2).fill = _INPUT_FILL
    for column, width in enumerate((23, 18, 18, 18, 18, 18), 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False


def _base_field_key(value: Any) -> str:
    return re.sub(r"__\d+$", "", _text(value)).casefold()


def _unique_values(values: Iterable[Any]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _text(value)
        folded = text.casefold()
        if not text or folded in seen:
            continue
        seen.add(folded)
        output.append(text)
    return output


def _split_wfx_multiselect(value: Any) -> list[str]:
    """Tách chuỗi WFX theo dấu phẩy nhưng giữ nguyên dấu phẩy trong ngoặc."""
    text = _text(value)
    if not text:
        return []
    output: list[str] = []
    start = 0
    depth = 0
    for index, character in enumerate(text):
        if character == "(":
            depth += 1
        elif character == ")" and depth:
            depth -= 1
        elif character in {",", "|"} and depth == 0:
            token = text[start:index].strip()
            if token:
                output.append(token)
            start = index + 1
    token = text[start:].strip()
    if token:
        output.append(token)
    return output


def _form_dropdown_options(
    document: Mapping[str, Any],
) -> dict[str, list[str]]:
    """Tạo lookup chung; option Material được gắn riêng theo từng item."""
    fields = list(document.get("fields") or ())

    def matching(field_key: str) -> list[Mapping[str, Any]]:
        wanted = field_key.casefold()
        return [
            field
            for field in fields
            if _base_field_key(field.get("field_key")) == wanted
        ]

    purchase_officers = matching("colPurchaseOfficer")
    return {
        "Color Dep.": _unique_values(
            [
                "[None]",
                "[Table]",
                "[Body Type]",
                "[Base Colors]",
            ]
        ),
        "Size Dep.": _unique_values(
            [
                "[None]",
                "[Table]",
                "[Body Type]",
            ]
        ),
        "Purchase Officer": _unique_values(
            [
                *(field.get("value") for field in purchase_officers),
                *(
                    option
                    for field in purchase_officers
                    for option in field.get("options") or ()
                ),
            ]
        ),
    }


def _add_form_dropdowns(
    ws: Any,
    document: Mapping[str, Any],
    *,
    last_row: int,
) -> int:
    """Gắn dropdown bằng lookup columns ẩn, không tạo sheet thứ ba."""
    option_sets = _form_dropdown_options(document)
    lookup_column = len(FORM_COLUMNS) + 1
    for label, options in option_sets.items():
        if not options:
            continue
        column_letter = get_column_letter(lookup_column)
        for row, option in enumerate(options, 2):
            ws.cell(row=row, column=lookup_column, value=_excel_safe(option))
        ws.column_dimensions[column_letter].hidden = True
        validation = DataValidation(
            type="list",
            formula1=f"=${column_letter}$2:${column_letter}${len(options) + 1}",
            allow_blank=True,
        )
        # Cho phép gõ mapping mới hoặc nhiều giá trị bằng dấu | khi lookup live
        # chưa có đủ màu/size. Dropdown vẫn là đường nhập mặc định.
        validation.showErrorMessage = False
        validation.promptTitle = "WFX Smart"
        validation.prompt = (
            "Chọn từ danh sách; có thể gõ nhiều giá trị, ngăn bằng dấu |."
        )
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        target_column = FORM_COLUMNS.index(label) + 1
        target_letter = get_column_letter(target_column)
        validation.add(f"{target_letter}2:{target_letter}{last_row}")
        lookup_column += 1
    return lookup_column


def _add_item_option_dropdowns(
    ws: Any,
    document: Mapping[str, Any],
    row_by_item: Mapping[tuple[str, str], int],
    *,
    lookup_column: int,
) -> None:
    """Gắn dropdown Material Color/Size đúng option của từng Article row."""
    wanted_fields = {
        "colmaterialcolorlist": "Material Color",
        "colmaterialsizelist": "Material Size",
    }
    for field in document.get("fields") or ():
        base_key = _base_field_key(field.get("field_key"))
        label = wanted_fields.get(base_key)
        if label is None:
            continue
        row = row_by_item.get(
            (
                _text(field.get("section_key")).casefold(),
                _text(field.get("item_key")).casefold(),
            )
        )
        options = _unique_values(
            [
                *(field.get("options") or ()),
                *_split_wfx_multiselect(field.get("value")),
            ]
        )
        if row is None or not options:
            continue
        lookup_letter = get_column_letter(lookup_column)
        for option_row, option in enumerate(options, 2):
            ws.cell(
                row=option_row,
                column=lookup_column,
                value=_excel_safe(option),
            )
        ws.column_dimensions[lookup_letter].hidden = True
        validation = DataValidation(
            type="list",
            formula1=(
                f"=${lookup_letter}$2:"
                f"${lookup_letter}${len(options) + 1}"
            ),
            allow_blank=True,
        )
        validation.showErrorMessage = False
        validation.promptTitle = f"{label} của Article"
        validation.prompt = (
            "Chọn một giá trị; nhiều giá trị có thể nhập theo đúng chuỗi WFX."
        )
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        target_letter = get_column_letter(FORM_COLUMNS.index(label) + 1)
        validation.add(f"{target_letter}{row}")
        lookup_column += 1


def _add_special_article_dropdowns(
    ws: Any,
    document: Mapping[str, Any],
    layout: _CostingFormLayout,
    *,
    lookup_column: int,
) -> int:
    """Gắn danh sách nhà máy/quy trình/chi phí cho từng block đặc biệt."""
    article_name_letter = get_column_letter(FORM_COLUMNS.index("Article Name") + 1)
    ordered_sections = sorted(
        document["sections"],
        key=lambda value: value["row_order"],
    )
    for section, rows in zip(ordered_sections, layout.section_rows, strict=True):
        if _section_item_type(section) != "cost_line":
            continue
        options = _unique_values(section.get("article_options") or ())
        if not options or not rows:
            continue
        lookup_letter = get_column_letter(lookup_column)
        for option_row, option in enumerate(options, 2):
            ws.cell(option_row, lookup_column, _excel_safe(option))
        ws.column_dimensions[lookup_letter].hidden = True
        validation = DataValidation(
            type="list",
            formula1=(
                f"=${lookup_letter}$2:"
                f"${lookup_letter}${len(options) + 1}"
            ),
            allow_blank=True,
        )
        validation.showErrorMessage = False
        validation.promptTitle = "Chọn từ WFX"
        validation.prompt = "Chọn tên đã quét; có thể gõ để tìm trong Excel."
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        validation.add(
            f"{article_name_letter}{min(rows)}:{article_name_letter}{max(rows)}"
        )
        lookup_column += 1
    return lookup_column


def _add_material_article_dropdowns(
    ws: Any,
    document: Mapping[str, Any],
    layout: _CostingFormLayout,
    *,
    lookup_column: int,
) -> int:
    """Dropdown Article Code/Name từ thư viện server/cache."""
    ordered_sections = sorted(
        document["sections"],
        key=lambda value: value["row_order"],
    )
    validation_cache: dict[
        tuple[tuple[str, str], ...],
        tuple[DataValidation, DataValidation, str, str, int],
    ] = {}
    code_column = FORM_COLUMNS.index("Article Code") + 1
    name_column = FORM_COLUMNS.index("Article Name") + 1
    code_target_letter = get_column_letter(code_column)
    name_target_letter = get_column_letter(name_column)
    for section, rows in zip(ordered_sections, layout.section_rows, strict=True):
        if _section_item_type(section) == "cost_line" or not rows:
            continue
        pairs = _article_lookup_options(section.get("article_lookup_options"))
        if not pairs:
            codes = _unique_values(section.get("article_code_options") or ())
            names = _unique_values(section.get("article_name_options") or ())
            pairs = [
                {"article_code": code, "article_name": names[index]}
                for index, code in enumerate(codes)
                if index < len(names)
            ]
        if not pairs:
            continue
        signature = tuple(
            (option["article_code"], option["article_name"])
            for option in pairs
        )
        target_code_range = (
            f"{code_target_letter}{min(rows)}:"
            f"{code_target_letter}{max(rows)}"
        )
        target_name_range = (
            f"{name_target_letter}{min(rows)}:"
            f"{name_target_letter}{max(rows)}"
        )
        cached = validation_cache.get(signature)
        if cached is None:
            lookup_code_letter = get_column_letter(lookup_column)
            lookup_name_letter = get_column_letter(lookup_column + 1)
            for option_row, option in enumerate(pairs, 2):
                ws.cell(
                    option_row,
                    lookup_column,
                    _excel_safe(option["article_code"]),
                )
                ws.cell(
                    option_row,
                    lookup_column + 1,
                    _excel_safe(option["article_name"]),
                )
            ws.column_dimensions[lookup_code_letter].hidden = True
            ws.column_dimensions[lookup_name_letter].hidden = True
            last_option_row = len(pairs) + 1

            def validation_for(
                letter: str,
                label: str,
                option_last_row: int = last_option_row,
            ) -> DataValidation:
                validation = DataValidation(
                    type="list",
                    formula1=(
                        f"=${letter}$2:${letter}${option_last_row}"
                    ),
                    allow_blank=True,
                )
                validation.showErrorMessage = False
                validation.promptTitle = f"{label} từ thư viện"
                validation.prompt = (
                    "Danh sách tự đồng bộ từ server; vẫn có thể nhập tay."
                )
                validation.showInputMessage = True
                ws.add_data_validation(validation)
                return validation

            code_validation = validation_for(
                lookup_code_letter,
                "Article Code",
            )
            name_validation = validation_for(
                lookup_name_letter,
                "Article Name",
            )
            cached = (
                code_validation,
                name_validation,
                lookup_code_letter,
                lookup_name_letter,
                last_option_row,
            )
            validation_cache[signature] = cached
            lookup_column += 2
        (
            code_validation,
            name_validation,
            lookup_code_letter,
            lookup_name_letter,
            last_option_row,
        ) = cached
        code_validation.add(target_code_range)
        name_validation.add(target_name_range)
        available_codes = {
            option["article_code"].casefold() for option in pairs
        }
        for row in rows:
            current_code = _text(ws.cell(row, code_column).value)
            current_name = _text(ws.cell(row, name_column).value)
            if current_name and current_code.casefold() not in available_codes:
                continue
            ws.cell(row, name_column).value = (
                f'=IFERROR(INDEX(${lookup_name_letter}$2:'
                f'${lookup_name_letter}${last_option_row},'
                f'MATCH({code_target_letter}{row},'
                f'${lookup_code_letter}$2:'
                f'${lookup_code_letter}${last_option_row},0)),"")'
            )
    return lookup_column


@dataclass(frozen=True)
class _CostingFormLayout:
    section_rows: list[list[int]]
    template_rows: list[tuple[int, str]]
    row_by_item: dict[tuple[str, str], int]


def _form_field_index(
    document: Mapping[str, Any],
) -> dict[tuple[str, str, str], Mapping[str, Any]]:
    return {
        (
            field["section_key"].casefold(),
            field["item_key"].casefold(),
            _base_field_key(field["field_key"]),
        ): field
        for field in document["fields"]
    }


def _form_items_by_section(
    document: Mapping[str, Any],
) -> dict[str, list[Mapping[str, Any]]]:
    grouped: dict[str, list[Mapping[str, Any]]] = {}
    for item in document["items"]:
        grouped.setdefault(item["section_key"].casefold(), []).append(item)
    for items in grouped.values():
        items.sort(key=lambda value: value["row_order"])
    return grouped


def _form_row_values(
    section: Mapping[str, Any],
    item: Mapping[str, Any] | None,
    row_order: int,
    field_by_item: Mapping[tuple[str, str, str], Mapping[str, Any]],
) -> list[Any]:
    section_key = str(section["section_key"])
    item_key = "" if item is None else str(item["item_key"])
    values: list[Any] = [
        section["name"],
        "" if item is None or item["action"] == "UPSERT" else item["action"],
        "" if item is None else item["article_code"],
        "" if item is None else item["article_name"],
    ]
    for definition in STANDARD_ITEM_FIELDS:
        field_key = _form_field_key(definition, section)
        field = field_by_item.get(
            (
                section_key.casefold(),
                item_key.casefold(),
                field_key.casefold(),
            )
        )
        value = "" if field is None else field["value"]
        token = _standard_section_token(section)
        if definition["label"] == "Minutes" and token == "productioncosts":
            value = 1
        elif definition["label"] == "Curr." and token in {
            "cmcosts",
            "indirectcosts",
        }:
            value = "USD"
        values.append(value)
    values.extend([section_key, item_key, row_order, _section_item_type(section)])
    return values


def _write_form_data_rows(
    ws: Any,
    document: Mapping[str, Any],
) -> _CostingFormLayout:
    field_by_item = _form_field_index(document)
    items_by_section = _form_items_by_section(document)
    section_rows: list[list[int]] = []
    template_rows: list[tuple[int, str]] = []
    row_by_item: dict[tuple[str, str], int] = {}
    output_row = 2
    for section in sorted(
        document["sections"],
        key=lambda value: value["row_order"],
    ):
        section_key = str(section["section_key"])
        section_items = items_by_section.get(section_key.casefold(), [])
        highest_row_order = max(
            (int(item["row_order"]) for item in section_items),
            default=0,
        )
        current_section_rows: list[int] = []
        rows_to_write = [
            *section_items,
            *([None] * _template_row_count(section)),
        ]
        for template_index, item in enumerate(rows_to_write, 1):
            is_template = item is None
            row_order = (
                highest_row_order + template_index - len(section_items)
                if is_template
                else int(item["row_order"])
            )
            for column, value in enumerate(
                _form_row_values(
                    section,
                    item,
                    row_order,
                    field_by_item,
                ),
                1,
            ):
                ws.cell(
                    row=output_row,
                    column=column,
                    value=_excel_safe(value),
                )
            if is_template:
                template_rows.append((output_row, str(section["name"])))
            else:
                row_by_item[
                    (section_key.casefold(), str(item["item_key"]).casefold())
                ] = output_row
            current_section_rows.append(output_row)
            output_row += 1
        section_rows.append(current_section_rows)
    return _CostingFormLayout(section_rows, template_rows, row_by_item)


def _form_column_widths() -> dict[int, int]:
    widths = {1: 24, 2: 12, 3: 20, 4: 30}
    special_widths = {
        "Material Color": 32,
        "Material Size": 28,
        "Color Mapping": 44,
        "Size Mapping": 44,
        "Cons. Qty. Incl. Waste": 24,
        "Value in (USD)": 20,
        "Remarks": 28,
        "Supplier": 25,
        "Placement": 20,
        "Purchase Officer": 20,
        "Shrinkage %(LxW)": 20,
    }
    for offset, definition in enumerate(
        STANDARD_ITEM_FIELDS,
        len(FORM_BASE_COLUMNS) + 1,
    ):
        widths[offset] = special_widths.get(str(definition["label"]), 16)
    return widths


def _style_formula_columns(ws: Any, last_row: int) -> set[int]:
    read_only_definitions = [
        definition
        for definition in STANDARD_ITEM_FIELDS
        if definition.get("read_only")
    ]
    read_only_columns: set[int] = set()
    for definition in read_only_definitions:
        column = FORM_COLUMNS.index(str(definition["label"])) + 1
        read_only_columns.add(column)
        header = ws.cell(row=1, column=column)
        header.fill = _READ_ONLY_HEADER_FILL
        header.font = _HEADER_FONT
        for row in range(2, last_row + 1):
            cell = ws.cell(row=row, column=column)
            cell.fill = _READ_ONLY_FILL
            cell.font = Font(color="991B1B")
        formula = (
            "Cons. Qty. × (1 + Waste %/100)"
            if definition["field_key"] == "colConsPlusWastageQty"
            else "Rate × Cons. Qty. Incl. Waste"
        )
        header.comment = Comment(
            f"Cột công thức WFX, chỉ đọc: {formula}.",
            "WFX Smart",
        )
    return read_only_columns


def _write_costing_formulas(ws: Any, last_row: int) -> None:
    columns = {
        label: FORM_COLUMNS.index(label) + 1
        for label in (
            "Cons. Qty.",
            "Waste %",
            "Cons. Qty. Incl. Waste",
            "Rate",
            "Value in (USD)",
        )
    }
    for row in range(2, last_row + 1):
        cons_ref = f"{get_column_letter(columns['Cons. Qty.'])}{row}"
        waste_ref = f"{get_column_letter(columns['Waste %'])}{row}"
        cons_incl_ref = (
            f"{get_column_letter(columns['Cons. Qty. Incl. Waste'])}{row}"
        )
        rate_ref = f"{get_column_letter(columns['Rate'])}{row}"
        cons_incl_cell = ws.cell(
            row=row,
            column=columns["Cons. Qty. Incl. Waste"],
        )
        value_cell = ws.cell(row=row, column=columns["Value in (USD)"])
        cons_incl_cell.value = (
            f'=IF({cons_ref}="","",{cons_ref}*'
            f'(1+IF({waste_ref}="",0,{waste_ref})/100))'
        )
        value_cell.value = (
            f'=IF(OR({rate_ref}="",{cons_incl_ref}=""),"",'
            f"{rate_ref}*{cons_incl_ref})"
        )
        cons_incl_cell.number_format = "0.0000"
        value_cell.number_format = "0.0000"


def _add_dependency_mapping_comments(
    ws: Any,
    document: Mapping[str, Any],
    row_by_item: Mapping[tuple[str, str], int],
) -> None:
    mapping_columns = {
        "colcolordependencymapping": FORM_COLUMNS.index("Color Mapping") + 1,
        "colsizedependencymapping": FORM_COLUMNS.index("Size Mapping") + 1,
    }
    for field in document.get("fields") or ():
        column = mapping_columns.get(_base_field_key(field.get("field_key")))
        if column is None:
            continue
        row = row_by_item.get(
            (
                _text(field.get("section_key")).casefold(),
                _text(field.get("item_key")).casefold(),
            )
        )
        if row is None:
            continue
        choices = _unique_values(field.get("options") or ())
        detail = "\n".join(choices[:100]) or "Chưa scan được option Style."
        cell = ws.cell(row=row, column=column)
        cell.comment = Comment(
            "Mỗi dòng: Material => Style 1 | Style 2\n\n"
            "Các lựa chọn Style đã scan:\n" + detail,
            "WFX Smart",
        )
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_form_sections(
    ws: Any,
    section_rows: Sequence[Sequence[int]],
    visible_column_count: int,
) -> None:
    for rows in section_rows:
        if not rows:
            continue
        for row in rows:
            section_cell = ws.cell(row=row, column=1)
            section_cell.fill = _SUBHEADER_FILL
            section_cell.font = Font(color="0F5660", bold=True)
        for column in range(1, visible_column_count + 1):
            ws.cell(row=rows[0], column=column).border = _SECTION_BORDER


def _style_form_templates(
    ws: Any,
    template_rows: Sequence[tuple[int, str]],
    visible_column_count: int,
    read_only_columns: set[int],
) -> None:
    article_code_column = FORM_COLUMNS.index("Article Code") + 1
    article_name_column = FORM_COLUMNS.index("Article Name") + 1
    for row, section_name in template_rows:
        ws.cell(row=row, column=1).font = Font(
            color="0F5660",
            bold=True,
            italic=True,
        )
        for column in range(2, visible_column_count + 1):
            if column not in read_only_columns:
                ws.cell(row=row, column=column).fill = _TEMPLATE_INPUT_FILL
        section = {"section_key": section_name, "name": section_name}
        is_cost_line = _section_item_type(section) == "cost_line"
        hint = (
            f"Chọn Article Name cho {section_name}; để trống nếu không dùng."
            if is_cost_line
            else (
                f"Dòng thêm Article mới cho {section_name}. "
                "Nhập Article Code hoặc Article Name; để trống nếu không dùng."
            )
        )
        ws.cell(row=row, column=article_code_column).comment = Comment(
            hint,
            "WFX Smart",
        )
        ws.cell(row=row, column=article_name_column).comment = Comment(
            hint,
            "WFX Smart",
        )


def _finish_costing_form(
    ws: Any,
    layout: _CostingFormLayout,
    visible_column_count: int,
) -> None:
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(visible_column_count)}{max(1, ws.max_row)}"
    )
    ws.row_dimensions[1].height = 34
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 23
    for mapping_label in ("Color Mapping", "Size Mapping"):
        column = FORM_COLUMNS.index(mapping_label) + 1
        for row in layout.row_by_item.values():
            line_count = str(ws.cell(row=row, column=column).value or "").count(
                "\n"
            ) + 1
            if line_count > 1:
                ws.row_dimensions[row].height = max(
                    ws.row_dimensions[row].height or 23,
                    min(120, 16 * line_count),
                )
    for column in range(visible_column_count + 1, len(FORM_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(column)].hidden = True
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0F766E"


def _write_costing_form(
    workbook: Workbook,
    document: Mapping[str, Any],
) -> None:
    """Ghi form một sheet với bộ cột cố định và dòng thêm Article sẵn có."""
    ws = workbook.create_sheet(FORM_SHEET)
    _set_header(ws, 1, FORM_COLUMNS)
    visible_column_count = len(FORM_COLUMNS) - len(FORM_TECH_COLUMNS)
    layout = _write_form_data_rows(ws, document)

    form_data_last_row = ws.max_row
    last_row = max(form_data_last_row, 200)
    action_validation = DataValidation(
        type="list",
        formula1='"DELETE"',
        allow_blank=True,
    )
    ws.add_data_validation(action_validation)
    action_validation.add(f"B2:B{last_row}")
    next_lookup_column = _add_form_dropdowns(
        ws,
        document,
        last_row=last_row,
    )
    next_lookup_column = _add_special_article_dropdowns(
        ws,
        document,
        layout,
        lookup_column=next_lookup_column,
    )
    next_lookup_column = _add_material_article_dropdowns(
        ws,
        document,
        layout,
        lookup_column=next_lookup_column,
    )
    _add_item_option_dropdowns(
        ws,
        document,
        layout.row_by_item,
        lookup_column=next_lookup_column,
    )
    for offset, definition in enumerate(
        STANDARD_ITEM_FIELDS,
        len(FORM_BASE_COLUMNS) + 1,
    ):
        ws.cell(row=1, column=offset).comment = Comment(
            f"WFX Field Key: {definition['field_key']}",
            "WFX Smart",
        )
    _finish_sheet(
        ws,
        _form_column_widths(),
        editable_columns=range(2, visible_column_count + 1),
    )
    read_only_columns = _style_formula_columns(ws, form_data_last_row)
    _write_costing_formulas(ws, form_data_last_row)
    _add_dependency_mapping_comments(ws, document, layout.row_by_item)
    _style_form_sections(ws, layout.section_rows, visible_column_count)
    _style_form_templates(
        ws,
        layout.template_rows,
        visible_column_count,
        read_only_columns,
    )
    _finish_costing_form(ws, layout, visible_column_count)


def write_costing_xlsx(document: Mapping[str, Any], path: str | Path) -> Path:
    normalized = workbook_document(document)
    target = _preflight_path(path, must_exist=False)
    if target.suffix.casefold() != ".xlsx":
        raise CostingWorkbookError(
            "COSTING_FILE_TYPE_UNSUPPORTED",
            "Đường dẫn export XLSX phải kết thúc bằng .xlsx.",
        )
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    _write_guide(workbook, normalized)
    _write_costing_form(workbook, normalized)
    workbook.save(target)
    return target


def write_costing_file(document: Mapping[str, Any], path: str | Path) -> Path:
    target = _preflight_path(path, must_exist=False)
    return write_costing_xlsx(document, target)


def _worksheet_rows(
    ws: Any,
    *,
    ignored_columns: set[str] | None = None,
) -> Iterable[dict[str, Any]]:
    ignored = ignored_columns or set()
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    columns = [_text(value) for value in header]
    for row_index, values in enumerate(rows, 2):
        if not any(value not in (None, "") for value in values):
            continue
        output: dict[str, Any] = {}
        for column_index, key in enumerate(columns):
            if not key or key in ignored:
                continue
            value = values[column_index] if column_index < len(values) else None
            output[key] = _clean_cell(
                value,
                f"{ws.title}!{get_column_letter(column_index + 1)}{row_index}",
            )
        output["__Excel Row"] = row_index
        yield output


def _read_guide_meta(workbook: Any) -> dict[str, Any]:
    if workbook.sheetnames != [GUIDE_SHEET, FORM_SHEET]:
        raise CostingWorkbookError(
            "COSTING_FORMAT_UNSUPPORTED",
            "Workbook Costing phải chỉ có hai sheet: Hướng dẫn và Costing.",
        )
    values = {
        _text(row[0]): _clean_cell(row[1], f"{GUIDE_SHEET}!B{index}")
        for index, row in enumerate(
            workbook[GUIDE_SHEET].iter_rows(values_only=True),
            1,
        )
        if len(row) >= 2 and _text(row[0])
    }
    version = _text(values.get("Phiên bản file")).removeprefix("v")
    return {
        "format_version": version,
        "style_code": _text(values.get("Style Code")),
        "style_name": _text(values.get("Style Name")),
        "cost_sheet_status": _text(values.get("Trạng thái khi tải")),
        "title": "",
        "cost_sheet_type": "Internal Cost Sheets",
        "order_execution_type": "Trading",
        "season": "",
        "template": "FOB",
        "signature": "",
    }


@dataclass
class _CostingFormReadState:
    sections: list[dict[str, Any]] = dataclass_field(default_factory=list)
    items: list[dict[str, Any]] = dataclass_field(default_factory=list)
    fields: list[dict[str, Any]] = dataclass_field(default_factory=list)
    errors: list[str] = dataclass_field(default_factory=list)
    section_keys: set[str] = dataclass_field(default_factory=set)
    item_locations: dict[tuple[str, str], int] = dataclass_field(
        default_factory=dict
    )


@dataclass(frozen=True)
class _FormItemMetadata:
    section_key: str
    item_key: str
    action: str
    item_type: str
    row_index: int


def _missing_form_columns(ws: Any) -> list[str]:
    header = [
        _text(cell.value)
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    return [
        column
        for column in FORM_COLUMNS
        if column not in header and column not in OPTIONAL_FORM_COLUMNS
    ]


_ARTICLE_NAME_FORMULA = re.compile(
    r'^=IFERROR\(INDEX\(\$([A-Z]+)\$2:\$\1\$(\d+),'
    r'MATCH\(([A-Z]+)(\d+),\$([A-Z]+)\$2:\$\5\$\2,0\)\),""\)$'
)
_ARTICLE_VALIDATION_RANGE = re.compile(
    r"^=\$([A-Z]+)\$2:\$\1\$(\d+)$"
)


def _resolve_generated_article_name_formulas(ws: Any) -> None:
    """Đổi riêng công thức lookup do app tạo thành text trước khi validate."""
    code_column = FORM_COLUMNS.index("Article Code") + 1
    name_column = FORM_COLUMNS.index("Article Name") + 1
    code_letter = get_column_letter(code_column)
    lookup_cache: dict[tuple[str, str, int], dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, name_column)
        formula = cell.value
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        match = _ARTICLE_NAME_FORMULA.fullmatch(formula)
        if match is None:
            continue
        (
            lookup_name_letter,
            last_row_text,
            target_code_letter,
            target_row_text,
            lookup_code_letter,
        ) = match.groups()
        last_row = int(last_row_text)
        if (
            target_code_letter != code_letter
            or int(target_row_text) != row
            or last_row < 2
            or last_row > ws.max_row
            or not ws.column_dimensions[lookup_code_letter].hidden
            or not ws.column_dimensions[lookup_name_letter].hidden
        ):
            continue
        cache_key = (
            lookup_code_letter,
            lookup_name_letter,
            last_row,
        )
        lookup = lookup_cache.get(cache_key)
        if lookup is None:
            lookup = {
                _text(ws[f"{lookup_code_letter}{lookup_row}"].value).casefold():
                _text(ws[f"{lookup_name_letter}{lookup_row}"].value)
                for lookup_row in range(2, last_row + 1)
                if _text(ws[f"{lookup_code_letter}{lookup_row}"].value)
            }
            lookup_cache[cache_key] = lookup
        article_code = _text(ws.cell(row, code_column).value).casefold()
        cell.value = lookup.get(article_code, "")


def _resolve_article_codes_selected_by_name(ws: Any) -> None:
    """Đồng bộ ngược Code khi người dùng chọn Name từ dropdown của app."""
    code_column = FORM_COLUMNS.index("Article Code") + 1
    name_column = FORM_COLUMNS.index("Article Name") + 1
    errors: list[str] = []
    for validation in ws.data_validations.dataValidation:
        formula = str(validation.formula1 or "")
        match = _ARTICLE_VALIDATION_RANGE.fullmatch(formula)
        if match is None:
            continue
        lookup_name_letter, last_row_text = match.groups()
        lookup_name_column = column_index_from_string(lookup_name_letter)
        lookup_code_column = lookup_name_column - 1
        last_row = int(last_row_text)
        if (
            lookup_code_column < 1
            or last_row < 2
            or last_row > ws.max_row
            or not ws.column_dimensions[lookup_name_letter].hidden
            or not ws.column_dimensions[
                get_column_letter(lookup_code_column)
            ].hidden
        ):
            continue
        codes_by_name: dict[str, list[str]] = {}
        names_by_code: dict[str, str] = {}
        for lookup_row in range(2, last_row + 1):
            name = _text(ws.cell(lookup_row, lookup_name_column).value)
            code = _text(ws.cell(lookup_row, lookup_code_column).value)
            if not name or not code:
                continue
            names_by_code.setdefault(code.casefold(), name)
            matching_codes = codes_by_name.setdefault(name.casefold(), [])
            if code.casefold() not in {
                value.casefold() for value in matching_codes
            }:
                matching_codes.append(code)
        for target_range in validation.ranges.ranges:
            if not (
                target_range.min_col <= name_column <= target_range.max_col
            ):
                continue
            for row in range(
                max(2, target_range.min_row),
                min(ws.max_row, target_range.max_row) + 1,
            ):
                name_cell = ws.cell(row, name_column)
                article_name = _text(name_cell.value)
                if not article_name:
                    continue
                if article_name.startswith("="):
                    # Excel/WPS có thể tự viết lại công thức app sinh (thêm @,
                    # _xlfn hoặc đổi cách đặt ngoặc). Chỉ cho phép ở đúng ô
                    # Article Name có validation trỏ vào hai cột lookup ẩn;
                    # không chạy công thức mà thay bằng text từ Code cùng dòng.
                    current_code = _text(ws.cell(row, code_column).value)
                    name_cell.value = names_by_code.get(
                        current_code.casefold(),
                        "",
                    )
                    continue
                matching_codes = codes_by_name.get(article_name.casefold(), [])
                if len(matching_codes) == 1:
                    ws.cell(row, code_column).value = matching_codes[0]
                    continue
                current_code = _text(ws.cell(row, code_column).value)
                if len(matching_codes) > 1 and current_code.casefold() not in {
                    value.casefold() for value in matching_codes
                }:
                    errors.append(
                        f"{FORM_SHEET}!{name_cell.coordinate}: Article Name "
                        f"“{article_name}” trùng {len(matching_codes)} mã; "
                        "hãy chọn Article Code."
                    )
    if errors:
        raise CostingWorkbookError(
            "COSTING_VALIDATION_FAILED",
            "File Costing có Article Name chưa xác định được mã.",
            details=errors[:100],
        )


def _costing_form_rows(ws: Any) -> list[dict[str, Any]]:
    _resolve_generated_article_name_formulas(ws)
    _resolve_article_codes_selected_by_name(ws)
    return list(
        _worksheet_rows(
            ws,
            ignored_columns={
                str(definition["label"])
                for definition in STANDARD_ITEM_FIELDS
                if definition.get("read_only")
            },
        )
    )


def _form_section_key_index(
    rows: Sequence[Mapping[str, Any]],
) -> dict[str, str]:
    return {
        _text(row.get("Section")).casefold(): _text(row.get("__Section Key"))
        for row in rows
        if _text(row.get("Section")) and _text(row.get("__Section Key"))
    }


def _register_form_section(
    state: _CostingFormReadState,
    section_key: str,
    section_name: str,
) -> None:
    normalized_key = section_key.casefold()
    if not section_key or normalized_key in state.section_keys:
        return
    state.section_keys.add(normalized_key)
    state.sections.append(
        {
            "section_key": section_key,
            "name": section_name or section_key,
            "row_order": len(state.sections) + 1,
        }
    )


def _form_row_has_data(row: Mapping[str, Any]) -> bool:
    if _text(row.get("__Item Type")).casefold() == "cost_line":
        # Các dòng mẫu đặc biệt có sẵn Minutes/Curr.; chỉ tạo dòng khi người
        # dùng thật sự chọn tên nhà máy/quy trình/chi phí.
        return bool(_text(row.get("Article Name")))
    if _text(row.get("Article Code")) or _text(row.get("Article Name")):
        return True
    return any(
        row.get(str(definition["label"]), "") not in (None, "")
        for definition in STANDARD_ITEM_FIELDS
    )


def _form_item_key(
    row: Mapping[str, Any],
    section_key: str,
    article_code: str,
    article_name: str,
    row_index: int,
) -> str:
    existing_key = _text(row.get("__Item Key"))
    if existing_key:
        return existing_key
    identity = article_code or article_name or f"row-{row_index}"
    # Adjacent duplicate Articles are valid Splitter requests. Excel row keeps
    # each new row uniquely addressable for planner occurrence matching.
    return f"new:{section_key}:{identity}:row-{row_index}"


def _validate_form_item_row(
    state: _CostingFormReadState,
    metadata: _FormItemMetadata,
) -> None:
    if not metadata.section_key:
        state.errors.append(
            f"{FORM_SHEET}!A{metadata.row_index}: Section không thuộc form chuẩn."
        )
    if metadata.action not in ITEM_ACTIONS:
        state.errors.append(
            f"{FORM_SHEET}!B{metadata.row_index}: "
            f"Action “{metadata.action}” không hợp lệ."
        )
    if metadata.item_type not in ITEM_TYPES:
        item_type_column = FORM_COLUMNS.index("__Item Type") + 1
        state.errors.append(
            f"{FORM_SHEET}!{get_column_letter(item_type_column)}"
            f"{metadata.row_index}: Item Type “{metadata.item_type}” "
            "không hợp lệ."
        )
    composite = (
        metadata.section_key.casefold(),
        metadata.item_key.casefold(),
    )
    if all(composite) and composite in state.item_locations:
        item_key_column = FORM_COLUMNS.index("__Item Key") + 1
        state.errors.append(
            f"{FORM_SHEET}!{get_column_letter(item_key_column)}"
            f"{metadata.row_index}: Item Key trùng dòng "
            f"{state.item_locations[composite]}."
        )
        return
    state.item_locations[composite] = metadata.row_index


def _form_fields_for_item(
    row: Mapping[str, Any],
    section_key: str,
    section_name: str,
    item_key: str,
) -> list[dict[str, Any]]:
    fields: list[dict[str, Any]] = []
    section = {"section_key": section_key, "name": section_name}
    for field_order, definition in enumerate(STANDARD_ITEM_FIELDS):
        if definition.get("read_only"):
            continue
        value = row.get(str(definition["label"]), "")
        if value in (None, ""):
            continue
        field_data = {
            "scope": "item",
            "section_key": section_key,
            "item_key": item_key,
            "field_key": _form_field_key(definition, section),
            "label": definition["label"],
            "value": value,
            "data_type": definition["data_type"],
            "editable": True,
            "required": False,
            "options": [],
            "row_order": field_order,
        }
        fields.append(_normalized_field(field_data, field_order))
        if (
            _standard_section_token(section) == "productioncosts"
            and definition["label"] == "Minutes"
        ):
            fields.append(
                _normalized_field(
                    {
                        **field_data,
                        "field_key": "ProductionHeaderMinutes",
                        "row_order": field_order - 1,
                    },
                    field_order,
                )
            )
    return fields


def _read_costing_form_row(
    row: Mapping[str, Any],
    fallback_row_index: int,
    section_key_by_name: Mapping[str, str],
    state: _CostingFormReadState,
) -> None:
    row_index = _order(row.get("__Excel Row"), fallback_row_index)
    section_name = _text(row.get("Section"))
    section_key = _text(row.get("__Section Key")) or section_key_by_name.get(
        section_name.casefold(),
        "",
    )
    _register_form_section(state, section_key, section_name)
    if not _form_row_has_data(row):
        return
    article_code = _text(row.get("Article Code"))
    article_name = _text(row.get("Article Name"))
    item_key = _form_item_key(
        row,
        section_key,
        article_code,
        article_name,
        row_index,
    )
    action = _text(row.get("Action") or "UPSERT").upper()
    item_type = _text(row.get("__Item Type") or "article").casefold()
    item_fields = _form_fields_for_item(
        row,
        section_key,
        section_name,
        item_key,
    )
    if (
        action == "UPSERT"
        and _text(row.get("__Item Key"))
        and not item_fields
    ):
        # Bản export cũ có thể gắn subtotal chỉ đọc vào một dòng ``>>`` ẩn.
        # Sau khi bỏ field công thức, dòng này chỉ còn identity và không có gì
        # để apply; bỏ qua để file cũ không tạo split/Purchase Officer giả.
        return
    _validate_form_item_row(
        state,
        _FormItemMetadata(
            section_key=section_key,
            item_key=item_key,
            action=action,
            item_type=item_type,
            row_index=row_index,
        ),
    )
    state.items.append(
        _normalized_item(
            {
                "section_key": section_key,
                "section_name": section_name,
                "item_key": item_key,
                "row_order": _order(row.get("__Row Order"), row_index),
                "action": action,
                "item_type": item_type,
                "article_code": article_code,
                "article_name": article_name,
            },
            row_index,
        )
    )
    state.fields.extend(item_fields)


def _read_costing_form(
    workbook: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    ws = workbook[FORM_SHEET]
    missing = _missing_form_columns(ws)
    if missing:
        raise CostingWorkbookError(
            "COSTING_FORMAT_UNSUPPORTED",
            "Sheet Costing thiếu cột chuẩn của WFX Smart.",
            details=[
                f"{FORM_SHEET}!hàng 1: thiếu cột “{column}”."
                for column in missing
            ],
        )
    rows = _costing_form_rows(ws)
    section_key_by_name = _form_section_key_index(rows)
    state = _CostingFormReadState()
    for fallback_row_index, row in enumerate(rows, 2):
        _read_costing_form_row(
            row,
            fallback_row_index,
            section_key_by_name,
            state,
        )
    if state.errors:
        raise CostingWorkbookError(
            "COSTING_VALIDATION_FAILED",
            "File Costing có dữ liệu chưa hợp lệ.",
            details=state.errors[:100],
        )
    return state.sections, state.items, state.fields


def read_costing_xlsx(path: str | Path) -> dict[str, Any]:
    target = _preflight_path(path, must_exist=True)
    try:
        workbook = load_workbook(
            target,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except CostingWorkbookError:
        raise
    except Exception as error:
        raise CostingWorkbookError(
            "COSTING_VALIDATION_FAILED",
            "Không đọc được workbook Costing.",
        ) from error
    meta = _read_guide_meta(workbook)
    if not _text(meta.get("style_code")):
        raise CostingWorkbookError(
            "COSTING_VALIDATION_FAILED",
            "File Costing có dữ liệu chưa hợp lệ.",
            details=[f"{GUIDE_SHEET}!B2: thiếu Style Code."],
        )
    sections, items, fields = _read_costing_form(workbook)
    document = {
        **meta,
        "fields": fields,
        "sections": sections,
        "items": items,
    }
    return workbook_document(document)


def read_costing_file(path: str | Path) -> dict[str, Any]:
    target = _preflight_path(path, must_exist=True)
    return read_costing_xlsx(target)


def costing_file_summary(
    document: Mapping[str, Any],
    path: str | Path,
) -> dict[str, Any]:
    normalized = workbook_document(document)
    target = Path(path)
    return {
        "file_name": target.name,
        "file_format": target.suffix.casefold().lstrip("."),
        "style_code": normalized["style_code"],
        "section_count": len(normalized["sections"]),
        "item_count": len(normalized["items"]),
        "field_count": len(normalized["fields"]),
    }
