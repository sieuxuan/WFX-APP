"""Form Excel và validation cho luồng tạo Sale ASN từ nhiều PO."""

from __future__ import annotations

import math
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

SALE_ASN_SHEET = "SALE ASN"
SALE_ASN_COLUMNS = (
    "Style No",
    "PO No",
    "Qty",
    "Price",
    "Carton",
    "NW",
    "GW",
    "CBM",
    "FOB Price",
    "Service Price",
    "Cargo Ready Date",
    "HS CODE",
    "Goods Description",
    "Invoice No",
    "Invoice Date",
    "Shipping Bill No",
    "Shipping Bill Date",
    "Destination",
    "FTY",
    "Consignee Address",
    "Ship To",
    "Shipping Mode",
)

SHIPPING_MODES = ("AIR", "SEA", "COURIER")

_TEMPLATE_NUMBER_COLUMNS = frozenset(
    {"Carton", "NW", "GW", "CBM", "FOB Price", "Service Price"}
)

MAX_SALE_ASN_ROWS = 2_000
MAX_SALE_ASN_BYTES = 20 * 1024 * 1024
# Số mũ tối đa của chữ số đầu; 15 tương đương 16 chữ số phần nguyên, thừa sức
# cho Qty/Carton/NW/GW/CBM/giá và vẫn an toàn cho quantize() ở lớp automation.
MAX_SALE_ASN_NUMBER_EXPONENT = 15

_OPTIONAL_COLUMNS = frozenset(
    {
        "HS CODE",
        "Goods Description",
        "Qty",
        "Price",
        "Carton",
        "NW",
        "GW",
        "CBM",
        "FOB Price",
        "Service Price",
        "Cargo Ready Date",
        "Consignee Address",
        "Ship To",
    }
)
_DATE_COLUMNS = frozenset(
    {"Invoice Date", "Shipping Bill Date", "Cargo Ready Date"}
)
_NUMBER_COLUMNS = frozenset(
    {"Qty", "Price", "Carton", "NW", "GW", "CBM", "FOB Price", "Service Price"}
)


class SaleASNWorkbookError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        errors: tuple[str, ...] = (),
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.errors = errors


@dataclass(frozen=True)
class SaleASNRow:
    source_row: int
    invoice_no: str
    invoice_date: str
    shipping_bill_no: str
    shipping_bill_date: str
    style_no: str
    po_no: str
    hs_code: str
    goods_description: str
    qty: str
    price: str
    carton: str
    nw: str
    gw: str
    cbm: str
    destination: str
    factory: str
    fob_price: str
    service_price: str
    cargo_ready_date: str
    consignee_address: str
    ship_to: str
    shipping_mode: str

    def automation_payload(self) -> dict[str, str | int]:
        return {
            field: getattr(self, field)
            for field in self.__dataclass_fields__
        }


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def _identifier(value: object) -> str:
    if isinstance(value, bool):
        return _text(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _text(value)


def _number(value: object, *, cell: str, integer: bool = False) -> str:
    text = _text(value)
    if not text:
        return ""
    try:
        number = Decimal(text.replace(",", ""))
    except InvalidOperation as error:
        raise ValueError(f"{cell}: phải là số.") from error
    if not number.is_finite() or number < 0:
        raise ValueError(f"{cell}: phải là số không âm.")
    # Decimal context chỉ giữ 28 chữ số. Một ô dán nhầm kiểu 1E+50 vẫn "hợp lệ"
    # ở đây nhưng sẽ làm quantize() lúc điền lên WFX ném InvalidOperation, biến
    # lỗi nhập liệu thành lỗi automation có gửi telemetry. Chặn ngay tại file.
    if number != 0 and number.adjusted() > MAX_SALE_ASN_NUMBER_EXPONENT:
        raise ValueError(f"{cell}: giá trị quá lớn, hãy kiểm tra lại ô này.")
    if integer and number != number.to_integral_value():
        raise ValueError(f"{cell}: phải là số nguyên.")
    normalized = format(number.normalize(), "f")
    return normalized.rstrip("0").rstrip(".") if "." in normalized else normalized


def _date_text(value: object, *, cell: str) -> str:
    if value is None or _text(value) == "":
        return ""
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        # openpyxl thường trả datetime khi ô có date format. Chỉ giữ fallback
        # cho workbook cũ đã mất style ngày.
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value)
            parsed = converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            parsed = None
    else:
        raw = _text(value)
        for pattern in (
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d %b %Y",
        ):
            try:
                parsed = datetime.strptime(raw, pattern).date()
                break
            except ValueError:
                continue
    if parsed is None:
        raise ValueError(f"{cell}: ngày không hợp lệ.")
    return parsed.isoformat()


def _assert_safe_xlsx(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_TYPE_UNSUPPORTED",
            "Tạo Sale ASN chỉ hỗ trợ file .xlsx.",
        )
    if not path.is_file():
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_NOT_FOUND",
            "Không tìm thấy file Sale ASN đã chọn.",
        )
    if path.stat().st_size > MAX_SALE_ASN_BYTES:
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_TOO_LARGE",
            "File Sale ASN vượt quá giới hạn 20 MB.",
        )
    try:
        with ZipFile(path) as archive:
            names = {name.casefold() for name in archive.namelist()}
    except (BadZipFile, OSError) as error:
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_INVALID",
            "File đã chọn không phải workbook Excel hợp lệ.",
        ) from error
    if any(name.endswith("vbaproject.bin") for name in names):
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_UNSAFE",
            "File Sale ASN không được chứa macro.",
        )


def _header_identity(value: object) -> str:
    return " ".join(_text(value).split()).casefold()


def _find_input_sheet(workbook) -> object:
    # Giữ đọc được form cũ chưa có Goods Description/Price; hai cột mới đều
    # tùy chọn để lượt tạo Sale ASN cũ không bị chặn chỉ vì muốn Check giá.
    expected_columns = tuple(
        value
        for value in SALE_ASN_COLUMNS
        if value not in {"Goods Description", "Price"}
    )
    expected = {_header_identity(value) for value in expected_columns}
    missing_by_sheet: list[str] = []
    for sheet in workbook.worksheets:
        actual = {_header_identity(cell.value) for cell in sheet[1] if cell.value}
        if expected <= actual:
            return sheet
        missing = [
            column
            for column in expected_columns
            if _header_identity(column) not in actual
        ]
        if missing:
            missing_by_sheet.append(
                f'Sheet "{sheet.title}": thiếu cột {", ".join(missing)}.'
            )
    raise SaleASNWorkbookError(
        "SALE_ASN_FILE_HEADERS_INVALID",
        "Không tìm thấy hàng tiêu đề Sale ASN đủ các cột chuẩn.",
        tuple(missing_by_sheet[:10]),
    )


def read_sale_asn_workbook(
    input_path: str | Path,
    required_stages: tuple[str, ...] | list[str] | None = None,
) -> dict:
    """Đọc workbook và chỉ bắt buộc dữ liệu cho các bước được chọn."""

    strict_create = required_stages is None
    stages = set(required_stages or ())
    require_style = strict_create or bool(stages & {"po", "style_details"})
    require_factory = strict_create or "shipping_info" in stages
    require_invoice = strict_create or "shipping_info" in stages
    require_shipping = strict_create or "shipping_info" in stages
    path = Path(input_path).expanduser().resolve()
    _assert_safe_xlsx(path)
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
    except Exception as error:
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_INVALID",
            "Không đọc được workbook Sale ASN.",
        ) from error

    try:
        sheet = _find_input_sheet(workbook)
        header_map = {
            _header_identity(cell.value): cell.column
            for cell in sheet[1]
            if cell.value
        }
        if sheet.max_row - 1 > MAX_SALE_ASN_ROWS:
            raise SaleASNWorkbookError(
                "SALE_ASN_FILE_TOO_MANY_ROWS",
                f"File Sale ASN chỉ hỗ trợ tối đa {MAX_SALE_ASN_ROWS} dòng.",
            )

        raw_rows: list[dict[str, object]] = []
        for source_row in range(2, sheet.max_row + 1):
            values = {
                column: (
                    sheet.cell(source_row, header_map[identity]).value
                    if (identity := _header_identity(column)) in header_map
                    else None
                )
                for column in SALE_ASN_COLUMNS
            }
            # PO No quyết định một dòng có phải dữ liệu ASN hay không. Dòng
            # tổng, ghi chú hoặc format thừa không có PO được bỏ qua hoàn toàn.
            if not _identifier(values["PO No"]):
                continue
            formula_cells = [
                f"{sheet.cell(source_row, header_map[_header_identity(column)]).coordinate}"
                for column, value in values.items()
                if (
                    _header_identity(column) in header_map
                    and isinstance(value, str)
                    and value.startswith("=")
                )
            ]
            if formula_cells:
                raise SaleASNWorkbookError(
                    "SALE_ASN_FILE_FORMULA_ERROR",
                    "File Sale ASN không được dùng công thức trong vùng nhập liệu.",
                    tuple(f"{cell}: đang chứa công thức." for cell in formula_cells),
                )
            raw_rows.append({"source_row": source_row, **values})
    finally:
        workbook.close()

    if not raw_rows:
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_EMPTY",
            "File Sale ASN chưa có dòng dữ liệu.",
        )

    first_values: dict[str, str] = {}
    errors: list[str] = []
    for column in (
        "Invoice No",
        "Invoice Date",
        "Shipping Bill No",
        "Shipping Bill Date",
        "Destination",
        "FTY",
        "Cargo Ready Date",
        "Consignee Address",
        "Ship To",
    ):
        raw = next((row[column] for row in raw_rows if _text(row[column])), None)
        try:
            first_values[column] = (
                _date_text(raw, cell=f"{column} dòng đầu")
                if column in _DATE_COLUMNS
                else _identifier(raw)
            )
        except ValueError as error:
            errors.append(str(error))
            first_values[column] = ""

    first_shipping_row = int(raw_rows[0]["source_row"])
    first_values["Shipping Mode"] = _text(
        raw_rows[0]["Shipping Mode"]
    ).upper()
    if (
        require_shipping
        and first_values["Shipping Mode"] not in SHIPPING_MODES
    ):
        errors.append(
            f"T{first_shipping_row}: Shipping Mode ở dòng dữ liệu đầu tiên "
            "bắt buộc chọn AIR, SEA hoặc COURIER."
        )

    # Form tạo mới giữ fallback ngày cũ. Khi tiếp tục một ASN đã có PO, ô ngày
    # trống phải thật sự được bỏ qua thay vì vô tình ghi ngày hôm nay lên WFX.
    if strict_create or "po" in stages:
        today = date.today().isoformat()
        for column in ("Invoice Date", "Shipping Bill Date"):
            if not first_values[column]:
                first_values[column] = today
    if require_invoice and not first_values["Invoice No"]:
        errors.append("Invoice No: bắt buộc nhập ở ít nhất một dòng.")
    if require_factory and not first_values["FTY"]:
        errors.append("FTY: bắt buộc nhập ở ít nhất một dòng.")
    if not first_values["Shipping Bill No"]:
        first_values["Shipping Bill No"] = first_values["Invoice No"]

    rows: list[SaleASNRow] = []
    seen_po_styles: set[tuple[str, str]] = set()
    invoice_values: set[str] = set()
    for raw in raw_rows:
        source_row = int(raw["source_row"])
        invoice_no = _identifier(raw["Invoice No"]) or first_values["Invoice No"]
        if invoice_no:
            invoice_values.add(invoice_no.casefold())
        style_no = _identifier(raw["Style No"])
        po_no = _identifier(raw["PO No"])
        destination = _text(raw["Destination"]) or first_values["Destination"]
        factory = _text(raw["FTY"]) or first_values["FTY"]
        consignee_address = (
            _text(raw["Consignee Address"])
            or first_values["Consignee Address"]
        )
        ship_to = _text(raw["Ship To"]) or first_values["Ship To"]
        shipping_mode = first_values["Shipping Mode"]
        if require_style and not style_no:
            errors.append(f"A{source_row}: Style No bắt buộc.")
        po_style_key = (po_no.casefold(), style_no.casefold())
        if po_style_key in seen_po_styles:
            errors.append(
                f"A{source_row}: PO No + Style No bị trùng "
                f"({po_no} + {style_no})."
            )
        else:
            seen_po_styles.add(po_style_key)
        if require_factory and not factory:
            errors.append(f"Q{source_row}: FTY bắt buộc.")

        numbers: dict[str, str] = {}
        for column in _NUMBER_COLUMNS:
            try:
                numbers[column] = _number(
                    raw[column],
                    cell=f"{column} dòng {source_row}",
                    integer=column in {"Qty", "Carton"},
                )
            except ValueError as error:
                errors.append(str(error))
                numbers[column] = ""
        dates: dict[str, str] = {}
        for column in _DATE_COLUMNS:
            try:
                parsed_date = _date_text(
                    raw[column], cell=f"{column} dòng {source_row}"
                )
                dates[column] = parsed_date or first_values[column]
            except ValueError as error:
                errors.append(str(error))
                dates[column] = first_values[column]

        rows.append(
            SaleASNRow(
                source_row=source_row,
                invoice_no=invoice_no,
                invoice_date=dates["Invoice Date"],
                shipping_bill_no=(
                    _identifier(raw["Shipping Bill No"])
                    or first_values["Shipping Bill No"]
                    or invoice_no
                ),
                shipping_bill_date=dates["Shipping Bill Date"],
                style_no=style_no,
                po_no=po_no,
                hs_code=_identifier(raw["HS CODE"]),
                goods_description=_text(raw["Goods Description"]),
                qty=numbers["Qty"],
                price=numbers["Price"],
                carton=numbers["Carton"],
                nw=numbers["NW"],
                gw=numbers["GW"],
                cbm=numbers["CBM"],
                destination=destination,
                factory=factory,
                fob_price=numbers["FOB Price"],
                service_price=numbers["Service Price"],
                cargo_ready_date=dates["Cargo Ready Date"],
                consignee_address=consignee_address,
                ship_to=ship_to,
                shipping_mode=shipping_mode,
            )
        )

    if require_invoice and len(invoice_values) > 1:
        errors.append("File chỉ được chứa một Invoice No.")
    factories = {row.factory.casefold() for row in rows if row.factory}
    if require_factory and len(factories) > 1:
        errors.append("File chỉ được chứa một FTY cho mỗi Sale ASN.")
    if errors:
        raise SaleASNWorkbookError(
            "SALE_ASN_FILE_VALIDATION_FAILED",
            f"File Sale ASN có {len(errors)} lỗi cần sửa.",
            tuple(errors[:100]),
        )

    return {
        "file_name": path.name,
        "invoice_no": rows[0].invoice_no,
        "invoice_date": rows[0].invoice_date,
        "shipping_bill_no": rows[0].shipping_bill_no,
        "shipping_bill_date": rows[0].shipping_bill_date,
        "destination": rows[0].destination,
        "factory": rows[0].factory,
        "po_count": len(rows),
        "style_count": len(
            {row.style_no.casefold() for row in rows if row.style_no}
        ),
        "rows": [row.automation_payload() for row in rows],
    }


def write_sale_asn_template(
    output_path: str | Path,
    rows: list[dict] | tuple[dict, ...] = (),
) -> Path:
    """Tạo form Sale ASN gọn với schema Shipping Info hiện hành."""
    target = Path(output_path).expanduser().resolve()
    if target.suffix.casefold() != ".xlsx":
        target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SALE_ASN_SHEET
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(list(SALE_ASN_COLUMNS))
    key_map = {
        "Invoice No": "invoice_no",
        "Invoice Date": "invoice_date",
        "Shipping Bill No": "shipping_bill_no",
        "Shipping Bill Date": "shipping_bill_date",
        "Style No": "style_no",
        "PO No": "po_no",
        "HS CODE": "hs_code",
        "Goods Description": "goods_description",
        "Qty": "qty",
        "Price": "price",
        "Carton": "carton",
        "NW": "nw",
        "GW": "gw",
        "CBM": "cbm",
        "Destination": "destination",
        "FTY": "factory",
        "FOB Price": "fob_price",
        "Service Price": "service_price",
        "Cargo Ready Date": "cargo_ready_date",
        "Consignee Address": "consignee_address",
        "Ship To": "ship_to",
        "Shipping Mode": "shipping_mode",
    }
    for row in rows:
        sheet.append(
            [
                _template_export_value(column, row.get(key_map[column]))
                for column in SALE_ASN_COLUMNS
            ]
        )
    reserved_rows = max(20, len(rows))
    for _ in range(len(rows), reserved_rows):
        sheet.append([None] * len(SALE_ASN_COLUMNS))

    widths = (
        34, 22, 12, 14, 12, 12, 12, 12, 14, 14, 18,
        15, 34, 20, 14, 20, 18, 20, 34, 34, 34, 16,
    )
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width

    required_fill = PatternFill("solid", fgColor="FDE68A")
    optional_fill = PatternFill("solid", fgColor="DBEAFE")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="172033")
        cell.fill = optional_fill if cell.value in _OPTIONAL_COLUMNS else required_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        note = (
            "Có thể để trống; app sẽ bỏ qua trường này."
            if cell.value in _OPTIONAL_COLUMNS
            else "Cột bắt buộc hoặc được app kế thừa từ dòng có dữ liệu đầu tiên."
        )
        cell.comment = Comment(note, "WFX Smart")
    sheet.row_dimensions[1].height = 34

    for row in range(2, reserved_rows + 2):
        for column in (11, 15, 17):
            sheet.cell(row, column).number_format = "dd/mm/yyyy"
        for column in (3, 4, 5, 6, 7, 8, 9, 10):
            sheet.cell(row, column).number_format = "#,##0.00"

    for cell_range in ("K2:K2001", "O2:O2001", "Q2:Q2001"):
        date_validation = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(1900,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
        )
        date_validation.error = "Nhập ngày hợp lệ trong khoảng 01/01/1900–31/12/2100."
        date_validation.errorTitle = "Ngày không hợp lệ"
        date_validation.prompt = "Chọn hoặc nhập ngày theo định dạng dd/mm/yyyy."
        date_validation.promptTitle = "Ngày"
        sheet.add_data_validation(date_validation)
        date_validation.add(cell_range)

    shipping_mode_validation = DataValidation(
        type="list",
        formula1='"AIR,SEA,COURIER"',
        allow_blank=False,
    )
    shipping_mode_validation.error = "Chọn AIR, SEA hoặc COURIER."
    shipping_mode_validation.errorTitle = "Shipping Mode không hợp lệ"
    shipping_mode_validation.prompt = "Chọn AIR, SEA hoặc COURIER."
    shipping_mode_validation.promptTitle = "Shipping Mode"
    sheet.add_data_validation(shipping_mode_validation)
    shipping_mode_validation.add("V2")
    table_ref = f"A1:V{reserved_rows + 1}"
    table = Table(displayName="SaleASNInput", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(target)
    workbook.close()
    return target


def write_sale_asn_price_check_workbook(
    output_path: str | Path,
    price_check: dict,
) -> Path:
    """Xuất kết quả đối chiếu Sale ASN thành workbook dễ kiểm tra lại."""

    target = Path(output_path).expanduser().resolve()
    if target.suffix.casefold() != ".xlsx":
        target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    details = workbook.active
    details.title = "PO + Style"
    details.sheet_view.showGridLines = False
    details.freeze_panes = "A4"
    details.merge_cells("A1:J1")
    details["A1"] = "KẾT QUẢ CHECK GIÁ / QTY SALE ASN"
    details["A1"].font = Font(bold=True, size=14, color="172033")
    details["A1"].alignment = Alignment(horizontal="center")
    details.merge_cells("A2:J2")
    details["A2"] = str(price_check.get("message") or "")
    details["A2"].alignment = Alignment(wrap_text=True)
    headers = (
        "Dòng file",
        "PO trong file",
        "Style trong file",
        "Order No. WFX (mã hệ thống/PO)",
        "Qty file",
        "Price file",
        "Shipping Qty WFX",
        "Price (USD) WFX",
        "Kết quả",
        "Chi tiết",
    )
    details.append([])
    details.append(list(headers))
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    ok_fill = PatternFill("solid", fgColor="DCFCE7")
    attention_fill = PatternFill("solid", fgColor="FEE2E2")
    for cell in details[4]:
        cell.font = Font(bold=True, color="172033")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    comparisons = price_check.get("comparisons")
    if not isinstance(comparisons, list):
        comparisons = []
    for item in comparisons:
        if not isinstance(item, dict):
            continue
        status = str(item.get("status") or "mismatch")
        details.append(
            [
                ", ".join(str(value) for value in item.get("source_rows") or ()),
                str(item.get("po_no") or ""),
                str(item.get("style_no") or ""),
                "\n".join(str(value) for value in item.get("system_order_nos") or ()),
                str(item.get("file_qty") or ""),
                str(item.get("file_price") or ""),
                str(item.get("system_qty") or ""),
                ", ".join(str(value) for value in item.get("system_prices") or ()),
                "KHỚP" if status == "ok" else "CẦN KIỂM TRA",
                str(item.get("message") or ""),
            ]
        )
        fill = ok_fill if status == "ok" else attention_fill
        for cell in details[details.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate((12, 20, 36, 38, 14, 14, 18, 18, 18, 42), 1):
        details.column_dimensions[details.cell(4, index).column_letter].width = width
    details.row_dimensions[1].height = 25
    details.row_dimensions[2].height = 30
    details.auto_filter.ref = f"A4:J{max(4, details.max_row)}"

    summary = workbook.create_sheet("Summary Total")
    summary.sheet_view.showGridLines = False
    summary.append(["SUMMARY TOTAL — ĐỐI CHIẾU FILE VÀ WFX"])
    summary.merge_cells("A1:D1")
    summary["A1"].font = Font(bold=True, size=14, color="172033")
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.append([])
    summary.append(["Chỉ tiêu", "File", "WFX", "Kết quả"])
    for cell in summary[3]:
        cell.font = Font(bold=True, color="172033")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    summary_labels = {
        "total_quantity": "Total Quantity",
        "value_in_doc_currency": "Value In Doc Currency",
        "net_value_in_doc_currency": "Net Value In Doc Currency",
    }
    summary_data = price_check.get("summary")
    checks = summary_data.get("checks", {}) if isinstance(summary_data, dict) else {}
    if not isinstance(checks, dict):
        checks = {}
    for key, label in summary_labels.items():
        check = checks.get(key) if isinstance(checks.get(key), dict) else {}
        ok = bool(check.get("ok"))
        summary.append(
            [
                label,
                str(check.get("expected") or ""),
                str(check.get("actual") or ""),
                "KHỚP" if ok else "CẦN KIỂM TRA",
            ]
        )
        for cell in summary[summary.max_row]:
            cell.fill = ok_fill if ok else attention_fill
    for column, width in {"A": 32, "B": 22, "C": 22, "D": 20}.items():
        summary.column_dimensions[column].width = width
    summary.auto_filter.ref = "A3:D6"
    workbook.save(target)
    workbook.close()
    return target


def _template_export_value(column: str, value: object) -> object:
    """Giữ ô export là number/date thật để người dùng tiếp tục tính toán."""

    if value is None or _text(value) == "":
        return None
    if column in _TEMPLATE_NUMBER_COLUMNS:
        try:
            normalized = _number(
                value,
                cell=column,
                integer=column == "Carton",
            )
            number = Decimal(normalized)
            return int(number) if number == number.to_integral_value() else float(number)
        except (ValueError, InvalidOperation):
            return _text(value)
    if column == "Cargo Ready Date":
        try:
            return date.fromisoformat(_date_text(value, cell=column))
        except ValueError:
            return _text(value)
    return _identifier(value)
