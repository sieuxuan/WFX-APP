"""Ghép hai workbook report Sale ASN mà không làm biến dạng report WFX."""

from __future__ import annotations

import math
import re
import xml.etree.ElementTree as ET
from copy import deepcopy
from pathlib import Path, PurePosixPath
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from openpyxl import load_workbook

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPE_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
WORKSHEET_REL_TYPE = f"{DOC_REL_NS}/worksheet"
WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", DOC_REL_NS)


class ASNWorkbookError(ValueError):
    """Workbook report WFX không thể đọc hoặc ghép an toàn."""


def _tag(namespace: str, name: str) -> str:
    return f"{{{namespace}}}{name}"


def _xml(data: bytes) -> ET.Element:
    return ET.fromstring(data)


def _xml_bytes(root: ET.Element, *, default_namespace: str = MAIN_NS) -> bytes:
    ET.register_namespace("", default_namespace)
    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    ET.register_namespace("", MAIN_NS)
    return data


def _safe_sheet_title(value: str, existing: set[str]) -> str:
    cleaned = "".join(
        " " if character in "[]:*?/\\" else character for character in value
    )
    cleaned = " ".join(cleaned.split()).strip(" '") or "Report"
    cleaned = cleaned[:31].rstrip()
    candidate = cleaned
    suffix = 2
    while candidate.casefold() in existing:
        tail = f" ({suffix})"
        candidate = f"{cleaned[: 31 - len(tail)].rstrip()}{tail}"
        suffix += 1
    existing.add(candidate.casefold())
    return candidate


_STYLE_SECTION_ORDER = (
    "numFmts",
    "fonts",
    "fills",
    "borders",
    "cellStyleXfs",
    "cellXfs",
    "cellStyles",
    "dxfs",
    "tableStyles",
    "colors",
    "extLst",
)


def _style_section(root: ET.Element, name: str) -> ET.Element:
    section = root.find(_tag(MAIN_NS, name))
    if section is not None:
        return section
    section = ET.Element(_tag(MAIN_NS, name))
    wanted = _STYLE_SECTION_ORDER.index(name)
    for index, child in enumerate(root):
        local = child.tag.rsplit("}", 1)[-1]
        if local in _STYLE_SECTION_ORDER and _STYLE_SECTION_ORDER.index(local) > wanted:
            root.insert(index, section)
            break
    else:
        root.append(section)
    return section


def _set_count(section: ET.Element) -> None:
    section.set("count", str(len(section)))


def _append_components(
    base_styles: ET.Element,
    source_styles: ET.Element,
    section_name: str,
) -> dict[int, int]:
    base = _style_section(base_styles, section_name)
    source = source_styles.find(_tag(MAIN_NS, section_name))
    offset = len(base)
    if source is None:
        return {}
    mapping: dict[int, int] = {}
    for source_index, item in enumerate(source):
        mapping[source_index] = offset + source_index
        base.append(deepcopy(item))
    _set_count(base)
    return mapping


def _merge_number_formats(
    base_styles: ET.Element,
    source_styles: ET.Element,
) -> dict[int, int]:
    base = _style_section(base_styles, "numFmts")
    source = source_styles.find(_tag(MAIN_NS, "numFmts"))
    formats_by_code: dict[str, int] = {}
    used_ids: set[int] = set()
    for item in base:
        number_id = int(item.get("numFmtId", "0"))
        used_ids.add(number_id)
        formats_by_code[str(item.get("formatCode") or "")] = number_id
    mapping: dict[int, int] = {}
    if source is None:
        return mapping
    next_id = max([163, *used_ids]) + 1
    for item in source:
        old_id = int(item.get("numFmtId", "0"))
        code = str(item.get("formatCode") or "")
        if code in formats_by_code:
            mapping[old_id] = formats_by_code[code]
            continue
        while next_id in used_ids:
            next_id += 1
        copied = deepcopy(item)
        copied.set("numFmtId", str(next_id))
        base.append(copied)
        mapping[old_id] = next_id
        formats_by_code[code] = next_id
        used_ids.add(next_id)
        next_id += 1
    _set_count(base)
    return mapping


def _remap_xf(
    item: ET.Element,
    *,
    font_map: dict[int, int],
    fill_map: dict[int, int],
    border_map: dict[int, int],
    number_format_map: dict[int, int],
    style_xf_map: dict[int, int] | None = None,
) -> ET.Element:
    copied = deepcopy(item)
    component_maps = {
        "fontId": font_map,
        "fillId": fill_map,
        "borderId": border_map,
        "numFmtId": number_format_map,
    }
    for attribute, mapping in component_maps.items():
        if attribute not in copied.attrib:
            continue
        old_value = int(copied.get(attribute, "0"))
        copied.set(attribute, str(mapping.get(old_value, old_value)))
    if style_xf_map is not None and "xfId" in copied.attrib:
        old_value = int(copied.get("xfId", "0"))
        copied.set("xfId", str(style_xf_map.get(old_value, old_value)))
    return copied


def _merge_styles(base_data: bytes, source_data: bytes) -> tuple[bytes, dict[int, int], dict[int, int]]:
    base_styles = _xml(base_data)
    source_styles = _xml(source_data)
    number_format_map = _merge_number_formats(base_styles, source_styles)
    font_map = _append_components(base_styles, source_styles, "fonts")
    fill_map = _append_components(base_styles, source_styles, "fills")
    border_map = _append_components(base_styles, source_styles, "borders")

    base_style_xfs = _style_section(base_styles, "cellStyleXfs")
    source_style_xfs = source_styles.find(_tag(MAIN_NS, "cellStyleXfs"))
    style_xf_map: dict[int, int] = {}
    if source_style_xfs is not None:
        offset = len(base_style_xfs)
        for source_index, item in enumerate(source_style_xfs):
            style_xf_map[source_index] = offset + source_index
            base_style_xfs.append(
                _remap_xf(
                    item,
                    font_map=font_map,
                    fill_map=fill_map,
                    border_map=border_map,
                    number_format_map=number_format_map,
                )
            )
        _set_count(base_style_xfs)

    base_cell_xfs = _style_section(base_styles, "cellXfs")
    source_cell_xfs = source_styles.find(_tag(MAIN_NS, "cellXfs"))
    cell_style_map: dict[int, int] = {}
    if source_cell_xfs is not None:
        offset = len(base_cell_xfs)
        for source_index, item in enumerate(source_cell_xfs):
            cell_style_map[source_index] = offset + source_index
            base_cell_xfs.append(
                _remap_xf(
                    item,
                    font_map=font_map,
                    fill_map=fill_map,
                    border_map=border_map,
                    number_format_map=number_format_map,
                    style_xf_map=style_xf_map,
                )
            )
        _set_count(base_cell_xfs)

    base_dxfs = _style_section(base_styles, "dxfs")
    source_dxfs = source_styles.find(_tag(MAIN_NS, "dxfs"))
    dxf_map: dict[int, int] = {}
    if source_dxfs is not None:
        offset = len(base_dxfs)
        for source_index, item in enumerate(source_dxfs):
            dxf_map[source_index] = offset + source_index
            base_dxfs.append(deepcopy(item))
        _set_count(base_dxfs)
    return _xml_bytes(base_styles), cell_style_map, dxf_map


def _remap_sheet_styles(
    source_data: bytes,
    cell_style_map: dict[int, int],
    dxf_map: dict[int, int],
) -> bytes:
    root = _xml(source_data)
    for item in root.iter():
        local = item.tag.rsplit("}", 1)[-1]
        attribute = "style" if local == "col" else "s"
        if local in {"c", "row", "col"} and attribute in item.attrib:
            old_value = int(item.get(attribute, "0"))
            item.set(attribute, str(cell_style_map.get(old_value, old_value)))
        if "dxfId" in item.attrib:
            old_value = int(item.get("dxfId", "0"))
            item.set("dxfId", str(dxf_map.get(old_value, old_value)))
    return _xml_bytes(root)


def _relationship_member(target: str) -> str:
    cleaned = target.replace("\\", "/").lstrip("/")
    if cleaned.startswith("xl/"):
        return cleaned
    return str(PurePosixPath("xl") / cleaned)


def _sheet_records(
    workbook_root: ET.Element,
    relationships_root: ET.Element,
) -> list[tuple[ET.Element, str]]:
    targets = {
        str(item.get("Id") or ""): _relationship_member(str(item.get("Target") or ""))
        for item in relationships_root.findall(_tag(PACKAGE_REL_NS, "Relationship"))
        if item.get("Type") == WORKSHEET_REL_TYPE
    }
    sheets = workbook_root.find(_tag(MAIN_NS, "sheets"))
    if sheets is None:
        return []
    records: list[tuple[ET.Element, str]] = []
    for sheet in sheets.findall(_tag(MAIN_NS, "sheet")):
        relationship_id = str(sheet.get(_tag(DOC_REL_NS, "id")) or "")
        member = targets.get(relationship_id)
        if member:
            records.append((sheet, member))
    return records


def _next_relationship_id(used: set[str]) -> str:
    index = 1
    while f"rId{index}" in used:
        index += 1
    value = f"rId{index}"
    used.add(value)
    return value


def _update_defined_names(
    base_workbook: ET.Element,
    source_workbook: ET.Element,
    invoice_positions: dict[int, int],
    packing_positions: dict[int, int],
) -> None:
    base = base_workbook.find(_tag(MAIN_NS, "definedNames"))
    source = source_workbook.find(_tag(MAIN_NS, "definedNames"))
    if base is not None:
        for item in base:
            if "localSheetId" in item.attrib:
                old_value = int(item.get("localSheetId", "0"))
                item.set("localSheetId", str(invoice_positions.get(old_value, old_value)))
    if source is None:
        return
    if base is None:
        base = ET.Element(_tag(MAIN_NS, "definedNames"))
        sheets = base_workbook.find(_tag(MAIN_NS, "sheets"))
        insert_at = list(base_workbook).index(sheets) + 1 if sheets is not None else 0
        base_workbook.insert(insert_at, base)
    for item in source:
        copied = deepcopy(item)
        if "localSheetId" not in copied.attrib:
            continue
        old_value = int(copied.get("localSheetId", "0"))
        copied.set("localSheetId", str(packing_positions.get(old_value, old_value)))
        base.append(copied)


def _merge_packages(
    packing_path: Path,
    buyer_path: Path,
    target: Path,
    invoice_label: str,
) -> None:
    with ZipFile(buyer_path) as buyer_zip, ZipFile(packing_path) as packing_zip:
        buyer_workbook = _xml(buyer_zip.read("xl/workbook.xml"))
        packing_workbook = _xml(packing_zip.read("xl/workbook.xml"))
        buyer_relationships = _xml(buyer_zip.read("xl/_rels/workbook.xml.rels"))
        packing_relationships = _xml(packing_zip.read("xl/_rels/workbook.xml.rels"))
        buyer_records = _sheet_records(buyer_workbook, buyer_relationships)
        packing_records = _sheet_records(packing_workbook, packing_relationships)
        if not buyer_records or not packing_records:
            raise ASNWorkbookError("Hai report Sale ASN phải có ít nhất một sheet.")

        for _sheet, member in packing_records:
            name = PurePosixPath(member).name
            relationship_member = f"xl/worksheets/_rels/{name}.rels"
            if relationship_member in packing_zip.namelist():
                raise ASNWorkbookError(
                    "Packing List có đối tượng liên kết chưa thể ghép nguyên trạng."
                )

        merged_styles, cell_style_map, dxf_map = _merge_styles(
            buyer_zip.read("xl/styles.xml"),
            packing_zip.read("xl/styles.xml"),
        )

        buyer_names = [str(item.get("name") or "") for item, _member in buyer_records]
        packing_names = [str(item.get("name") or "") for item, _member in packing_records]
        collisions = {name.casefold() for name in buyer_names} & {
            name.casefold() for name in packing_names
        }
        existing = {
            name.casefold()
            for name in [*buyer_names, *packing_names]
            if name.casefold() not in collisions
        }
        for sheet, _member in buyer_records:
            if str(sheet.get("name") or "").casefold() in collisions:
                sheet.set(
                    "name",
                    _safe_sheet_title(f"INVOICE {invoice_label}", existing),
                )

        used_relationship_ids = {
            str(item.get("Id") or "")
            for item in buyer_relationships.findall(
                _tag(PACKAGE_REL_NS, "Relationship")
            )
        }
        used_sheet_ids = [
            int(item.get("sheetId", "0")) for item, _member in buyer_records
        ]
        used_sheet_numbers = [
            int(match.group(1))
            for name in buyer_zip.namelist()
            if (match := re.fullmatch(r"xl/worksheets/sheet(\d+)\.xml", name))
        ]
        next_sheet_id = max([0, *used_sheet_ids]) + 1
        next_sheet_number = max([0, *used_sheet_numbers]) + 1
        packing_output: list[tuple[ET.Element, str, bytes]] = []
        for source_sheet, source_member in packing_records:
            source_name = str(source_sheet.get("name") or "")
            if source_name.casefold() in collisions:
                target_name = _safe_sheet_title(
                    f"PKL {invoice_label}",
                    existing,
                )
            else:
                target_name = source_name
            relationship_id = _next_relationship_id(used_relationship_ids)
            target_member = f"xl/worksheets/sheet{next_sheet_number}.xml"
            sheet = ET.Element(
                _tag(MAIN_NS, "sheet"),
                {
                    "name": target_name,
                    "sheetId": str(next_sheet_id),
                    _tag(DOC_REL_NS, "id"): relationship_id,
                },
            )
            buyer_relationships.append(
                ET.Element(
                    _tag(PACKAGE_REL_NS, "Relationship"),
                    {
                        "Type": WORKSHEET_REL_TYPE,
                        "Target": f"/{target_member}",
                        "Id": relationship_id,
                    },
                )
            )
            packing_output.append(
                (
                    sheet,
                    target_member,
                    _remap_sheet_styles(
                        packing_zip.read(source_member),
                        cell_style_map,
                        dxf_map,
                    ),
                )
            )
            next_sheet_id += 1
            next_sheet_number += 1

        sequence: list[tuple[str, int, ET.Element]] = []
        for index in range(max(len(buyer_records), len(packing_output))):
            if index < len(buyer_records):
                sequence.append(("invoice", index, buyer_records[index][0]))
            if index < len(packing_output):
                sequence.append(("packing", index, packing_output[index][0]))
        sheets = buyer_workbook.find(_tag(MAIN_NS, "sheets"))
        if sheets is None:
            raise ASNWorkbookError("Buyer Invoice không có danh sách sheet.")
        sheets[:] = [item for _kind, _index, item in sequence]
        invoice_positions = {
            old_index: final_index
            for final_index, (kind, old_index, _item) in enumerate(sequence)
            if kind == "invoice"
        }
        packing_positions = {
            old_index: final_index
            for final_index, (kind, old_index, _item) in enumerate(sequence)
            if kind == "packing"
        }
        _update_defined_names(
            buyer_workbook,
            packing_workbook,
            invoice_positions,
            packing_positions,
        )

        content_types = _xml(buyer_zip.read("[Content_Types].xml"))
        for _sheet, member, _data in packing_output:
            content_types.append(
                ET.Element(
                    _tag(CONTENT_TYPE_NS, "Override"),
                    {
                        "PartName": f"/{member}",
                        "ContentType": WORKSHEET_CONTENT_TYPE,
                    },
                )
            )

        replacements = {
            "[Content_Types].xml": _xml_bytes(
                content_types,
                default_namespace=CONTENT_TYPE_NS,
            ),
            "xl/workbook.xml": _xml_bytes(buyer_workbook),
            "xl/_rels/workbook.xml.rels": _xml_bytes(
                buyer_relationships,
                default_namespace=PACKAGE_REL_NS,
            ),
            "xl/styles.xml": merged_styles,
        }
        with ZipFile(target, "w", compression=ZIP_DEFLATED) as output_zip:
            for info in buyer_zip.infolist():
                output_zip.writestr(info, replacements.get(info.filename, buyer_zip.read(info)))
            for _sheet, member, data in packing_output:
                output_zip.writestr(member, data)


def _shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = _xml(archive.read("xl/sharedStrings.xml"))
    return ["".join(item.itertext()) for item in root.findall(_tag(MAIN_NS, "si"))]


def _wrapped_style_ids(archive: ZipFile) -> set[int]:
    root = _xml(archive.read("xl/styles.xml"))
    cell_xfs = root.find(_tag(MAIN_NS, "cellXfs"))
    if cell_xfs is None:
        return set()
    wrapped: set[int] = set()
    for index, item in enumerate(cell_xfs):
        alignment = item.find(_tag(MAIN_NS, "alignment"))
        if alignment is not None and alignment.get("wrapText") in {"1", "true"}:
            wrapped.add(index)
    return wrapped


def _column_index(reference: str) -> int:
    match = re.match(r"([A-Z]+)", reference.upper())
    if match is None:
        return 1
    value = 0
    for character in match.group(1):
        value = value * 26 + ord(character) - ord("A") + 1
    return value


def _cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.get("t")
    if cell_type == "inlineStr":
        inline = cell.find(_tag(MAIN_NS, "is"))
        return "".join(inline.itertext()) if inline is not None else ""
    value = cell.find(_tag(MAIN_NS, "v"))
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        try:
            return shared_strings[int(value.text)]
        except (IndexError, ValueError):
            return ""
    return value.text


def _sheet_column_widths(root: ET.Element) -> tuple[float, list[tuple[int, int, float]]]:
    sheet_format = root.find(_tag(MAIN_NS, "sheetFormatPr"))
    default = float(sheet_format.get("defaultColWidth", "8.43")) if sheet_format is not None else 8.43
    ranges: list[tuple[int, int, float]] = []
    columns = root.find(_tag(MAIN_NS, "cols"))
    if columns is not None:
        for column in columns.findall(_tag(MAIN_NS, "col")):
            try:
                ranges.append(
                    (
                        int(column.get("min", "1")),
                        int(column.get("max", "1")),
                        float(column.get("width", str(default))),
                    )
                )
            except ValueError:
                continue
    return default, ranges


def _column_width(
    column: int,
    default: float,
    ranges: list[tuple[int, int, float]],
) -> float:
    for first, last, width in reversed(ranges):
        if first <= column <= last:
            return width
    return default


def _merged_cell_ends(root: ET.Element) -> dict[str, int]:
    merged = root.find(_tag(MAIN_NS, "mergeCells"))
    if merged is None:
        return {}
    ends: dict[str, int] = {}
    for item in merged.findall(_tag(MAIN_NS, "mergeCell")):
        reference = str(item.get("ref") or "")
        if ":" not in reference:
            continue
        start, end = reference.split(":", 1)
        ends[start.upper()] = _column_index(end)
    return ends


def _display_units(value: str) -> float:
    # Ký tự full-width/CJK chiếm gần gấp đôi ký tự Latin trong Excel.
    return sum(2 if ord(character) > 0x2E7F else 1 for character in value)


def _needed_row_lines(value: str, capacity: float) -> int:
    return sum(
        max(1, math.ceil(_display_units(line) / max(1.0, capacity)))
        for line in value.replace("\r", "").split("\n")
    )


def _fit_wrapped_report_rows(target: Path) -> None:
    """Tăng đúng các hàng wrap text để Excel không cắt nội dung report WFX."""
    with ZipFile(target, "r") as source:
        shared_strings = _shared_strings(source)
        wrapped_styles = _wrapped_style_ids(source)
        replacements: dict[str, bytes] = {}
        for info in source.infolist():
            if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", info.filename):
                continue
            root = _xml(source.read(info.filename))
            default_width, width_ranges = _sheet_column_widths(root)
            merged_ends = _merged_cell_ends(root)
            sheet_format = root.find(_tag(MAIN_NS, "sheetFormatPr"))
            try:
                default_height = float(
                    sheet_format.get("defaultRowHeight", "15")
                    if sheet_format is not None
                    else "15"
                )
            except ValueError:
                default_height = 15.0
            changed = False
            for row in root.findall(f".//{_tag(MAIN_NS, 'row')}"):
                row_style = int(row.get("s", "0"))
                needed_lines = 1
                for cell in row.findall(_tag(MAIN_NS, "c")):
                    text = _cell_text(cell, shared_strings)
                    if not text:
                        continue
                    style = int(cell.get("s", str(row_style)))
                    if style not in wrapped_styles and "\n" not in text:
                        continue
                    reference = str(cell.get("r") or "A1").upper()
                    start = _column_index(reference)
                    end = merged_ends.get(reference, start)
                    width = sum(
                        _column_width(column, default_width, width_ranges)
                        for column in range(start, end + 1)
                    )
                    # Width trong OOXML là số ký tự chuẩn; chừa một ít khoảng
                    # đệm để text wrap không bị sát mép/bị cắt dòng cuối.
                    needed_lines = max(
                        needed_lines,
                        _needed_row_lines(text, max(1.0, width * 0.95)),
                    )
                if needed_lines <= 1:
                    continue
                try:
                    current_height = float(row.get("ht", str(default_height)))
                except ValueError:
                    current_height = default_height
                desired = default_height * needed_lines + 2
                if desired > current_height:
                    row.set("ht", f"{desired:g}")
                    row.set("customHeight", "1")
                    changed = True
            if changed:
                replacements[info.filename] = _xml_bytes(root)
        if not replacements:
            return
        temp_target = target.with_suffix(".height-adjusting.xlsx")
        with ZipFile(temp_target, "w", compression=ZIP_DEFLATED) as output:
            for info in source.infolist():
                output.writestr(
                    info,
                    replacements.get(info.filename, source.read(info.filename)),
                )
    temp_target.replace(target)


_PACKING_COLUMN_MINIMUM_WIDTHS = {
    "no of pcs": 11.0,
    "qty unit": 11.0,
    "net wt": 11.0,
    "net weight": 11.0,
    "gross wt": 12.0,
    "gross weight": 12.0,
    "no of carton": 13.0,
    "qty cartons": 13.0,
    "cbm": 8.0,
}


def _set_column_minimum_width(
    root: ET.Element,
    column: int,
    minimum_width: float,
) -> bool:
    default_width, width_ranges = _sheet_column_widths(root)
    if _column_width(column, default_width, width_ranges) >= minimum_width:
        return False
    columns = root.find(_tag(MAIN_NS, "cols"))
    if columns is None:
        columns = ET.Element(_tag(MAIN_NS, "cols"))
        sheet_data = root.find(_tag(MAIN_NS, "sheetData"))
        root.insert(
            list(root).index(sheet_data) if sheet_data is not None else len(root),
            columns,
        )
    for index, item in enumerate(list(columns)):
        try:
            start = int(item.get("min", "0"))
            end = int(item.get("max", "0"))
        except ValueError:
            continue
        if not start <= column <= end:
            continue
        attributes = dict(item.attrib)
        replacements: list[ET.Element] = []
        if start < column:
            before = ET.Element(_tag(MAIN_NS, "col"), attributes | {"max": str(column - 1)})
            replacements.append(before)
        adjusted = ET.Element(
            _tag(MAIN_NS, "col"),
            attributes
            | {
                "min": str(column),
                "max": str(column),
                "width": f"{minimum_width:g}",
                "customWidth": "1",
            },
        )
        replacements.append(adjusted)
        if column < end:
            after = ET.Element(_tag(MAIN_NS, "col"), attributes | {"min": str(column + 1)})
            replacements.append(after)
        columns.remove(item)
        for offset, replacement in enumerate(replacements):
            columns.insert(index + offset, replacement)
        return True
    columns.append(
        ET.Element(
            _tag(MAIN_NS, "col"),
            {
                "min": str(column),
                "max": str(column),
                "width": f"{minimum_width:g}",
                "customWidth": "1",
            },
        )
    )
    return True


def _fit_packing_measurement_columns(target: Path) -> None:
    """Nới cột số PKL để header và số liệu luôn đọc đủ khi mở Excel."""
    with ZipFile(target, "r") as source:
        shared_strings = _shared_strings(source)
        replacements: dict[str, bytes] = {}
        for info in source.infolist():
            if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", info.filename):
                continue
            root = _xml(source.read(info.filename))
            changed = False
            for row in root.findall(f".//{_tag(MAIN_NS, 'row')}"):
                for column, cell in _cell_map(row).items():
                    label = _report_header_text(_cell_text(cell, shared_strings))
                    minimum_width = _PACKING_COLUMN_MINIMUM_WIDTHS.get(label)
                    if minimum_width is not None:
                        changed = _set_column_minimum_width(
                            root,
                            column,
                            minimum_width,
                        ) or changed
            if changed:
                replacements[info.filename] = _xml_bytes(root)
        if not replacements:
            return
        temp_target = target.with_suffix(".packing-width-adjusting.xlsx")
        with ZipFile(temp_target, "w", compression=ZIP_DEFLATED) as output:
            for info in source.infolist():
                output.writestr(
                    info,
                    replacements.get(info.filename, source.read(info.filename)),
                )
    temp_target.replace(target)


def _set_a4_page_setup(root: ET.Element) -> bool:
    """Đặt khổ A4, giữ hướng in gốc và cho phép phân trang theo chiều dọc."""
    changed = False
    sheet_properties = root.find(_tag(MAIN_NS, "sheetPr"))
    if sheet_properties is None:
        sheet_properties = ET.Element(_tag(MAIN_NS, "sheetPr"))
        root.insert(0, sheet_properties)
        changed = True
    page_setup_properties = sheet_properties.find(_tag(MAIN_NS, "pageSetUpPr"))
    if page_setup_properties is None:
        page_setup_properties = ET.Element(_tag(MAIN_NS, "pageSetUpPr"))
        sheet_properties.append(page_setup_properties)
        changed = True
    if page_setup_properties.get("fitToPage") != "1":
        page_setup_properties.set("fitToPage", "1")
        changed = True

    page_setup = root.find(_tag(MAIN_NS, "pageSetup"))
    if page_setup is None:
        page_setup = ET.Element(_tag(MAIN_NS, "pageSetup"))
        # pageSetup đứng sau pageMargins/printOptions và trước headerFooter.
        insert_before = next(
            (
                index
                for index, child in enumerate(root)
                if child.tag
                in {
                    _tag(MAIN_NS, "headerFooter"),
                    _tag(MAIN_NS, "drawing"),
                    _tag(MAIN_NS, "legacyDrawing"),
                    _tag(MAIN_NS, "tableParts"),
                    _tag(MAIN_NS, "extLst"),
                }
            ),
            len(root),
        )
        root.insert(insert_before, page_setup)
        changed = True
    desired = {
        "paperSize": "9",
        "fitToWidth": "1",
        "fitToHeight": "0",
    }
    for key, value in desired.items():
        if page_setup.get(key) != value:
            page_setup.set(key, value)
            changed = True
    return changed


def _fit_reports_to_a4(target: Path) -> None:
    """Chuẩn hóa mọi sheet đã ghép sang A4 mà không đổi hướng in WFX."""
    with ZipFile(target, "r") as source:
        replacements: dict[str, bytes] = {}
        for info in source.infolist():
            if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", info.filename):
                continue
            root = _xml(source.read(info.filename))
            if _set_a4_page_setup(root):
                replacements[info.filename] = _xml_bytes(root)
        if not replacements:
            return
        temp_target = target.with_suffix(".a4-adjusting.xlsx")
        with ZipFile(temp_target, "w", compression=ZIP_DEFLATED) as output:
            for info in source.infolist():
                output.writestr(
                    info,
                    replacements.get(info.filename, source.read(info.filename)),
                )
    temp_target.replace(target)


_JL_PACKING_HEADERS = {
    "jl po#": "po",
    "style no": "style",
    "net wt": "net_wt",
    "gross wt": "gross_wt",
    "no of carton": "carton",
    "cbm": "cbm",
}
_JL_PACKING_MEASUREMENTS = ("net_wt", "gross_wt", "carton", "cbm")
_TRUEWERK_PACKING_HEADERS = {
    "qty unit": "quantity",
    "net weight": "net_wt",
    "gross weight": "gross_wt",
    "qty cartons": "carton",
    "cbm": "cbm",
}
_TRUEWERK_PACKING_DETAIL_HEADERS = {"style": "style", "po no": "po"}
_TRUEWERK_PACKING_MEASUREMENTS = ("net_wt", "gross_wt", "carton", "cbm")


def _cell_row(reference: str) -> int:
    match = re.search(r"(\d+)$", reference)
    return int(match.group(1)) if match is not None else 0


def _cell_map(row: ET.Element) -> dict[int, ET.Element]:
    return {
        _column_index(str(cell.get("r") or "A1")): cell
        for cell in row.findall(_tag(MAIN_NS, "c"))
    }


def _header_text(value: str) -> str:
    return " ".join(value.replace("\n", " ").split()).casefold()


def _report_header_text(value: str) -> str:
    """Chuẩn hoá các biến thể dấu nối/chấm trong header report WFX."""
    return " ".join(re.sub(r"[^a-z0-9]+", " ", _header_text(value)).split())


def _existing_merge_ranges(root: ET.Element) -> list[tuple[int, int, int, int]]:
    merge_cells = root.find(_tag(MAIN_NS, "mergeCells"))
    if merge_cells is None:
        return []
    ranges: list[tuple[int, int, int, int]] = []
    for item in merge_cells.findall(_tag(MAIN_NS, "mergeCell")):
        reference = str(item.get("ref") or "")
        start, separator, end = reference.partition(":")
        if not separator:
            end = start
        start_column, end_column = _column_index(start), _column_index(end)
        start_row, end_row = _cell_row(start), _cell_row(end)
        if start_row and end_row:
            ranges.append((start_column, end_column, start_row, end_row))
    return ranges


def _merge_range_is_free(
    existing: list[tuple[int, int, int, int]],
    column: int,
    first_row: int,
    last_row: int,
) -> bool:
    return not any(
        start_column <= column <= end_column
        and not (last_row < start_row or first_row > end_row)
        for start_column, end_column, start_row, end_row in existing
    )


def _clear_merged_cell_value(cell: ET.Element) -> None:
    for child in tuple(cell):
        if child.tag in {
            _tag(MAIN_NS, "f"),
            _tag(MAIN_NS, "v"),
            _tag(MAIN_NS, "is"),
        }:
            cell.remove(child)
    cell.attrib.pop("t", None)


def _copy_cell_value(source: ET.Element, target: ET.Element) -> None:
    """Chuyển nội dung ô nguồn sang ô đầu của vùng merge, giữ style ô đích."""
    _clear_merged_cell_value(target)
    if "t" in source.attrib:
        target.set("t", source.attrib["t"])
    for child in source:
        if child.tag in {
            _tag(MAIN_NS, "f"),
            _tag(MAIN_NS, "v"),
            _tag(MAIN_NS, "is"),
        }:
            target.append(deepcopy(child))


def _merge_jl_packing_sheet(root: ET.Element, shared_strings: list[str]) -> bool:
    """Gộp cột tổng của Packing List J.Lindeberg theo JL PO# + Style No."""
    sheet_data = root.find(_tag(MAIN_NS, "sheetData"))
    if sheet_data is None:
        return False
    rows = sorted(
        sheet_data.findall(_tag(MAIN_NS, "row")),
        key=lambda row: int(row.get("r", "0")),
    )
    header_index = -1
    columns: dict[str, int] = {}
    for index, row in enumerate(rows):
        labels = {
            _header_text(_cell_text(cell, shared_strings)): column
            for column, cell in _cell_map(row).items()
        }
        candidate = {
            key: labels[label]
            for label, key in _JL_PACKING_HEADERS.items()
            if label in labels
        }
        if len(candidate) == len(_JL_PACKING_HEADERS):
            header_index, columns = index, candidate
            break
    if header_index < 0:
        return False

    existing = _existing_merge_ranges(root)
    new_ranges: list[tuple[int, int, int]] = []

    def merge_group(group: list[tuple[ET.Element, dict[int, ET.Element]]]) -> None:
        if len(group) < 2:
            return
        first_row = int(group[0][0].get("r", "0"))
        last_row = int(group[-1][0].get("r", "0"))
        for name in _JL_PACKING_MEASUREMENTS:
            column = columns[name]
            cells = [cell_map.get(column) for _row, cell_map in group]
            if any(cell is None for cell in cells):
                continue
            values = [_cell_text(cell, shared_strings).strip() for cell in cells]
            if not values[0] or any(value != values[0] for value in values[1:]):
                continue
            if not _merge_range_is_free(existing, column, first_row, last_row):
                continue
            new_ranges.append((column, first_row, last_row))
            existing.append((column, column, first_row, last_row))
            for cell in cells[1:]:
                _clear_merged_cell_value(cell)

    group: list[tuple[ET.Element, dict[int, ET.Element]]] = []
    previous_row = 0
    previous_key: tuple[str, str] | None = None
    for row in rows[header_index + 1 :]:
        row_number = int(row.get("r", "0"))
        cell_map = _cell_map(row)
        po_cell, style_cell = cell_map.get(columns["po"]), cell_map.get(columns["style"])
        key = (
            _cell_text(po_cell, shared_strings).strip() if po_cell is not None else "",
            _cell_text(style_cell, shared_strings).strip() if style_cell is not None else "",
        )
        if not key[0] or not key[1] or row_number != previous_row + 1 or key != previous_key:
            merge_group(group)
            group = []
        if key[0] and key[1]:
            group.append((row, cell_map))
            previous_key = key
            previous_row = row_number
        else:
            previous_key = None
            previous_row = 0
    merge_group(group)
    if not new_ranges:
        return False

    merge_cells = root.find(_tag(MAIN_NS, "mergeCells"))
    if merge_cells is None:
        merge_cells = ET.Element(_tag(MAIN_NS, "mergeCells"))
        root.insert(list(root).index(sheet_data) + 1, merge_cells)
    for column, first_row, last_row in new_ranges:
        letter = ""
        value = column
        while value:
            value, remainder = divmod(value - 1, 26)
            letter = chr(ord("A") + remainder) + letter
        merge_cells.append(
            ET.Element(
                _tag(MAIN_NS, "mergeCell"),
                {"ref": f"{letter}{first_row}:{letter}{last_row}"},
            )
        )
    merge_cells.set("count", str(len(merge_cells)))
    return True


def _merge_jl_packing_measurements(target: Path) -> None:
    """Áp dụng gộp tổng J.Lindeberg, chỉ cho sheet có đủ header đặc trưng."""
    with ZipFile(target, "r") as source:
        shared_strings = _shared_strings(source)
        replacements: dict[str, bytes] = {}
        for info in source.infolist():
            if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", info.filename):
                continue
            root = _xml(source.read(info.filename))
            if _merge_jl_packing_sheet(root, shared_strings):
                replacements[info.filename] = _xml_bytes(root)
        if not replacements:
            return
        temp_target = target.with_suffix(".jl-merge-adjusting.xlsx")
        with ZipFile(temp_target, "w", compression=ZIP_DEFLATED) as output:
            for info in source.infolist():
                output.writestr(
                    info,
                    replacements.get(info.filename, source.read(info.filename)),
                )
    temp_target.replace(target)


def _truewerk_po_base(value: str) -> tuple[str, bool]:
    """Trả PO gốc và cờ dòng phụ ADD của Packing List TRUEWERK."""
    normalized = " ".join(value.split())
    base = re.sub(r"\s*-?\s*ADD$", "", normalized, flags=re.IGNORECASE).strip()
    return base.casefold(), base != normalized


def _is_zero_measurement(value: str) -> bool:
    return bool(re.fullmatch(r"[+-]?0+(?:[.,]0+)?", value.replace(" ", "")))


def _merge_truewerk_packing_sheet(
    root: ET.Element,
    shared_strings: list[str],
) -> bool:
    """Gộp số liệu kiện hàng TRUEWERK từ PO gốc xuống dòng PO ADD."""
    sheet_data = root.find(_tag(MAIN_NS, "sheetData"))
    if sheet_data is None:
        return False
    rows = sorted(
        sheet_data.findall(_tag(MAIN_NS, "row")),
        key=lambda row: int(row.get("r", "0")),
    )
    body_index = -1
    columns: dict[str, int] = {}
    for index, row in enumerate(rows):
        labels = {
            _report_header_text(_cell_text(cell, shared_strings)): column
            for column, cell in _cell_map(row).items()
        }
        candidate = {
            key: labels[label]
            for label, key in _TRUEWERK_PACKING_HEADERS.items()
            if label in labels
        }
        if len(candidate) != len(_TRUEWERK_PACKING_HEADERS):
            continue
        for detail_index, detail_row in enumerate(rows[index + 1 : index + 3], index + 1):
            detail_labels = {
                _report_header_text(_cell_text(cell, shared_strings)): column
                for column, cell in _cell_map(detail_row).items()
            }
            detail_candidate = {
                key: detail_labels[label]
                for label, key in _TRUEWERK_PACKING_DETAIL_HEADERS.items()
                if label in detail_labels
            }
            if len(detail_candidate) == len(_TRUEWERK_PACKING_DETAIL_HEADERS):
                columns = candidate | detail_candidate
                body_index = detail_index + 1
                break
        if body_index >= 0:
            break
    if body_index < 0:
        return False

    existing = _existing_merge_ranges(root)
    new_ranges: list[tuple[int, int, int, int]] = []
    replaced_ranges: set[tuple[int, int, int, int]] = set()

    def merge_group(group: list[tuple[ET.Element, dict[int, ET.Element], bool]]) -> None:
        nonlocal existing
        if len(group) < 2 or not any(is_add for _row, _cells, is_add in group):
            return
        if not any(not is_add for _row, _cells, is_add in group):
            return
        first_row = int(group[0][0].get("r", "0"))
        last_row = int(group[-1][0].get("r", "0"))
        for name in _TRUEWERK_PACKING_MEASUREMENTS:
            column = columns[name]
            cells = [cell_map.get(column) for _row, cell_map, _is_add in group]
            if any(cell is None for cell in cells):
                continue
            values = [_cell_text(cell, shared_strings).strip() for cell in cells]
            non_zero_indices = [
                index
                for index, value in enumerate(values)
                if value and not _is_zero_measurement(value)
            ]
            if len(non_zero_indices) != 1:
                continue
            anchor_range = next(
                (
                    item
                    for item in existing
                    if item[0] <= column <= item[1]
                    and item[2] == first_row
                    and item[3] == first_row
                ),
                None,
            )
            start_column, end_column = (
                anchor_range[:2] if anchor_range is not None else (column, column)
            )
            target_range = (start_column, end_column, first_row, last_row)
            row_ranges = {
                (start_column, end_column, row_number, row_number)
                for row_number in range(first_row, last_row + 1)
            }
            conflicts = [
                item
                for item in existing
                if not (
                    end_column < item[0]
                    or start_column > item[1]
                    or last_row < item[2]
                    or first_row > item[3]
                )
            ]
            if (
                any(item not in row_ranges for item in conflicts)
            ):
                continue
            new_ranges.append(target_range)
            replaced_ranges.update(conflicts)
            existing = [item for item in existing if item not in conflicts]
            existing.append(target_range)
            source_index = non_zero_indices[0]
            if source_index:
                _copy_cell_value(cells[source_index], cells[0])
            for cell in cells[1:]:
                _clear_merged_cell_value(cell)

    group: list[tuple[ET.Element, dict[int, ET.Element], bool]] = []
    previous_row = 0
    previous_key: tuple[str, str] | None = None
    for row in rows[body_index:]:
        row_number = int(row.get("r", "0"))
        cell_map = _cell_map(row)
        style_cell, po_cell = cell_map.get(columns["style"]), cell_map.get(columns["po"])
        style = _cell_text(style_cell, shared_strings).strip() if style_cell is not None else ""
        po_value = _cell_text(po_cell, shared_strings).strip() if po_cell is not None else ""
        po_base, is_add = _truewerk_po_base(po_value)
        key = (style.casefold(), po_base)
        if not all(key) or row_number != previous_row + 1 or key != previous_key:
            merge_group(group)
            group = []
        if all(key):
            group.append((row, cell_map, is_add))
            previous_key = key
            previous_row = row_number
        else:
            previous_key = None
            previous_row = 0
    merge_group(group)
    if not new_ranges:
        return False

    merge_cells = root.find(_tag(MAIN_NS, "mergeCells"))
    if merge_cells is None:
        merge_cells = ET.Element(_tag(MAIN_NS, "mergeCells"))
        root.insert(list(root).index(sheet_data) + 1, merge_cells)
    for item in list(merge_cells):
        reference = str(item.get("ref") or "")
        start, separator, end = reference.partition(":")
        if not separator:
            end = start
        range_tuple = (
            _column_index(start),
            _column_index(end),
            _cell_row(start),
            _cell_row(end),
        )
        if range_tuple in replaced_ranges:
            merge_cells.remove(item)
    for start_column, end_column, first_row, last_row in new_ranges:
        start_letter = ""
        value = start_column
        while value:
            value, remainder = divmod(value - 1, 26)
            start_letter = chr(ord("A") + remainder) + start_letter
        end_letter = ""
        value = end_column
        while value:
            value, remainder = divmod(value - 1, 26)
            end_letter = chr(ord("A") + remainder) + end_letter
        merge_cells.append(
            ET.Element(
                _tag(MAIN_NS, "mergeCell"),
                {"ref": f"{start_letter}{first_row}:{end_letter}{last_row}"},
            )
        )
    merge_cells.set("count", str(len(merge_cells)))
    return True


def _merge_truewerk_packing_measurements(target: Path) -> None:
    """Áp dụng gộp bốn cột tổng PO/PO ADD cho Packing List TRUEWERK."""
    with ZipFile(target, "r") as source:
        shared_strings = _shared_strings(source)
        replacements: dict[str, bytes] = {}
        for info in source.infolist():
            if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", info.filename):
                continue
            root = _xml(source.read(info.filename))
            if _merge_truewerk_packing_sheet(root, shared_strings):
                replacements[info.filename] = _xml_bytes(root)
        if not replacements:
            return
        temp_target = target.with_suffix(".truewerk-merge-adjusting.xlsx")
        with ZipFile(temp_target, "w", compression=ZIP_DEFLATED) as output:
            for info in source.infolist():
                output.writestr(
                    info,
                    replacements.get(info.filename, source.read(info.filename)),
                )
    temp_target.replace(target)


def merge_sale_asn_reports(
    packing_list_path: str | Path,
    buyer_invoice_path: str | Path,
    output_path: str | Path,
    *,
    invoice_no: str = "",
) -> Path:
    """Ghép report ở cấp OOXML để giữ nguyên khung của cả Invoice và PKL."""
    packing_path = Path(packing_list_path).expanduser().resolve()
    buyer_path = Path(buyer_invoice_path).expanduser().resolve()
    target = Path(output_path).expanduser().resolve()
    if target.suffix.casefold() != ".xlsx":
        raise ASNWorkbookError("File Sale ASN phải có đuôi .xlsx.")
    for source in (packing_path, buyer_path):
        if not source.is_file() or source.stat().st_size <= 0:
            raise ASNWorkbookError(f"Không đọc được report: {source.name}.")
    invoice_label = str(invoice_no or target.stem).strip() or "Invoice"
    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        _merge_packages(packing_path, buyer_path, target, invoice_label)
        _merge_jl_packing_measurements(target)
        _merge_truewerk_packing_measurements(target)
        _fit_packing_measurement_columns(target)
        _fit_wrapped_report_rows(target)
        _fit_reports_to_a4(target)
        verified = load_workbook(target, read_only=True, data_only=False)
        verified.close()
    except ASNWorkbookError:
        target.unlink(missing_ok=True)
        raise
    except (BadZipFile, KeyError, ET.ParseError) as error:
        target.unlink(missing_ok=True)
        raise ASNWorkbookError(
            f"Report WFX không phải workbook Excel hợp lệ: {error}"
        ) from error
    except Exception as error:
        target.unlink(missing_ok=True)
        raise ASNWorkbookError(f"Không ghép được hai report Sale ASN: {error}") from error
    return target


def sale_asn_sheet_names(path: str | Path) -> list[str]:
    """Đọc tên sheet thực tế sau khi ghép để UI không báo tên giả định."""
    source = Path(path).expanduser().resolve()
    try:
        workbook = load_workbook(source, read_only=True, data_only=False)
    except Exception as error:
        raise ASNWorkbookError(f"Không đọc được tên sheet Sale ASN: {error}") from error
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()
