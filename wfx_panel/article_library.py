"""Đồng bộ thư viện Article dùng cho dropdown Costing từ server.

Server phát hành một manifest nhỏ trỏ tới JSON dữ liệu kèm SHA-256. Ứng dụng
kiểm tra nền, chỉ tải khi version đổi và luôn giữ bản cache cuối cùng để dùng
offline. Đây là dữ liệu tham chiếu đọc-only; user không phải scan WFX hay nhập
file thủ công trên từng máy.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import threading
import time
from collections.abc import Callable, Mapping
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from wfx_panel.atomic_io import write_json_atomic

ENV_MANIFEST_URL = "WFX_ARTICLE_LIBRARY_MANIFEST_URL"
DEFAULT_MANIFEST_URL = (
    "https://raw.githubusercontent.com/sieuxuan/WFX-APP/main/"
    "data/article-library-manifest.json"
)
SCHEMA_VERSION = 1
REQUEST_TIMEOUT_SECONDS = 20
MAX_MANIFEST_BYTES = 128 * 1024
MAX_LIBRARY_BYTES = 25 * 1024 * 1024
MAX_SECTIONS = 100
MAX_OPTIONS_PER_SECTION = 100_000
ARTICLE_CODE_HEADERS = frozenset(
    {"articlecode", "code", "materialcode", "itemcode"}
)
ARTICLE_NAME_HEADERS = frozenset(
    {"articlename", "name", "materialname", "itemname", "description"}
)
BUYER_REFERENCE_HEADERS = frozenset(
    {"buyerreference", "buyerref", "buyerarticle", "buyercode"}
)
ARTICLE_CATEGORY_HEADERS = frozenset(
    {"articlecategory", "category", "categoryname"}
)
_MEMORY_LOCK = threading.Lock()
_MEMORY_CACHE: dict[Path, tuple[int, int, dict[str, Any]]] = {}


def _cache_path(base_dir: Path) -> Path:
    return Path(base_dir) / "article-library.json"


def _safe_https_url(value: str, *, base_url: str | None = None) -> str:
    url = urljoin(base_url or "", str(value or "").strip())
    parsed = urlparse(url)
    if parsed.scheme.casefold() != "https" or not parsed.hostname:
        raise ValueError("Article Library chỉ chấp nhận URL HTTPS.")
    return url


def _download_bytes(url: str, *, maximum: int) -> bytes:
    request = Request(
        _safe_https_url(url),
        headers={
            "Accept": "application/json",
            "User-Agent": "WFX-Smart-Article-Library/1",
        },
    )
    with urlopen(request, timeout=REQUEST_TIMEOUT_SECONDS) as response:
        content_length = int(response.headers.get("Content-Length") or 0)
        if content_length > maximum:
            raise ValueError("Article Library vượt giới hạn dung lượng.")
        payload = response.read(maximum + 1)
    if len(payload) > maximum:
        raise ValueError("Article Library vượt giới hạn dung lượng.")
    return payload


def _option(raw: object) -> dict[str, str] | None:
    if not isinstance(raw, Mapping):
        return None
    code = str(raw.get("article_code") or raw.get("code") or "").strip()[:160]
    name = str(raw.get("article_name") or raw.get("name") or "").strip()[:300]
    buyer_reference = str(
        raw.get("buyer_reference") or raw.get("buyer_ref") or ""
    ).strip()[:300]
    article_category = str(
        raw.get("article_category") or raw.get("category") or ""
    ).strip()[:180]
    if not code and not name:
        return None
    return {
        "article_code": code,
        "article_name": name,
        "buyer_reference": buyer_reference,
        "article_category": article_category,
    }


def _header_token(value: object) -> str:
    return "".join(character for character in str(value or "").casefold() if character.isalnum())


def _normalise_articles(raw_articles: object) -> list[dict[str, str]]:
    if not isinstance(raw_articles, list):
        return []
    options: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for raw in raw_articles[:MAX_OPTIONS_PER_SECTION]:
        option = _option(raw)
        if option is None:
            continue
        identity = (
            option["article_code"].casefold(),
            option["article_name"].casefold(),
            option["buyer_reference"].casefold(),
            option["article_category"].casefold(),
        )
        if identity in seen:
            continue
        seen.add(identity)
        options.append(option)
    return options


def _columns_from_headers(
    values: tuple[object, ...] | list[object],
) -> dict[str, int | None]:
    tokens = [_header_token(value) for value in values]

    def find(headers: frozenset[str]) -> int | None:
        return next(
            (index for index, token in enumerate(tokens) if token in headers),
            None,
        )

    return {
        "article_code": find(ARTICLE_CODE_HEADERS),
        "article_name": find(ARTICLE_NAME_HEADERS),
        "buyer_reference": find(BUYER_REFERENCE_HEADERS),
        "article_category": find(ARTICLE_CATEGORY_HEADERS),
    }


def _option_from_row(
    row: tuple[object, ...] | list[object],
    columns: Mapping[str, int | None],
) -> dict[str, str]:
    def value(key: str) -> str:
        column = columns.get(key)
        if column is None or column >= len(row):
            return ""
        return str(row[column] or "").strip()

    return {
        "article_code": value("article_code"),
        "article_name": value("article_name"),
        "buyer_reference": value("buyer_reference"),
        "article_category": value("article_category"),
    }


def _articles_from_xlsx(payload: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(
        BytesIO(payload),
        read_only=True,
        data_only=True,
    )
    articles: list[dict[str, str]] = []
    try:
        for worksheet in workbook.worksheets:
            columns: dict[str, int | None] = {}
            header_row = None
            for row_index, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=20, values_only=True),
                start=1,
            ):
                columns = _columns_from_headers(list(row))
                if (
                    columns["article_code"] is not None
                    and columns["article_name"] is not None
                ):
                    header_row = row_index
                    break
            if header_row is None:
                continue
            for row in worksheet.iter_rows(
                min_row=header_row + 1,
                values_only=True,
            ):
                articles.append(_option_from_row(list(row), columns))
                if len(articles) >= MAX_OPTIONS_PER_SECTION:
                    return _normalise_articles(articles)
    finally:
        workbook.close()
    return _normalise_articles(articles)


def _articles_from_csv(payload: bytes) -> list[dict[str, str]]:
    text = payload.decode("utf-8-sig")
    sample = text[:16_384]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = csv.reader(StringIO(text), dialect)
    try:
        headers = next(rows)
    except StopIteration:
        return []
    columns = _columns_from_headers(headers)
    if (
        columns["article_code"] is None
        or columns["article_name"] is None
    ):
        return []
    articles = []
    for row in rows:
        articles.append(_option_from_row(row, columns))
        if len(articles) >= MAX_OPTIONS_PER_SECTION:
            break
    return _normalise_articles(articles)


def normalise_sections(raw_sections: object) -> list[dict[str, Any]]:
    if not isinstance(raw_sections, list):
        return []
    sections: list[dict[str, Any]] = []
    seen_sections: set[tuple[str, str]] = set()
    for raw in raw_sections[:MAX_SECTIONS]:
        if not isinstance(raw, Mapping):
            continue
        key = str(raw.get("section_key") or "").strip()[:180]
        name = str(raw.get("section_name") or raw.get("name") or "").strip()[:180]
        if not key or not name:
            continue
        section_identity = (key.casefold(), name.casefold())
        if section_identity in seen_sections:
            continue
        options: list[dict[str, str]] = []
        seen_options: set[tuple[str, str, str, str]] = set()
        raw_options = raw.get("options")
        if not isinstance(raw_options, list):
            continue
        for value in raw_options[:MAX_OPTIONS_PER_SECTION]:
            option = _option(value)
            if option is None:
                continue
            identity = (
                option["article_code"].casefold(),
                option["article_name"].casefold(),
                option["buyer_reference"].casefold(),
                option["article_category"].casefold(),
            )
            if identity in seen_options:
                continue
            seen_options.add(identity)
            options.append(option)
        if not options:
            continue
        seen_sections.add(section_identity)
        sections.append(
            {
                "section_key": key,
                "section_name": name,
                "options": options,
            }
        )
    return sections


def _sections_from_payload(
    payload: bytes,
    *,
    data_format: str,
) -> tuple[list[dict[str, Any]], str]:
    if data_format == "xlsx":
        articles = _articles_from_xlsx(payload)
        generated_at = ""
    elif data_format == "csv":
        articles = _articles_from_csv(payload)
        generated_at = ""
    else:
        document = json.loads(payload.decode("utf-8-sig"))
        if not isinstance(document, Mapping):
            raise ValueError("Dữ liệu Article Library không hợp lệ.")
        if int(document.get("schema_version") or 0) != SCHEMA_VERSION:
            raise ValueError("Schema dữ liệu Article Library chưa được hỗ trợ.")
        articles = _normalise_articles(document.get("articles"))
        generated_at = str(document.get("generated_at") or "")[:120]
        sections = normalise_sections(document.get("sections"))
        if sections:
            return sections, generated_at
    if not articles:
        return [], generated_at
    return [
        {
            "section_key": "*",
            "section_name": "All Categories",
            "options": articles,
        }
    ], generated_at


def load_cached(base_dir: Path) -> dict[str, Any] | None:
    path = _cache_path(base_dir)
    try:
        file_state = path.stat()
    except (OSError, json.JSONDecodeError):
        with _MEMORY_LOCK:
            _MEMORY_CACHE.pop(path, None)
        return None
    signature = (file_state.st_mtime_ns, file_state.st_size)
    with _MEMORY_LOCK:
        cached = _MEMORY_CACHE.get(path)
        if cached is not None and cached[:2] == signature:
            return cached[2]
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(raw, Mapping) or raw.get("schema_version") != SCHEMA_VERSION:
        return None
    sections = normalise_sections(raw.get("sections"))
    if not sections:
        return None
    document = {
        "schema_version": SCHEMA_VERSION,
        "remote_version": str(raw.get("remote_version") or "")[:120],
        "generated_at": str(raw.get("generated_at") or "")[:120],
        "synced_at": float(raw.get("synced_at") or 0),
        "sha256": str(raw.get("sha256") or "").casefold(),
        "sections": sections,
    }
    with _MEMORY_LOCK:
        _MEMORY_CACHE[path] = (*signature, document)
    return document


def status(base_dir: Path) -> dict[str, Any]:
    cached = load_cached(base_dir)
    if cached is None:
        return {
            "available": False,
            "version": "",
            "generated_at": "",
            "synced_at": 0,
            "section_count": 0,
            "article_count": 0,
        }
    sections = cached["sections"]
    return {
        "available": True,
        "version": cached["remote_version"],
        "generated_at": cached["generated_at"],
        "synced_at": cached["synced_at"],
        "section_count": len(sections),
        "article_count": sum(len(section["options"]) for section in sections),
    }


def seed_bundled(base_dir: Path, source: Path) -> bool:
    """Tạo cache lần đầu từ CSV đóng gói; không ghi đè cache user đã có."""
    target = _cache_path(base_dir)
    if target.is_file() or not Path(source).is_file():
        return False
    payload = Path(source).read_bytes()
    sections, generated_at = _sections_from_payload(
        payload,
        data_format="csv",
    )
    if not sections:
        return False
    digest = hashlib.sha256(payload).hexdigest()
    target.parent.mkdir(parents=True, exist_ok=True)
    write_json_atomic(
        target,
        {
            "schema_version": SCHEMA_VERSION,
            "remote_version": f"bundled-{digest[:12]}",
            "generated_at": generated_at,
            "synced_at": time.time(),
            "sha256": digest,
            "sections": sections,
        },
        separators=(",", ":"),
    )
    with _MEMORY_LOCK:
        _MEMORY_CACHE.pop(target, None)
    return True


def save_server_articles(
    base_dir: Path,
    raw_articles: list[object],
    *,
    version: str,
    generated_at: str = "",
) -> dict[str, Any]:
    """Lưu list 3 cột từ PostgreSQL và giữ Category từ cache đóng gói.

    API chia sẻ cố ý chỉ có Code/Name/Buyer Reference. Category vẫn cần cho
    bộ lọc Catalog, nên app ghép lại từ CSV đóng gói/cache cũ theo mã Article.
    """
    existing = load_cached(base_dir)
    category_by_code: dict[str, str] = {}
    if existing is not None:
        for section in existing.get("sections") or ():
            for option in section.get("options") or ():
                code = str(option.get("article_code") or "").strip().casefold()
                category = str(option.get("article_category") or "").strip()
                if code and category:
                    category_by_code.setdefault(code, category)
    enriched = []
    for raw in raw_articles:
        if not isinstance(raw, Mapping):
            continue
        row = dict(raw)
        code = str(row.get("article_code") or row.get("code") or "").strip()
        if not str(row.get("article_category") or "").strip():
            row["article_category"] = category_by_code.get(code.casefold(), "")
        enriched.append(row)
    articles = _normalise_articles(enriched)
    if not articles:
        raise ValueError("Server không trả Article hợp lệ.")
    digest_payload = json.dumps(
        articles,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    target = _cache_path(base_dir)
    target.parent.mkdir(parents=True, exist_ok=True)
    write_json_atomic(
        target,
        {
            "schema_version": SCHEMA_VERSION,
            "remote_version": str(version or "")[:120],
            "generated_at": str(generated_at or "")[:120],
            "synced_at": time.time(),
            "sha256": hashlib.sha256(digest_payload).hexdigest(),
            "sections": [
                {
                    "section_key": "*",
                    "section_name": "All Categories",
                    "options": articles,
                }
            ],
        },
        separators=(",", ":"),
    )
    with _MEMORY_LOCK:
        _MEMORY_CACHE.pop(target, None)
    return status(base_dir)


def sync(
    base_dir: Path,
    log: Callable[[str], None] = print,
    *,
    manifest_url: str | None = None,
) -> dict[str, Any]:
    """Đồng bộ server → cache; lỗi mạng không xóa bản cache đang dùng."""
    source_url = _safe_https_url(
        manifest_url
        or os.getenv(ENV_MANIFEST_URL, "").strip()
        or DEFAULT_MANIFEST_URL
    )
    try:
        manifest_payload = _download_bytes(
            source_url,
            maximum=MAX_MANIFEST_BYTES,
        )
        manifest = json.loads(manifest_payload.decode("utf-8-sig"))
        if not isinstance(manifest, Mapping):
            raise ValueError("Manifest Article Library không hợp lệ.")
        if int(manifest.get("schema_version") or 0) != SCHEMA_VERSION:
            raise ValueError("Schema Article Library chưa được hỗ trợ.")
        version = str(manifest.get("version") or "").strip()[:120]
        expected_sha = str(manifest.get("sha256") or "").strip().casefold()
        data_url = _safe_https_url(
            str(manifest.get("data_url") or ""),
            base_url=source_url,
        )
        data_format = str(manifest.get("format") or "").strip().casefold()
        if not data_format:
            suffix = Path(urlparse(data_url).path).suffix.casefold()
            data_format = {
                ".xlsx": "xlsx",
                ".csv": "csv",
            }.get(suffix, "json")
        if data_format not in {"json", "xlsx", "csv"}:
            raise ValueError("Định dạng Article Library chưa được hỗ trợ.")
        if not version or len(expected_sha) != 64:
            raise ValueError("Manifest Article Library thiếu version/checksum.")

        cached = load_cached(base_dir)
        if (
            cached is not None
            and cached["remote_version"] == version
            and cached["sha256"] == expected_sha
        ):
            return {
                "ok": True,
                "code": "ARTICLE_LIBRARY_CURRENT",
                "message": "Thư viện Article đã là bản mới nhất.",
                **status(base_dir),
            }

        payload = _download_bytes(data_url, maximum=MAX_LIBRARY_BYTES)
        actual_sha = hashlib.sha256(payload).hexdigest()
        if actual_sha != expected_sha:
            raise ValueError("Checksum Article Library không khớp.")
        sections, generated_at = _sections_from_payload(
            payload,
            data_format=data_format,
        )
        if not sections:
            raise ValueError("Article Library không có dữ liệu hợp lệ.")

        synced_at = time.time()
        write_json_atomic(
            _cache_path(base_dir),
            {
                "schema_version": SCHEMA_VERSION,
                "remote_version": version,
                "generated_at": generated_at,
                "synced_at": synced_at,
                "sha256": actual_sha,
                "sections": sections,
            },
            separators=(",", ":"),
        )
        result = status(base_dir)
        log(
            "[ARTICLE LIBRARY] Đã cập nhật "
            f"{result['article_count']} Article từ server."
        )
        return {
            "ok": True,
            "code": "ARTICLE_LIBRARY_UPDATED",
            "message": (
                f"Đã cập nhật {result['article_count']} Article từ server."
            ),
            **result,
        }
    except (
        OSError,
        ValueError,
        BadZipFile,
        InvalidFileException,
        json.JSONDecodeError,
        UnicodeError,
    ) as error:
        cached_status = status(base_dir)
        log(
            "[ARTICLE LIBRARY] Chưa đồng bộ được; "
            f"tiếp tục dùng cache: {type(error).__name__}."
        )
        return {
            "ok": False,
            "code": "ARTICLE_LIBRARY_SYNC_FAILED",
            "message": (
                "Chưa tải được thư viện Article mới; "
                "app tiếp tục dùng bản gần nhất."
            ),
            "error_type": type(error).__name__,
            **cached_status,
        }
