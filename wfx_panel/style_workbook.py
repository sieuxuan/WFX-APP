"""Form Excel và validation cho luồng tạo Style Apparel hàng loạt."""

from __future__ import annotations

import unicodedata
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

STYLE_SHEET = "Tạo Style"
GUIDE_SHEET = "Hướng dẫn"
LIST_SHEET = "_Danh sách"
MAX_STYLE_ROWS = 500
MAX_STYLE_FILE_BYTES = 15 * 1024 * 1024

STYLE_COLUMNS = (
    "Type",
    "Style copy",
    "Material Type",
    "Buyer",
    "Division",
    "Product Group",
    "Sub-Category",
    "Color Card",
    "Size Range",
    "Season",
    "Buyer Style Ref.",
    "Internal Style Ref",
)

_FIELD_KEYS = (
    "type",
    "style_copy",
    "material_type",
    "buyer",
    "division",
    "product_group",
    "sub_category",
    "color_card",
    "size_range",
    "season",
    "buyer_style_ref",
    "internal_style_ref",
)


class StyleWorkbookError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        errors: tuple[str, ...] = (),
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.errors = errors


@dataclass(frozen=True)
class StyleImportRow:
    source_row: int
    type: str
    style_copy: str
    material_type: str
    buyer: str
    division: str
    product_group: str
    sub_category: str
    color_card: str
    size_range: str
    season: str
    buyer_style_ref: str
    internal_style_ref: str

    def automation_payload(self) -> dict[str, str | int]:
        return {
            key: getattr(self, key)
            for key in ("source_row", *_FIELD_KEYS)
        }


def _text(value: object) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def _assert_safe_xlsx(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        raise StyleWorkbookError(
            "STYLE_FILE_TYPE_UNSUPPORTED",
            "Tạo Style hàng loạt chỉ hỗ trợ file .xlsx.",
        )
    if not path.is_file():
        raise StyleWorkbookError(
            "STYLE_FILE_NOT_FOUND",
            "Không tìm thấy file Tạo Style đã chọn.",
        )
    if path.stat().st_size > MAX_STYLE_FILE_BYTES:
        raise StyleWorkbookError(
            "STYLE_FILE_TOO_LARGE",
            "File Tạo Style vượt quá giới hạn 15 MB.",
        )
    try:
        with ZipFile(path) as archive:
            names = {name.casefold() for name in archive.namelist()}
    except (BadZipFile, OSError) as error:
        raise StyleWorkbookError(
            "STYLE_FILE_INVALID",
            "File đã chọn không phải workbook Excel hợp lệ.",
        ) from error
    if any(name.endswith("vbaproject.bin") for name in names):
        raise StyleWorkbookError(
            "STYLE_FILE_UNSAFE",
            "File Tạo Style không được chứa macro.",
        )


def _option_values(options: Mapping | None, key: str) -> list[str]:
    fields = options.get("fields") if isinstance(options, Mapping) else None
    raw = fields.get(key) if isinstance(fields, Mapping) else None
    if not isinstance(raw, list):
        return []
    values: list[str] = []
    seen: set[str] = set()
    for item in raw:
        value = _text(item.get("label") if isinstance(item, Mapping) else item)
        identity = value.casefold()
        if value and identity not in seen:
            seen.add(identity)
            values.append(value)
    return values


def _add_named_list(
    workbook: Workbook,
    worksheet,
    name: str,
    column: int,
    values: list[str],
) -> str:
    worksheet.cell(1, column, name)
    for row, value in enumerate(values, start=2):
        worksheet.cell(row, column, value)
    if not values:
        return ""
    letter = get_column_letter(column)
    reference = f"'{LIST_SHEET}'!${letter}$2:${letter}${len(values) + 1}"
    workbook.defined_names.add(DefinedName(name, attr_text=reference))
    return name


def write_style_template(
    output_path: str | Path,
    options: Mapping | None = None,
) -> Path:
    """Tạo form nhập liệu có dropdown server/cache, không chứa macro."""
    target = Path(output_path).expanduser().resolve()
    if target.suffix.casefold() != ".xlsx":
        target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    guide = workbook.active
    guide.title = GUIDE_SHEET
    guide.sheet_view.showGridLines = False
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 92
    guide.append(["WFX Smart", "Tạo Style Apparel hàng loạt"])
    guide.append(["Bước 1", "Trong tab Catalog > Tạo Style, chọn đúng một Group."])
    guide.append(["Bước 2", "Điền dữ liệu trong sheet Tạo Style rồi chọn Import."])
    guide.append([
        "Bước 3",
        "Mặc định app dừng trước Save. Có thể bật Tự động Save trong app sau "
        "khi đã kiểm tra kỹ dữ liệu Excel.",
    ])
    guide.append([
        "Type",
        "New để tạo mới hoàn toàn; Copy để tìm Style nguồn và Copy as Variant.",
    ])
    guide.append([
        "Style copy",
        "Bắt buộc với Copy. Chọn Article Name từ danh sách Apparel; "
        "app tìm trên WFX bằng ArticleCode/Name.",
    ])
    guide.append([
        "Ô trống",
        "Với Copy, ô trống giữ nguyên dữ liệu nguồn. Với New, các cột dữ liệu "
        "từ Material Type đến Internal Style Ref đều bắt buộc.",
    ])
    guide.append([
        "Mặc định",
        "Purchase UOM = Pcs; Price Per = Article; "
        "Color Definition = Single Colors.",
    ])
    guide.append([
        "An toàn",
        "Tự động Save luôn mặc định tắt. Không đóng/chuyển màn WFX khi app đang chạy.",
    ])
    guide.append([
        "Dropdown",
        "Danh sách được dùng chung từ server và chỉ quét lại WFX khi cache quá 30 ngày. "
        "Sub-Category đổi theo Product Group đã chọn.",
    ])
    for cell in guide[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="214A66")
    for row in guide.iter_rows(min_row=2, max_col=2):
        row[0].font = Font(bold=True, color="214A66")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    guide.freeze_panes = "A2"

    sheet = workbook.create_sheet(STYLE_SHEET)
    sheet.sheet_view.showGridLines = False
    sheet.append(STYLE_COLUMNS)
    # Phát hành sẵn các dòng trống để dropdown/format hoạt động ngay trong Excel.
    for _ in range(100):
        sheet.append([""] * len(STYLE_COLUMNS))
    widths = (13, 22, 18, 24, 20, 22, 22, 22, 22, 18, 24, 24)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="214A66")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"
    table = Table(displayName="StyleUpload", ref="A1:L101")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    lists = workbook.create_sheet(LIST_SHEET)
    lists.sheet_state = "veryHidden"
    list_specs = (
        ("style_copy", "StyleCopy", 2),
        ("material_type", "StyleMaterialType", 3),
        ("buyer", "StyleBuyer", 4),
        ("division", "StyleDivision", 5),
        ("product_group", "StyleProductGroup", 6),
        ("color_card", "StyleColorCard", 8),
        ("size_range", "StyleSizeRange", 9),
        ("season", "StyleSeason", 10),
    )
    named_lists: dict[str, str] = {}
    list_column = 1
    for key, range_name, _style_column in list_specs:
        values = _option_values(options, key)
        if key == "material_type" and not values:
            values = ["KNIT", "WOVEN"]
        named_lists[key] = _add_named_list(
            workbook, lists, range_name, list_column, values
        )
        list_column += 1

    dependencies_raw = (
        options.get("subcategories_by_product_group")
        if isinstance(options, Mapping)
        else None
    )
    dependencies = dependencies_raw if isinstance(dependencies_raw, Mapping) else {}
    dependency_map_column = list_column
    lists.cell(1, dependency_map_column, "Product Group")
    lists.cell(1, dependency_map_column + 1, "Range")
    all_subcategories: list[str] = []
    seen_subcategories: set[str] = set()
    dependency_row = 2
    for index, (product_group, raw_values) in enumerate(dependencies.items(), start=1):
        values = []
        for item in raw_values if isinstance(raw_values, list) else []:
            value = _text(item.get("label") if isinstance(item, Mapping) else item)
            if value and value.casefold() not in {item.casefold() for item in values}:
                values.append(value)
            if value and value.casefold() not in seen_subcategories:
                seen_subcategories.add(value.casefold())
                all_subcategories.append(value)
        if not values:
            continue
        range_name = f"StyleSub{index:04d}"
        _add_named_list(workbook, lists, range_name, list_column + 2 + index, values)
        lists.cell(dependency_row, dependency_map_column, _text(product_group))
        lists.cell(dependency_row, dependency_map_column + 1, range_name)
        dependency_row += 1
    fallback_subcategory = _add_named_list(
        workbook,
        lists,
        "StyleSubcategoryAll",
        list_column + 2,
        all_subcategories,
    )

    type_validation = DataValidation(
        type="list",
        formula1='"New,Copy"',
        allow_blank=False,
    )
    sheet.add_data_validation(type_validation)
    type_validation.add(f"A2:A{MAX_STYLE_ROWS + 1}")
    for key, _range_name, style_column in list_specs:
        if not named_lists.get(key):
            continue
        validation = DataValidation(
            type="list",
            formula1=f"={named_lists[key]}",
            allow_blank=True,
        )
        sheet.add_data_validation(validation)
        letter = get_column_letter(style_column)
        validation.add(f"{letter}2:{letter}{MAX_STYLE_ROWS + 1}")
    if fallback_subcategory:
        map_end = max(2, dependency_row - 1)
        formula = (
            "=INDIRECT(IFERROR(VLOOKUP($F2,'_Danh sách'!"
            f"${get_column_letter(dependency_map_column)}$2:"
            f"${get_column_letter(dependency_map_column + 1)}${map_end},2,FALSE),"
            '"StyleSubcategoryAll"))'
        )
        subcategory_validation = DataValidation(
            type="list", formula1=formula, allow_blank=True
        )
        sheet.add_data_validation(subcategory_validation)
        subcategory_validation.add(f"G2:G{MAX_STYLE_ROWS + 1}")
    warning_fill = PatternFill("solid", fgColor="FDE9D9")
    sheet.conditional_formatting.add(
        f"B2:B{MAX_STYLE_ROWS + 1}",
        FormulaRule(formula=['AND($A2="Copy",$B2="")'], fill=warning_fill),
    )
    sheet.conditional_formatting.add(
        f"C2:L{MAX_STYLE_ROWS + 1}",
        FormulaRule(formula=['AND($A2="New",C2="")'], fill=warning_fill),
    )
    workbook.save(target)
    workbook.close()
    return target


def read_style_workbook(source_path: str | Path) -> list[StyleImportRow]:
    """Đọc và validate toàn bộ hàng trước khi mở WFX."""
    source = Path(source_path).expanduser().resolve()
    _assert_safe_xlsx(source)
    try:
        workbook = load_workbook(source, read_only=True, data_only=False)
    except Exception as error:
        raise StyleWorkbookError(
            "STYLE_FILE_INVALID",
            "Không đọc được workbook Tạo Style.",
        ) from error
    try:
        if STYLE_SHEET not in workbook.sheetnames:
            raise StyleWorkbookError(
                "STYLE_TEMPLATE_SHEET_MISSING",
                f"File phải có sheet {STYLE_SHEET}.",
            )
        sheet = workbook[STYLE_SHEET]
        headers = tuple(_text(cell.value) for cell in sheet[1][: len(STYLE_COLUMNS)])
        if headers != STYLE_COLUMNS:
            raise StyleWorkbookError(
                "STYLE_FILE_HEADERS_INVALID",
                "Header sheet Tạo Style không đúng form chuẩn.",
                ("Cần đúng thứ tự: " + " | ".join(STYLE_COLUMNS),),
            )
        rows: list[StyleImportRow] = []
        errors: list[str] = []
        for source_row, cells in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(STYLE_COLUMNS)),
            start=2,
        ):
            values: list[str] = []
            formula = False
            for cell in cells:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = True
                values.append(_text(cell.value))
            if not any(values):
                continue
            if len(rows) >= MAX_STYLE_ROWS:
                errors.append(f"Vượt quá {MAX_STYLE_ROWS} dòng dữ liệu.")
                break
            if formula:
                errors.append(f"Dòng {source_row}: không được dùng công thức.")
                continue
            kind = values[0].casefold()
            if kind not in {"new", "copy"}:
                errors.append(f"Dòng {source_row}: Type phải là New hoặc Copy.")
                continue
            normalized_kind = "New" if kind == "new" else "Copy"
            if normalized_kind == "Copy" and not values[1]:
                errors.append(f"Dòng {source_row}: Copy cần có Style copy.")
                continue
            if normalized_kind == "New":
                missing = [
                    STYLE_COLUMNS[index]
                    for index, value in enumerate(values[2:], start=2)
                    if not value
                ]
                if missing:
                    errors.append(
                        f"Dòng {source_row}: New còn thiếu {', '.join(missing)}."
                    )
                    continue
            material = values[2].upper()
            if material and material not in {"KNIT", "WOVEN"}:
                errors.append(
                    f"Dòng {source_row}: Material Type phải là KNIT hoặc WOVEN."
                )
                continue
            payload = [normalized_kind, values[1], material, *values[3:]]
            rows.append(
                StyleImportRow(
                    source_row=source_row,
                    **dict(zip(_FIELD_KEYS, payload, strict=True)),
                )
            )
        if errors:
            raise StyleWorkbookError(
                "STYLE_FILE_VALIDATION_FAILED",
                f"File Tạo Style có {len(errors)} lỗi.",
                tuple(errors[:50]),
            )
        if not rows:
            raise StyleWorkbookError(
                "STYLE_FILE_EMPTY",
                "File Tạo Style chưa có dòng dữ liệu.",
            )
        return rows
    finally:
        workbook.close()
