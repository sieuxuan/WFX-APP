"""Chuẩn hóa và kiểm tra workbook Upload OC trước khi gửi sang WFX EDI."""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.datavalidation import DataValidation

MAX_XLSX_BYTES = 100 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 2_000
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_OC_ROWS = 10_000

FORM_HEADERS = (
    "Factory",
    "Ship Under PO Ref",
    "Article Code",
    "Buyer Style Ref",
    "Buyer PO Num",
    "Summary Buyer Order Ref",
    "Buyer Order Date",
    "Order/Buyer Delivery Date",
    "Raw Matetrial ETA Date",
    "Payment Terms",
    "Country of Final Destination",
    "Color code",
    "Color name",
    "Size code",
    "Selling Price",
    "Units",
    "Internal Lot No.",
    "PO Type",
    "Extra Production",
    "Buyer Lot No.",
)

INPUT_SHEET_NAME = "OC INPUT"
REFERENCE_SHEET_NAME = "REFERENCES"
INPUT_HEADERS = (
    "Buyer",
    "Season",
    "Order Type",
    "Currency",
    "Factory",
    "Ship Under PO Ref",
    "Article Code",
    "Buyer Style Ref",
    "Buyer PO Num",
    "Summary Buyer Order Ref",
    "Buyer Order Date",
    "Buyer Delivery Date",
    "Raw Material ETA Date",
    "Payment Terms",
    "Country of Final Destination",
    "Color Code",
    "Color Name",
    "Size Code",
    "Selling Price",
    "Units",
    "Internal Lot No.",
    "PO Type (Zone)",
    "Extra Production %",
    "Buyer Lot No.",
)

FACTORY_OPTIONS = (
    "888 COMPANY LTD",
    "CELEBRITY FASHION VINA COMPANY LIMITED",
    "GARMENT 10 CORPORATION-JOINT STOCK COMPANY",
    "PRO SPORTS GIAO YEN GARMENT JOINT STOCK COMPANY",
    "HABAC EXPORT GARMENT JOINT STOCK COMPANY",
    "HANSOL VINA LTD COMPANY",
    "HUNG BINH GARMENT JOINT STOCK COMPANY",
    "PHU THO GARMENT JOINT STOCK COMPANY",
    "PROSPORTS GIAO THUY JOINT STOCK COMPANY",
    "PRO SPORTS HA NOI JSC",
    "SON HA GARMENT JOINT STOCK COMPANY",
    "TNG PHU BINH 1 BRANCH",
    "X20 JOINT STOCK COMPANY (HEAD OFFICE)",
    "FACTORY GARMENT BIM SON",
    "S&D THANH HOA CO.LTD",
    "X20 NGHE AN ONE MEMBER COMPANY LTD",
    "THANH TRI JOINT STOCK COMPANY",
    "THAGACO INTERNATIONAL INVESTMENT JSC",
    "HANA KOVI INC.",
    "THIEN AN PHU TEXTILE GARMENT JOINT STOCK COMPANY",
    "VIET THAI GARMENT EXPORT JOINT STOCK COMPANY",
    "S-LIFE JOINT STOCK COMPANY",
)

BUYER_OPTIONS = (
    "BIRDDOGS",
    "CORPORATE OFFICE - TRUEWERK",
    "DOITE",
    "FAM BRANDS",
    "FORTUNE SWIMWEAR LLC",
    "J.LINDEBERG",
    "JOCKEY",
    "ONEILL",
    "PARAGON FITWEAR, LLC",
    "PREMIER EXIM (HK) LTD.,",
    "REVOLUTIONRACE",
    "SWIM RX",
    "UFPRO",
)

ORDER_TYPE_OPTIONS = (
    "Confirmed",
    "Forecast",
    "SMS",
)

PAYMENT_TERM_OPTIONS = (
    "15% Deposit After Contract - 85% TT Before Shipment",
    "20% Deposit, Balance TT at Sight",
    "30 Days At Month End",
    "30% Advanced Before Shipment - 70% TT After Shipment",
    "30% Deposit + ROG + 30 Days",
    "30% Once Order Committed - 70% LC Irrevocable 30 Days",
    "40% Advanced Before Shipment - 60% TT After Shipment",
    "50% Deposit / 50% TT After 30 Days",
    "After Finished 30-45 Days",
    "By Bank Draft or TT Before Shipment",
    "Cash Before Delivery First 3 Shipments And Then 30 Days",
    "Credit of T/T 30 days",
    "Deposit 30% - 70% TT Against Shipment",
    "LC 45 Days",
    "LC 60 Days",
    "LC At Sight",
    "LC At Sight 30 Days",
    "OA 15 Days",
    "Payment 60 Days After Ex Works Date",
    "Payment Within 90 Days",
    "ROG 30 Days",
    "TT 30% Deposit - 70% Before Shipping",
    "TT After Shipment",
    "TT After Shipment 10 Days",
    "TT After Shipment 15 Days",
    "TT After Shipment 20 Days",
    "TT After Shipment 30 Days",
    "TT After Shipment 40 Days",
    "TT After Shipment 45 Days",
    "TT After Shipment 60 Days",
    "TT After Shipment 90 Days",
    "TT Against Documents",
    "TT Before ETA",
    "TT Before Shipment",
    "TT Before Shipment 30 Days",
    "TT In Advance For First Order / TT After 30 Days For Next Order",
    "TT Payment",
    "Wire Payment 90 Days",
)

PO_TYPE_OPTIONS = ("CM", "CMT", "FOB", "DDP")

COUNTRY_MARKET = {
    "Australia": "Australia",
    "Austria": "Europe",
    "Canada": "America",
    "Chile": "America",
    "China": "Asia",
    "Dominican Republic": "America",
    "France": "Europe",
    "Germany": "Europe",
    "Hong Kong": "Asia",
    "Italy": "Europe",
    "Japan": "Asia",
    "New Zealand": "Australia",
    "Norway": "Europe",
    "Singapore": "Asia",
    "South Korea": "Asia",
    "Spain": "Europe",
    "Sweden": "Europe",
    "Switzerland": "Europe",
    "Taiwan": "Asia",
    "Thailand": "Asia",
    "United Arab Emirates": "Asia",
    "United Kingdom": "Europe",
    "United States": "America",
    "United States-US-Flat": "America",
    "United States-US-Hanger": "America",
    "Vietnam": "Asia",
}

INPUT_COMMENTS = {
    "Buyer": "Chọn Buyer đúng với Buyer sẽ chọn tại EDI Buyer PO.",
    "Season": "Season phải giống Season trong Techpack Style.",
    "Order Type": "Chọn Confirmed, Forecast hoặc SMS.",
    "Ship Under PO Ref": "Mã PO dùng để gom các dòng Color/Size cùng đơn hàng.",
    "Article Code": "Article Code lấy trên WFX.",
    "Buyer Style Ref": "Buyer Style Ref phải giống Techpack Style.",
    "Buyer PO Num": "Thường giống Summary Buyer Order Ref.",
    "Buyer Order Date": "Nhập ngày theo dd-mm-yyyy; phải trước Raw Material ETA.",
    "Buyer Delivery Date": "Nhập ngày theo dd-mm-yyyy; phải sau Raw Material ETA.",
    "Raw Material ETA Date": (
        "Nhập ngày theo dd-mm-yyyy; phải sau Buyer Order Date và trước "
        "Buyer Delivery Date."
    ),
    "Payment Terms": "Chọn đúng điều khoản thanh toán trong danh sách WFX.",
    "Country of Final Destination": "App tự mapping Final Destination và Market.",
    "Color Code": "Color Code lấy trên WFX.",
    "Color Name": "Tên màu lấy trên WFX.",
    "Size Code": "Size Code lấy trên WFX.",
    "Selling Price": "Phải lớn hơn 0 và không nhỏ hơn Costing.",
    "Units": "Số nguyên; dòng có Units = 0 sẽ được app tự bỏ qua.",
    "Internal Lot No.": "Chia theo Buy hoặc số nội bộ.",
    "PO Type (Zone)": "Để trống nếu dùng FOB; hoặc chọn CM/CMT/FOB/DDP.",
    "Extra Production %": "Có thể để trống; app tự xuất 0.",
    "Buyer Lot No.": "Tuỳ chọn; dùng theo quy định Buyer.",
}

EDI_HEADERS = (
    "Factory",
    "Ship Under PO Ref",
    "Article",
    "Buyer",
    "Buyer Division/Dept",
    "Currency",
    "Season",
    "Country of Origin",
    "Place of Receipt by Pre-Carrier",
    "Prod. Capacity Booking No",
    "Order Initiation Date",
    "Payment Terms",
    "Buyer PO Num",
    "Summary Buyer Order Ref",
    "Market Buyer Order Ref",
    "Destination Buyer Order Ref",
    "Delivery Buyer Order Ref",
    "Buyer Order Date",
    "Order Type",
    "Mode of Shipment",
    "Buyer Delivery Date",
    "OC Delivery Date",
    "PCD Date",
    "Original GAC Date",
    "GAC Date",
    "Raw Matetrial ETA",
    "Country of Final Destination",
    "Final Destination",
    "Market",
    "Buyer Style Ref.",
    "Packing Type",
    "Packing Option/Flat Pack)",
    "Color",
    "Size",
    "Total Qty",
    "Price",
    "Units",
    "Delivery Terms",
    "Zone",
    "Internal Lot No.",
    "Buyer Lot No.",
    "DeliveryOCID",
    "Fulfillment Type",
    "Initial PCD Date",
    "FirstBuyerDeliveryDate",
    "Packing Code(SKU)",
    "Make to Stock",
    "Split",
    "Other Instruction",
    "Extra Production %",
    "Upcharge",
)

DATE_HEADERS = frozenset(
    {
        "Buyer Order Date",
        "Buyer Delivery Date",
        "OC Delivery Date",
        "PCD Date",
        "Original GAC Date",
        "GAC Date",
        "Raw Matetrial ETA",
        "Initial PCD Date",
        "FirstBuyerDeliveryDate",
    }
)

NEW_REQUIRED_FORM_COLUMNS = frozenset(range(17))
SIMPLE_NEW_OPTIONAL_COLUMNS = frozenset({21, 22, 23})
REVISE_REQUIRED_HEADERS = frozenset(
    {
        "Factory",
        "Ship Under PO Ref",
        "Article",
        "Buyer",
        "Currency",
        "Season",
        "Payment Terms",
        "Buyer PO Num",
        "Summary Buyer Order Ref",
        "Market Buyer Order Ref",
        "Destination Buyer Order Ref",
        "Delivery Buyer Order Ref",
        "Buyer Order Date",
        "Order Type",
        "Mode of Shipment",
        "Buyer Delivery Date",
        "OC Delivery Date",
        "Raw Matetrial ETA",
        "Country of Final Destination",
        "Final Destination",
        "Market",
        "Buyer Style Ref.",
        "Color",
        "Size",
        "Price",
        "Units",
        "Internal Lot No.",
        "DeliveryOCID",
        "Fulfillment Type",
    }
)


class OCWorkbookError(ValueError):
    def __init__(self, code: str, message: str, errors: Iterable[str] = ()):
        super().__init__(message)
        self.code = code
        self.message = message
        self.errors = tuple(errors)


@dataclass(frozen=True)
class PreparedOCUpload:
    mode: str
    buyer: str
    row_count: int
    upload_path: Path
    seasons: tuple[str, ...] = ()
    po_count: int = 0
    style_count: int = 0
    total_units: int | float = 0
    warnings: tuple[str, ...] = ()


def _normalise_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _normalise_header(value: Any) -> str:
    return _normalise_text(value).casefold().rstrip(".")


def _validate_xlsx_archive(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        raise OCWorkbookError(
            "OC_FILE_TYPE_UNSUPPORTED",
            "Upload OC chỉ hỗ trợ file .xlsx.",
        )
    if not path.is_file():
        raise OCWorkbookError(
            "OC_FILE_NOT_FOUND",
            "Không tìm thấy file Upload OC đã chọn.",
        )
    if path.stat().st_size > MAX_XLSX_BYTES:
        raise OCWorkbookError(
            "OC_FILE_TOO_LARGE",
            "File Upload OC vượt quá giới hạn 100 MB.",
        )
    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                raise OCWorkbookError(
                    "OC_FILE_UNSAFE",
                    "Workbook có quá nhiều thành phần nội bộ.",
                )
            total = 0
            for entry in entries:
                parts = Path(entry.filename.replace("\\", "/")).parts
                if entry.filename.startswith(("/", "\\")) or ".." in parts:
                    raise OCWorkbookError(
                        "OC_FILE_UNSAFE",
                        "Workbook chứa đường dẫn nội bộ không an toàn.",
                    )
                total += entry.file_size
                if total > MAX_UNCOMPRESSED_BYTES:
                    raise OCWorkbookError(
                        "OC_FILE_UNSAFE",
                        "Workbook nén vượt giới hạn giải nén an toàn.",
                    )
            names = {entry.filename for entry in entries}
            if "[Content_Types].xml" not in names or not any(
                name.startswith("xl/workbook") for name in names
            ):
                raise OCWorkbookError(
                    "OC_FILE_INVALID",
                    "File không phải workbook Excel hợp lệ.",
                )
    except BadZipFile as error:
        raise OCWorkbookError(
            "OC_FILE_INVALID",
            "File .xlsx bị hỏng hoặc không phải workbook Excel.",
        ) from error


def _ensure_input_values_only(path: Path, mode: str) -> None:
    """Reject formulas in the sheet users are allowed to edit."""
    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
    except (OSError, ValueError, BadZipFile) as error:
        raise OCWorkbookError(
            "OC_FILE_INVALID",
            f"Không đọc được workbook: {type(error).__name__}: {error}",
        ) from error
    try:
        if mode == "revise":
            sheet_name = "Sheet1"
            width = len(EDI_HEADERS)
        elif INPUT_SHEET_NAME in workbook.sheetnames:
            sheet_name = INPUT_SHEET_NAME
            width = len(INPUT_HEADERS)
        else:
            sheet_name = "FORM"
            width = len(FORM_HEADERS)
        if sheet_name not in workbook.sheetnames:
            return
        sheet = workbook[sheet_name]
        formulas = [
            cell.coordinate
            for row in sheet.iter_rows(max_col=width)
            for cell in row
            if cell.data_type == "f"
        ]
        if formulas:
            shown = ", ".join(formulas[:12])
            suffix = "…" if len(formulas) > 12 else "."
            raise OCWorkbookError(
                "OC_FILE_FORMULA_ERROR",
                f"Sheet {sheet_name} phải chỉ chứa giá trị nhập, không dùng công thức.",
                (f"Ô có công thức: {shown}{suffix}",),
            )
    finally:
        workbook.close()


def _decimal(value: Any, label: str, row_number: int) -> Decimal:
    if isinstance(value, bool):
        raise InvalidOperation
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError, ValueError) as error:
        raise OCWorkbookError(
            "OC_FILE_VALIDATION_FAILED",
            "Workbook Upload OC có dữ liệu chưa hợp lệ.",
            (f"Dòng {row_number}: {label} phải là số.",),
        ) from error
    if not number.is_finite():
        raise OCWorkbookError(
            "OC_FILE_VALIDATION_FAILED",
            "Workbook Upload OC có dữ liệu chưa hợp lệ.",
            (f"Dòng {row_number}: {label} phải là số hữu hạn.",),
        )
    return number


def _date_value(value: Any, label: str, row_number: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            pass
    raw = _normalise_text(value)
    for pattern in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    raise OCWorkbookError(
        "OC_FILE_VALIDATION_FAILED",
        "Workbook Upload OC có dữ liệu chưa hợp lệ.",
        (f"Dòng {row_number}: {label} phải là ngày hợp lệ.",),
    )


def _safe_text(value: Any, label: str, row_number: int) -> str:
    text = _normalise_text(value)
    if text.startswith(("=", "+", "@")):
        raise OCWorkbookError(
            "OC_FILE_VALIDATION_FAILED",
            "Workbook Upload OC có dữ liệu chưa hợp lệ.",
            (f"Dòng {row_number}: {label} không được bắt đầu bằng công thức.",),
        )
    if text.startswith("#") and text.upper() in {
        "#REF!",
        "#VALUE!",
        "#N/A",
        "#NAME?",
        "#DIV/0!",
        "#NUM!",
        "#NULL!",
    }:
        raise OCWorkbookError(
            "OC_FILE_FORMULA_ERROR",
            "Workbook còn lỗi công thức Excel.",
            (f"Dòng {row_number}: {label} đang là {text}.",),
        )
    return text


def _is_zero_quantity(value: Any) -> bool:
    if isinstance(value, bool) or value in (None, ""):
        return False
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError, ValueError):
        return False
    return number.is_finite() and number == 0


def _canonical_option(
    value: Any,
    label: str,
    row_number: int,
    options: tuple[str, ...],
    *,
    default: str = "",
) -> str:
    text = _safe_text(value, label, row_number)
    if not text and default:
        return default
    matches = {option.casefold(): option for option in options}
    canonical = matches.get(text.casefold())
    if canonical:
        return canonical
    raise OCWorkbookError(
        "OC_FILE_VALIDATION_FAILED",
        "Workbook Upload OC có dữ liệu chưa hợp lệ.",
        (f"Dòng {row_number}: {label} '{text or '[trống]'} không có trong danh sách.",),
    )


def _validate_delivery_dates(
    buyer_order_date: date,
    raw_material_eta: date,
    buyer_delivery_date: date,
    row_number: int,
    *,
    oc_delivery_date: date | None = None,
) -> list[str]:
    errors: list[str] = []
    if not buyer_order_date < raw_material_eta < buyer_delivery_date:
        errors.append(
            f"Dòng {row_number}: ngày phải theo Buyer Order Date "
            "< Raw Material ETA < Buyer Delivery Date."
        )
    if oc_delivery_date is not None and buyer_delivery_date != oc_delivery_date:
        errors.append(
            f"Dòng {row_number}: Buyer Delivery Date phải bằng OC Delivery Date."
        )
    return errors


def _ensure_headers(actual: list[Any], expected: tuple[str, ...], label: str) -> None:
    actual_normalised = tuple(_normalise_header(item) for item in actual)
    expected_normalised = tuple(_normalise_header(item) for item in expected)
    if actual_normalised == expected_normalised:
        return
    errors: list[str] = []
    for index, expected_header in enumerate(expected):
        actual_header = _normalise_text(actual[index]) if index < len(actual) else ""
        if _normalise_header(actual_header) != _normalise_header(expected_header):
            errors.append(
                f"Cột {index + 1}: cần '{expected_header}', đang là "
                f"'{actual_header or '[trống]'}'."
            )
        if len(errors) >= 8:
            break
    raise OCWorkbookError(
        "OC_FILE_HEADERS_INVALID",
        f"Header {label} không đúng mẫu Upload OC.",
        errors,
    )


def _lookup_lists(workbook: Any) -> tuple[set[str], set[str], dict[str, tuple[str, str]]]:
    if "THONG TIN" not in workbook.sheetnames:
        raise OCWorkbookError(
            "OC_TEMPLATE_SHEET_MISSING",
            "Thiếu sheet THONG TIN trong UPLOAD FORM.",
        )
    sheet = workbook["THONG TIN"]
    factories = {
        _normalise_text(sheet.cell(row, 1).value).casefold()
        for row in range(2, sheet.max_row + 1)
        if _normalise_text(sheet.cell(row, 1).value)
    }
    buyers: set[str] = set()
    for row in range(2, sheet.max_row + 1):
        value = _normalise_text(sheet.cell(row, 2).value)
        if not value:
            continue
        if value.casefold() == "po type":
            break
        buyers.add(value.casefold())
    countries: dict[str, tuple[str, str]] = {}
    for row in range(2, sheet.max_row + 1):
        destination = _normalise_text(sheet.cell(row, 5).value)
        market = _normalise_text(sheet.cell(row, 6).value)
        key = destination.casefold()
        if key and market:
            countries[key] = (destination, market)
    return factories, buyers, countries


def _add_list_validation(
    sheet: Any,
    reference_sheet: Any,
    header: str,
    reference_column: int,
    option_count: int,
) -> None:
    input_header = {
        "Country": "Country of Final Destination",
    }.get(header, header)
    column = INPUT_HEADERS.index(input_header) + 1
    reference_letter = reference_sheet.cell(1, reference_column).column_letter
    validation = DataValidation(
        type="list",
        formula1=(
            f"'{REFERENCE_SHEET_NAME}'!${reference_letter}$2:"
            f"${reference_letter}${option_count + 1}"
        ),
        allow_blank=header == "PO Type (Zone)",
    )
    validation.error = f"Hãy chọn {input_header} từ danh sách."
    validation.errorTitle = "Giá trị không hợp lệ"
    validation.prompt = f"Chọn {input_header}."
    validation.promptTitle = "Upload OC"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    if header in {"Buyer", "Factory"}:
        validation.errorStyle = "warning"
        validation.error = (
            f"{input_header} chưa có trong danh sách gợi ý của form. "
            "Chỉ tiếp tục nếu giá trị này đang tồn tại trên WFX."
        )
    sheet.add_data_validation(validation)
    validation.add(f"{sheet.cell(2, column).column_letter}2:{sheet.cell(2, column).column_letter}10001")


def write_oc_input_template(path: str | Path) -> Path:
    """Create the simplified one-header workbook shown to end users."""
    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = INPUT_SHEET_NAME
    sheet.append(list(INPUT_HEADERS))
    required_fill = PatternFill("solid", fgColor="FFFF00")
    optional_fill = PatternFill("solid", fgColor="F4B183")
    optional_headers = {"PO Type (Zone)", "Extra Production %", "Buyer Lot No."}
    for cell in sheet[1]:
        cell.fill = optional_fill if cell.value in optional_headers else required_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        comment = INPUT_COMMENTS.get(str(cell.value))
        if comment:
            cell.comment = Comment(comment, "WFX Smart")
    sheet.row_dimensions[1].height = 48
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:X10001"
    widths = {
        "Buyer": 24,
        "Season": 12,
        "Order Type": 14,
        "Currency": 10,
        "Factory": 34,
        "Ship Under PO Ref": 28,
        "Article Code": 16,
        "Buyer Style Ref": 18,
        "Buyer PO Num": 28,
        "Summary Buyer Order Ref": 28,
        "Buyer Order Date": 17,
        "Buyer Delivery Date": 18,
        "Raw Material ETA Date": 20,
        "Payment Terms": 42,
        "Country of Final Destination": 25,
        "Color Code": 14,
        "Color Name": 20,
        "Size Code": 12,
        "Selling Price": 14,
        "Units": 11,
        "Internal Lot No.": 16,
        "PO Type (Zone)": 16,
        "Extra Production %": 18,
        "Buyer Lot No.": 18,
    }
    for index, header in enumerate(INPUT_HEADERS, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = widths[header]
    for row_number in range(2, 202):
        for column in (11, 12, 13):
            sheet.cell(row_number, column).number_format = "dd-mm-yyyy"
        sheet.cell(row_number, 19).number_format = "0.00"
        sheet.cell(row_number, 20).number_format = "0"
        sheet.cell(row_number, 23).number_format = "0.00"

    references = workbook.create_sheet(REFERENCE_SHEET_NAME)
    reference_lists = (
        ("Buyer", BUYER_OPTIONS),
        ("Factory", FACTORY_OPTIONS),
        ("Order Type", ORDER_TYPE_OPTIONS),
        ("Currency", ("USD", "EUR", "GBP")),
        ("Country", tuple(COUNTRY_MARKET)),
        ("PO Type (Zone)", PO_TYPE_OPTIONS),
        ("Payment Terms", PAYMENT_TERM_OPTIONS),
    )
    for column, (heading, options) in enumerate(reference_lists, start=1):
        references.cell(1, column, heading)
        for row, option in enumerate(options, start=2):
            references.cell(row, column, option)
        _add_list_validation(sheet, references, heading, column, len(options))
    references.sheet_state = "veryHidden"
    workbook.save(target)
    return target


def _nonempty_rows(sheet: Any, start_row: int, width: int) -> list[tuple[int, list[Any]]]:
    rows: list[tuple[int, list[Any]]] = []
    for row_number in range(start_row, min(sheet.max_row, MAX_OC_ROWS + start_row) + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, width + 1)]
        if any(value not in (None, "") for value in values):
            rows.append((row_number, values))
    if sheet.max_row > MAX_OC_ROWS + start_row:
        for row_number in range(MAX_OC_ROWS + start_row + 1, sheet.max_row + 1):
            if any(
                sheet.cell(row_number, column).value not in (None, "")
                for column in range(1, width + 1)
            ):
                raise OCWorkbookError(
                    "OC_FILE_TOO_MANY_ROWS",
                    f"Upload OC hỗ trợ tối đa {MAX_OC_ROWS:,} dòng dữ liệu.",
                )
    return rows


def _simple_new_rows(workbook: Any) -> tuple[str, list[list[Any]], tuple[str, ...]]:
    sheet = workbook[INPUT_SHEET_NAME]
    _ensure_headers(
        [sheet.cell(1, column).value for column in range(1, len(INPUT_HEADERS) + 1)],
        INPUT_HEADERS,
        INPUT_SHEET_NAME,
    )
    source_rows = _nonempty_rows(sheet, 2, len(INPUT_HEADERS))
    if not source_rows:
        raise OCWorkbookError(
            "OC_FILE_EMPTY",
            f"{INPUT_SHEET_NAME} chưa có dòng đơn hàng nào từ dòng 2.",
        )
    errors: list[str] = []
    warnings: list[str] = []
    prepared: list[dict[str, Any]] = []
    buyers: dict[str, str] = {}
    seen_keys: set[tuple[str, ...]] = set()
    known_buyers = {item.casefold() for item in BUYER_OPTIONS}
    known_factories = {item.casefold() for item in FACTORY_OPTIONS}
    skipped_zero_units = 0
    countries = {
        country.casefold(): (country, market)
        for country, market in COUNTRY_MARKET.items()
    }
    for row_number, values in source_rows:
        if _is_zero_quantity(values[19]):
            skipped_zero_units += 1
            continue
        missing = [
            INPUT_HEADERS[index]
            for index in range(len(INPUT_HEADERS))
            if index not in SIMPLE_NEW_OPTIONAL_COLUMNS
            if values[index] in (None, "")
        ]
        if missing:
            errors.append(
                f"Dòng {row_number}: thiếu {', '.join(missing[:5])}"
                + ("…" if len(missing) > 5 else ".")
            )
            continue
        try:
            text = [
                _safe_text(value, INPUT_HEADERS[index], row_number)
                for index, value in enumerate(values)
            ]
            buyer = text[0]
            factory = text[4]
            order_type = _canonical_option(
                values[2], "Order Type", row_number, ORDER_TYPE_OPTIONS
            )
            payment_terms = _canonical_option(
                values[13], "Payment Terms", row_number, PAYMENT_TERM_OPTIONS
            )
            zone = _canonical_option(
                values[21],
                "PO Type (Zone)",
                row_number,
                PO_TYPE_OPTIONS,
                default="FOB",
            )
            buyers.setdefault(buyer.casefold(), buyer)
            if buyer.casefold() not in known_buyers:
                warning = (
                    f"Buyer '{buyer}' chưa có trong danh sách gợi ý; "
                    "app sẽ yêu cầu khớp chính xác trên WFX."
                )
                if warning not in warnings:
                    warnings.append(warning)
            if factory.casefold() not in known_factories:
                warning = (
                    f"Factory '{factory}' chưa có trong danh sách gợi ý; "
                    "WFX sẽ kiểm tra khi Process Package."
                )
                if warning not in warnings:
                    warnings.append(warning)
            country_key = text[14].casefold()
            if country_key not in countries:
                errors.append(
                    f"Dòng {row_number}: Country '{text[14]}' chưa có mapping Market."
                )
                destination, market = text[14], ""
            else:
                destination, market = countries[country_key]
            buyer_order_date = _date_value(values[10], INPUT_HEADERS[10], row_number)
            delivery_date = _date_value(values[11], INPUT_HEADERS[11], row_number)
            raw_material_eta = _date_value(values[12], INPUT_HEADERS[12], row_number)
            price = _decimal(values[18], INPUT_HEADERS[18], row_number)
            units = _decimal(values[19], INPUT_HEADERS[19], row_number)
            extra = (
                Decimal(0)
                if values[22] in (None, "")
                else _decimal(values[22], INPUT_HEADERS[22], row_number)
            )
            errors.extend(
                _validate_delivery_dates(
                    buyer_order_date,
                    raw_material_eta,
                    delivery_date,
                    row_number,
                )
            )
            if price <= 0:
                errors.append(f"Dòng {row_number}: Selling Price phải lớn hơn 0.")
            if units <= 0 or units != units.to_integral_value():
                errors.append(f"Dòng {row_number}: Units phải là số nguyên lớn hơn 0.")
            if extra < 0:
                errors.append(f"Dòng {row_number}: Extra Production % không được âm.")
            duplicate_key = tuple(
                item.casefold()
                for item in (text[5], text[7], text[8], text[15], text[17])
            )
            if duplicate_key in seen_keys:
                errors.append(
                    f"Dòng {row_number}: trùng PO/Style/Color/Size với dòng trước."
                )
            seen_keys.add(duplicate_key)
            prepared.append(
                {
                    "buyer": buyer,
                    "season": text[1],
                    "order_type": order_type,
                    "currency": text[3],
                    "factory": factory,
                    "ship_ref": text[5],
                    "article": text[6],
                    "buyer_style": text[7],
                    "buyer_po": text[8],
                    "summary_ref": text[9],
                    "buyer_order_date": buyer_order_date,
                    "delivery_date": delivery_date,
                    "raw_material_eta": raw_material_eta,
                    "payment_terms": payment_terms,
                    "destination": destination,
                    "market": market,
                    "color": f"{text[15]}^{text[16]}",
                    "size": text[17],
                    "price": price,
                    "units": units,
                    "internal_lot": text[20],
                    "zone": zone,
                    "extra": extra,
                    "buyer_lot": text[23],
                }
            )
        except OCWorkbookError as error:
            errors.extend(error.errors or (f"Dòng {row_number}: {error.message}",))
    if len(buyers) > 1:
        errors.append(
            "File có nhiều Buyer: " + ", ".join(sorted(buyers.values()))
            + ". Mỗi lần EDI chỉ upload một Buyer."
        )
    if errors:
        raise OCWorkbookError(
            "OC_FILE_VALIDATION_FAILED",
            f"Workbook có {len(errors)} lỗi cần sửa trước khi upload.",
            errors[:100],
        )
    if not prepared:
        raise OCWorkbookError(
            "OC_FILE_EMPTY",
            "Không còn dòng Upload OC nào sau khi bỏ các dòng có Units = 0.",
        )
    if skipped_zero_units:
        warnings.append(
            f"App đã bỏ qua {skipped_zero_units} dòng có Units = 0."
        )

    totals: defaultdict[tuple[str, str, str], Decimal] = defaultdict(Decimal)
    for item in prepared:
        totals[
            (
                item["ship_ref"].casefold(),
                item["summary_ref"].casefold(),
                item["buyer_style"].casefold(),
            )
        ] += item["units"]
    output_rows: list[list[Any]] = []
    for item in prepared:
        key = (
            item["ship_ref"].casefold(),
            item["summary_ref"].casefold(),
            item["buyer_style"].casefold(),
        )
        row = [None] * len(EDI_HEADERS)
        output_values = {
            "Factory": item["factory"],
            "Ship Under PO Ref": item["ship_ref"],
            "Article": item["article"],
            "Buyer": item["buyer"],
            "Currency": item["currency"],
            "Season": item["season"],
            "Country of Origin": "Vietnam",
            "Payment Terms": item["payment_terms"],
            "Buyer PO Num": item["buyer_po"],
            "Summary Buyer Order Ref": item["summary_ref"],
            "Market Buyer Order Ref": item["summary_ref"],
            "Destination Buyer Order Ref": item["summary_ref"],
            "Delivery Buyer Order Ref": item["summary_ref"],
            "Buyer Order Date": item["buyer_order_date"],
            "Order Type": item["order_type"],
            "Mode of Shipment": "AIR/SEA",
            "Buyer Delivery Date": item["delivery_date"],
            "OC Delivery Date": item["delivery_date"],
            "Raw Matetrial ETA": item["raw_material_eta"],
            "Country of Final Destination": item["destination"],
            "Final Destination": item["destination"],
            "Market": item["market"],
            "Buyer Style Ref.": item["buyer_style"],
            "Color": item["color"],
            "Size": item["size"],
            "Total Qty": totals[key],
            "Price": item["price"],
            "Units": item["units"],
            "Zone": item["zone"],
            "Internal Lot No.": item["internal_lot"],
            "Buyer Lot No.": item["buyer_lot"] or None,
            "Fulfillment Type": "Back Order",
            "FirstBuyerDeliveryDate": item["delivery_date"],
            "Extra Production %": item["extra"],
        }
        for header, value in output_values.items():
            row[EDI_HEADERS.index(header)] = value
        output_rows.append(row)
    return next(iter(buyers.values())), output_rows, tuple(warnings)


def _new_rows(workbook: Any) -> tuple[str, list[list[Any]], tuple[str, ...]]:
    if INPUT_SHEET_NAME in workbook.sheetnames:
        return _simple_new_rows(workbook)
    if "FORM" not in workbook.sheetnames:
        raise OCWorkbookError(
            "OC_TEMPLATE_SHEET_MISSING",
            f"Thiếu sheet {INPUT_SHEET_NAME} hoặc FORM trong workbook Upload OC.",
        )
    sheet = workbook["FORM"]
    _ensure_headers(
        [sheet.cell(5, column).value for column in range(1, 21)],
        FORM_HEADERS,
        "FORM",
    )
    buyer = _safe_text(sheet["B1"].value, "Buyer", 1)
    season = _safe_text(sheet["B2"].value, "Season", 2)
    order_type = _safe_text(sheet["B3"].value, "Order Type", 3)
    currency = _safe_text(sheet["B4"].value, "Currency", 4)
    metadata_errors = [
        f"{label} (ô {cell}) không được để trống."
        for label, cell, value in (
            ("Buyer", "B1", buyer),
            ("Season", "B2", season),
            ("Order Type", "B3", order_type),
            ("Currency", "B4", currency),
        )
        if not value
    ]
    if metadata_errors:
        raise OCWorkbookError(
            "OC_FILE_VALIDATION_FAILED",
            "Thông tin chung trong FORM chưa đầy đủ.",
            metadata_errors,
        )
    order_type = _canonical_option(
        order_type, "Order Type", 3, ORDER_TYPE_OPTIONS
    )
    factories, buyers, countries = _lookup_lists(workbook)
    if buyer.casefold() not in buyers:
        metadata_errors.append(f"Buyer '{buyer}' không có trong THONG TIN.")

    source_rows = _nonempty_rows(sheet, 6, 20)
    if not source_rows:
        raise OCWorkbookError(
            "OC_FILE_EMPTY",
            "FORM chưa có dòng đơn hàng nào từ dòng 6.",
        )

    errors = list(metadata_errors)
    prepared: list[dict[str, Any]] = []
    seen_keys: set[tuple[str, ...]] = set()
    skipped_zero_units = 0
    for row_number, values in source_rows:
        if _is_zero_quantity(values[15]):
            skipped_zero_units += 1
            continue
        missing = [
            FORM_HEADERS[index]
            for index in NEW_REQUIRED_FORM_COLUMNS
            if values[index] in (None, "")
        ]
        if missing:
            errors.append(
                f"Dòng {row_number}: thiếu {', '.join(missing[:5])}"
                + ("…" if len(missing) > 5 else ".")
            )
            continue
        try:
            text = [
                _safe_text(value, FORM_HEADERS[index], row_number)
                for index, value in enumerate(values)
            ]
            factory = text[0]
            country_key = text[10].casefold()
            if factory.casefold() not in factories:
                errors.append(
                    f"Dòng {row_number}: Factory '{factory}' không có trong THONG TIN."
                )
            if country_key not in countries:
                errors.append(
                    f"Dòng {row_number}: Country '{text[10]}' không có mapping Market."
                )
                destination, market = text[10], ""
            else:
                destination, market = countries[country_key]
            buyer_order_date = _date_value(values[6], FORM_HEADERS[6], row_number)
            delivery_date = _date_value(values[7], FORM_HEADERS[7], row_number)
            raw_material_eta = _date_value(values[8], FORM_HEADERS[8], row_number)
            price = _decimal(values[14], FORM_HEADERS[14], row_number)
            units = _decimal(values[15], FORM_HEADERS[15], row_number)
            extra = (
                Decimal(0)
                if values[18] in (None, "")
                else _decimal(values[18], FORM_HEADERS[18], row_number)
            )
            payment_terms = _canonical_option(
                values[9], "Payment Terms", row_number, PAYMENT_TERM_OPTIONS
            )
            zone = _canonical_option(
                values[17],
                "PO Type",
                row_number,
                PO_TYPE_OPTIONS,
                default="FOB",
            )
            errors.extend(
                _validate_delivery_dates(
                    buyer_order_date,
                    raw_material_eta,
                    delivery_date,
                    row_number,
                )
            )
            if price <= 0:
                errors.append(f"Dòng {row_number}: Selling Price phải lớn hơn 0.")
            if units <= 0 or units != units.to_integral_value():
                errors.append(f"Dòng {row_number}: Units phải là số nguyên lớn hơn 0.")
            if extra < 0:
                errors.append(f"Dòng {row_number}: Extra Production không được âm.")
            duplicate_key = tuple(
                item.casefold()
                for item in (text[1], text[3], text[4], text[11], text[13])
            )
            if duplicate_key in seen_keys:
                errors.append(
                    f"Dòng {row_number}: trùng PO/Style/Color/Size với dòng trước."
                )
            seen_keys.add(duplicate_key)
            prepared.append(
                {
                    "source_row": row_number,
                    "factory": factory,
                    "ship_ref": text[1],
                    "article": text[2],
                    "buyer_style": text[3],
                    "buyer_po": text[4],
                    "summary_ref": text[5],
                    "buyer_order_date": buyer_order_date,
                    "delivery_date": delivery_date,
                    "raw_material_eta": raw_material_eta,
                    "payment_terms": payment_terms,
                    "destination": destination,
                    "market": market,
                    "color": f"{text[11]}^{text[12]}",
                    "size": text[13],
                    "price": price,
                    "units": units,
                    "internal_lot": text[16],
                    "zone": zone,
                    "extra": extra,
                    "buyer_lot": text[19],
                }
            )
        except OCWorkbookError as error:
            errors.extend(error.errors or (f"Dòng {row_number}: {error.message}",))

    if errors:
        raise OCWorkbookError(
            "OC_FILE_VALIDATION_FAILED",
            f"Workbook có {len(errors)} lỗi cần sửa trước khi upload.",
            errors[:100],
        )
    if not prepared:
        raise OCWorkbookError(
            "OC_FILE_EMPTY",
            "Không còn dòng Upload OC nào sau khi bỏ các dòng có Units = 0.",
        )

    totals: defaultdict[tuple[str, str, str], Decimal] = defaultdict(Decimal)
    for item in prepared:
        totals[
            (
                item["ship_ref"].casefold(),
                item["summary_ref"].casefold(),
                item["buyer_style"].casefold(),
            )
        ] += item["units"]

    output_rows: list[list[Any]] = []
    for item in prepared:
        key = (
            item["ship_ref"].casefold(),
            item["summary_ref"].casefold(),
            item["buyer_style"].casefold(),
        )
        row = [None] * len(EDI_HEADERS)
        values = {
            "Factory": item["factory"],
            "Ship Under PO Ref": item["ship_ref"],
            "Article": item["article"],
            "Buyer": buyer,
            "Currency": currency,
            "Season": season,
            "Country of Origin": "Vietnam",
            "Payment Terms": item["payment_terms"],
            "Buyer PO Num": item["buyer_po"],
            "Summary Buyer Order Ref": item["summary_ref"],
            "Market Buyer Order Ref": item["summary_ref"],
            "Destination Buyer Order Ref": item["summary_ref"],
            "Delivery Buyer Order Ref": item["summary_ref"],
            "Buyer Order Date": item["buyer_order_date"],
            "Order Type": order_type,
            "Mode of Shipment": "AIR/SEA",
            "Buyer Delivery Date": item["delivery_date"],
            "OC Delivery Date": item["delivery_date"],
            "Raw Matetrial ETA": item["raw_material_eta"],
            "Country of Final Destination": item["destination"],
            "Final Destination": item["destination"],
            "Market": item["market"],
            "Buyer Style Ref.": item["buyer_style"],
            "Color": item["color"],
            "Size": item["size"],
            "Total Qty": totals[key],
            "Price": item["price"],
            "Units": item["units"],
            "Zone": item["zone"],
            "Internal Lot No.": item["internal_lot"],
            "Buyer Lot No.": item["buyer_lot"] or None,
            "Fulfillment Type": "Back Order",
            "FirstBuyerDeliveryDate": item["delivery_date"],
            "Extra Production %": item["extra"],
        }
        for header, value in values.items():
            row[EDI_HEADERS.index(header)] = value
        output_rows.append(row)
    warnings = (
        (f"App đã bỏ qua {skipped_zero_units} dòng có Units = 0.",)
        if skipped_zero_units
        else ()
    )
    return buyer, output_rows, warnings


def _revise_rows(workbook: Any) -> tuple[str, list[list[Any]], tuple[str, ...]]:
    if "Sheet1" not in workbook.sheetnames:
        raise OCWorkbookError(
            "OC_TEMPLATE_SHEET_MISSING",
            "File Revise OC phải có sheet Sheet1.",
        )
    sheet = workbook["Sheet1"]
    _ensure_headers(
        [sheet.cell(1, column).value for column in range(1, len(EDI_HEADERS) + 1)],
        EDI_HEADERS,
        "Sheet1",
    )
    source_rows = _nonempty_rows(sheet, 2, len(EDI_HEADERS))
    if not source_rows:
        raise OCWorkbookError("OC_FILE_EMPTY", "Sheet1 chưa có dữ liệu Revise OC.")

    indexes = {header: index for index, header in enumerate(EDI_HEADERS)}
    errors: list[str] = []
    buyers: dict[str, str] = {}
    prepared: list[list[Any]] = []
    seen_keys: set[tuple[str, ...]] = set()
    skipped_zero_units = 0
    for row_number, values in source_rows:
        if _is_zero_quantity(values[indexes["Units"]]):
            skipped_zero_units += 1
            continue
        text_values: dict[str, str] = {}
        try:
            for header in EDI_HEADERS:
                value = values[indexes[header]]
                if header not in DATE_HEADERS and not isinstance(value, (int, float, Decimal)):
                    text_values[header] = _safe_text(value, header, row_number)
            missing = [
                header
                for header in REVISE_REQUIRED_HEADERS
                if values[indexes[header]] in (None, "")
            ]
            if missing:
                errors.append(
                    f"Dòng {row_number}: thiếu {', '.join(sorted(missing)[:5])}"
                    + ("…" if len(missing) > 5 else ".")
                )
                continue
            buyer = _safe_text(values[indexes["Buyer"]], "Buyer", row_number)
            buyers.setdefault(buyer.casefold(), buyer)
            for header in DATE_HEADERS:
                value = values[indexes[header]]
                if value not in (None, ""):
                    values[indexes[header]] = _date_value(value, header, row_number)
            values[indexes["Order Type"]] = _canonical_option(
                values[indexes["Order Type"]],
                "Order Type",
                row_number,
                ORDER_TYPE_OPTIONS,
            )
            values[indexes["Payment Terms"]] = _canonical_option(
                values[indexes["Payment Terms"]],
                "Payment Terms",
                row_number,
                PAYMENT_TERM_OPTIONS,
            )
            values[indexes["Zone"]] = _canonical_option(
                values[indexes["Zone"]],
                "Zone",
                row_number,
                PO_TYPE_OPTIONS,
                default="FOB",
            )
            errors.extend(
                _validate_delivery_dates(
                    values[indexes["Buyer Order Date"]],
                    values[indexes["Raw Matetrial ETA"]],
                    values[indexes["Buyer Delivery Date"]],
                    row_number,
                    oc_delivery_date=values[indexes["OC Delivery Date"]],
                )
            )
            units = _decimal(values[indexes["Units"]], "Units", row_number)
            price = _decimal(values[indexes["Price"]], "Price", row_number)
            extra = (
                Decimal(0)
                if values[indexes["Extra Production %"]] in (None, "")
                else _decimal(
                    values[indexes["Extra Production %"]],
                    "Extra Production %",
                    row_number,
                )
            )
            if units <= 0 or units != units.to_integral_value():
                errors.append(f"Dòng {row_number}: Units phải là số nguyên lớn hơn 0.")
            if price <= 0:
                errors.append(f"Dòng {row_number}: Price phải lớn hơn 0.")
            if extra < 0:
                errors.append(f"Dòng {row_number}: Extra Production % không được âm.")
            values[indexes["Units"]] = units
            values[indexes["Price"]] = price
            values[indexes["Extra Production %"]] = extra
            duplicate_key = tuple(
                _normalise_text(values[indexes[header]]).casefold()
                for header in (
                    "Ship Under PO Ref",
                    "Delivery Buyer Order Ref",
                    "Buyer Style Ref.",
                    "Color",
                    "Size",
                )
            )
            if duplicate_key in seen_keys:
                errors.append(f"Dòng {row_number}: trùng Delivery/Style/Color/Size.")
            seen_keys.add(duplicate_key)
            prepared.append(values)
        except OCWorkbookError as error:
            errors.extend(error.errors or (f"Dòng {row_number}: {error.message}",))

    if len(buyers) > 1:
        errors.append(
            "File có nhiều Buyer: " + ", ".join(sorted(buyers.values()))
            + ". EDI chỉ cho chọn một Buyer mỗi lần upload."
        )
    if errors:
        raise OCWorkbookError(
            "OC_FILE_VALIDATION_FAILED",
            f"Workbook Revise OC có {len(errors)} lỗi cần sửa.",
            errors[:100],
        )
    if not prepared:
        raise OCWorkbookError(
            "OC_FILE_EMPTY",
            "Không còn dòng Revise OC nào sau khi bỏ các dòng có Units = 0.",
        )

    totals: defaultdict[tuple[str, str, str], Decimal] = defaultdict(Decimal)
    for row in prepared:
        key = tuple(
            _normalise_text(row[indexes[header]]).casefold()
            for header in (
                "Ship Under PO Ref",
                "Delivery Buyer Order Ref",
                "Buyer Style Ref.",
            )
        )
        totals[key] += row[indexes["Units"]]
    corrected = 0
    for row in prepared:
        key = tuple(
            _normalise_text(row[indexes[header]]).casefold()
            for header in (
                "Ship Under PO Ref",
                "Delivery Buyer Order Ref",
                "Buyer Style Ref.",
            )
        )
        expected = totals[key]
        current = row[indexes["Total Qty"]]
        try:
            current_number = Decimal(str(current).replace(",", ""))
        except (InvalidOperation, AttributeError, ValueError):
            current_number = Decimal("NaN")
        if current_number != expected:
            corrected += 1
        row[indexes["Total Qty"]] = expected
    warnings: list[str] = []
    if skipped_zero_units:
        warnings.append(
            f"App đã bỏ qua {skipped_zero_units} dòng có Units = 0."
        )
    if corrected:
        warnings.append(
            f"App đã tính lại Total Qty cho {corrected} dòng từ cột Units."
        )
    buyer = next(iter(buyers.values()))
    return buyer, prepared, tuple(warnings)


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value


def _write_static_workbook(rows: list[list[Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(list(EDI_HEADERS))
    header_fill = PatternFill("solid", fgColor="FFFF00")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for source in rows:
        sheet.append([_excel_value(value) for value in source])
    date_indexes = [EDI_HEADERS.index(header) + 1 for header in DATE_HEADERS]
    for column in date_indexes:
        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row_number, column).number_format = "dd-mm-yyyy"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:AY{sheet.max_row}"
    sheet.row_dimensions[1].height = 42
    workbook.save(output_path)


def _verify_static_output(path: Path, expected_rows: int) -> None:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise OCWorkbookError(
                "OC_OUTPUT_INVALID",
                "File EDI sinh ra phải chỉ có một sheet Sheet1.",
            )
        sheet = workbook["Sheet1"]
        _ensure_headers(
            [sheet.cell(1, column).value for column in range(1, len(EDI_HEADERS) + 1)],
            EDI_HEADERS,
            "Sheet1",
        )
        if sheet.max_row - 1 != expected_rows:
            raise OCWorkbookError(
                "OC_OUTPUT_INVALID",
                "Số dòng trong file EDI sinh ra không khớp dữ liệu nguồn.",
            )
        for row in sheet.iter_rows(min_row=2, max_col=len(EDI_HEADERS)):
            for cell in row:
                if cell.data_type == "f":
                    raise OCWorkbookError(
                        "OC_OUTPUT_INVALID",
                        "File EDI sinh ra vẫn còn công thức Excel.",
                    )
    finally:
        workbook.close()


def prepare_oc_workbook(
    input_path: str | Path,
    mode: str,
    output_path: str | Path,
) -> PreparedOCUpload:
    """Validate source then create a value-only 51-column EDI workbook."""
    source = Path(input_path).expanduser().resolve()
    target = Path(output_path).expanduser().resolve()
    selected_mode = str(mode or "").strip().casefold()
    if selected_mode not in {"new", "revise"}:
        raise OCWorkbookError(
            "OC_MODE_INVALID",
            "Chế độ Upload OC phải là New hoặc Revise.",
        )
    _validate_xlsx_archive(source)
    _ensure_input_values_only(source, selected_mode)
    try:
        workbook = load_workbook(source, data_only=True, read_only=False)
    except (OSError, ValueError, BadZipFile) as error:
        raise OCWorkbookError(
            "OC_FILE_INVALID",
            f"Không đọc được workbook: {type(error).__name__}: {error}",
        ) from error
    try:
        if selected_mode == "new":
            buyer, rows, warnings = _new_rows(workbook)
        else:
            buyer, rows, warnings = _revise_rows(workbook)
    finally:
        workbook.close()
    _write_static_workbook(rows, target)
    _verify_static_output(target, len(rows))
    indexes = {header: index for index, header in enumerate(EDI_HEADERS)}
    seasons = tuple(
        sorted(
            {
                _normalise_text(row[indexes["Season"]])
                for row in rows
                if _normalise_text(row[indexes["Season"]])
            },
            key=str.casefold,
        )
    )
    po_refs = {
        _normalise_text(row[indexes["Summary Buyer Order Ref"]]).casefold()
        for row in rows
        if _normalise_text(row[indexes["Summary Buyer Order Ref"]])
    }
    articles = {
        _normalise_text(row[indexes["Article"]]).casefold()
        for row in rows
        if _normalise_text(row[indexes["Article"]])
    }
    total_units = sum(
        (Decimal(str(row[indexes["Units"]])) for row in rows),
        Decimal(0),
    )
    return PreparedOCUpload(
        mode=selected_mode,
        buyer=buyer,
        row_count=len(rows),
        upload_path=target,
        seasons=seasons,
        po_count=len(po_refs),
        style_count=len(articles),
        total_units=_excel_value(total_units),
        warnings=warnings,
    )
